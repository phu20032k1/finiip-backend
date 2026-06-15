"""Finiip V25-V40: productization pack for the self-made AI accounting prototype.

This module intentionally uses a small JSON-backed store so it can run inside the
existing prototype without requiring a destructive database migration. The APIs
are designed as safe MVP endpoints: AI creates drafts and recommendations, while
posting, approval, permissions and backup remain explicit user actions.
"""
from __future__ import annotations

from datetime import datetime
import csv
import hashlib
import json
import os
import re
import secrets
import unicodedata
import shutil
import zipfile
from collections import defaultdict
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, HTTPException, UploadFile, File, Form
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel, Field

try:  # openpyxl is already a project dependency in this prototype
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
except Exception:  # pragma: no cover
    Workbook = None
    Font = Alignment = None

router = APIRouter(tags=["Finiip V25-V40"])
BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
FRONTEND_DIR = BASE_DIR / "frontend"
STORE_PATH = DATA_DIR / "v25_v40_store.json"
BACKUP_DIR = DATA_DIR / "backups"
INTENTS_PATH = BASE_DIR / "ai_intents.json"
ACCOUNTING_RULES_PATH = BASE_DIR / "accounting_rules.json"
KNOWLEDGE_BASE_DIR = BASE_DIR / "knowledge_base"


def _load_dotenv_local(path: Path = BASE_DIR / ".env") -> None:
    """Tiny .env loader so local VS Code runs can read Supabase/OpenAI keys without extra deps."""
    try:
        if not path.exists():
            return
        for raw_line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
            line = raw_line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            if key and key not in os.environ:
                os.environ[key] = value
    except Exception:
        # Never block app startup because of .env parsing.
        pass

_load_dotenv_local()

ROLE_PERMISSIONS = {
    "admin": ["*"] ,
    "chief_accountant": ["journal.approve", "journal.post", "reports.read", "backup.manage", "users.read"],
    "accountant": ["journal.create", "journal.review", "reports.read"],
    "viewer": ["reports.read"],
}
REVENUE_PREFIXES = ("511", "515", "711")
EXPENSE_PREFIXES = ("632", "635", "641", "642", "811")
VAT_INPUT_CODES = ("133", "1331")
VAT_OUTPUT_CODES = ("3331",)


class JournalDraftCreate(BaseModel):
    description: str = Field(..., min_length=2)
    amount: float = Field(..., ge=0)
    vat_rate: float = Field(0.1, ge=0, le=1)
    payment_method: str = "cash"
    category: Optional[str] = None
    debit_account: Optional[str] = None
    credit_account: Optional[str] = None
    source: str = "manual"


class JournalDraftDecision(BaseModel):
    note: Optional[str] = None


class PostEntryRequest(BaseModel):
    draft_id: Optional[str] = None
    description: Optional[str] = None
    lines: Optional[List[Dict[str, Any]]] = None
    note: Optional[str] = None


class UserCreate(BaseModel):
    name: str
    email: str
    role: str = "accountant"
    password: str = "finiip123"


class LoginRequest(BaseModel):
    email: str
    password: str


class OCRTextRequest(BaseModel):
    text: str


class BulkTransactionItem(BaseModel):
    date: Optional[str] = None
    description: str
    amount: float = Field(..., ge=0)
    transaction_type: Optional[str] = None
    category: Optional[str] = None
    vat_rate: float = 0.1
    payment_method: str = "cash"
    source: str = "bulk_manual"


class BulkImportRequest(BaseModel):
    items: List[BulkTransactionItem]
    auto_create_drafts: bool = True
    auto_approve_safe: bool = False
    source: str = "manual_bulk"


class ChatRequest(BaseModel):
    message: str
    period: Optional[str] = None
    confirm_token: Optional[str] = None


class ChatActionConfirmRequest(BaseModel):
    confirmation_id: str
    confirm: bool = True


class ConfirmJournalRequest(BaseModel):
    description: str = Field(..., min_length=2)
    amount: Optional[float] = Field(None, ge=0)
    category: Optional[str] = None
    debit_account: Optional[str] = None
    credit_account: Optional[str] = None
    payment_method: str = "bank"
    vat_rate: float = Field(0.1, ge=0, le=1)
    risk_note: Optional[str] = None
    user_correction: Optional[str] = None
    post_immediately: bool = False


class AIFeedbackRequest(BaseModel):
    user_message: str
    ai_intent: Optional[str] = None
    ai_prediction: Optional[Any] = None
    user_correction: Optional[str] = None
    final_result: Optional[Any] = None
    rating: Optional[int] = Field(None, ge=1, le=5)
    note: Optional[str] = None


class ExplainRequest(BaseModel):
    description: str
    lines: List[Dict[str, Any]]


class MissingInfoRequest(BaseModel):
    description: str
    amount: Optional[float] = None
    payment_method: Optional[str] = None
    vat_rate: Optional[float] = None
    supplier_or_customer: Optional[str] = None


def _now() -> str:
    return datetime.utcnow().isoformat(timespec="seconds") + "Z"


def _empty_store() -> Dict[str, Any]:
    return {
        "journal_drafts": [],
        "journal_entries": [],
        "raw_transactions": [],
        "import_batches": [],
        "learning_examples": [],
        "ai_feedback": [],
        "chat_confirmations": [],
        "users": [],
        "sessions": [],
        "audit_logs": [],
        "counters": {},
    }


def _load_store() -> Dict[str, Any]:
    DATA_DIR.mkdir(exist_ok=True)
    if not STORE_PATH.exists():
        store = _empty_store()
        _save_store(store)
        return store
    try:
        with STORE_PATH.open("r", encoding="utf-8") as f:
            loaded = json.load(f)
    except json.JSONDecodeError:
        loaded = _empty_store()
    store = _empty_store()
    store.update(loaded if isinstance(loaded, dict) else {})
    for key, default in _empty_store().items():
        store.setdefault(key, default)
    return store


def _save_store(store: Dict[str, Any]) -> None:
    DATA_DIR.mkdir(exist_ok=True)
    with STORE_PATH.open("w", encoding="utf-8") as f:
        json.dump(store, f, ensure_ascii=False, indent=2)


def _next_id(store: Dict[str, Any], prefix: str) -> str:
    counters = store.setdefault("counters", {})
    counters[prefix] = int(counters.get(prefix, 0)) + 1
    return f"{prefix}-{counters[prefix]:05d}"


def _audit(store: Dict[str, Any], action: str, entity: str, entity_id: str, detail: Optional[Dict[str, Any]] = None, actor: str = "system") -> None:
    store.setdefault("audit_logs", []).append({
        "id": _next_id(store, "AUDIT"),
        "created_at": _now(),
        "actor": actor,
        "action": action,
        "entity": entity,
        "entity_id": entity_id,
        "detail": detail or {},
    })


def _hash_password(password: str, salt: Optional[str] = None) -> Dict[str, str]:
    salt = salt or secrets.token_hex(8)
    digest = hashlib.sha256(f"{salt}:{password}".encode("utf-8")).hexdigest()
    return {"salt": salt, "password_hash": digest}


def _match_keyword(text: str, words: List[str]) -> bool:
    low = text.lower()
    return any(w in low for w in words)


def _suggest_accounts(description: str, payment_method: str, debit: Optional[str] = None, credit: Optional[str] = None) -> Dict[str, str]:
    text = description.lower()
    if debit:
        debit_account = debit
    elif _match_keyword(text, ["bán", "doanh thu", "thu tiền", "sales", "revenue"]):
        debit_account = "111" if payment_method == "cash" else "112"
    elif _match_keyword(text, ["máy tính", "thiết bị", "tài sản", "computer", "asset"]):
        debit_account = "211"
    elif _match_keyword(text, ["quảng cáo", "marketing", "facebook", "google ads"]):
        debit_account = "641"
    elif _match_keyword(text, ["lương", "salary", "nhân viên"]):
        debit_account = "642"
    elif _match_keyword(text, ["mua", "chi phí", "dịch vụ", "văn phòng", "điện", "nước"]):
        debit_account = "642"
    else:
        debit_account = "642"

    if credit:
        credit_account = credit
    elif _match_keyword(text, ["chưa thanh toán", "công nợ", "trả sau", "payable"]):
        credit_account = "331"
    elif payment_method in {"bank", "transfer", "ngân hàng", "112"}:
        credit_account = "112"
    elif _match_keyword(text, ["bán", "doanh thu", "sales", "revenue"]):
        credit_account = "511"
    else:
        credit_account = "111"
    return {"debit": debit_account, "credit": credit_account}


def _build_lines(req: JournalDraftCreate) -> List[Dict[str, Any]]:
    accounts = _suggest_accounts(req.description, req.payment_method, req.debit_account, req.credit_account)
    amount = round(float(req.amount), 2)
    vat = round(amount * float(req.vat_rate or 0), 2)
    lines: List[Dict[str, Any]] = []
    if accounts["credit"].startswith("511"):
        total = round(amount + vat, 2)
        lines.append({"account": accounts["debit"], "debit": total, "credit": 0, "memo": "Thu tiền / ghi nhận phải thu"})
        lines.append({"account": accounts["credit"], "debit": 0, "credit": amount, "memo": "Doanh thu chưa VAT"})
        if vat > 0:
            lines.append({"account": "3331", "debit": 0, "credit": vat, "memo": "VAT đầu ra"})
    else:
        total = round(amount + vat, 2)
        lines.append({"account": accounts["debit"], "debit": amount, "credit": 0, "memo": "Chi phí/tài sản chưa VAT"})
        if vat > 0:
            lines.append({"account": "1331", "debit": vat, "credit": 0, "memo": "VAT đầu vào được khấu trừ"})
        lines.append({"account": accounts["credit"], "debit": 0, "credit": total, "memo": "Thanh toán hoặc công nợ"})
    return lines


def _line_total(lines: List[Dict[str, Any]], side: str) -> float:
    return round(sum(float(line.get(side, 0) or 0) for line in lines), 2)


def _quality_score(store: Dict[str, Any]) -> int:
    drafts = store.get("journal_drafts", [])
    posted = store.get("journal_entries", [])
    approved = len([d for d in drafts if d.get("status") in {"approved", "posted"}])
    rejected = len([d for d in drafts if d.get("status") == "rejected"])
    total = max(len(drafts), 1)
    score = 55 + min(25, len(posted) * 2) + round((approved / total) * 15) - round((rejected / total) * 10)
    return max(0, min(100, int(score)))


def _detect_errors(lines: List[Dict[str, Any]], description: str = "") -> List[Dict[str, str]]:
    errors: List[Dict[str, str]] = []
    debit = _line_total(lines, "debit")
    credit = _line_total(lines, "credit")
    if debit != credit:
        errors.append({"level": "high", "code": "UNBALANCED", "message": f"Tổng Nợ {debit} khác tổng Có {credit}."})
    for idx, line in enumerate(lines, start=1):
        account = str(line.get("account", "")).strip()
        if not account:
            errors.append({"level": "high", "code": "MISSING_ACCOUNT", "message": f"Dòng {idx} thiếu tài khoản."})
        if float(line.get("debit", 0) or 0) > 0 and float(line.get("credit", 0) or 0) > 0:
            errors.append({"level": "medium", "code": "BOTH_SIDES", "message": f"Dòng {idx} vừa có Nợ vừa có Có."})
    if "vat" in description.lower() and not any(str(l.get("account", "")).startswith(("133", "3331")) for l in lines):
        errors.append({"level": "medium", "code": "VAT_MISSING", "message": "Mô tả có VAT nhưng bút toán chưa có tài khoản VAT."})
    return errors




def _infer_category(description: str) -> str:
    text = description.lower()
    if _match_keyword(text, ["bán", "doanh thu", "thu tiền", "sales", "revenue"]):
        return "Doanh thu"
    if _match_keyword(text, ["quảng cáo", "marketing", "facebook", "google ads"]):
        return "Chi phí marketing"
    if _match_keyword(text, ["lương", "salary", "nhân viên"]):
        return "Chi phí nhân sự"
    if _match_keyword(text, ["máy tính", "thiết bị", "tài sản", "computer", "asset"]):
        return "Tài sản cố định"
    if _match_keyword(text, ["điện", "nước", "internet", "văn phòng", "dịch vụ", "mua", "chi phí"]):
        return "Chi phí quản lý"
    return "Cần kế toán kiểm tra"


def _confidence_for(description: str, lines: List[Dict[str, Any]]) -> float:
    base = 0.62
    if _infer_category(description) != "Cần kế toán kiểm tra":
        base += 0.18
    if not _detect_errors(lines, description):
        base += 0.12
    if _line_total(lines, "debit") == _line_total(lines, "credit"):
        base += 0.08
    return round(min(base, 0.98), 2)


def _remember_learning(store: Dict[str, Any], *, source: str, description: str, lines: List[Dict[str, Any]], label: str = "approved", note: Optional[str] = None) -> None:
    example_id = _next_id(store, "LEARN")
    accounts = [str(line.get("account")) for line in lines if line.get("account")]
    store.setdefault("learning_examples", []).append({
        "id": example_id,
        "created_at": _now(),
        "source": source,
        "label": label,
        "description": description,
        "category": _infer_category(description),
        "accounts": accounts,
        "lines": lines,
        "note": note,
        "keywords": sorted(set(re.findall(r"[A-Za-zÀ-ỹ0-9]+", description.lower())))[:20],
    })


def _create_draft_from_item(store: Dict[str, Any], item: BulkTransactionItem, batch_id: Optional[str] = None, auto_approve: bool = False) -> Dict[str, Any]:
    req = JournalDraftCreate(
        description=item.description,
        amount=item.amount,
        vat_rate=item.vat_rate,
        payment_method=item.payment_method,
        category=item.category or _infer_category(item.description),
        source=item.source,
    )
    lines = _build_lines(req)
    errors = _detect_errors(lines, req.description)
    draft_id = _next_id(store, "DRAFT")
    draft = {
        "id": draft_id,
        "version": "V40.5/V25+",
        "created_at": _now(),
        "updated_at": _now(),
        "status": "approved" if auto_approve and not errors else "draft",
        "source": item.source,
        "batch_id": batch_id,
        "description": req.description,
        "amount": req.amount,
        "vat_rate": req.vat_rate,
        "payment_method": req.payment_method,
        "category": req.category,
        "ai_confidence": _confidence_for(req.description, lines),
        "lines": lines,
        "ai_confidence": _confidence_for(req.description, lines),
        "debit_total": _line_total(lines, "debit"),
        "credit_total": _line_total(lines, "credit"),
        "balanced": _line_total(lines, "debit") == _line_total(lines, "credit"),
        "risk_flags": errors,
        "safe_policy": "AI chỉ tạo nháp. Kế toán phải approve/post trước khi vào sổ.",
    }
    store.setdefault("journal_drafts", []).append(draft)
    _audit(store, "bulk_create_draft", "journal_draft", draft_id, {"batch_id": batch_id, "description": item.description})
    return draft


def _sample_transactions() -> List[BulkTransactionItem]:
    return [
        BulkTransactionItem(date="2026-05-01", description="Bán hàng cho khách A chuyển khoản", amount=52000000, transaction_type="income", vat_rate=0.1, payment_method="bank", source="sample_seed"),
        BulkTransactionItem(date="2026-05-03", description="Thanh toán quảng cáo Facebook tháng 5", amount=6500000, transaction_type="expense", vat_rate=0.1, payment_method="bank", source="sample_seed"),
        BulkTransactionItem(date="2026-05-07", description="Mua máy tính cho phòng kế toán", amount=20000000, transaction_type="expense", vat_rate=0.1, payment_method="bank", source="sample_seed"),
        BulkTransactionItem(date="2026-05-12", description="Chi tiền điện văn phòng", amount=1800000, transaction_type="expense", vat_rate=0.1, payment_method="cash", source="sample_seed"),
        BulkTransactionItem(date="2026-05-18", description="Bán dịch vụ tư vấn cho khách B", amount=38000000, transaction_type="income", vat_rate=0.1, payment_method="bank", source="sample_seed"),
        BulkTransactionItem(date="2026-05-22", description="Trả lương nhân viên tháng 5", amount=24000000, transaction_type="expense", vat_rate=0, payment_method="bank", source="sample_seed"),
    ]


def _parse_import_row(row: Dict[str, Any], source: str) -> BulkTransactionItem:
    def pick(*names, default=None):
        for n in names:
            for k, v in row.items():
                if str(k).strip().lower() == n:
                    return v
        return default
    amount_raw = pick("amount", "số tiền", "so tien", "tien", "value", default=0)
    try:
        amount = float(str(amount_raw).replace(".", "").replace(",", ".")) if isinstance(amount_raw, str) else float(amount_raw or 0)
    except Exception:
        amount = 0
    vat_raw = pick("vat_rate", "vat", "thuế", "thue", default=0.1)
    try:
        vat = float(str(vat_raw).replace("%", ""))
        vat = vat / 100 if vat > 1 else vat
    except Exception:
        vat = 0.1
    return BulkTransactionItem(
        date=str(pick("date", "ngày", "ngay", default="") or "") or None,
        description=str(pick("description", "mô tả", "mo ta", "nội dung", "noi dung", default="") or "Giao dịch import"),
        amount=amount,
        transaction_type=str(pick("type", "loại", "loai", default="") or ""),
        category=str(pick("category", "danh mục", "danh muc", default="") or "") or None,
        vat_rate=vat,
        payment_method=str(pick("payment_method", "thanh toán", "thanh toan", default="cash") or "cash"),
        source=source,
    )


def _chat_summary(store: Dict[str, Any]) -> Dict[str, Any]:
    vat = v28_vat_report()
    income = v29_income_statement()
    dashboard = v30_financial_dashboard()
    return {"vat": vat, "income": income, "dashboard": dashboard}


def _requires_confirmation(intent: str) -> bool:
    return intent in {"post_approved_drafts", "backup", "approve_all_drafts", "restore"}


@router.post("/ai/v25/journal-draft/create")
def v25_create_journal_draft(req: JournalDraftCreate):
    store = _load_store()
    lines = _build_lines(req)
    errors = _detect_errors(lines, req.description)
    draft_id = _next_id(store, "DRAFT")
    draft = {
        "id": draft_id,
        "version": "V25",
        "created_at": _now(),
        "updated_at": _now(),
        "status": "draft",
        "source": req.source,
        "description": req.description,
        "amount": req.amount,
        "vat_rate": req.vat_rate,
        "payment_method": req.payment_method,
        "category": req.category,
        "lines": lines,
        "debit_total": _line_total(lines, "debit"),
        "credit_total": _line_total(lines, "credit"),
        "balanced": _line_total(lines, "debit") == _line_total(lines, "credit"),
        "risk_flags": errors,
        "safe_policy": "AI chỉ tạo nháp. Kế toán phải approve/post trước khi vào sổ.",
    }
    store["journal_drafts"].append(draft)
    _audit(store, "create", "journal_draft", draft_id, {"description": req.description})
    _save_store(store)
    return draft


@router.get("/ai/v25/journal-draft/list")
def v25_list_journal_drafts(status: Optional[str] = None):
    store = _load_store()
    drafts = store.get("journal_drafts", [])
    if status:
        drafts = [d for d in drafts if d.get("status") == status]
    return {"count": len(drafts), "items": drafts}


def _find_draft(store: Dict[str, Any], draft_id: str) -> Dict[str, Any]:
    for draft in store.get("journal_drafts", []):
        if draft.get("id") == draft_id:
            return draft
    raise HTTPException(status_code=404, detail="Không tìm thấy journal draft")


@router.post("/ai/v25/journal-draft/{draft_id}/approve")
def v25_approve_journal_draft(draft_id: str, req: JournalDraftDecision = JournalDraftDecision()):
    store = _load_store()
    draft = _find_draft(store, draft_id)
    draft["status"] = "approved"
    draft["updated_at"] = _now()
    draft["approval_note"] = req.note
    _remember_learning(store, source="v25_approve", description=draft.get("description", ""), lines=draft.get("lines", []), label="approved", note=req.note)
    _audit(store, "approve", "journal_draft", draft_id, {"note": req.note})
    _save_store(store)
    return draft


@router.post("/ai/v25/journal-draft/{draft_id}/reject")
def v25_reject_journal_draft(draft_id: str, req: JournalDraftDecision = JournalDraftDecision()):
    store = _load_store()
    draft = _find_draft(store, draft_id)
    draft["status"] = "rejected"
    draft["updated_at"] = _now()
    draft["reject_note"] = req.note
    _remember_learning(store, source="v25_reject", description=draft.get("description", ""), lines=draft.get("lines", []), label="rejected", note=req.note)
    _audit(store, "reject", "journal_draft", draft_id, {"note": req.note})
    _save_store(store)
    return draft


@router.post("/ledger/post-entry")
def v26_post_entry(req: PostEntryRequest):
    store = _load_store()
    if req.draft_id:
        draft = _find_draft(store, req.draft_id)
        if draft.get("status") not in {"approved", "draft"}:
            raise HTTPException(status_code=400, detail="Chỉ draft/approved mới được post")
        lines = draft.get("lines", [])
        description = draft.get("description")
    else:
        if not req.lines or not req.description:
            raise HTTPException(status_code=400, detail="Cần draft_id hoặc description + lines")
        lines = req.lines
        description = req.description
    errors = _detect_errors(lines, description or "")
    if any(e["code"] == "UNBALANCED" for e in errors):
        raise HTTPException(status_code=400, detail={"message": "Bút toán chưa cân", "errors": errors})
    entry_id = _next_id(store, "JE")
    entry = {
        "id": entry_id,
        "version": "V26",
        "posted_at": _now(),
        "description": description,
        "source_draft_id": req.draft_id,
        "lines": lines,
        "debit_total": _line_total(lines, "debit"),
        "credit_total": _line_total(lines, "credit"),
        "note": req.note,
    }
    store["journal_entries"].append(entry)
    if req.draft_id:
        draft = _find_draft(store, req.draft_id)
        draft["status"] = "posted"
        draft["posted_entry_id"] = entry_id
        draft["updated_at"] = _now()
    _audit(store, "post", "journal_entry", entry_id, {"source_draft_id": req.draft_id})
    _save_store(store)
    return entry


@router.get("/ledger/general-journal")
def v26_general_journal():
    store = _load_store()
    return {"count": len(store.get("journal_entries", [])), "items": store.get("journal_entries", [])}


@router.get("/ledger/general-ledger")
def v26_general_ledger():
    store = _load_store()
    ledger = defaultdict(lambda: {"account": "", "debit": 0.0, "credit": 0.0, "balance": 0.0, "lines": []})
    for entry in store.get("journal_entries", []):
        for line in entry.get("lines", []):
            account = str(line.get("account", ""))
            debit = float(line.get("debit", 0) or 0)
            credit = float(line.get("credit", 0) or 0)
            row = ledger[account]
            row["account"] = account
            row["debit"] = round(row["debit"] + debit, 2)
            row["credit"] = round(row["credit"] + credit, 2)
            row["balance"] = round(row["debit"] - row["credit"], 2)
            row["lines"].append({"entry_id": entry.get("id"), "description": entry.get("description"), **line})
    return {"count": len(ledger), "accounts": list(ledger.values())}


@router.get("/reports/v27/export-excel")
def v27_export_excel():
    if Workbook is None:
        raise HTTPException(status_code=500, detail="openpyxl chưa được cài")
    store = _load_store()
    wb = Workbook()
    ws = wb.active
    ws.title = "General Journal"
    headers = ["Entry ID", "Posted At", "Description", "Account", "Debit", "Credit", "Memo"]
    ws.append(headers)
    for h in ws[1]:
        h.font = Font(bold=True)
        h.alignment = Alignment(horizontal="center")
    for entry in store.get("journal_entries", []):
        for line in entry.get("lines", []):
            ws.append([entry.get("id"), entry.get("posted_at"), entry.get("description"), line.get("account"), line.get("debit", 0), line.get("credit", 0), line.get("memo")])
    ws2 = wb.create_sheet("Journal Drafts")
    ws2.append(["Draft ID", "Status", "Description", "Amount", "VAT Rate", "Debit Total", "Credit Total"])
    for draft in store.get("journal_drafts", []):
        ws2.append([draft.get("id"), draft.get("status"), draft.get("description"), draft.get("amount"), draft.get("vat_rate"), draft.get("debit_total"), draft.get("credit_total")])
    ws3 = wb.create_sheet("Audit Logs")
    ws3.append(["ID", "Created At", "Actor", "Action", "Entity", "Entity ID"])
    for log in store.get("audit_logs", []):
        ws3.append([log.get("id"), log.get("created_at"), log.get("actor"), log.get("action"), log.get("entity"), log.get("entity_id")])
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=finiip_v27_accounting_export.xlsx"})


@router.get("/reports/vat")
def v28_vat_report():
    store = _load_store()
    vat_input = 0.0
    vat_output = 0.0
    details = []
    for entry in store.get("journal_entries", []):
        for line in entry.get("lines", []):
            account = str(line.get("account", ""))
            debit = float(line.get("debit", 0) or 0)
            credit = float(line.get("credit", 0) or 0)
            if account.startswith(VAT_INPUT_CODES):
                vat_input += debit - credit
                details.append({"type": "vat_input", "entry_id": entry.get("id"), **line})
            if account.startswith(VAT_OUTPUT_CODES):
                vat_output += credit - debit
                details.append({"type": "vat_output", "entry_id": entry.get("id"), **line})
    return {"version": "V28", "vat_input": round(vat_input, 2), "vat_output": round(vat_output, 2), "vat_payable": round(vat_output - vat_input, 2), "details": details}


@router.get("/reports/vat/export")
def v28_vat_export():
    return v27_export_excel()


@router.get("/reports/income-statement")
def v29_income_statement():
    store = _load_store()
    revenue = 0.0
    expenses = 0.0
    rows = []
    for entry in store.get("journal_entries", []):
        for line in entry.get("lines", []):
            account = str(line.get("account", ""))
            debit = float(line.get("debit", 0) or 0)
            credit = float(line.get("credit", 0) or 0)
            if account.startswith(REVENUE_PREFIXES):
                amount = credit - debit
                revenue += amount
                rows.append({"type": "revenue", "entry_id": entry.get("id"), "account": account, "amount": round(amount, 2)})
            if account.startswith(EXPENSE_PREFIXES):
                amount = debit - credit
                expenses += amount
                rows.append({"type": "expense", "entry_id": entry.get("id"), "account": account, "amount": round(amount, 2)})
    return {"version": "V29", "revenue": round(revenue, 2), "expenses": round(expenses, 2), "profit_before_tax": round(revenue - expenses, 2), "rows": rows}


@router.get("/dashboard/v30/financial")
def v30_financial_dashboard():
    store = _load_store()
    vat = v28_vat_report()
    income = v29_income_statement()
    drafts = store.get("journal_drafts", [])
    return {
        "version": "V30",
        "cards": {
            "revenue": income["revenue"],
            "expenses": income["expenses"],
            "profit_before_tax": income["profit_before_tax"],
            "vat_payable": vat["vat_payable"],
            "journal_entries": len(store.get("journal_entries", [])),
            "drafts_total": len(drafts),
            "drafts_pending": len([d for d in drafts if d.get("status") == "draft"]),
            "drafts_posted": len([d for d in drafts if d.get("status") == "posted"]),
            "quality_score": _quality_score(store),
        },
        "next_actions": [
            "Post các draft đã approve",
            "Kiểm tra VAT nếu vat_payable âm/dương bất thường",
            "Backup dữ liệu trước khi import hoặc chỉnh sửa lớn",
        ],
    }


@router.get("/v30/financial-dashboard-ui")
def v30_financial_dashboard_ui():
    path = FRONTEND_DIR / "v30_financial_dashboard.html"
    if not path.exists():
        raise HTTPException(status_code=404, detail="Không tìm thấy frontend/v30_financial_dashboard.html")
    return FileResponse(path, media_type="text/html")


@router.post("/ai/v31/explain-journal")
def v31_explain_journal(req: ExplainRequest):
    explanations = []
    for line in req.lines:
        account = str(line.get("account", ""))
        side = "Nợ" if float(line.get("debit", 0) or 0) > 0 else "Có"
        if account.startswith("111"):
            meaning = "tiền mặt"
        elif account.startswith("112"):
            meaning = "tiền gửi ngân hàng"
        elif account.startswith("133"):
            meaning = "VAT đầu vào được khấu trừ"
        elif account.startswith("3331"):
            meaning = "VAT đầu ra phải nộp"
        elif account.startswith("511"):
            meaning = "doanh thu bán hàng/cung cấp dịch vụ"
        elif account.startswith("211"):
            meaning = "tài sản cố định"
        elif account.startswith("641"):
            meaning = "chi phí bán hàng/marketing"
        elif account.startswith("642"):
            meaning = "chi phí quản lý doanh nghiệp"
        else:
            meaning = "tài khoản kế toán cần kiểm tra thêm"
        explanations.append({"account": account, "side": side, "meaning": meaning, "reason": f"Ghi {side} TK {account} vì nghiệp vụ: {req.description}."})
    return {"version": "V31", "description": req.description, "explanations": explanations, "summary": "AI giải thích để kế toán kiểm tra, không thay thế xét đoán chuyên môn."}


@router.post("/ai/v32/detect-accounting-errors")
def v32_detect_accounting_errors(req: ExplainRequest):
    return {"version": "V32", "errors": _detect_errors(req.lines, req.description), "checked_at": _now()}


@router.post("/ai/v33/missing-info-questions")
def v33_missing_info_questions(req: MissingInfoRequest):
    questions = []
    if req.amount is None:
        questions.append("Số tiền chưa VAT là bao nhiêu?")
    if not req.payment_method:
        questions.append("Thanh toán bằng tiền mặt, ngân hàng hay công nợ?")
    if req.vat_rate is None and _match_keyword(req.description, ["hóa đơn", "vat", "thuế"]):
        questions.append("Thuế VAT là 0%, 5%, 8% hay 10%?")
    if not req.supplier_or_customer and _match_keyword(req.description, ["mua", "bán", "hóa đơn"]):
        questions.append("Tên nhà cung cấp/khách hàng là gì?")
    return {"version": "V33", "need_more_info": bool(questions), "questions": questions or ["Thông tin hiện tại đủ để tạo bút toán nháp."]}


@router.post("/ocr/v34/invoice-enhanced/text")
def v34_invoice_enhanced_text(req: OCRTextRequest):
    text = req.text
    amount_matches = [float(x.replace(".", "").replace(",", ".")) for x in re.findall(r"\b\d{1,3}(?:[.]\d{3})+(?:,\d+)?\b|\b\d+(?:,\d+)?\b", text)]
    tax_code = re.search(r"(?:MST|Mã số thuế|Tax code)[:\s]*([0-9\-]{8,20})", text, re.I)
    date_match = re.search(r"(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})", text)
    invoice_no = re.search(r"(?:Số|No\.?|Invoice No)[:\s]*([A-Z0-9\-]+)", text, re.I)
    company = None
    for line in [l.strip() for l in text.splitlines() if l.strip()]:
        if any(k in line.lower() for k in ["công ty", "company", "co.,", "tnhh", "jsc"]):
            company = line
            break
    total = max(amount_matches) if amount_matches else 0.0
    vat_guess = round(total / 11, 2) if total else 0.0
    return {"version": "V34", "company": company, "tax_code": tax_code.group(1) if tax_code else None, "invoice_date": date_match.group(1) if date_match else None, "invoice_no": invoice_no.group(1) if invoice_no else None, "total_amount": total, "vat_guess": vat_guess, "raw_text": text}


@router.post("/ai/v35/invoice-to-journal-draft")
def v35_invoice_to_journal_draft(req: OCRTextRequest):
    extracted = v34_invoice_enhanced_text(req)
    description = f"Hóa đơn {extracted.get('invoice_no') or ''} {extracted.get('company') or ''}".strip()
    total = float(extracted.get("total_amount") or 0)
    before_vat = round(total / 1.1, 2) if total else 0.0
    draft_req = JournalDraftCreate(description=description or "Hóa đơn OCR", amount=before_vat, vat_rate=0.1, payment_method="payable", source="v35_invoice_to_journal")
    draft = v25_create_journal_draft(draft_req)
    draft["ocr"] = extracted
    return {"version": "V35", "ocr": extracted, "journal_draft": draft}


@router.post("/admin/v36/users")
def v36_create_user(req: UserCreate):
    if req.role not in ROLE_PERMISSIONS:
        raise HTTPException(status_code=400, detail="Role không hợp lệ")
    store = _load_store()
    if any(u.get("email") == req.email for u in store.get("users", [])):
        raise HTTPException(status_code=400, detail="Email đã tồn tại")
    user_id = _next_id(store, "USER")
    hashed = _hash_password(req.password)
    user = {"id": user_id, "created_at": _now(), "name": req.name, "email": req.email, "role": req.role, "permissions": ROLE_PERMISSIONS[req.role], **hashed}
    store["users"].append(user)
    _audit(store, "create", "user", user_id, {"email": req.email, "role": req.role})
    _save_store(store)
    safe = {k: v for k, v in user.items() if k not in {"password_hash", "salt"}}
    return safe


@router.get("/admin/v36/users")
def v36_list_users():
    store = _load_store()
    return {"roles": ROLE_PERMISSIONS, "users": [{k: v for k, v in u.items() if k not in {"password_hash", "salt"}} for u in store.get("users", [])]}


@router.post("/auth/v37/login")
def v37_login(req: LoginRequest):
    store = _load_store()
    user = next((u for u in store.get("users", []) if u.get("email") == req.email), None)
    if not user:
        raise HTTPException(status_code=401, detail="Sai email hoặc mật khẩu")
    hashed = _hash_password(req.password, user.get("salt"))
    if hashed["password_hash"] != user.get("password_hash"):
        raise HTTPException(status_code=401, detail="Sai email hoặc mật khẩu")
    token = secrets.token_urlsafe(24)
    session = {"token": token, "user_id": user.get("id"), "email": user.get("email"), "role": user.get("role"), "created_at": _now()}
    store.setdefault("sessions", []).append(session)
    _audit(store, "login", "user", user.get("id"), {"email": req.email})
    _save_store(store)
    return {"token": token, "user": {k: v for k, v in user.items() if k not in {"password_hash", "salt"}}}


@router.get("/system/v38/database-health")
def v38_database_health():
    store = _load_store()
    return {"version": "V38", "storage": "json_mvp_store", "path": str(STORE_PATH), "collections": {k: len(v) if isinstance(v, list) else len(v) if isinstance(v, dict) else 1 for k, v in store.items()}, "migration_note": "MVP dùng JSON store; giai đoạn sản phẩm thật có thể chuyển sang PostgreSQL/SQLAlchemy."}


@router.get("/audit/v39/logs")
def v39_audit_logs(limit: int = 100):
    store = _load_store()
    logs = list(reversed(store.get("audit_logs", [])))[:limit]
    return {"version": "V39", "count": len(logs), "items": logs}


@router.post("/backup/v40/create")
def v40_create_backup():
    store = _load_store()
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    backup_name = f"finiip_backup_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.zip"
    backup_path = BACKUP_DIR / backup_name
    manifest = {"created_at": _now(), "version": "V40", "files": ["v25_v40_store.json"]}
    with zipfile.ZipFile(backup_path, "w", zipfile.ZIP_DEFLATED) as zf:
        if STORE_PATH.exists():
            zf.write(STORE_PATH, arcname="v25_v40_store.json")
        zf.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))
    _audit(store, "create", "backup", backup_name, {"path": str(backup_path)})
    _save_store(store)
    return {"version": "V40", "backup_file": backup_name, "path": str(backup_path), "download_url": f"/backup/v40/download/{backup_name}"}


@router.get("/backup/v40/download/{backup_name}")
def v40_download_backup(backup_name: str):
    path = BACKUP_DIR / backup_name
    if not path.exists():
        raise HTTPException(status_code=404, detail="Không tìm thấy backup")
    return FileResponse(path, media_type="application/zip", filename=backup_name)


@router.post("/backup/v40/restore")
async def v40_restore_backup(file: UploadFile = File(...)):
    if not file.filename.endswith(".zip"):
        raise HTTPException(status_code=400, detail="Chỉ nhận file .zip")
    raw = await file.read()
    with zipfile.ZipFile(BytesIO(raw), "r") as zf:
        if "v25_v40_store.json" not in zf.namelist():
            raise HTTPException(status_code=400, detail="Backup thiếu v25_v40_store.json")
        restored = json.loads(zf.read("v25_v40_store.json").decode("utf-8"))
    current = _load_store()
    restore_id = _next_id(current, "RESTORE")
    _save_store(restored)
    store = _load_store()
    _audit(store, "restore", "backup", restore_id, {"filename": file.filename})
    _save_store(store)
    return {"version": "V40", "restored": True, "restore_id": restore_id}




@router.post("/import/v40-5/transactions/bulk")
def v405_import_transactions_bulk(req: BulkImportRequest):
    store = _load_store()
    batch_id = _next_id(store, "BATCH")
    drafts = []
    raw_rows = []
    for item in req.items:
        row_id = _next_id(store, "RAW")
        raw = item.model_dump()
        raw.update({"id": row_id, "batch_id": batch_id, "created_at": _now(), "category_inferred": item.category or _infer_category(item.description)})
        store.setdefault("raw_transactions", []).append(raw)
        raw_rows.append(raw)
        if req.auto_create_drafts:
            drafts.append(_create_draft_from_item(store, item, batch_id=batch_id, auto_approve=req.auto_approve_safe))
    batch = {
        "id": batch_id,
        "version": "V40.5",
        "created_at": _now(),
        "source": req.source,
        "raw_count": len(raw_rows),
        "draft_count": len(drafts),
        "auto_create_drafts": req.auto_create_drafts,
        "auto_approve_safe": req.auto_approve_safe,
    }
    store.setdefault("import_batches", []).append(batch)
    _audit(store, "import_bulk", "import_batch", batch_id, {"raw_count": len(raw_rows), "draft_count": len(drafts)})
    _save_store(store)
    return {"version": "V40.5", "batch": batch, "raw_transactions": raw_rows, "journal_drafts": drafts}


@router.post("/import/v40-5/sample-data")
def v405_seed_sample_data(auto_post: bool = False):
    store = _load_store()
    batch_id = _next_id(store, "BATCH")
    drafts = []
    entries = []
    for item in _sample_transactions():
        raw_id = _next_id(store, "RAW")
        raw = item.model_dump()
        raw.update({"id": raw_id, "batch_id": batch_id, "created_at": _now(), "category_inferred": _infer_category(item.description)})
        store.setdefault("raw_transactions", []).append(raw)
        draft = _create_draft_from_item(store, item, batch_id=batch_id, auto_approve=True)
        drafts.append(draft)
        if auto_post:
            errors = _detect_errors(draft.get("lines", []), draft.get("description", ""))
            if not errors:
                entry_id = _next_id(store, "JE")
                entry = {
                    "id": entry_id,
                    "version": "V40.5/V26+",
                    "posted_at": _now(),
                    "description": draft.get("description"),
                    "source_draft_id": draft.get("id"),
                    "lines": draft.get("lines", []),
                    "debit_total": draft.get("debit_total"),
                    "credit_total": draft.get("credit_total"),
                    "note": "Auto posted sample data for demo/testing",
                }
                store.setdefault("journal_entries", []).append(entry)
                draft["status"] = "posted"
                draft["posted_entry_id"] = entry_id
                entries.append(entry)
                _remember_learning(store, source="sample_seed_auto_post", description=draft.get("description", ""), lines=draft.get("lines", []), label="approved_sample")
    batch = {"id": batch_id, "version": "V40.5", "created_at": _now(), "source": "sample_seed", "raw_count": len(drafts), "draft_count": len(drafts), "posted_count": len(entries)}
    store.setdefault("import_batches", []).append(batch)
    _audit(store, "seed_sample_data", "import_batch", batch_id, {"auto_post": auto_post, "draft_count": len(drafts), "posted_count": len(entries)})
    _save_store(store)
    return {"version": "V40.5", "batch": batch, "journal_drafts": drafts, "journal_entries": entries, "tip": "Dùng auto_post=true để có ngay dữ liệu cho chatbot/báo cáo."}


@router.post("/import/v40-5/transactions/file")
async def v405_import_transactions_file(file: UploadFile = File(...), auto_create_drafts: bool = True):
    raw = await file.read()
    rows: List[Dict[str, Any]] = []
    name = (file.filename or "").lower()
    if name.endswith(".csv"):
        text = raw.decode("utf-8-sig")
        rows = list(csv.DictReader(text.splitlines()))
    elif name.endswith(".xlsx"):
        if Workbook is None:
            raise HTTPException(status_code=500, detail="openpyxl chưa được cài")
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(raw), data_only=True)
        ws = wb.active
        values = list(ws.iter_rows(values_only=True))
        if not values:
            raise HTTPException(status_code=400, detail="File rỗng")
        headers = [str(h or "").strip() for h in values[0]]
        for row in values[1:]:
            rows.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
    else:
        raise HTTPException(status_code=400, detail="Chỉ hỗ trợ .csv hoặc .xlsx")
    items = [_parse_import_row(row, source=f"file:{file.filename}") for row in rows if any(str(v or "").strip() for v in row.values())]
    return v405_import_transactions_bulk(BulkImportRequest(items=items, auto_create_drafts=auto_create_drafts, source=f"file:{file.filename}"))


@router.get("/import/v40-5/status")
def v405_import_status():
    store = _load_store()
    return {
        "version": "V40.5",
        "raw_transactions": len(store.get("raw_transactions", [])),
        "import_batches": len(store.get("import_batches", [])),
        "learning_examples": len(store.get("learning_examples", [])),
        "journal_drafts": len(store.get("journal_drafts", [])),
        "journal_entries": len(store.get("journal_entries", [])),
        "supported_import": ["manual bulk JSON", "CSV", "XLSX", "sample demo data"],
    }


@router.get("/ai/learning/v40-5/memory")
def v405_learning_memory(limit: int = 100):
    store = _load_store()
    examples = list(reversed(store.get("learning_examples", [])))[:limit]
    categories: Dict[str, int] = defaultdict(int)
    accounts: Dict[str, int] = defaultdict(int)
    for ex in store.get("learning_examples", []):
        categories[ex.get("category", "unknown")] += 1
        for account in ex.get("accounts", []):
            accounts[account] += 1
    return {"version": "V40.5", "count": len(store.get("learning_examples", [])), "category_stats": dict(categories), "account_stats": dict(accounts), "items": examples}





def _strip_accents(text: str) -> str:
    normalized = unicodedata.normalize("NFD", text or "")
    stripped = "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")
    return stripped.replace("đ", "d").replace("Đ", "D")


def _norm_text(text: str) -> str:
    return _strip_accents(text).lower().strip()


def _load_json_file(path: Path, default: Any) -> Any:
    try:
        if path.exists():
            with path.open("r", encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        return default
    return default


def _contains_any_normalized(message: str, keywords: List[str]) -> bool:
    low = _norm_text(message)
    return any(_norm_text(k) in low for k in keywords)


def _extract_amount_from_text(message: str) -> Optional[float]:
    low = _norm_text(message).replace(",", ".")
    # Examples: 5 triệu, 800 nghìn, 12000000, 2.5 ty
    patterns = [
        (r"(\d+(?:\.\d+)?)\s*(ty|tỷ)", 1_000_000_000),
        (r"(\d+(?:\.\d+)?)\s*(trieu|triệu)", 1_000_000),
        (r"(\d+(?:\.\d+)?)\s*(nghin|ngàn|k)\b", 1_000),
        (r"(\d{4,})", 1),
    ]
    for pattern, multiplier in patterns:
        match = re.search(pattern, low)
        if match:
            try:
                return round(float(match.group(1)) * multiplier, 2)
            except ValueError:
                continue
    return None


def _detect_payment_method_from_text(message: str) -> str:
    if _contains_any_normalized(message, ["chuyển khoản", "ngân hàng", "bank", "112", "thẻ", "tai khoan"]):
        return "bank"
    if _contains_any_normalized(message, ["tiền mặt", "cash", "111"]):
        return "cash"
    if _contains_any_normalized(message, ["chưa thanh toán", "trả sau", "công nợ", "chua tra"]):
        return "payable"
    return "unknown"


def _choose_credit_account(rule: Dict[str, Any], payment_method: str, message: str) -> str:
    options = [str(x) for x in rule.get("credit_account_options", [])]
    if payment_method == "bank" and "112" in options:
        return "112"
    if payment_method == "cash" and "111" in options:
        return "111"
    if payment_method == "payable" and "331" in options:
        return "331"
    if _contains_any_normalized(message, ["khách", "khach", "phải thu", "phai thu"]) and "131" in options:
        return "131"
    return str(rule.get("default_credit_account") or (options[0] if options else "112"))


def _match_accounting_rule(message: str) -> Optional[Dict[str, Any]]:
    rules = _load_json_file(ACCOUNTING_RULES_PATH, [])
    best: Optional[Dict[str, Any]] = None
    best_score = 0
    norm_msg = _norm_text(message)
    for rule in rules:
        score = 0
        for kw in rule.get("keywords", []):
            nkw = _norm_text(str(kw))
            if nkw and nkw in norm_msg:
                score += max(1, len(nkw.split()))
        if score > best_score:
            best = rule
            best_score = score
    return best if best_score > 0 else None


def _detect_intent_v41_smart(message: str) -> str:
    msg = _norm_text(message)
    msg_words = set(re.findall(r"\b[a-z0-9]+\b", msg))
    if msg in {"xin chao", "chao", "chao ban", "hello", "hi", "alo"} or msg_words & {"hello", "alo"}:
        return "greeting"
    if _contains_any_normalized(msg, ["ban lam duoc gi", "giup toi", "huong dan", "chuc nang gi", "lam duoc gi"]):
        return "help"
    if _contains_any_normalized(msg, ["rui ro", "hop le", "hoa don", "chung tu", "thieu hoa don", "tien mat 30 trieu"]):
        return "tax_risk_check"
    if _contains_any_normalized(msg, ["la gi", "dung khi nao", "khau tru", "quy trinh", "tai khoan 641", "tai khoan 642", "tai khoan 131"]):
        return "knowledge_question"
    if _match_accounting_rule(message):
        if _contains_any_normalized(msg, ["but toan", "dinh khoan", "no co", "mua", "ban", "thu", "tra", "thanh toan", "chi"]):
            return "journal_entry"
        return "transaction_classification"
    return "general"


def _rule_based_transaction_answer(message: str) -> Dict[str, Any]:
    rule = _match_accounting_rule(message)
    amount = _extract_amount_from_text(message)
    payment_method = _detect_payment_method_from_text(message)
    if not rule:
        return {
            "matched": False,
            "answer": "Mình chưa nhận ra nghiệp vụ này. Bạn hãy bổ sung: mô tả giao dịch, số tiền, thanh toán tiền mặt/chuyển khoản, có hóa đơn hay chưa.",
            "called": ["accounting_rules.json"],
        }
    debit = str(rule.get("debit_account", "642"))
    credit = _choose_credit_account(rule, payment_method, message)
    amount_text = f"{amount:,.0f}đ" if amount is not None else "chưa rõ"
    confidence = float(rule.get("confidence", 0.75)) * 100
    lines = [
        f"Loại giao dịch: {rule.get('category', 'Chưa phân loại')}",
        f"Số tiền nhận diện: {amount_text}",
        f"Gợi ý bút toán: Nợ {debit} / Có {credit}",
        f"Diễn giải Nợ: {rule.get('debit_name', 'Tài khoản phù hợp theo nghiệp vụ')}",
        f"Cảnh báo nghiệp vụ/thuế: {rule.get('risk', 'Kế toán cần kiểm tra chứng từ trước khi ghi sổ.')}",
        f"Độ tin cậy rule: {confidence:.0f}%",
        "Trạng thái: Chờ kế toán xác nhận, chưa tự ghi sổ.",
    ]
    if amount is not None and payment_method == "cash" and amount >= 5_000_000:
        lines.append("Cảnh báo thêm: giao dịch tiền mặt giá trị lớn, cần kiểm tra điều kiện chứng từ và thanh toán trước khi hạch toán chính thức.")
    return {"matched": True, "answer": "\n".join(lines), "called": ["accounting_rules.json"], "rule_id": rule.get("rule_id")}


def _knowledge_base_answer(message: str) -> Dict[str, Any]:
    if not KNOWLEDGE_BASE_DIR.exists():
        return {"answer": "Chưa có thư mục knowledge_base để tra cứu.", "called": []}
    query_words = {w for w in re.findall(r"[a-zA-Z0-9À-ỹ]+", _norm_text(message)) if len(w) >= 3}
    best_para = ""
    best_file = ""
    best_score = 0
    for file in KNOWLEDGE_BASE_DIR.glob("*.md"):
        text = file.read_text(encoding="utf-8", errors="ignore")
        paras = [p.strip() for p in re.split(r"\n\s*\n", text) if p.strip()]
        for para in paras:
            words = {w for w in re.findall(r"[a-zA-Z0-9À-ỹ]+", _norm_text(para)) if len(w) >= 3}
            score = len(query_words & words)
            if score > best_score:
                best_para = para
                best_file = file.name
                best_score = score
    if best_para and best_score >= 1:
        return {"answer": f"Theo kho tri thức `{best_file}`:\n{best_para}", "called": [f"knowledge_base/{best_file}"]}
    return {
        "answer": "Mình chưa tìm thấy đoạn kiến thức đủ sát trong knowledge_base. Bạn nên bổ sung tài liệu hoặc hỏi cụ thể hơn, ví dụ: 'Tài khoản 641 dùng khi nào?' hoặc 'VAT đầu vào được khấu trừ khi nào?'.",
        "called": ["knowledge_base/*.md"],
    }

@router.get("/v41/chatbot-ui")
def v41_chatbot_ui():
    path = FRONTEND_DIR / "v41_chatbot.html"
    if not path.exists():
        raise HTTPException(status_code=404, detail="Không tìm thấy frontend/v41_chatbot.html")
    return FileResponse(path, media_type="text/html")


@router.post("/ai/v41/chat")
def v41_chat(req: ChatRequest):
    store = _load_store()
    msg = req.message.lower().strip()
    data_ready = len(store.get("journal_entries", [])) > 0
    summary = _chat_summary(store)
    answer_lines: List[str] = []
    called: List[str] = []
    intent = "general"

    smart_intent = _detect_intent_v41_smart(req.message)

    if smart_intent == "greeting":
        intent = "greeting"
        answer_lines.append("Xin chào! Tôi là trợ lý AI kế toán Finiip. Tôi có thể giúp bạn phân loại giao dịch, gợi ý bút toán, kiểm tra rủi ro thuế/chứng từ, xem VAT, lợi nhuận và dashboard tổng hợp. Bạn muốn làm gì hôm nay?")
    elif smart_intent == "help":
        intent = "help"
        called.append("ai_intents.json")
        answer_lines.append("Tôi có thể hỗ trợ 6 nhóm việc:")
        answer_lines.append("1. Phân loại giao dịch: 'tiền điện 2 triệu', 'quảng cáo Facebook 5 triệu'.")
        answer_lines.append("2. Gợi ý bút toán Nợ/Có: 'mua máy tính 20 triệu bằng chuyển khoản'.")
        answer_lines.append("3. Kiểm tra rủi ro thuế/chứng từ: 'chi tiền mặt 30 triệu có rủi ro không?'.")
        answer_lines.append("4. Xem báo cáo VAT, doanh thu, chi phí, lợi nhuận từ dữ liệu đã ghi sổ.")
        answer_lines.append("5. Hỏi kiến thức kế toán: 'Tài khoản 641 dùng khi nào?'.")
        answer_lines.append("6. Hướng dẫn import dữ liệu mẫu hoặc xuất Excel.")
    elif smart_intent in {"transaction_classification", "journal_entry"}:
        intent = smart_intent
        result = _rule_based_transaction_answer(req.message)
        called.extend(result.get("called", []))
        answer_lines.append(result["answer"])
    elif smart_intent == "tax_risk_check":
        intent = "tax_risk_check"
        result = _rule_based_transaction_answer(req.message)
        called.extend(result.get("called", []))
        if result.get("matched"):
            answer_lines.append(result["answer"])
        else:
            amount = _extract_amount_from_text(req.message)
            payment_method = _detect_payment_method_from_text(req.message)
            if amount is not None and payment_method == "cash" and amount >= 5_000_000:
                called.append("generic_tax_risk_rule")
                answer_lines.append(f"Cảnh báo: giao dịch tiền mặt {amount:,.0f}đ là giao dịch giá trị lớn. Kế toán cần kiểm tra hóa đơn/chứng từ, mục đích chi, người nhận tiền và điều kiện thanh toán trước khi hạch toán chính thức.")
            else:
                answer_lines.append("Để kiểm tra rủi ro, tôi cần ít nhất: mô tả giao dịch, số tiền, phương thức thanh toán và tình trạng hóa đơn/chứng từ.")
        if _contains_any_normalized(req.message, ["không có hóa đơn", "thieu hoa don", "thiếu hóa đơn", "hoa don"]):
            kb = _knowledge_base_answer("thiếu hóa đơn chứng từ")
            called.extend(kb.get("called", []))
            answer_lines.append(kb["answer"])
    elif smart_intent == "knowledge_question":
        intent = "knowledge_question"
        kb = _knowledge_base_answer(req.message)
        called.extend(kb.get("called", []))
        answer_lines.append(kb["answer"])
    elif any(k in msg for k in ["vat", "thuế", "thue"]):
        intent = "vat_report"
        vat = summary["vat"]
        called.append("GET /reports/vat")
        answer_lines.append(f"VAT đầu ra: {vat['vat_output']:,.0f}đ")
        answer_lines.append(f"VAT đầu vào: {vat['vat_input']:,.0f}đ")
        answer_lines.append(f"VAT phải nộp: {vat['vat_payable']:,.0f}đ")
    elif any(k in msg for k in ["lãi", "lỗ", "lợi nhuận", "doanh thu", "chi phí", "ket qua", "kết quả"]):
        intent = "income_statement"
        income = summary["income"]
        called.append("GET /reports/income-statement")
        answer_lines.append(f"Doanh thu: {income['revenue']:,.0f}đ")
        answer_lines.append(f"Chi phí: {income['expenses']:,.0f}đ")
        answer_lines.append(f"Lợi nhuận trước thuế: {income['profit_before_tax']:,.0f}đ")
    elif any(k in msg for k in ["dashboard", "tổng hợp", "bao cao", "báo cáo", "tổng quan"]):
        intent = "summary_report"
        cards = summary["dashboard"]["cards"]
        called.extend(["GET /dashboard/v30/financial", "GET /reports/vat", "GET /reports/income-statement"])
        answer_lines.append("Báo cáo tổng hợp hiện tại:")
        answer_lines.append(f"Doanh thu: {cards['revenue']:,.0f}đ")
        answer_lines.append(f"Chi phí: {cards['expenses']:,.0f}đ")
        answer_lines.append(f"Lợi nhuận: {cards['profit_before_tax']:,.0f}đ")
        answer_lines.append(f"VAT phải nộp: {cards['vat_payable']:,.0f}đ")
        answer_lines.append(f"Bút toán đã ghi sổ: {cards['journal_entries']}")
        answer_lines.append(f"Draft đang chờ: {cards['drafts_pending']}")
    elif any(k in msg for k in ["excel", "xuất", "xuat"]):
        intent = "export_excel"
        called.append("GET /reports/v27/export-excel")
        answer_lines.append("Bạn có thể tải file Excel tại endpoint /reports/v27/export-excel.")
    elif any(k in msg for k in ["dữ liệu mẫu", "du lieu mau", "sample"]):
        intent = "seed_sample_hint"
        called.append("POST /import/v40-5/sample-data?auto_post=true")
        answer_lines.append("Để tạo dữ liệu mẫu cho báo cáo/chatbot, gọi POST /import/v40-5/sample-data?auto_post=true.")
    else:
        answer_lines.append("Mình có thể làm báo cáo VAT, báo cáo lợi nhuận, dashboard tổng hợp, xuất Excel, xem draft hoặc hướng dẫn import dữ liệu.")

    if not data_ready and intent in {"vat_report", "income_statement", "summary_report"}:
        answer_lines.append("Lưu ý: hiện chưa có bút toán đã ghi sổ, nên số liệu có thể đang bằng 0. Hãy import dữ liệu hoặc tạo sample data trước.")

    _audit(store, "chat", "v41", _next_id(store, "CHAT"), {"message": req.message, "intent": intent, "called": called})
    _save_store(store)
    return {"version": "V41", "intent": intent, "answer": "\n".join(answer_lines), "called_tools": called, "need_confirmation": False}


def _proposal_from_message_v42(message: str) -> Dict[str, Any]:
    """Turn a free-text transaction into a structured proposal for UI confirmation."""
    rule = _match_accounting_rule(message)
    amount = _extract_amount_from_text(message)
    payment_method = _detect_payment_method_from_text(message)
    if payment_method == "unknown":
        payment_method = "bank"
    proposal: Dict[str, Any] = {
        "description": message.strip(),
        "amount": amount,
        "payment_method": payment_method,
        "vat_rate": 0.1,
        "category": None,
        "debit_account": None,
        "credit_account": None,
        "risk_note": "Kế toán cần kiểm tra chứng từ trước khi ghi sổ.",
        "confidence": 0.45,
        "matched_rule_id": None,
    }
    if rule:
        proposal.update({
            "category": rule.get("category"),
            "debit_account": str(rule.get("debit_account", "642")),
            "credit_account": _choose_credit_account(rule, payment_method, message),
            "risk_note": rule.get("risk") or proposal["risk_note"],
            "confidence": float(rule.get("confidence", 0.75)),
            "matched_rule_id": rule.get("rule_id"),
        })
    return proposal


@router.post("/ai/v42/transaction-proposal")
def v42_transaction_proposal(req: ChatRequest):
    proposal = _proposal_from_message_v42(req.message)
    answer = _rule_based_transaction_answer(req.message)
    return {
        "version": "V42",
        "intent": _detect_intent_v41_smart(req.message),
        "proposal": proposal,
        "answer": answer.get("answer"),
        "called_tools": answer.get("called", ["accounting_rules.json"]),
        "next_step": "POST /ai/v42/confirm-journal để lưu bút toán nháp sau khi người dùng xác nhận.",
    }


@router.post("/ai/v42/confirm-journal")
def v42_confirm_journal(req: ConfirmJournalRequest):
    store = _load_store()
    amount = req.amount if req.amount is not None else _extract_amount_from_text(req.description)
    if amount is None:
        raise HTTPException(status_code=400, detail="Chưa nhận diện được số tiền. Vui lòng gửi amount hoặc viết số tiền trong description.")
    payment_method = req.payment_method or _detect_payment_method_from_text(req.description)
    if payment_method == "unknown":
        payment_method = "bank"
    draft_req = JournalDraftCreate(
        description=req.description,
        amount=float(amount),
        vat_rate=req.vat_rate,
        payment_method=payment_method,
        category=req.category,
        debit_account=req.debit_account,
        credit_account=req.credit_account,
        source="v42_confirm_journal",
    )
    lines = _build_lines(draft_req)
    errors = _detect_errors(lines, req.description)
    draft_id = _next_id(store, "DRAFT")
    draft = {
        "id": draft_id,
        "version": "V42",
        "created_at": _now(),
        "updated_at": _now(),
        "status": "approved" if req.post_immediately else "draft",
        "source": "v42_confirm_journal",
        "description": req.description,
        "amount": float(amount),
        "vat_rate": req.vat_rate,
        "payment_method": payment_method,
        "category": req.category,
        "lines": lines,
        "debit_total": _line_total(lines, "debit"),
        "credit_total": _line_total(lines, "credit"),
        "balanced": _line_total(lines, "debit") == _line_total(lines, "credit"),
        "risk_flags": errors,
        "risk_note": req.risk_note,
        "user_correction": req.user_correction,
        "safe_policy": "AI chỉ lưu nháp/approved theo xác nhận. Kế toán chịu trách nhiệm kiểm tra trước khi ghi sổ chính thức.",
    }
    store.setdefault("journal_drafts", []).append(draft)
    store.setdefault("ai_feedback", []).append({
        "id": _next_id(store, "FB"),
        "created_at": _now(),
        "user_message": req.description,
        "ai_intent": "confirm_journal",
        "ai_prediction": {"category": req.category, "debit_account": req.debit_account, "credit_account": req.credit_account, "amount": amount},
        "user_correction": req.user_correction,
        "final_result": {"draft_id": draft_id, "status": draft["status"], "lines": lines},
        "rating": None,
        "note": req.risk_note,
    })
    _remember_learning(store, source="v42_confirm_journal", description=req.description, lines=lines, label="confirmed", note=req.user_correction)
    _audit(store, "confirm_journal", "journal_draft", draft_id, {"post_immediately": req.post_immediately})

    posted_entry = None
    if req.post_immediately:
        post_errors = _detect_errors(lines, req.description)
        if any(e["code"] == "UNBALANCED" for e in post_errors):
            raise HTTPException(status_code=400, detail={"message": "Bút toán chưa cân", "errors": post_errors})
        entry_id = _next_id(store, "JE")
        posted_entry = {
            "id": entry_id,
            "version": "V42",
            "posted_at": _now(),
            "description": req.description,
            "source_draft_id": draft_id,
            "lines": lines,
            "debit_total": draft["debit_total"],
            "credit_total": draft["credit_total"],
            "note": "Posted immediately by V42 confirm-journal",
        }
        store.setdefault("journal_entries", []).append(posted_entry)
        draft["status"] = "posted"
        draft["posted_entry_id"] = entry_id
        _audit(store, "post", "journal_entry", entry_id, {"source_draft_id": draft_id})

    _save_store(store)
    return {
        "version": "V42",
        "status": "saved",
        "message": "Đã lưu bút toán nháp chờ kế toán kiểm tra." if not posted_entry else "Đã lưu và ghi sổ bút toán theo xác nhận.",
        "journal_draft": draft,
        "posted_entry": posted_entry,
        "next_steps": ["GET /ai/v25/journal-draft/list", "POST /ai/v25/journal-draft/{draft_id}/approve", "POST /ledger/post-entry"],
    }


@router.post("/ai/v47/feedback")
def v47_save_ai_feedback(req: AIFeedbackRequest):
    store = _load_store()
    feedback = {
        "id": _next_id(store, "FB"),
        "created_at": _now(),
        "user_message": req.user_message,
        "ai_intent": req.ai_intent,
        "ai_prediction": req.ai_prediction,
        "user_correction": req.user_correction,
        "final_result": req.final_result,
        "rating": req.rating,
        "note": req.note,
    }
    store.setdefault("ai_feedback", []).append(feedback)
    _audit(store, "save_feedback", "ai_feedback", feedback["id"], {"rating": req.rating})
    _save_store(store)
    return {"version": "V47", "status": "saved", "feedback": feedback, "message": "Đã lưu feedback để cải thiện rule/ML sau này."}


@router.get("/ai/v47/feedback/list")
def v47_list_ai_feedback(limit: int = 50):
    store = _load_store()
    items = list(reversed(store.get("ai_feedback", [])))[: max(1, min(limit, 200))]
    return {"version": "V47", "count": len(store.get("ai_feedback", [])), "items": items}


@router.get("/v43/working-chatbot-ui")
def v43_working_chatbot_ui():
    path = FRONTEND_DIR / "v43_working_chatbot.html"
    if not path.exists():
        raise HTTPException(status_code=404, detail="Không tìm thấy frontend/v43_working_chatbot.html")
    return FileResponse(path, media_type="text/html")


@router.post("/ai/v42/chat-action")
def v42_chat_action(req: ChatRequest):
    store = _load_store()
    msg = req.message.lower().strip()
    intent = "unknown"
    plan: Dict[str, Any] = {"steps": []}
    if any(k in msg for k in ["backup", "sao lưu", "sao luu"]):
        intent = "backup"
        plan = {"steps": ["Tạo file backup dữ liệu V25-V42", "Ghi audit log", "Trả link download"], "endpoint": "POST /backup/v40/create"}
    elif any(k in msg for k in ["post", "ghi sổ", "ghi so"]):
        intent = "post_approved_drafts"
        approved = [d for d in store.get("journal_drafts", []) if d.get("status") == "approved"]
        plan = {"steps": [f"Tìm {len(approved)} draft đã approved", "Ghi từng draft vào journal_entries", "Đổi trạng thái draft sang posted"], "draft_ids": [d.get("id") for d in approved]}
    elif any(k in msg for k in ["duyệt hết", "approve all", "duyet het"]):
        intent = "approve_all_drafts"
        drafts = [d for d in store.get("journal_drafts", []) if d.get("status") == "draft" and not d.get("risk_flags")]
        plan = {"steps": [f"Duyệt {len(drafts)} draft an toàn không có risk flag"], "draft_ids": [d.get("id") for d in drafts]}
    else:
        return v41_chat(req)

    confirmation_id = _next_id(store, "CONFIRM")
    store.setdefault("chat_confirmations", []).append({"id": confirmation_id, "created_at": _now(), "intent": intent, "plan": plan, "status": "pending", "message": req.message})
    _audit(store, "prepare_action", "v42_confirmation", confirmation_id, {"intent": intent})
    _save_store(store)
    return {"version": "V42", "intent": intent, "need_confirmation": True, "confirmation_id": confirmation_id, "plan": plan, "answer": "Hành động này có thể thay đổi dữ liệu. Gọi /ai/v42/confirm-action với confirmation_id để xác nhận."}


@router.post("/ai/v42/confirm-action")
def v42_confirm_action(req: ChatActionConfirmRequest):
    store = _load_store()
    confirmation = next((c for c in store.get("chat_confirmations", []) if c.get("id") == req.confirmation_id), None)
    if not confirmation:
        raise HTTPException(status_code=404, detail="Không tìm thấy yêu cầu xác nhận")
    if confirmation.get("status") != "pending":
        raise HTTPException(status_code=400, detail="Yêu cầu này đã được xử lý")
    if not req.confirm:
        confirmation["status"] = "cancelled"
        _audit(store, "cancel_action", "v42_confirmation", req.confirmation_id, {"intent": confirmation.get("intent")})
        _save_store(store)
        return {"version": "V42", "cancelled": True, "confirmation_id": req.confirmation_id}

    intent = confirmation.get("intent")
    result: Dict[str, Any] = {}
    if intent == "backup":
        _save_store(store)
        result = v40_create_backup()
        store = _load_store()
    elif intent == "approve_all_drafts":
        approved = []
        for draft in store.get("journal_drafts", []):
            if draft.get("id") in confirmation.get("plan", {}).get("draft_ids", []):
                draft["status"] = "approved"
                draft["updated_at"] = _now()
                approved.append(draft.get("id"))
                _remember_learning(store, source="v42_approve_all", description=draft.get("description", ""), lines=draft.get("lines", []), label="approved")
        result = {"approved_count": len(approved), "draft_ids": approved}
    elif intent == "post_approved_drafts":
        posted = []
        for draft in store.get("journal_drafts", []):
            if draft.get("id") in confirmation.get("plan", {}).get("draft_ids", []) and draft.get("status") == "approved":
                errors = _detect_errors(draft.get("lines", []), draft.get("description", ""))
                if errors:
                    continue
                entry_id = _next_id(store, "JE")
                entry = {"id": entry_id, "version": "V42/V26+", "posted_at": _now(), "description": draft.get("description"), "source_draft_id": draft.get("id"), "lines": draft.get("lines", []), "debit_total": draft.get("debit_total"), "credit_total": draft.get("credit_total"), "note": "Posted by V42 confirmed chatbot action"}
                store.setdefault("journal_entries", []).append(entry)
                draft["status"] = "posted"
                draft["posted_entry_id"] = entry_id
                posted.append(entry_id)
                _remember_learning(store, source="v42_post", description=draft.get("description", ""), lines=draft.get("lines", []), label="posted")
        result = {"posted_count": len(posted), "entry_ids": posted}
    else:
        raise HTTPException(status_code=400, detail="Intent chưa hỗ trợ")
    confirmation["status"] = "confirmed"
    confirmation["confirmed_at"] = _now()
    confirmation["result"] = result
    _audit(store, "confirm_action", "v42_confirmation", req.confirmation_id, {"intent": intent, "result": result})
    _save_store(store)
    return {"version": "V42", "confirmation_id": req.confirmation_id, "intent": intent, "result": result}


@router.get("/ai/v25-v40/upgrade-status")
def v25_v40_upgrade_status():
    store = _load_store()
    return {
        "stage": "Finiip V25-V40 - Prototype chuyển sang app kế toán AI dùng thử",
        "completed": [
            "V25 Auto Journal Draft",
            "V26 Sổ cái / Nhật ký chung",
            "V27 Xuất Excel",
            "V28 Báo cáo VAT",
            "V29 Báo cáo kết quả kinh doanh",
            "V30 Dashboard tài chính",
            "V31 AI giải thích bút toán",
            "V32 AI phát hiện lỗi kế toán",
            "V33 AI hỏi lại khi thiếu thông tin",
            "V34 OCR hóa đơn nâng cao",
            "V35 Mapping hóa đơn → bút toán",
            "V36 User / Role permission MVP",
            "V37 Login / token MVP",
            "V38 Database health / migration readiness",
            "V39 Audit Log",
            "V40 Backup / Restore",
            "V40 Backup / Restore nâng cấp",
            "V40.5 Import dữ liệu Excel/CSV/bulk + dữ liệu mẫu",
            "V41 Chatbot báo cáo/tính toán từ API",
            "V42 Chatbot hành động có xác nhận",
        ],
        "current_level": "Cấp 5 prototype an toàn + chatbot điều khiển báo cáo/hành động có xác nhận",
        "counts": {
            "journal_drafts": len(store.get("journal_drafts", [])),
            "journal_entries": len(store.get("journal_entries", [])),
            "users": len(store.get("users", [])),
            "audit_logs": len(store.get("audit_logs", [])),
            "raw_transactions": len(store.get("raw_transactions", [])),
            "import_batches": len(store.get("import_batches", [])),
            "learning_examples": len(store.get("learning_examples", [])),
            "chat_confirmations": len(store.get("chat_confirmations", [])),
            "quality_score": _quality_score(store),
        },
        "main_new_apis": [
            "POST /import/v40-5/sample-data?auto_post=true",
            "POST /import/v40-5/transactions/bulk",
            "POST /import/v40-5/transactions/file",
            "GET  /ai/learning/v40-5/memory",
            "POST /ai/v41/chat",
            "POST /ai/v42/transaction-proposal",
            "POST /ai/v42/confirm-journal",
            "POST /ai/v42/chat-action",
            "POST /ai/v42/confirm-action",
            "POST /ai/v47/feedback",
            "GET /v43/working-chatbot-ui"
        ],
        "important_note": "V40.5-V42 đã có import, học từ approve/reject/post và chatbot gọi báo cáo. Đây vẫn là MVP; trước production cần bảo mật, phân quyền thật và kiểm thử kế toán.",
    }

# ---------------------------------------------------------------------------
# V43 + V43.5 + V45 MVP: RAG accounting knowledge, problem solver and Q&A learning
# ---------------------------------------------------------------------------

ACCOUNTING_FORMULAS_V43 = {
    "vat_payable": {
        "name": "VAT phải nộp",
        "formula": "VAT phải nộp = VAT đầu ra - VAT đầu vào",
        "example": "VAT đầu ra 8.000.000, VAT đầu vào 10.000.000 → còn được khấu trừ 2.000.000",
    },
    "profit": {
        "name": "Lợi nhuận",
        "formula": "Lợi nhuận = Doanh thu - Chi phí",
        "example": "Doanh thu 80.000.000, chi phí 50.000.000 → lợi nhuận 30.000.000",
    },
    "gross_profit": {
        "name": "Lợi nhuận gộp",
        "formula": "Lợi nhuận gộp = Doanh thu thuần - Giá vốn hàng bán",
        "example": "Doanh thu 80.000.000, giá vốn 50.000.000 → lãi gộp 30.000.000",
    },
    "straight_line_depreciation": {
        "name": "Khấu hao đường thẳng",
        "formula": "Khấu hao kỳ = Nguyên giá / Thời gian sử dụng",
        "example": "Máy tính 24.000.000 dùng 24 tháng → khấu hao 1.000.000/tháng",
    },
}

ACCOUNTING_ACCOUNTS_V43 = {
    "111": "Tiền mặt",
    "112": "Tiền gửi ngân hàng",
    "131": "Phải thu của khách hàng",
    "1331": "Thuế GTGT được khấu trừ của hàng hóa, dịch vụ",
    "152": "Nguyên liệu, vật liệu",
    "153": "Công cụ, dụng cụ",
    "156": "Hàng hóa",
    "211": "Tài sản cố định hữu hình",
    "214": "Hao mòn tài sản cố định",
    "242": "Chi phí trả trước",
    "331": "Phải trả cho người bán",
    "3331": "Thuế GTGT phải nộp",
    "511": "Doanh thu bán hàng và cung cấp dịch vụ",
    "632": "Giá vốn hàng bán",
    "641": "Chi phí bán hàng",
    "642": "Chi phí quản lý doanh nghiệp",
}

class V43DocumentUpload(BaseModel):
    title: str = Field(..., min_length=2)
    content: str = Field(..., min_length=5)
    source: str = "manual"
    tags: List[str] = []

class V43SearchRequest(BaseModel):
    query: str = Field(..., min_length=2)
    limit: int = Field(5, ge=1, le=20)

class V43AccountingQARequest(BaseModel):
    question: str = Field(..., min_length=2)
    use_rag: bool = True
    save_learning: bool = True

class V435ProblemRequest(BaseModel):
    question: str = Field(..., min_length=5)
    standard: str = "TT200"
    mode: str = "step_by_step"
    save_learning: bool = True

class V435CheckAnswerRequest(BaseModel):
    question: str
    user_answer: str
    standard: str = "TT200"


def _normalize_v43(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").lower()).strip()


def _tokenize_v43(text: str) -> List[str]:
    tokens = re.findall(r"[a-zA-ZÀ-ỹ0-9]+", _normalize_v43(text))
    stop = {"la", "là", "cua", "của", "the", "thi", "thì", "va", "và", "cho", "toi", "tôi", "hoi", "hỏi", "ve", "về", "nay", "này", "gi", "gì", "tài", "tai", "khoản", "khoan"}
    return [t for t in tokens if len(t) > 1 and t not in stop]


def _money_values_v43(text: str) -> List[float]:
    values: List[float] = []
    # 100 triệu, 20 tr, 1.5 tỷ
    for raw, unit in re.findall(r"(\d+(?:[\.,]\d+)?)\s*(tỷ|ty|triệu|tr|nghìn|ngàn|k)?", text.lower()):
        try:
            num = float(raw.replace(".", "").replace(",", "."))
        except ValueError:
            continue
        mult = 1.0
        if unit in {"tỷ", "ty"}:
            mult = 1_000_000_000
        elif unit in {"triệu", "tr"}:
            mult = 1_000_000
        elif unit in {"nghìn", "ngàn", "k"}:
            mult = 1_000
        # Ignore common account codes and percentages in plain numbers
        if mult > 1 or num >= 1000:
            values.append(num * mult)
    return values


def _vat_rate_v43(text: str) -> float:
    m = re.search(r"vat\s*(\d+(?:[\.,]\d+)?)\s*%", text.lower()) or re.search(r"gtgt\s*(\d+(?:[\.,]\d+)?)\s*%", text.lower())
    if not m:
        m = re.search(r"(\d+(?:[\.,]\d+)?)\s*%", text.lower())
    if not m:
        return 0.1
    try:
        return float(m.group(1).replace(",", ".")) / 100
    except ValueError:
        return 0.1


def _payment_credit_account_v43(text: str, default_payable: str = "331") -> str:
    msg = _normalize_v43(text)
    if any(k in msg for k in ["chuyển khoản", "ngân hàng", "bank", "112"]):
        return "112"
    if any(k in msg for k in ["tiền mặt", "cash", "111"]):
        return "111"
    if any(k in msg for k in ["chưa thanh toán", "công nợ", "phải trả", "331"]):
        return "331"
    return default_payable


def _payment_debit_account_v43(text: str) -> str:
    msg = _normalize_v43(text)
    if any(k in msg for k in ["chưa trả", "khách hàng chưa", "phải thu", "131"]):
        return "131"
    if any(k in msg for k in ["chuyển khoản", "ngân hàng", "bank", "112"]):
        return "112"
    if any(k in msg for k in ["tiền mặt", "cash", "111"]):
        return "111"
    return "131"


def _append_qa_learning(store: Dict[str, Any], question: str, answer: str, intent: str, sources: Optional[List[Dict[str, Any]]] = None) -> None:
    store.setdefault("qa_learning_examples", []).append({
        "id": _next_id(store, "QA"),
        "created_at": _now(),
        "intent": intent,
        "question": question,
        "answer_preview": answer[:800],
        "sources": sources or [],
    })
    _audit(store, "learn_qa", "qa_learning", "latest", {"intent": intent, "question": question[:120]})


def _rag_search_v43(store: Dict[str, Any], query: str, limit: int = 5) -> List[Dict[str, Any]]:
    q_tokens = set(_tokenize_v43(query))
    results: List[Dict[str, Any]] = []
    for doc in store.setdefault("rag_documents", []):
        content = doc.get("content", "")
        title = doc.get("title", "")
        d_tokens = set(_tokenize_v43(title + " " + content))
        overlap = len(q_tokens & d_tokens)
        bonus = 2 if any(t in _normalize_v43(title) for t in q_tokens) else 0
        score = overlap + bonus
        if score <= 0:
            continue
        # Extract small snippets around first matched token
        norm_content = content
        snippet = norm_content[:500]
        for tok in q_tokens:
            idx = _normalize_v43(norm_content).find(tok)
            if idx >= 0:
                start = max(0, idx - 160)
                end = min(len(norm_content), idx + 360)
                snippet = norm_content[start:end]
                break
        results.append({
            "document_id": doc.get("id"),
            "title": title,
            "source": doc.get("source"),
            "score": score,
            "snippet": snippet,
            "tags": doc.get("tags", []),
        })
    results.sort(key=lambda x: x["score"], reverse=True)
    return results[:limit]


def _answer_formula_or_account_v43(question: str) -> Optional[Dict[str, Any]]:
    msg = _normalize_v43(question)
    if any(k in msg for k in ["công thức", "cong thuc", "tính", "tinh"]):
        if "vat" in msg or "gtgt" in msg:
            f = ACCOUNTING_FORMULAS_V43["vat_payable"]
            return {"intent": "formula", "answer": f"{f['name']}: {f['formula']}. Ví dụ: {f['example']}", "data": f}
        if any(k in msg for k in ["lợi nhuận gộp", "loi nhuan gop", "giá vốn", "gia von"]):
            f = ACCOUNTING_FORMULAS_V43["gross_profit"]
            return {"intent": "formula", "answer": f"{f['name']}: {f['formula']}. Ví dụ: {f['example']}", "data": f}
        if any(k in msg for k in ["lợi nhuận", "loi nhuan", "lãi", "lai"]):
            f = ACCOUNTING_FORMULAS_V43["profit"]
            return {"intent": "formula", "answer": f"{f['name']}: {f['formula']}. Ví dụ: {f['example']}", "data": f}
        if any(k in msg for k in ["khấu hao", "khau hao"]):
            f = ACCOUNTING_FORMULAS_V43["straight_line_depreciation"]
            return {"intent": "formula", "answer": f"{f['name']}: {f['formula']}. Ví dụ: {f['example']}", "data": f}
    for code, name in ACCOUNTING_ACCOUNTS_V43.items():
        if re.search(rf"\b{re.escape(code)}\b", msg):
            return {"intent": "account_lookup", "answer": f"Tài khoản {code}: {name}.", "data": {"code": code, "name": name}}
    if any(k in msg for k in ["quảng cáo", "quang cao", "facebook", "marketing", "bán hàng", "ban hang"]):
        return {"intent": "account_lookup", "answer": "Tài khoản thường dùng là 641: Chi phí bán hàng. Nếu đây là chi phí quản lý chung, có thể cân nhắc 642 theo quy định nội bộ.", "data": {"code": "641", "name": ACCOUNTING_ACCOUNTS_V43["641"]}}
    if any(k in msg for k in ["điện", "dien", "nước", "nuoc", "văn phòng", "van phong", "quản lý", "quan ly"]):
        return {"intent": "account_lookup", "answer": "Tài khoản thường dùng là 642: Chi phí quản lý doanh nghiệp.", "data": {"code": "642", "name": ACCOUNTING_ACCOUNTS_V43["642"]}}
    if any(k in msg for k in ["tài khoản", "tai khoan", "tk "]):
        q_tokens = set(_tokenize_v43(msg))
        best = None
        best_score = 0
        for code, name in ACCOUNTING_ACCOUNTS_V43.items():
            name_tokens = set(_tokenize_v43(name))
            score = len(q_tokens & name_tokens)
            if score > best_score:
                best = (code, name)
                best_score = score
        if best and best_score >= 2:
            code, name = best
            return {"intent": "account_lookup", "answer": f"Tài khoản thường dùng cho nội dung này là {code}: {name}.", "data": {"code": code, "name": name}}
    return None


def _solve_accounting_problem_v43(question: str) -> Dict[str, Any]:
    q = _normalize_v43(question)
    amounts = _money_values_v43(question)
    vat_rate = _vat_rate_v43(question)
    steps: List[Dict[str, Any]] = []
    calculations: Dict[str, Any] = {"vat_rate": vat_rate}
    assumptions: List[str] = []
    warnings: List[str] = []

    purchase_amount = amounts[0] if amounts else 0.0
    sale_amount = 0.0
    if len(amounts) >= 2:
        sale_amount = amounts[1]

    # Try to detect cost of goods sold when selling half/one half.
    sale_ratio = 1.0
    if any(k in q for k in ["một nửa", "mot nua", "1/2", "50%"]):
        sale_ratio = 0.5
    elif m := re.search(r"bán\s+(\d+(?:[\.,]\d+)?)\s*%", q):
        try:
            sale_ratio = float(m.group(1).replace(",", ".")) / 100
        except ValueError:
            sale_ratio = 1.0

    if purchase_amount:
        purchase_vat = round(purchase_amount * vat_rate, 2)
        credit = _payment_credit_account_v43(question, default_payable="331")
        asset_account = "156"
        if any(k in q for k in ["máy tính", "may tinh", "thiết bị", "thiet bi", "tài sản", "tai san"]):
            asset_account = "211" if purchase_amount >= 30_000_000 else "242"
        steps.append({
            "step": "Mua hàng/tài sản",
            "journal_entry": [
                {"debit": asset_account, "account_name": ACCOUNTING_ACCOUNTS_V43.get(asset_account), "amount": purchase_amount},
                {"debit": "1331", "account_name": ACCOUNTING_ACCOUNTS_V43["1331"], "amount": purchase_vat},
                {"credit": credit, "account_name": ACCOUNTING_ACCOUNTS_V43.get(credit), "amount": purchase_amount + purchase_vat},
            ],
            "explanation": f"Ghi nhận giá trị mua chưa VAT và VAT đầu vào; Có {credit} theo phương thức thanh toán/công nợ.",
        })
        calculations["vat_input"] = purchase_vat

    if sale_amount or "bán" in q or "ban" in q:
        if not sale_amount:
            warnings.append("Đề có nghiệp vụ bán nhưng chưa thấy rõ giá bán; cần bổ sung để tính doanh thu/VAT đầu ra.")
        sale_vat = round(sale_amount * vat_rate, 2)
        debit = _payment_debit_account_v43(question)
        steps.append({
            "step": "Bán hàng/ghi nhận doanh thu",
            "journal_entry": [
                {"debit": debit, "account_name": ACCOUNTING_ACCOUNTS_V43.get(debit), "amount": sale_amount + sale_vat},
                {"credit": "511", "account_name": ACCOUNTING_ACCOUNTS_V43["511"], "amount": sale_amount},
                {"credit": "3331", "account_name": ACCOUNTING_ACCOUNTS_V43["3331"], "amount": sale_vat},
            ],
            "explanation": "Ghi nhận doanh thu bán hàng và VAT đầu ra.",
        })
        calculations["vat_output"] = sale_vat
        if purchase_amount:
            cogs = round(purchase_amount * sale_ratio, 2)
            steps.append({
                "step": "Ghi nhận giá vốn",
                "journal_entry": [
                    {"debit": "632", "account_name": ACCOUNTING_ACCOUNTS_V43["632"], "amount": cogs},
                    {"credit": "156", "account_name": ACCOUNTING_ACCOUNTS_V43["156"], "amount": cogs},
                ],
                "explanation": f"Giả định giá vốn theo tỷ lệ hàng đã bán: {sale_ratio:.0%} của giá mua.",
            })
            calculations["cost_of_goods_sold"] = cogs
            calculations["gross_profit"] = round(sale_amount - cogs, 2)

    vat_payable = round(calculations.get("vat_output", 0) - calculations.get("vat_input", 0), 2)
    calculations["vat_payable"] = vat_payable
    if not purchase_amount and not sale_amount:
        assumptions.append("Chưa phát hiện được số tiền rõ ràng trong đề; module sẽ trả lời dạng hướng dẫn tổng quát.")
    if "chưa" not in q and "ngân hàng" not in q and "tiền mặt" not in q and "bank" not in q:
        assumptions.append("Đề chưa nói rõ phương thức thanh toán; với mua hàng mặc định là công nợ 331, với bán hàng mặc định là phải thu 131.")

    total_debit = 0.0
    total_credit = 0.0
    for step in steps:
        for line in step.get("journal_entry", []):
            if "debit" in line:
                total_debit += float(line.get("amount") or 0)
            if "credit" in line:
                total_credit += float(line.get("amount") or 0)

    answer_parts = ["Mình đã tách đề thành các nghiệp vụ và giải từng bước:"]
    for idx, step in enumerate(steps, start=1):
        answer_parts.append(f"\n{idx}. {step['step']}")
        for line in step.get("journal_entry", []):
            if "debit" in line:
                answer_parts.append(f"- Nợ {line['debit']} {line.get('account_name') or ''}: {line.get('amount'):,.0f}đ")
            if "credit" in line:
                answer_parts.append(f"- Có {line['credit']} {line.get('account_name') or ''}: {line.get('amount'):,.0f}đ")
        answer_parts.append(f"Giải thích: {step['explanation']}")
    if calculations:
        answer_parts.append("\nTính toán chính:")
        if "vat_input" in calculations:
            answer_parts.append(f"- VAT đầu vào: {calculations['vat_input']:,.0f}đ")
        if "vat_output" in calculations:
            answer_parts.append(f"- VAT đầu ra: {calculations['vat_output']:,.0f}đ")
        answer_parts.append(f"- VAT phải nộp = VAT đầu ra - VAT đầu vào = {vat_payable:,.0f}đ")
        if "gross_profit" in calculations:
            answer_parts.append(f"- Lợi nhuận gộp = Doanh thu - Giá vốn = {calculations['gross_profit']:,.0f}đ")
    if assumptions:
        answer_parts.append("\nGiả định: " + " ".join(assumptions))
    if warnings:
        answer_parts.append("\nCần kiểm tra thêm: " + " ".join(warnings))

    return {
        "steps": steps,
        "calculations": calculations,
        "assumptions": assumptions,
        "warnings": warnings,
        "check": {"total_debit": total_debit, "total_credit": total_credit, "balanced": round(total_debit, 2) == round(total_credit, 2)},
        "answer": "\n".join(answer_parts),
    }


@router.get("/ai/v43/status")
def v43_status():
    store = _load_store()
    return {
        "version": "V43/V43.5/V45 MVP",
        "completed": [
            "V43 Accounting formulas + account knowledge base",
            "V43 RAG tài liệu kế toán/luật/quy trình nội bộ dạng keyword retrieval MVP",
            "V43.5 Accounting Problem Solver cho câu hỏi dài/bài tập",
            "V45 Q&A learning memory từ câu hỏi, đáp án và nguồn được dùng",
        ],
        "counts": {
            "rag_documents": len(store.setdefault("rag_documents", [])),
            "qa_learning_examples": len(store.setdefault("qa_learning_examples", [])),
            "formulas": len(ACCOUNTING_FORMULAS_V43),
            "accounts": len(ACCOUNTING_ACCOUNTS_V43),
        },
        "main_apis": [
            "POST /rag/v43/documents/upload",
            "POST /rag/v43/documents/upload-file",
            "POST /rag/v43/search",
            "POST /ai/v43/accounting-qa",
            "POST /ai/v43-5/problem-solver",
            "POST /ai/v43-5/check-answer",
            "GET /ai/v45/qa-learning/memory",
        ],
    }


@router.get("/ai/v43/formulas")
def v43_formulas():
    return {"version": "V43", "formulas": ACCOUNTING_FORMULAS_V43}


@router.get("/ai/v43/accounts")
def v43_accounts():
    return {"version": "V43", "accounts": ACCOUNTING_ACCOUNTS_V43}


@router.post("/rag/v43/documents/upload")
def v43_upload_document(req: V43DocumentUpload):
    store = _load_store()
    doc_id = _next_id(store, "DOC")
    doc = {"id": doc_id, "title": req.title, "content": req.content, "source": req.source, "tags": req.tags, "created_at": _now()}
    store.setdefault("rag_documents", []).append(doc)
    _audit(store, "upload", "rag_document", doc_id, {"title": req.title, "source": req.source})
    _save_store(store)
    return {"version": "V43", "document": {k: v for k, v in doc.items() if k != "content"}, "content_length": len(req.content)}


@router.post("/rag/v43/documents/upload-file")
async def v43_upload_document_file(file: UploadFile = File(...), source: str = "upload", tags: str = ""):
    raw = await file.read()
    try:
        content = raw.decode("utf-8")
    except UnicodeDecodeError:
        content = raw.decode("utf-8", errors="ignore")
    if not content.strip():
        raise HTTPException(status_code=400, detail="File không đọc được text. MVP này ưu tiên .txt/.csv/.md; PDF/DOCX cần parser riêng.")
    tag_list = [t.strip() for t in tags.split(",") if t.strip()]
    return v43_upload_document(V43DocumentUpload(title=file.filename or "uploaded_document", content=content, source=source, tags=tag_list))


@router.get("/rag/v43/documents")
def v43_list_documents():
    store = _load_store()
    docs = [{k: v for k, v in d.items() if k != "content"} | {"content_length": len(d.get("content", ""))} for d in store.setdefault("rag_documents", [])]
    return {"version": "V43", "count": len(docs), "documents": docs}


@router.post("/rag/v43/search")
def v43_search(req: V43SearchRequest):
    store = _load_store()
    results = _rag_search_v43(store, req.query, req.limit)
    return {"version": "V43", "query": req.query, "count": len(results), "results": results}


@router.post("/ai/v43/accounting-qa")
def v43_accounting_qa(req: V43AccountingQARequest):
    store = _load_store()
    direct = _answer_formula_or_account_v43(req.question)
    sources: List[Dict[str, Any]] = []
    if direct:
        answer = direct["answer"]
        intent = direct["intent"]
    else:
        results = _rag_search_v43(store, req.question, 3) if req.use_rag else []
        sources = results
        if results:
            context_lines = [f"- {r['title']}: {r['snippet']}" for r in results]
            answer = "Theo tài liệu hiện có trong hệ thống, mình tìm thấy các đoạn liên quan:\n" + "\n".join(context_lines)
            answer += "\n\nKết luận tham khảo: hãy đối chiếu chứng từ thực tế và quy định đang áp dụng trước khi ghi sổ/quyết toán."
            intent = "rag_accounting_qa"
        else:
            answer = "Mình chưa tìm thấy tài liệu phù hợp trong kho RAG. Với câu hỏi về luật/thông tư, hãy upload tài liệu liên quan vào /rag/v43/documents/upload để mình tra chính xác hơn."
            intent = "needs_documents"
    if req.save_learning:
        _append_qa_learning(store, req.question, answer, intent, sources)
        _save_store(store)
    return {"version": "V43", "intent": intent, "answer": answer, "sources": sources}


@router.post("/ai/v43/chat-with-docs")
def v43_chat_with_docs(req: V43AccountingQARequest):
    return v43_accounting_qa(req)


@router.get("/ai/v43-5/problem-templates")
def v435_problem_templates():
    return {
        "version": "V43.5",
        "templates": [
            "Công ty mua hàng hóa 100 triệu, VAT 10%, chưa thanh toán. Sau đó bán một nửa với giá 80 triệu, VAT 10%, khách chưa trả tiền. Hãy định khoản, tính VAT và lợi nhuận.",
            "Mua máy tính 20 triệu, VAT 10%, thanh toán chuyển khoản. Hãy định khoản và giải thích tài khoản sử dụng.",
            "Bán hàng 50 triệu, VAT 10%, khách hàng chưa trả tiền. Hãy định khoản doanh thu và VAT đầu ra.",
        ],
        "answer_structure": ["Tóm tắt đề", "Tách nghiệp vụ", "Định khoản", "Tính VAT/lợi nhuận", "Kiểm tra Nợ = Có", "Giả định/cảnh báo"],
    }


@router.post("/ai/v43-5/problem-solver")
def v435_problem_solver(req: V435ProblemRequest):
    store = _load_store()
    solution = _solve_accounting_problem_v43(req.question)
    result = {"version": "V43.5", "standard": req.standard, "mode": req.mode, **solution}
    if req.save_learning:
        _append_qa_learning(store, req.question, solution["answer"], "accounting_problem_solver", [])
        _save_store(store)
    return result


@router.post("/ai/v43-5/check-answer")
def v435_check_answer(req: V435CheckAnswerRequest):
    expected = _solve_accounting_problem_v43(req.question)
    user = _normalize_v43(req.user_answer)
    expected_codes = set()
    for step in expected.get("steps", []):
        for line in step.get("journal_entry", []):
            if "debit" in line:
                expected_codes.add(str(line["debit"]))
            if "credit" in line:
                expected_codes.add(str(line["credit"]))
    matched = sorted([code for code in expected_codes if re.search(rf"\b{re.escape(code)}\b", user)])
    missing = sorted(expected_codes - set(matched))
    score = round(100 * len(matched) / max(1, len(expected_codes)), 2)
    feedback = []
    if missing:
        feedback.append("Bạn còn thiếu hoặc chưa ghi rõ các tài khoản: " + ", ".join(missing))
    if expected.get("check", {}).get("balanced") is False:
        feedback.append("Lời giải mẫu hiện chưa cân đối do đề có dữ liệu thiếu; cần kiểm tra giả định.")
    if not feedback:
        feedback.append("Câu trả lời của bạn khớp các tài khoản chính trong lời giải mẫu MVP.")
    store = _load_store()
    _append_qa_learning(store, req.question, "Check answer score " + str(score), "check_answer", [])
    _save_store(store)
    return {"version": "V43.5", "score": score, "matched_accounts": matched, "missing_accounts": missing, "feedback": feedback, "expected_answer": expected.get("answer")}


@router.get("/ai/v45/qa-learning/memory")
def v45_qa_learning_memory():
    store = _load_store()
    examples = store.setdefault("qa_learning_examples", [])
    intent_counts: Dict[str, int] = defaultdict(int)
    for ex in examples:
        intent_counts[ex.get("intent", "unknown")] += 1
    return {"version": "V45 MVP", "count": len(examples), "intent_counts": dict(intent_counts), "examples": examples[-50:]}

# ---------------------------------------------------------------------------
# V44 + V46 + V47: AI CFO mini, real vector RAG MVP, multi-format document upload
# ---------------------------------------------------------------------------
import math
import xml.etree.ElementTree as ET


class V44ScenarioRequest(BaseModel):
    revenue_change_percent: float = 0.0
    expense_change_percent: float = 0.0
    extra_revenue: float = 0.0
    extra_expense: float = 0.0
    note: Optional[str] = None


class V44AskRequest(BaseModel):
    question: str


class V46VectorSearchRequest(BaseModel):
    query: str
    limit: int = Field(5, ge=1, le=20)
    min_score: float = 0.0


class V47TextUploadRequest(BaseModel):
    title: str
    content: str
    source: str = "manual"
    tags: List[str] = []
    auto_chunk: bool = True


class V47ChatDocsRequest(BaseModel):
    question: str
    limit: int = Field(5, ge=1, le=10)
    save_learning: bool = True


_VI_STOPWORDS_V46 = {
    "la", "là", "cua", "của", "va", "và", "cho", "the", "thi", "thì", "co", "có", "khong", "không",
    "trong", "ngoai", "ngoài", "mot", "một", "cac", "các", "nhung", "những", "duoc", "được", "khi", "neu", "nếu",
    "toi", "tôi", "ban", "bạn", "hoi", "hỏi", "ve", "về", "nay", "này", "do", "đó", "gi", "gì",
}


def _v46_tokens(text: str) -> List[str]:
    raw = _normalize_v43(text)
    tokens = re.findall(r"[a-zA-ZÀ-ỹ0-9]{2,}", raw)
    return [t for t in tokens if t not in _VI_STOPWORDS_V46]


def _v46_term_vector(text: str) -> Dict[str, float]:
    vector: Dict[str, float] = defaultdict(float)
    for tok in _v46_tokens(text):
        vector[tok] += 1.0
    # light sublinear tf to avoid long docs overpowering everything
    return {k: 1.0 + math.log(v) for k, v in vector.items()}


def _v46_cosine(a: Dict[str, float], b: Dict[str, float]) -> float:
    if not a or not b:
        return 0.0
    dot = sum(v * b.get(k, 0.0) for k, v in a.items())
    norm_a = math.sqrt(sum(v * v for v in a.values()))
    norm_b = math.sqrt(sum(v * v for v in b.values()))
    if not norm_a or not norm_b:
        return 0.0
    return dot / (norm_a * norm_b)


def _chunk_text_v47(text: str, max_chars: int = 900, overlap: int = 120) -> List[str]:
    cleaned = re.sub(r"\r\n?", "\n", text or "").strip()
    if not cleaned:
        return []
    paras = [p.strip() for p in re.split(r"\n\s*\n", cleaned) if p.strip()]
    chunks: List[str] = []
    current = ""
    for para in paras:
        if len(current) + len(para) + 2 <= max_chars:
            current = (current + "\n\n" + para).strip()
        else:
            if current:
                chunks.append(current)
            if len(para) <= max_chars:
                current = para
            else:
                start = 0
                while start < len(para):
                    chunks.append(para[start:start + max_chars])
                    start += max(1, max_chars - overlap)
                current = ""
    if current:
        chunks.append(current)
    return chunks


def _v47_extract_docx(raw: bytes) -> str:
    with zipfile.ZipFile(BytesIO(raw)) as zf:
        xml = zf.read("word/document.xml")
    root = ET.fromstring(xml)
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    paragraphs = []
    for p in root.findall(".//w:p", ns):
        texts = [t.text or "" for t in p.findall(".//w:t", ns)]
        line = "".join(texts).strip()
        if line:
            paragraphs.append(line)
    return "\n".join(paragraphs)


def _v47_extract_xlsx(raw: bytes) -> str:
    if Workbook is None:
        return ""
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(raw), data_only=True, read_only=True)
    lines = []
    for ws in wb.worksheets:
        lines.append(f"# Sheet: {ws.title}")
        for row in ws.iter_rows(values_only=True):
            values = [str(v) for v in row if v is not None]
            if values:
                lines.append(" | ".join(values))
    return "\n".join(lines)


def _v47_extract_pdf_mvp(raw: bytes) -> str:
    # MVP fallback: many simple PDFs contain text streams; this is not OCR.
    decoded = raw.decode("latin-1", errors="ignore")
    decoded = re.sub(r"\\[nrt]", " ", decoded)
    candidates = re.findall(r"\(([^()]{3,500})\)", decoded)
    text = "\n".join(candidates)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _v47_extract_text(filename: str, raw: bytes) -> Dict[str, Any]:
    suffix = Path(filename or "").suffix.lower()
    extraction = "plain_text"
    text = ""
    try:
        if suffix in {".txt", ".md", ".csv", ".json", ".xml", ".html"}:
            text = raw.decode("utf-8", errors="ignore")
        elif suffix == ".docx":
            extraction = "docx_xml"
            text = _v47_extract_docx(raw)
        elif suffix in {".xlsx", ".xlsm"}:
            extraction = "openpyxl"
            text = _v47_extract_xlsx(raw)
        elif suffix == ".pdf":
            extraction = "pdf_text_stream_mvp"
            text = _v47_extract_pdf_mvp(raw)
        else:
            text = raw.decode("utf-8", errors="ignore")
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Không đọc được file {filename}: {exc}")
    text = (text or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="File không trích xuất được text. PDF scan ảnh cần OCR riêng.")
    return {"text": text, "file_type": suffix.replace('.', '') or "unknown", "extraction": extraction}


def _v47_store_document(store: Dict[str, Any], title: str, content: str, source: str = "manual", tags: Optional[List[str]] = None, file_type: str = "text", extraction: str = "manual", auto_chunk: bool = True) -> Dict[str, Any]:
    doc_id = _next_id(store, "VDOC")
    chunks = _chunk_text_v47(content) if auto_chunk else [content]
    doc = {
        "id": doc_id,
        "version": "V47",
        "title": title,
        "source": source,
        "tags": tags or [],
        "file_type": file_type,
        "extraction": extraction,
        "created_at": _now(),
        "content": content,
        "content_length": len(content),
        "chunk_count": len(chunks),
    }
    store.setdefault("v47_documents", []).append(doc)
    for idx, chunk in enumerate(chunks, start=1):
        chunk_id = f"{doc_id}-CHUNK-{idx:04d}"
        vector = _v46_term_vector(f"{title}\n{chunk}")
        store.setdefault("v46_vector_chunks", []).append({
            "id": chunk_id,
            "doc_id": doc_id,
            "title": title,
            "source": source,
            "tags": tags or [],
            "file_type": file_type,
            "chunk_index": idx,
            "content": chunk,
            "tokens": sorted(vector.keys()),
            "vector": vector,
            "created_at": _now(),
        })
    # Keep V43 keyword RAG compatible too.
    store.setdefault("rag_documents", []).append({
        "id": doc_id,
        "title": title,
        "content": content,
        "source": source,
        "tags": tags or [],
        "created_at": _now(),
        "upgraded_by": "V47",
    })
    _audit(store, "upload", "v47_document", doc_id, {"title": title, "chunks": len(chunks), "file_type": file_type})
    return {k: v for k, v in doc.items() if k != "content"}


def _v46_vector_search(store: Dict[str, Any], query: str, limit: int = 5, min_score: float = 0.0) -> List[Dict[str, Any]]:
    qv = _v46_term_vector(query)
    q_tokens = set(qv.keys())
    results: List[Dict[str, Any]] = []
    for chunk in store.setdefault("v46_vector_chunks", []):
        cv = {k: float(v) for k, v in (chunk.get("vector") or {}).items()}
        score = _v46_cosine(qv, cv)
        # keyword boost for exact account codes / Vietnamese terms
        overlap = sorted(q_tokens & set(chunk.get("tokens", [])))
        if overlap:
            score += min(0.15, 0.02 * len(overlap))
        if score >= min_score and score > 0:
            content = chunk.get("content", "")
            results.append({
                "chunk_id": chunk.get("id"),
                "doc_id": chunk.get("doc_id"),
                "title": chunk.get("title"),
                "source": chunk.get("source"),
                "file_type": chunk.get("file_type"),
                "chunk_index": chunk.get("chunk_index"),
                "score": round(score, 4),
                "matched_terms": overlap[:12],
                "snippet": content[:700] + ("..." if len(content) > 700 else ""),
            })
    results.sort(key=lambda x: x["score"], reverse=True)
    return results[:limit]


def _v44_financial_metrics(store: Dict[str, Any]) -> Dict[str, Any]:
    income = v29_income_statement()
    vat = v28_vat_report()
    entries = store.get("journal_entries", [])
    drafts = store.get("journal_drafts", [])
    cash_bank_balance = 0.0
    receivables = 0.0
    payables = 0.0
    expense_by_account: Dict[str, float] = defaultdict(float)
    revenue_by_account: Dict[str, float] = defaultdict(float)
    for entry in entries:
        for line in entry.get("lines", []):
            account = str(line.get("account", ""))
            debit = float(line.get("debit", 0) or 0)
            credit = float(line.get("credit", 0) or 0)
            movement = debit - credit
            if account.startswith(("111", "112")):
                cash_bank_balance += movement
            if account.startswith("131"):
                receivables += movement
            if account.startswith("331"):
                payables += credit - debit
            if account.startswith(EXPENSE_PREFIXES):
                expense_by_account[account[:3]] += debit - credit
            if account.startswith(REVENUE_PREFIXES):
                revenue_by_account[account[:3]] += credit - debit
    revenue = float(income.get("revenue", 0) or 0)
    expenses = float(income.get("expenses", 0) or 0)
    profit = float(income.get("profit_before_tax", 0) or 0)
    profit_margin = round((profit / revenue * 100), 2) if revenue else 0.0
    burn_rate = max(expenses - revenue, 0.0)
    runway_days = None
    if burn_rate > 0:
        runway_days = round(max(cash_bank_balance, 0.0) / burn_rate * 30, 1)
    return {
        "revenue": round(revenue, 2),
        "expenses": round(expenses, 2),
        "profit_before_tax": round(profit, 2),
        "profit_margin_percent": profit_margin,
        "vat_payable": vat.get("vat_payable", 0),
        "cash_bank_balance_estimate": round(cash_bank_balance, 2),
        "receivables_estimate": round(receivables, 2),
        "payables_estimate": round(payables, 2),
        "runway_days_estimate": runway_days,
        "journal_entries": len(entries),
        "drafts_pending": len([d for d in drafts if d.get("status") == "draft"]),
        "drafts_approved": len([d for d in drafts if d.get("status") == "approved"]),
        "expense_by_account": dict(sorted(expense_by_account.items(), key=lambda kv: kv[1], reverse=True)),
        "revenue_by_account": dict(sorted(revenue_by_account.items(), key=lambda kv: kv[1], reverse=True)),
    }


def _v44_risk_alerts(metrics: Dict[str, Any]) -> List[Dict[str, Any]]:
    alerts = []
    if metrics["journal_entries"] == 0:
        alerts.append({"level": "medium", "code": "NO_POSTED_DATA", "message": "Chưa có bút toán đã ghi sổ; CFO chỉ có thể phân tích demo hoặc dữ liệu nháp."})
    if metrics["profit_before_tax"] < 0:
        alerts.append({"level": "high", "code": "LOSS", "message": "Lợi nhuận trước thuế đang âm; cần kiểm tra doanh thu và nhóm chi phí lớn."})
    if metrics["profit_margin_percent"] and metrics["profit_margin_percent"] < 10:
        alerts.append({"level": "medium", "code": "LOW_MARGIN", "message": "Biên lợi nhuận dưới 10%, nên rà soát giá bán hoặc chi phí."})
    if metrics["vat_payable"] < 0:
        alerts.append({"level": "low", "code": "VAT_CREDIT", "message": "VAT phải nộp âm, có thể đang còn VAT đầu vào được khấu trừ."})
    if metrics["drafts_pending"] > 0:
        alerts.append({"level": "medium", "code": "PENDING_DRAFTS", "message": f"Có {metrics['drafts_pending']} bút toán nháp chưa duyệt/ghi sổ."})
    if metrics["payables_estimate"] > metrics["receivables_estimate"] + max(metrics["cash_bank_balance_estimate"], 0):
        alerts.append({"level": "high", "code": "PAYABLE_PRESSURE", "message": "Công nợ phải trả ước tính lớn hơn phải thu + tiền hiện có; cần theo dõi dòng tiền."})
    return alerts


@router.get("/ai/v44/cfo/summary")
def v44_cfo_summary():
    store = _load_store()
    metrics = _v44_financial_metrics(store)
    alerts = _v44_risk_alerts(metrics)
    recommendations = []
    if metrics["drafts_pending"]:
        recommendations.append("Duyệt/post các draft còn pending để báo cáo CFO chính xác hơn.")
    if metrics["profit_before_tax"] > 0:
        recommendations.append("Doanh nghiệp đang có lãi theo dữ liệu đã ghi sổ; có thể phân tích sâu nhóm chi phí lớn nhất.")
    if metrics["expense_by_account"]:
        top_expense = next(iter(metrics["expense_by_account"].items()))
        recommendations.append(f"Chi phí lớn nhất đang ở TK {top_expense[0]}: {top_expense[1]:,.0f}đ.")
    return {"version": "V44", "metrics": metrics, "alerts": alerts, "recommendations": recommendations}


@router.post("/ai/v44/cfo/scenario")
def v44_cfo_scenario(req: V44ScenarioRequest):
    store = _load_store()
    metrics = _v44_financial_metrics(store)
    current_revenue = float(metrics["revenue"])
    current_expenses = float(metrics["expenses"])
    new_revenue = round(current_revenue * (1 + req.revenue_change_percent / 100) + req.extra_revenue, 2)
    new_expenses = round(current_expenses * (1 + req.expense_change_percent / 100) + req.extra_expense, 2)
    new_profit = round(new_revenue - new_expenses, 2)
    delta_profit = round(new_profit - float(metrics["profit_before_tax"]), 2)
    explanation = (
        f"Doanh thu kịch bản: {new_revenue:,.0f}đ; Chi phí kịch bản: {new_expenses:,.0f}đ; "
        f"Lợi nhuận dự kiến: {new_profit:,.0f}đ; Chênh lệch lợi nhuận: {delta_profit:,.0f}đ."
    )
    return {"version": "V44", "base_metrics": metrics, "scenario": req.dict(), "result": {"revenue": new_revenue, "expenses": new_expenses, "profit_before_tax": new_profit, "delta_profit": delta_profit}, "explanation": explanation}


@router.get("/ai/v44/cfo/cashflow-forecast")
def v44_cfo_cashflow_forecast(days: int = 30):
    store = _load_store()
    metrics = _v44_financial_metrics(store)
    daily_expense = float(metrics["expenses"] or 0) / 30
    daily_revenue = float(metrics["revenue"] or 0) / 30
    opening_cash = float(metrics["cash_bank_balance_estimate"] or 0)
    forecast = []
    for d in [7, 14, 30, days]:
        if d <= 0:
            continue
        ending = round(opening_cash + (daily_revenue - daily_expense) * d, 2)
        forecast.append({"days": d, "estimated_cash": ending})
    return {"version": "V44", "method": "MVP linear forecast from posted revenue/expense run-rate", "opening_cash_estimate": opening_cash, "forecast": forecast, "warning": "Dự báo MVP, chưa thay thế kế hoạch dòng tiền thật."}


@router.get("/ai/v44/cfo/risk-alerts")
def v44_cfo_risk_alerts():
    store = _load_store()
    metrics = _v44_financial_metrics(store)
    return {"version": "V44", "alerts": _v44_risk_alerts(metrics), "metrics_snapshot": metrics}


@router.post("/ai/v44/cfo/ask")
def v44_cfo_ask(req: V44AskRequest):
    msg = _normalize_v43(req.question)
    if any(k in msg for k in ["tang", "tăng", "%", "neu", "nếu", "kich ban", "kịch bản"]):
        m = re.search(r"(\d+(?:[\.,]\d+)?)\s*%", msg)
        pct = float(m.group(1).replace(",", ".")) if m else 0.0
        scenario = V44ScenarioRequest(revenue_change_percent=pct if "doanh thu" in msg else 0.0, expense_change_percent=pct if "chi phi" in msg or "chi phí" in msg else 0.0)
        res = v44_cfo_scenario(scenario)
        return {"version": "V44", "intent": "scenario", "answer": res["explanation"], "data": res}
    if any(k in msg for k in ["rui ro", "rủi ro", "canh bao", "cảnh báo", "bat thuong", "bất thường"]):
        res = v44_cfo_risk_alerts()
        answer = "\n".join([a["message"] for a in res["alerts"]]) or "Chưa thấy cảnh báo lớn từ dữ liệu hiện tại."
        return {"version": "V44", "intent": "risk_alerts", "answer": answer, "data": res}
    if any(k in msg for k in ["dong tien", "dòng tiền", "thieu tien", "thiếu tiền", "cashflow"]):
        res = v44_cfo_cashflow_forecast()
        answer = "Dự báo dòng tiền MVP: " + "; ".join([f"sau {x['days']} ngày: {x['estimated_cash']:,.0f}đ" for x in res["forecast"]])
        return {"version": "V44", "intent": "cashflow", "answer": answer, "data": res}
    res = v44_cfo_summary()
    m = res["metrics"]
    answer = f"Tóm tắt CFO: doanh thu {m['revenue']:,.0f}đ, chi phí {m['expenses']:,.0f}đ, lợi nhuận {m['profit_before_tax']:,.0f}đ, biên lợi nhuận {m['profit_margin_percent']}%."
    return {"version": "V44", "intent": "cfo_summary", "answer": answer, "data": res}


@router.get("/rag/v46/status")
def v46_status():
    store = _load_store()
    return {
        "version": "V46/V47",
        "rag_engine": "local TF cosine vector retrieval MVP, no external API",
        "documents": len(store.setdefault("v47_documents", [])),
        "chunks": len(store.setdefault("v46_vector_chunks", [])),
        "supported_uploads": ["txt", "md", "csv", "json", "docx", "xlsx", "pdf text-layer MVP"],
    }


@router.post("/rag/v46/search")
def v46_search(req: V46VectorSearchRequest):
    store = _load_store()
    results = _v46_vector_search(store, req.query, req.limit, req.min_score)
    return {"version": "V46", "query": req.query, "count": len(results), "results": results, "note": "Vector RAG MVP dùng TF-cosine local; có thể nâng lên Chroma/FAISS/Qdrant sau."}


@router.post("/rag/v47/documents/upload-text")
def v47_upload_text(req: V47TextUploadRequest):
    store = _load_store()
    doc = _v47_store_document(store, req.title, req.content, source=req.source, tags=req.tags, file_type="text", extraction="manual_text", auto_chunk=req.auto_chunk)
    _save_store(store)
    return {"version": "V47", "document": doc, "message": "Đã upload text, chunk và tạo vector index."}


@router.post("/rag/v47/documents/upload-file")
async def v47_upload_file(file: UploadFile = File(...), source: str = "upload", tags: str = "", auto_chunk: bool = True):
    raw = await file.read()
    extracted = _v47_extract_text(file.filename or "uploaded", raw)
    tag_list = [t.strip() for t in tags.split(",") if t.strip()]
    store = _load_store()
    doc = _v47_store_document(
        store,
        title=file.filename or "uploaded_document",
        content=extracted["text"],
        source=source,
        tags=tag_list,
        file_type=extracted["file_type"],
        extraction=extracted["extraction"],
        auto_chunk=auto_chunk,
    )
    _save_store(store)
    return {"version": "V47", "document": doc, "message": "Đã upload file nhiều định dạng và tạo vector chunks.", "extraction": extracted["extraction"]}


@router.get("/rag/v47/documents")
def v47_documents():
    store = _load_store()
    docs = [{k: v for k, v in d.items() if k != "content"} for d in store.setdefault("v47_documents", [])]
    return {"version": "V47", "count": len(docs), "documents": docs}


def _v49_question_intent(question: str) -> str:
    q = _normalize_v43(question)
    if any(k in q for k in ["hach toan", "hạch toán", "dinh khoan", "định khoản", "tai khoan", "tài khoản", "no ", "co "]):
        return "accounting_entry"
    if any(k in q for k in ["thue", "thuế", "vat", "hoa don", "hóa đơn", "khau tru", "khấu trừ"]):
        return "tax_invoice"
    if any(k in q for k in ["quy trinh", "quy trình", "buoc", "bước", "lo trinh", "lộ trình"]):
        return "workflow"
    if any(k in q for k in ["luat", "luật", "thong tu", "thông tư", "nghi dinh", "nghị định"]):
        return "legal_policy"
    return "general_accounting_qa"


def _v49_build_safe_rag_answer(question: str, results: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Build a structured accounting RAG answer using only retrieved chunks.

    This MVP intentionally does not invent law numbers or accounting entries. It
    gives a conservative answer and tells the accountant what evidence is missing.
    """
    if not results:
        return {
            "intent": "needs_vector_docs",
            "confidence": "low",
            "answer": (
                "Chưa đủ căn cứ để trả lời từ kho tài liệu đã upload.\n\n"
                "Bạn hãy upload thêm hợp đồng, hóa đơn, chứng từ, chính sách kế toán hoặc thông tư/quy định liên quan; "
                "sau đó hỏi lại. AI sẽ không tự bịa số thông tư, điều khoản hoặc tài khoản kế toán khi chưa có nguồn."
            ),
            "missing_info": [
                "Tài liệu nguồn liên quan đến câu hỏi",
                "Loại nghiệp vụ/chứng từ cụ thể",
                "Thời điểm phát sinh và quy định áp dụng nếu là câu hỏi về thuế/luật",
            ],
            "source_summary": [],
        }

    best_score = float(results[0].get("score") or 0)
    confidence = "high" if best_score >= 0.45 else "medium" if best_score >= 0.22 else "low"
    source_summary = []
    evidence_lines = []
    for idx, r in enumerate(results[:5], start=1):
        title = r.get("title", "Tài liệu")
        score = r.get("score", 0)
        snippet = (r.get("snippet") or "").strip()
        source_summary.append({
            "rank": idx,
            "title": title,
            "score": score,
            "chunk_id": r.get("chunk_id"),
            "matched_terms": r.get("matched_terms", []),
        })
        evidence_lines.append(f"{idx}. {title} [score {score}]: {snippet}")

    intent = _v49_question_intent(question)
    if confidence == "low":
        conclusion = (
            "Có vài đoạn tài liệu gần đúng, nhưng độ khớp còn thấp. Chỉ nên xem đây là gợi ý tra cứu, "
            "chưa nên dùng để ghi sổ/quyết toán ngay."
        )
    elif intent == "accounting_entry":
        conclusion = (
            "Có thể dùng các đoạn nguồn dưới đây để xác định hướng hạch toán, nhưng cần kiểm tra lại chứng từ, "
            "đối tượng chi phí/tài sản, thuế VAT và chính sách kế toán trước khi chốt bút toán."
        )
    elif intent == "tax_invoice":
        conclusion = (
            "Có căn cứ tài liệu để phân tích thuế/hóa đơn, nhưng cần đối chiếu ngày hiệu lực quy định, loại hóa đơn, "
            "phương thức thanh toán và điều kiện khấu trừ trước khi kết luận cuối cùng."
        )
    elif intent == "workflow":
        conclusion = "Có thể lập quy trình theo các nguồn tìm thấy; nên để kế toán/kiểm soát nội bộ duyệt bước cuối."
    elif intent == "legal_policy":
        conclusion = "Có nguồn liên quan đến luật/thông tư/quy định; không suy diễn số điều khoản ngoài nội dung tài liệu đã upload."
    else:
        conclusion = "Có nguồn liên quan để tham khảo; phần nào chưa có trong tài liệu thì cần bổ sung nguồn trước khi kết luận."

    missing_info = [
        "Bản scan/hóa đơn/hợp đồng hoặc chứng từ gốc nếu câu hỏi liên quan nghiệp vụ cụ thể",
        "Ngày phát sinh nghiệp vụ để kiểm tra quy định còn hiệu lực",
        "Mục đích sử dụng và bộ phận chịu chi phí nếu cần định khoản",
    ]
    answer = (
        f"Kết luận ngắn: {conclusion}\n\n"
        "Căn cứ tìm thấy:\n" + "\n".join(evidence_lines[:3]) + "\n\n"
        "Việc cần kiểm tra thêm:\n- " + "\n- ".join(missing_info) + "\n\n"
        "Lưu ý an toàn: AI chỉ dựa trên tài liệu đã upload, không tự tạo số thông tư/điều khoản/tài khoản nếu nguồn không nêu rõ."
    )
    return {
        "intent": f"v49_safe_{intent}",
        "confidence": confidence,
        "answer": answer,
        "missing_info": missing_info,
        "source_summary": source_summary,
    }


@router.post("/ai/v47/chat-with-vector-docs")
def v47_chat_with_vector_docs(req: V47ChatDocsRequest):
    store = _load_store()
    results = _v46_vector_search(store, req.question, req.limit, 0.0)
    built = _v49_build_safe_rag_answer(req.question, results)
    if req.save_learning:
        _append_qa_learning(store, req.question, built["answer"], built["intent"], results)
        _save_store(store)
    return {
        "version": "V49/V50/V51/V52 over V47/V46",
        "intent": built["intent"],
        "confidence": built["confidence"],
        "answer": built["answer"],
        "missing_info": built["missing_info"],
        "source_summary": built["source_summary"],
        "sources": results,
    }


@router.get("/ai/v49-v52/upgrade-status")
def v49_v52_upgrade_status():
    store = _load_store()
    return {
        "stage": "Finiip V49-V52 - AI kế toán upload UI + chatbot RAG an toàn",
        "completed": [
            "V49: HTML UI upload tài liệu và xem danh sách tài liệu",
            "V50: Chatbot kế toán gọi /ai/v47/chat-with-vector-docs",
            "V51: Chuẩn hóa câu trả lời thành kết luận, căn cứ, việc cần kiểm tra",
            "V52: Safety guard - không bịa luật/thông tư/tài khoản nếu tài liệu không có căn cứ",
        ],
        "ui_url": "/v49/accounting-ai-ui",
        "counts": {
            "documents": len(store.setdefault("v47_documents", [])),
            "chunks": len(store.setdefault("v46_vector_chunks", [])),
            "qa_learning_examples": len(store.setdefault("qa_learning_examples", [])),
        },
        "main_flow": [
            "Open /v49/accounting-ai-ui",
            "Upload PDF/DOCX/XLSX/TXT/CSV/JSON",
            "Ask accounting/tax/workflow questions",
            "Read answer with source snippets and missing-info warning",
        ],
    }


@router.get("/v49/accounting-ai-ui")
def v49_accounting_ai_ui():
    ui_path = FRONTEND_DIR / "v49_accounting_ai.html"
    if not ui_path.exists():
        raise HTTPException(status_code=404, detail="Không tìm thấy frontend/v49_accounting_ai.html")
    return FileResponse(str(ui_path), media_type="text/html")


@router.get("/ai/v44-v47/upgrade-status")
def v44_v46_v47_upgrade_status():
    store = _load_store()
    return {
        "stage": "Finiip V44 + V46 + V47 - CFO mini và RAG tài liệu thật hơn",
        "completed": [
            "V44 AI CFO mini: summary, scenario, cashflow forecast, risk alerts, CFO ask",
            "V46 Vector RAG thật hơn: local TF-cosine vector chunks, scoring, matched terms",
            "V47 Upload tài liệu nhiều định dạng: txt/md/csv/json/docx/xlsx/pdf text-layer MVP",
            "Tự chunk tài liệu và đồng bộ với V43 keyword RAG để tương thích ngược",
        ],
        "counts": {
            "journal_entries": len(store.get("journal_entries", [])),
            "v47_documents": len(store.setdefault("v47_documents", [])),
            "v46_chunks": len(store.setdefault("v46_vector_chunks", [])),
            "qa_learning_examples": len(store.setdefault("qa_learning_examples", [])),
        },
        "main_apis": [
            "GET  /ai/v44/cfo/summary",
            "POST /ai/v44/cfo/scenario",
            "GET  /ai/v44/cfo/cashflow-forecast",
            "GET  /ai/v44/cfo/risk-alerts",
            "POST /ai/v44/cfo/ask",
            "GET  /rag/v46/status",
            "POST /rag/v46/search",
            "POST /rag/v47/documents/upload-text",
            "POST /rag/v47/documents/upload-file",
            "GET  /rag/v47/documents",
            "POST /ai/v47/chat-with-vector-docs",
        ],
    }

# ---------------------------------------------------------------------------
# V53-V58: Backend-only accounting AI APIs for document review, journal suggestion,
# document classification, tax risk checklist, Excel export, and legal knowledge map.
# These endpoints intentionally return structured JSON for an external frontend.
# ---------------------------------------------------------------------------

class V53DocumentTextReviewRequest(BaseModel):
    title: str = "manual_document"
    content: str
    source: str = "manual"
    tags: List[str] = []
    save_document: bool = True
    save_review: bool = True


class V54JournalSuggestionRequest(BaseModel):
    description: str = ""
    content: Optional[str] = None
    review_id: Optional[str] = None
    document_id: Optional[str] = None
    amount_before_vat: Optional[float] = None
    vat_amount: Optional[float] = None
    total_amount: Optional[float] = None
    payment_method: Optional[str] = None
    document_type: Optional[str] = None
    purpose: Optional[str] = None


class V55ClassifyDocumentRequest(BaseModel):
    title: str = "document"
    content: str
    tags: List[str] = []


class V56TaxRiskChecklistRequest(BaseModel):
    description: str = ""
    content: Optional[str] = None
    review_id: Optional[str] = None
    document_type: Optional[str] = None
    amount_before_vat: Optional[float] = None
    vat_amount: Optional[float] = None
    total_amount: Optional[float] = None
    payment_method: Optional[str] = None


class V57ExportAnalysisExcelRequest(BaseModel):
    title: str = "AI ke toan export"
    rows: List[Dict[str, Any]] = []


class V58LegalKnowledgeUploadRequest(BaseModel):
    title: str
    content: str
    category: str = "general"
    source: str = "global_legal_knowledge"
    tags: List[str] = []
    auto_chunk: bool = True


class V58LegalSearchRequest(BaseModel):
    query: str
    category: Optional[str] = None
    limit: int = Field(5, ge=1, le=20)
    min_score: float = 0.0


_V55_DOCUMENT_TYPES = {
    "purchase_invoice": {
        "label": "Hóa đơn đầu vào",
        "keywords": ["hóa đơn", "hoa don", "vat", "gtgt", "người bán", "nguoi ban", "mua", "đầu vào", "dau vao", "supplier", "invoice"],
    },
    "sale_invoice": {
        "label": "Hóa đơn đầu ra",
        "keywords": ["người mua", "nguoi mua", "bán hàng", "ban hang", "đầu ra", "dau ra", "doanh thu", "customer", "output vat"],
    },
    "contract": {
        "label": "Hợp đồng",
        "keywords": ["hợp đồng", "hop dong", "bên a", "ben a", "bên b", "ben b", "điều khoản", "dieu khoan", "contract", "nghiệm thu"],
    },
    "cash_receipt": {
        "label": "Phiếu thu",
        "keywords": ["phiếu thu", "phieu thu", "thu tiền", "thu tien", "người nộp", "nguoi nop"],
    },
    "cash_payment": {
        "label": "Phiếu chi",
        "keywords": ["phiếu chi", "phieu chi", "chi tiền", "chi tien", "người nhận", "nguoi nhan"],
    },
    "bank_statement": {
        "label": "Sao kê ngân hàng",
        "keywords": ["sao kê", "sao ke", "ngân hàng", "ngan hang", "bank", "transaction", "chuyển khoản", "chuyen khoan", "tài khoản ngân hàng"],
    },
    "payroll": {
        "label": "Bảng lương / BHXH",
        "keywords": ["lương", "luong", "bảng lương", "bang luong", "bhxh", "bhyt", "bhtn", "thuế tncn", "tncn", "payroll"],
    },
    "tax_return": {
        "label": "Tờ khai thuế",
        "keywords": ["tờ khai", "to khai", "thuế gtgt", "thuế tndn", "tndn", "tncn", "kỳ tính thuế", "mẫu số", "mau so"],
    },
    "financial_statement": {
        "label": "Báo cáo tài chính",
        "keywords": ["bảng cân đối", "bang can doi", "kết quả kinh doanh", "ket qua kinh doanh", "lưu chuyển tiền tệ", "báo cáo tài chính", "bctc"],
    },
    "legal_policy": {
        "label": "Luật / thông tư / quy định",
        "keywords": ["thông tư", "thong tu", "nghị định", "nghi dinh", "luật", "luat", "điều", "khoản", "chuẩn mực", "chế độ kế toán"],
    },
}

_V58_KNOWLEDGE_CATEGORIES = {
    "thue_gtgt": "Thuế GTGT, hóa đơn VAT, khấu trừ, kê khai",
    "thue_tndn": "Thuế TNDN, chi phí được trừ/không được trừ",
    "hoa_don_chung_tu": "Hóa đơn, chứng từ, hợp đồng, thanh toán",
    "tai_san_co_dinh": "Tài sản cố định, khấu hao, điều kiện ghi nhận",
    "cong_cu_dung_cu": "Công cụ dụng cụ, phân bổ chi phí trả trước",
    "tien_luong_bhxh": "Tiền lương, BHXH, BHYT, BHTN, TNCN",
    "bao_cao_tai_chinh": "Báo cáo tài chính, thuyết minh, chỉ tiêu BCTC",
    "che_do_ke_toan": "Chế độ kế toán, hệ thống tài khoản, sổ sách",
    "general": "Kiến thức kế toán/thuế chung",
}


def _v53_money_to_float(raw: str) -> Optional[float]:
    if not raw:
        return None
    s = raw.strip().replace(" ", "")
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    else:
        s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


def _v53_extract_amount_candidates(text: str) -> List[float]:
    candidates: List[float] = []
    for m in re.finditer(r"(?<!\d)(\d{1,3}(?:[\.,]\d{3})+(?:[\.,]\d+)?|\d{5,})(?:\s*(?:vnd|vnđ|đ|dong|đồng))?", text or "", flags=re.I):
        val = _v53_money_to_float(m.group(1))
        if val is not None and val >= 1000:
            candidates.append(val)
    return sorted(set(candidates), reverse=True)[:10]


def _v53_extract_dates(text: str) -> List[str]:
    found = []
    patterns = [
        r"\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b",
        r"ngày\s+\d{1,2}\s+tháng\s+\d{1,2}\s+năm\s+\d{4}",
        r"ngay\s+\d{1,2}\s+thang\s+\d{1,2}\s+nam\s+\d{4}",
    ]
    for pat in patterns:
        for m in re.finditer(pat, text or "", flags=re.I):
            found.append(m.group(0))
    return list(dict.fromkeys(found))[:8]


def _v55_classify_accounting_document(text: str, title: str = "", tags: Optional[List[str]] = None) -> Dict[str, Any]:
    haystack = _normalize_v43("\n".join([title or "", text or "", " ".join(tags or [])]))
    scores = []
    for code, cfg in _V55_DOCUMENT_TYPES.items():
        matched = []
        score = 0
        for kw in cfg["keywords"]:
            nkw = _normalize_v43(kw)
            if nkw and nkw in haystack:
                matched.append(kw)
                score += 2 if " " in nkw else 1
        scores.append({"type": code, "label": cfg["label"], "score": score, "matched_keywords": matched[:12]})
    scores.sort(key=lambda x: x["score"], reverse=True)
    top = scores[0] if scores and scores[0]["score"] > 0 else {"type": "unknown", "label": "Chưa phân loại được", "score": 0, "matched_keywords": []}
    confidence = "high" if top["score"] >= 6 else "medium" if top["score"] >= 3 else "low"
    suggested_storage = "knowledge_base/global" if top["type"] == "legal_policy" else "company_documents"
    if top["type"] in {"purchase_invoice", "sale_invoice", "cash_receipt", "cash_payment", "bank_statement", "payroll"}:
        suggested_storage = "accounting_evidence"
    return {"document_type": top["type"], "label": top["label"], "confidence": confidence, "matched_keywords": top["matched_keywords"], "all_scores": scores[:5], "suggested_storage": suggested_storage}


def _v53_basic_document_review(title: str, text: str, source: str = "manual", tags: Optional[List[str]] = None) -> Dict[str, Any]:
    classification = _v55_classify_accounting_document(text, title, tags)
    norm = _normalize_v43(text)
    amounts = _v53_extract_amount_candidates(text)
    dates = _v53_extract_dates(text)
    has_vat = any(k in norm for k in ["vat", "gtgt", "thue gia tri gia tang", "thuế giá trị gia tăng"])
    has_invoice = any(k in norm for k in ["hoa don", "hóa đơn", "invoice"])
    has_contract = any(k in norm for k in ["hop dong", "hợp đồng", "contract"])
    has_bank_payment = any(k in norm for k in ["chuyen khoan", "chuyển khoản", "ngan hang", "ngân hàng", "bank"])
    has_cash_payment = any(k in norm for k in ["tien mat", "tiền mặt", "cash"])
    total_amount = amounts[0] if amounts else None
    vat_guess = None
    amount_before_vat = None
    if total_amount and has_vat:
        # MVP heuristic: common Vietnam VAT 10% included in total.
        amount_before_vat = round(total_amount / 1.1, 2)
        vat_guess = round(total_amount - amount_before_vat, 2)
    missing = []
    if not dates:
        missing.append("Chưa thấy ngày chứng từ/ngày phát sinh")
    if not amounts:
        missing.append("Chưa thấy số tiền rõ ràng")
    if classification["document_type"] in {"purchase_invoice", "sale_invoice"} and not has_vat:
        missing.append("Chưa thấy thông tin VAT/GTGT rõ ràng")
    if classification["document_type"] in {"purchase_invoice", "cash_payment"} and not (has_bank_payment or has_cash_payment):
        missing.append("Chưa thấy phương thức thanh toán")
    detected = {
        "dates": dates,
        "amount_candidates": amounts,
        "total_amount_guess": total_amount,
        "amount_before_vat_guess": amount_before_vat,
        "vat_amount_guess": vat_guess,
        "has_vat_keyword": has_vat,
        "has_invoice_keyword": has_invoice,
        "has_contract_keyword": has_contract,
        "has_bank_payment_keyword": has_bank_payment,
        "has_cash_payment_keyword": has_cash_payment,
    }
    summary = [
        f"Loại tài liệu dự đoán: {classification['label']} ({classification['confidence']}).",
        f"Số tiền lớn nhất phát hiện: {total_amount:,.0f}đ." if total_amount else "Chưa phát hiện số tiền rõ ràng.",
        "Có dấu hiệu VAT/GTGT." if has_vat else "Chưa thấy dấu hiệu VAT/GTGT rõ ràng.",
    ]
    return {
        "classification": classification,
        "detected_fields": detected,
        "missing_info": missing,
        "review_summary": " ".join(summary),
        "next_actions": [
            "Gọi /ai/v54/journal-suggestion để lấy bút toán đề xuất",
            "Gọi /ai/v56/tax-risk-checklist để lấy checklist rủi ro thuế",
            "Gọi /ai/v57/document-reviews/export-excel để xuất danh sách review ra Excel",
        ],
        "source": source,
        "tags": tags or [],
    }


def _v53_get_review(store: Dict[str, Any], review_id: str) -> Optional[Dict[str, Any]]:
    for r in store.setdefault("v53_document_reviews", []):
        if r.get("id") == review_id:
            return r
    return None


def _v53_get_document_content(store: Dict[str, Any], document_id: str) -> Optional[str]:
    for d in store.setdefault("v47_documents", []):
        if d.get("id") == document_id:
            return d.get("content") or ""
    return None


def _v54_suggest_journal_payload(payload: Dict[str, Any]) -> Dict[str, Any]:
    description = payload.get("description") or ""
    content = payload.get("content") or ""
    doc_type = payload.get("document_type") or _v55_classify_accounting_document(content, description).get("document_type")
    norm = _normalize_v43("\n".join([description, content, str(payload.get("purpose") or "")]))
    total = payload.get("total_amount")
    before_vat = payload.get("amount_before_vat")
    vat = payload.get("vat_amount")
    if total is None:
        amounts = _v53_extract_amount_candidates(content + "\n" + description)
        total = amounts[0] if amounts else 0.0
    if before_vat is None and vat is None and total:
        if any(k in norm for k in ["vat", "gtgt", "10%"]):
            before_vat = round(float(total) / 1.1, 2)
            vat = round(float(total) - before_vat, 2)
        else:
            before_vat = float(total)
            vat = 0.0
    before_vat = float(before_vat or 0.0)
    vat = float(vat or 0.0)
    total = float(total or before_vat + vat)
    payment = _normalize_v43(payload.get("payment_method") or "")
    credit = "112" if any(k in payment or k in norm for k in ["chuyen khoan", "chuyển khoản", "ngan hang", "ngân hàng", "bank", "112"]) else "111" if any(k in payment or k in norm for k in ["tien mat", "tiền mặt", "cash", "111"]) else "331"
    lines = []
    assumptions = []
    warnings = []
    if doc_type == "sale_invoice" or any(k in norm for k in ["ban hang", "doanh thu", "dau ra"]):
        debit = "131" if credit == "331" else credit
        lines.append({"debit_account": debit, "credit_account": "511", "amount": before_vat, "description": "Ghi nhận doanh thu bán hàng/cung cấp dịch vụ"})
        if vat:
            lines.append({"debit_account": debit, "credit_account": "3331", "amount": vat, "description": "Thuế GTGT đầu ra phải nộp"})
        assumptions.append("Mặc định là nghiệp vụ bán hàng chịu thuế GTGT nếu có VAT.")
    else:
        debit = "642"  # default office/admin expense
        if any(k in norm for k in ["laptop", "may tinh", "máy tính", "tai san co dinh", "tài sản cố định", "khau hao"]):
            debit = "211" if before_vat >= 30000000 else "242"
            assumptions.append("Laptop/máy tính dưới ngưỡng TSCĐ thường xem xét CCDC/chi phí trả trước; trên ngưỡng có thể xem xét TSCĐ.")
        elif any(k in norm for k in ["hang hoa", "hàng hóa", "nhap kho", "nhập kho", "vat tu", "vật tư"]):
            debit = "156"
            assumptions.append("Có dấu hiệu mua hàng hóa/vật tư nhập kho.")
        elif any(k in norm for k in ["quang cao", "quảng cáo", "marketing", "facebook", "google ads"]):
            debit = "641"
            assumptions.append("Có dấu hiệu chi phí bán hàng/quảng cáo.")
        elif any(k in norm for k in ["luong", "lương", "bhxh", "nhan vien", "nhân viên"]):
            debit = "334"
            credit = "338" if "bhxh" in norm else credit
            assumptions.append("Có dấu hiệu bảng lương/khoản phải trả người lao động.")
        lines.append({"debit_account": debit, "credit_account": credit, "amount": before_vat, "description": "Ghi nhận giá trị chưa VAT/chi phí/tài sản theo chứng từ"})
        if vat:
            lines.append({"debit_account": "1331", "credit_account": credit, "amount": vat, "description": "Thuế GTGT đầu vào được khấu trừ nếu đủ điều kiện"})
    if not total:
        warnings.append("Chưa có số tiền chắc chắn; bút toán chỉ là khung gợi ý.")
    if credit in {"111", "331"} and total >= 5_000_000:
        warnings.append("Giao dịch từ 5 triệu đồng trở lên cần kiểm tra điều kiện thanh toán không dùng tiền mặt để khấu trừ VAT và xem xét chi phí được trừ.")
    warnings.append("Cần kế toán kiểm tra chứng từ gốc, chính sách kế toán và quy định hiện hành trước khi ghi sổ.")
    return {
        "journal_type": "sale" if any(l.get("credit_account") == "511" for l in lines) else "purchase_or_expense",
        "document_type": doc_type,
        "currency": "VND",
        "amount_before_vat": before_vat,
        "vat_amount": vat,
        "total_amount": total,
        "suggested_lines": lines,
        "assumptions": assumptions,
        "warnings": warnings,
        "confidence": "medium" if total and lines else "low",
    }


def _v56_build_tax_risks(payload: Dict[str, Any]) -> Dict[str, Any]:
    description = payload.get("description") or ""
    content = payload.get("content") or ""
    norm = _normalize_v43(description + "\n" + content + "\n" + str(payload.get("payment_method") or ""))
    doc_type = payload.get("document_type") or _v55_classify_accounting_document(content, description).get("document_type")
    total = payload.get("total_amount")
    if total is None:
        amounts = _v53_extract_amount_candidates(content + "\n" + description)
        total = amounts[0] if amounts else 0
    total = float(total or 0)
    has_invoice = any(k in norm for k in ["hoa don", "hóa đơn", "invoice"])
    has_vat = any(k in norm for k in ["vat", "gtgt", "thuế giá trị gia tăng", "thue gia tri gia tang"])
    bank = any(k in norm for k in ["chuyen khoan", "chuyển khoản", "ngan hang", "ngân hàng", "bank", "112"])
    contract = any(k in norm for k in ["hop dong", "hợp đồng", "contract"])
    risks = []
    def add(level: str, code: str, message: str, required: List[str]):
        risks.append({"level": level, "code": code, "message": message, "required_evidence": required})
    if doc_type in {"purchase_invoice", "cash_payment", "unknown"} and not has_invoice:
        add("high", "missing_invoice", "Chưa thấy hóa đơn/chứng từ hợp lệ trong nội dung.", ["Hóa đơn VAT hoặc chứng từ thay thế hợp lệ", "Thông tin người bán/người mua", "Ngày và số hóa đơn"])
    if has_vat and total >= 5_000_000 and not bank:
        add("high", "vat_bank_payment", "Có VAT và số tiền từ 5 triệu đồng trở lên nhưng chưa thấy dấu hiệu thanh toán không dùng tiền mặt.", ["Ủy nhiệm chi/sao kê ngân hàng", "Chứng từ thanh toán không dùng tiền mặt"])
    if any(k in norm for k in ["dich vu", "dịch vụ", "nghiem thu", "nghiệm thu"]) and not contract:
        add("medium", "service_contract_acceptance", "Dịch vụ thường cần hợp đồng/biên bản nghiệm thu hoặc tài liệu chứng minh hoàn thành.", ["Hợp đồng", "Biên bản nghiệm thu", "Bảng kê/đề nghị thanh toán"])
    if any(k in norm for k in ["laptop", "may tinh", "máy tính", "tai san", "tài sản"]) and total:
        add("medium", "asset_or_tool_classification", "Cần xác định là TSCĐ, CCDC hay chi phí trả trước dựa trên giá trị, thời gian sử dụng và chính sách công ty.", ["Quyết định đưa vào sử dụng", "Biên bản bàn giao", "Chính sách ghi nhận TSCĐ/CCDC"])
    if not total:
        add("medium", "missing_amount", "Chưa xác định được số tiền để đánh giá rủi ro thuế/kế toán.", ["Số tiền trước VAT", "VAT", "Tổng thanh toán"])
    if not risks:
        add("low", "basic_evidence_ok", "Chưa thấy rủi ro lớn từ dữ liệu text hiện có, nhưng vẫn cần kiểm tra chứng từ gốc.", ["Chứng từ gốc", "Ngày hiệu lực quy định", "Phê duyệt nội bộ"])
    level_order = {"high": 3, "medium": 2, "low": 1}
    overall = max(risks, key=lambda r: level_order.get(r["level"], 0))["level"]
    return {"overall_risk": overall, "document_type": doc_type, "risks": risks, "safe_note": "Checklist MVP, không thay thế tư vấn thuế/kế toán chính thức."}


def _v57_reviews_to_rows(reviews: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    rows = []
    for r in reviews:
        review = r.get("review") or {}
        cls = review.get("classification") or {}
        detected = review.get("detected_fields") or {}
        risks = r.get("tax_risk") or {}
        journal = r.get("journal_suggestion") or {}
        rows.append({
            "review_id": r.get("id"),
            "title": r.get("title"),
            "document_type": cls.get("label"),
            "confidence": cls.get("confidence"),
            "dates": ", ".join(detected.get("dates") or []),
            "total_amount_guess": detected.get("total_amount_guess"),
            "vat_amount_guess": detected.get("vat_amount_guess"),
            "overall_risk": risks.get("overall_risk"),
            "journal_type": journal.get("journal_type"),
            "missing_info": "; ".join(review.get("missing_info") or []),
            "created_at": r.get("created_at"),
        })
    return rows


def _v57_make_excel(rows: List[Dict[str, Any]], title: str = "AI ke toan export") -> BytesIO:
    if Workbook is None:
        raise HTTPException(status_code=500, detail="openpyxl chưa sẵn sàng; hãy cài openpyxl trong requirements.txt")
    wb = Workbook()
    ws = wb.active
    ws.title = "AI accounting"
    headers = list(rows[0].keys()) if rows else ["note"]
    if not rows:
        rows = [{"note": "Không có dữ liệu để xuất"}]
        headers = ["note"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
    for row in rows:
        ws.append([row.get(h) for h in headers])
    for idx, header in enumerate(headers, start=1):
        width = min(max(len(str(header)) + 4, 14), 45)
        ws.column_dimensions[chr(64 + idx) if idx <= 26 else "A"].width = width
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


@router.post("/ai/v53/document-review/text")
def v53_document_review_text(req: V53DocumentTextReviewRequest):
    store = _load_store()
    doc_meta = None
    if req.save_document:
        doc_meta = _v47_store_document(store, req.title, req.content, source=req.source, tags=req.tags, file_type="text", extraction="manual_text", auto_chunk=True)
    review = _v53_basic_document_review(req.title, req.content, req.source, req.tags)
    journal = _v54_suggest_journal_payload({"description": req.title, "content": req.content, "document_type": review["classification"]["document_type"], "total_amount": review["detected_fields"].get("total_amount_guess"), "vat_amount": review["detected_fields"].get("vat_amount_guess"), "amount_before_vat": review["detected_fields"].get("amount_before_vat_guess")})
    tax_risk = _v56_build_tax_risks({"description": req.title, "content": req.content, "document_type": review["classification"]["document_type"], "total_amount": review["detected_fields"].get("total_amount_guess")})
    saved = None
    if req.save_review:
        saved = {"id": _next_id(store, "REV"), "version": "V53-V58", "title": req.title, "document_id": (doc_meta or {}).get("id"), "created_at": _now(), "review": review, "journal_suggestion": journal, "tax_risk": tax_risk}
        store.setdefault("v53_document_reviews", []).append(saved)
    _save_store(store)
    return {"version": "V53-V58", "review_id": (saved or {}).get("id"), "document": doc_meta, "review": review, "journal_suggestion": journal, "tax_risk": tax_risk}


@router.post("/ai/v53/document-review/upload-file")
async def v53_document_review_upload_file(file: UploadFile = File(...), source: str = "upload", tags: str = "", save_document: bool = True, save_review: bool = True):
    raw = await file.read()
    extracted = _v47_extract_text(file.filename or "uploaded", raw)
    req = V53DocumentTextReviewRequest(title=file.filename or "uploaded", content=extracted["text"], source=source, tags=[t.strip() for t in tags.split(",") if t.strip()], save_document=save_document, save_review=save_review)
    result = v53_document_review_text(req)
    result["extraction"] = extracted["extraction"]
    result["file_type"] = extracted["file_type"]
    return result


@router.get("/ai/v53/document-reviews")
def v53_document_reviews(limit: int = 50):
    store = _load_store()
    reviews = store.setdefault("v53_document_reviews", [])[-limit:]
    return {"version": "V53", "count": len(reviews), "reviews": reviews}


@router.post("/ai/v54/journal-suggestion")
def v54_journal_suggestion(req: V54JournalSuggestionRequest):
    store = _load_store()
    payload = req.dict()
    if req.review_id:
        review = _v53_get_review(store, req.review_id)
        if not review:
            raise HTTPException(status_code=404, detail="Không tìm thấy review_id")
        payload.update({
            "description": payload.get("description") or review.get("title") or "",
            "content": payload.get("content") or json.dumps(review.get("review", {}), ensure_ascii=False),
            "document_type": payload.get("document_type") or ((review.get("review") or {}).get("classification") or {}).get("document_type"),
            "total_amount": payload.get("total_amount") or ((review.get("review") or {}).get("detected_fields") or {}).get("total_amount_guess"),
            "vat_amount": payload.get("vat_amount") or ((review.get("review") or {}).get("detected_fields") or {}).get("vat_amount_guess"),
            "amount_before_vat": payload.get("amount_before_vat") or ((review.get("review") or {}).get("detected_fields") or {}).get("amount_before_vat_guess"),
        })
    if req.document_id and not payload.get("content"):
        content = _v53_get_document_content(store, req.document_id)
        if content is None:
            raise HTTPException(status_code=404, detail="Không tìm thấy document_id")
        payload["content"] = content
    suggestion = _v54_suggest_journal_payload(payload)
    return {"version": "V54", "input": {k: v for k, v in payload.items() if k != "content"}, "journal_suggestion": suggestion}


@router.post("/ai/v55/classify-document")
def v55_classify_document(req: V55ClassifyDocumentRequest):
    classification = _v55_classify_accounting_document(req.content, req.title, req.tags)
    return {"version": "V55", "classification": classification}


@router.post("/ai/v56/tax-risk-checklist")
def v56_tax_risk_checklist(req: V56TaxRiskChecklistRequest):
    store = _load_store()
    payload = req.dict()
    if req.review_id:
        review = _v53_get_review(store, req.review_id)
        if not review:
            raise HTTPException(status_code=404, detail="Không tìm thấy review_id")
        payload.update({
            "description": payload.get("description") or review.get("title") or "",
            "content": payload.get("content") or json.dumps(review.get("review", {}), ensure_ascii=False),
            "document_type": payload.get("document_type") or ((review.get("review") or {}).get("classification") or {}).get("document_type"),
            "total_amount": payload.get("total_amount") or ((review.get("review") or {}).get("detected_fields") or {}).get("total_amount_guess"),
        })
    checklist = _v56_build_tax_risks(payload)
    return {"version": "V56", "checklist": checklist}


@router.get("/ai/v57/document-reviews/export-excel")
def v57_document_reviews_export_excel(limit: int = 500):
    store = _load_store()
    rows = _v57_reviews_to_rows(store.setdefault("v53_document_reviews", [])[-limit:])
    bio = _v57_make_excel(rows, "document_reviews")
    filename = f"ai_ke_toan_document_reviews_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(bio, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.post("/ai/v57/export-analysis-excel")
def v57_export_analysis_excel(req: V57ExportAnalysisExcelRequest):
    rows = req.rows or []
    bio = _v57_make_excel(rows, req.title)
    filename = f"ai_ke_toan_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(bio, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/ai/v58/knowledge-map")
def v58_knowledge_map():
    store = _load_store()
    counts = defaultdict(int)
    for d in store.setdefault("v47_documents", []):
        for t in d.get("tags") or []:
            if t in _V58_KNOWLEDGE_CATEGORIES:
                counts[t] += 1
    return {"version": "V58", "categories": _V58_KNOWLEDGE_CATEGORIES, "document_counts_by_category_tag": dict(counts), "recommended_folder_structure": {k: f"knowledge_base/global/{k}/" for k in _V58_KNOWLEDGE_CATEGORIES}}


@router.post("/ai/v58/legal-knowledge/upload-text")
def v58_legal_knowledge_upload_text(req: V58LegalKnowledgeUploadRequest):
    category = req.category if req.category in _V58_KNOWLEDGE_CATEGORIES else "general"
    tags = list(dict.fromkeys([category, "legal_knowledge", *req.tags]))
    store = _load_store()
    doc = _v47_store_document(store, req.title, req.content, source=req.source, tags=tags, file_type="text", extraction="manual_legal_text", auto_chunk=req.auto_chunk)
    _save_store(store)
    return {"version": "V58", "category": category, "document": doc, "message": "Đã đưa tài liệu luật/thông tư vào vector RAG với tag danh mục."}


@router.post("/ai/v58/legal-search")
def v58_legal_search(req: V58LegalSearchRequest):
    store = _load_store()
    query = req.query if not req.category else f"{req.category} {req.query}"
    results = _v46_vector_search(store, query, req.limit, req.min_score)
    if req.category:
        results = [r for r in results if req.category in (r.get("tags") or []) or req.category in _normalize_v43(r.get("title") or "")]
    return {"version": "V58", "query": req.query, "category": req.category, "count": len(results), "results": results, "note": "Tìm kiếm trong kho RAG local; nên upload văn bản nguồn chính thức vào đúng category."}


@router.get("/ai/v53-v58/upgrade-status")
def v53_v58_upgrade_status():
    store = _load_store()
    return {
        "stage": "Finiip V53-V58 backend-only accounting AI",
        "frontend_policy": "Không thêm UI; frontend riêng có thể gọi API JSON/download Excel.",
        "completed": [
            "V53: AI kiểm tra chứng từ từ text/file upload",
            "V54: AI đề xuất bút toán cấu trúc JSON",
            "V55: Tự phân loại tài liệu kế toán",
            "V56: Checklist rủi ro thuế/chứng từ",
            "V57: Xuất kết quả review/analysis ra Excel",
            "V58: Knowledge map luật/thông tư theo danh mục và legal search",
        ],
        "counts": {
            "documents": len(store.setdefault("v47_documents", [])),
            "chunks": len(store.setdefault("v46_vector_chunks", [])),
            "document_reviews": len(store.setdefault("v53_document_reviews", [])),
        },
        "main_apis": [
            "POST /ai/v53/document-review/upload-file",
            "POST /ai/v53/document-review/text",
            "GET  /ai/v53/document-reviews",
            "POST /ai/v54/journal-suggestion",
            "POST /ai/v55/classify-document",
            "POST /ai/v56/tax-risk-checklist",
            "GET  /ai/v57/document-reviews/export-excel",
            "POST /ai/v57/export-analysis-excel",
            "GET  /ai/v58/knowledge-map",
            "POST /ai/v58/legal-knowledge/upload-text",
            "POST /ai/v58/legal-search",
        ],
    }


# ---------------------------------------------------------------------------
# V59-V65: Backend-only accounting system workflow APIs
# Company/tenant, chart of accounts, journal entries, approval workflow,
# basic reports, audit logs, and standard response helpers for frontend clients.
# ---------------------------------------------------------------------------

_V59_DEFAULT_COMPANY_ID = "COMP-00001"
_V60_DEFAULT_ACCOUNTS = [
    {"code": "111", "name": "Tiền mặt", "type": "asset", "parent_code": "11"},
    {"code": "112", "name": "Tiền gửi ngân hàng", "type": "asset", "parent_code": "11"},
    {"code": "131", "name": "Phải thu của khách hàng", "type": "asset", "parent_code": "13"},
    {"code": "1331", "name": "Thuế GTGT được khấu trừ của hàng hóa, dịch vụ", "type": "asset", "parent_code": "133"},
    {"code": "152", "name": "Nguyên liệu, vật liệu", "type": "asset", "parent_code": "15"},
    {"code": "153", "name": "Công cụ, dụng cụ", "type": "asset", "parent_code": "15"},
    {"code": "156", "name": "Hàng hóa", "type": "asset", "parent_code": "15"},
    {"code": "211", "name": "Tài sản cố định hữu hình", "type": "asset", "parent_code": "21"},
    {"code": "214", "name": "Hao mòn tài sản cố định", "type": "contra_asset", "parent_code": "21"},
    {"code": "242", "name": "Chi phí trả trước", "type": "asset", "parent_code": "24"},
    {"code": "331", "name": "Phải trả cho người bán", "type": "liability", "parent_code": "33"},
    {"code": "3331", "name": "Thuế GTGT phải nộp", "type": "liability", "parent_code": "333"},
    {"code": "334", "name": "Phải trả người lao động", "type": "liability", "parent_code": "33"},
    {"code": "338", "name": "Phải trả, phải nộp khác", "type": "liability", "parent_code": "33"},
    {"code": "411", "name": "Vốn đầu tư của chủ sở hữu", "type": "equity", "parent_code": "41"},
    {"code": "511", "name": "Doanh thu bán hàng và cung cấp dịch vụ", "type": "revenue", "parent_code": "51"},
    {"code": "515", "name": "Doanh thu hoạt động tài chính", "type": "revenue", "parent_code": "51"},
    {"code": "632", "name": "Giá vốn hàng bán", "type": "expense", "parent_code": "63"},
    {"code": "635", "name": "Chi phí tài chính", "type": "expense", "parent_code": "63"},
    {"code": "641", "name": "Chi phí bán hàng", "type": "expense", "parent_code": "64"},
    {"code": "642", "name": "Chi phí quản lý doanh nghiệp", "type": "expense", "parent_code": "64"},
    {"code": "711", "name": "Thu nhập khác", "type": "revenue", "parent_code": "71"},
    {"code": "811", "name": "Chi phí khác", "type": "expense", "parent_code": "81"},
    {"code": "911", "name": "Xác định kết quả kinh doanh", "type": "equity", "parent_code": "91"},
]


def _api_ok(data: Any = None, message: str = "OK") -> Dict[str, Any]:
    return {"success": True, "data": data, "message": message, "errors": []}


def _api_error(message: str, errors: Optional[List[Dict[str, Any]]] = None, status_code: int = 400) -> None:
    raise HTTPException(status_code=status_code, detail={"success": False, "data": None, "message": message, "errors": errors or []})


def _v59_bootstrap_store(store: Dict[str, Any]) -> Dict[str, Any]:
    store.setdefault("companies", [])
    store.setdefault("chart_of_accounts", [])
    store.setdefault("v61_journal_entries", [])
    store.setdefault("document_workflows", [])
    store.setdefault("audit_logs", [])
    if not store["companies"]:
        company = {
            "id": _V59_DEFAULT_COMPANY_ID,
            "company_name": "Default Company",
            "tax_code": "",
            "accounting_mode": "TT200",
            "currency": "VND",
            "fiscal_year": datetime.now().year,
            "is_active": True,
            "created_at": _now(),
            "updated_at": _now(),
        }
        store["companies"].append(company)
    if not store["chart_of_accounts"]:
        for acc in _V60_DEFAULT_ACCOUNTS:
            row = dict(acc)
            row.update({"company_id": _V59_DEFAULT_COMPANY_ID, "is_active": True, "created_at": _now(), "updated_at": _now(), "source": "V60_DEFAULT_TT200"})
            store["chart_of_accounts"].append(row)
    return store


def _v59_get_company(store: Dict[str, Any], company_id: str) -> Optional[Dict[str, Any]]:
    _v59_bootstrap_store(store)
    return next((c for c in store.setdefault("companies", []) if c.get("id") == company_id), None)


def _v60_accounts_for_company(store: Dict[str, Any], company_id: str) -> List[Dict[str, Any]]:
    _v59_bootstrap_store(store)
    return [a for a in store.setdefault("chart_of_accounts", []) if a.get("company_id") == company_id and a.get("is_active", True)]


def _v60_account_exists(store: Dict[str, Any], company_id: str, code: str) -> bool:
    return any(a.get("code") == code and a.get("company_id") == company_id and a.get("is_active", True) for a in store.setdefault("chart_of_accounts", []))


def _v61_normalize_lines(lines: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    normalized = []
    for i, line in enumerate(lines or [], start=1):
        account = str(line.get("account") or line.get("account_code") or "").strip()
        debit = float(line.get("debit") or 0)
        credit = float(line.get("credit") or 0)
        normalized.append({
            "line_no": int(line.get("line_no") or i),
            "account": account,
            "debit": round(debit, 2),
            "credit": round(credit, 2),
            "description": line.get("description") or line.get("dien_giai") or "",
            "object": line.get("object") or line.get("partner") or None,
        })
    return normalized


def _v61_validate_journal(store: Dict[str, Any], company_id: str, lines: List[Dict[str, Any]]) -> List[Dict[str, str]]:
    errors: List[Dict[str, str]] = []
    if not lines:
        errors.append({"field": "lines", "message": "Bút toán phải có ít nhất 2 dòng."})
        return errors
    total_debit = round(sum(float(l.get("debit") or 0) for l in lines), 2)
    total_credit = round(sum(float(l.get("credit") or 0) for l in lines), 2)
    if total_debit <= 0 and total_credit <= 0:
        errors.append({"field": "lines", "message": "Tổng Nợ/Có phải lớn hơn 0."})
    if abs(total_debit - total_credit) > 0.01:
        errors.append({"field": "lines", "message": f"Tổng Nợ phải bằng tổng Có. Hiện tại Nợ={total_debit}, Có={total_credit}."})
    for idx, line in enumerate(lines):
        code = line.get("account") or ""
        debit = float(line.get("debit") or 0)
        credit = float(line.get("credit") or 0)
        if not code:
            errors.append({"field": f"lines[{idx}].account", "message": "Thiếu mã tài khoản."})
        elif not _v60_account_exists(store, company_id, code):
            errors.append({"field": f"lines[{idx}].account", "message": f"Tài khoản {code} chưa có trong chart of accounts của công ty."})
        if debit > 0 and credit > 0:
            errors.append({"field": f"lines[{idx}]", "message": "Một dòng không nên vừa Nợ vừa Có."})
        if debit < 0 or credit < 0:
            errors.append({"field": f"lines[{idx}]", "message": "Số tiền Nợ/Có không được âm."})
    return errors


def _v64_audit_event(store: Dict[str, Any], action: str, entity: str, entity_id: str, detail: Optional[Dict[str, Any]] = None, actor: str = "api", company_id: Optional[str] = None) -> Dict[str, Any]:
    row = {
        "id": _next_id(store, "AUDIT"),
        "created_at": _now(),
        "actor": actor,
        "company_id": company_id,
        "action": action,
        "entity": entity,
        "entity_id": entity_id,
        "detail": detail or {},
        "version": "V64",
    }
    store.setdefault("audit_logs", []).append(row)
    return row


def _v62_find_workflow(store: Dict[str, Any], document_id: str) -> Optional[Dict[str, Any]]:
    return next((w for w in store.setdefault("document_workflows", []) if w.get("document_id") == document_id), None)


def _account_type_map(store: Dict[str, Any], company_id: str) -> Dict[str, str]:
    return {a.get("code"): a.get("type", "other") for a in _v60_accounts_for_company(store, company_id)}


def _v63_signed_delta(account_type: str, debit: float, credit: float) -> float:
    if account_type in {"asset", "expense", "contra_liability"}:
        return debit - credit
    return credit - debit


class V59CompanyCreate(BaseModel):
    company_name: str = Field(..., min_length=2)
    tax_code: str = ""
    accounting_mode: str = "TT200"
    currency: str = "VND"
    fiscal_year: int = Field(default_factory=lambda: datetime.now().year)


class V59CompanyUpdate(BaseModel):
    company_name: Optional[str] = None
    tax_code: Optional[str] = None
    accounting_mode: Optional[str] = None
    currency: Optional[str] = None
    fiscal_year: Optional[int] = None
    is_active: Optional[bool] = None


class V60AccountCreate(BaseModel):
    code: str = Field(..., min_length=2)
    name: str = Field(..., min_length=2)
    type: str = "asset"
    parent_code: Optional[str] = None
    company_id: str = _V59_DEFAULT_COMPANY_ID
    is_active: bool = True


class V60AccountUpdate(BaseModel):
    name: Optional[str] = None
    type: Optional[str] = None
    parent_code: Optional[str] = None
    is_active: Optional[bool] = None


class V61JournalEntryCreate(BaseModel):
    date: str = Field(default_factory=lambda: datetime.now().date().isoformat())
    description: str = Field(..., min_length=2)
    company_id: str = _V59_DEFAULT_COMPANY_ID
    source_document_id: Optional[str] = None
    lines: List[Dict[str, Any]]
    status: str = "draft"
    actor: str = "api"
    metadata: Dict[str, Any] = Field(default_factory=dict)


class V61JournalEntryUpdate(BaseModel):
    date: Optional[str] = None
    description: Optional[str] = None
    source_document_id: Optional[str] = None
    lines: Optional[List[Dict[str, Any]]] = None
    status: Optional[str] = None
    actor: str = "api"
    metadata: Optional[Dict[str, Any]] = None


class V62WorkflowAction(BaseModel):
    company_id: str = _V59_DEFAULT_COMPANY_ID
    actor: str = "api"
    note: Optional[str] = None
    journal_entry_id: Optional[str] = None
    rejection_reason: Optional[str] = None


@router.post("/companies")
def v59_create_company(req: V59CompanyCreate):
    store = _load_store(); _v59_bootstrap_store(store)
    company = {
        "id": _next_id(store, "COMP"),
        "company_name": req.company_name,
        "tax_code": req.tax_code,
        "accounting_mode": req.accounting_mode,
        "currency": req.currency,
        "fiscal_year": req.fiscal_year,
        "is_active": True,
        "created_at": _now(),
        "updated_at": _now(),
    }
    store.setdefault("companies", []).append(company)
    for acc in _V60_DEFAULT_ACCOUNTS:
        row = dict(acc)
        row.update({"company_id": company["id"], "is_active": True, "created_at": _now(), "updated_at": _now(), "source": "V60_DEFAULT_TT200"})
        store.setdefault("chart_of_accounts", []).append(row)
    _v64_audit_event(store, "company.create", "company", company["id"], {"company_name": company["company_name"]}, company_id=company["id"])
    _save_store(store)
    return _api_ok(company, "Đã tạo company/tenant và seed hệ thống tài khoản mặc định.")


@router.get("/companies")
def v59_list_companies(active_only: bool = False):
    store = _load_store(); _v59_bootstrap_store(store); _save_store(store)
    rows = store.setdefault("companies", [])
    if active_only:
        rows = [c for c in rows if c.get("is_active", True)]
    return _api_ok({"count": len(rows), "companies": rows})


@router.get("/companies/{company_id}")
def v59_get_company(company_id: str):
    store = _load_store(); company = _v59_get_company(store, company_id)
    if not company:
        _api_error("Không tìm thấy company.", [{"field": "company_id", "message": company_id}], 404)
    return _api_ok(company)


@router.put("/companies/{company_id}")
def v59_update_company(company_id: str, req: V59CompanyUpdate):
    store = _load_store(); company = _v59_get_company(store, company_id)
    if not company:
        _api_error("Không tìm thấy company.", [{"field": "company_id", "message": company_id}], 404)
    for key, value in req.dict(exclude_unset=True).items():
        company[key] = value
    company["updated_at"] = _now()
    _v64_audit_event(store, "company.update", "company", company_id, req.dict(exclude_unset=True), company_id=company_id)
    _save_store(store)
    return _api_ok(company, "Đã cập nhật company.")


@router.get("/accounting/accounts")
def v60_list_accounts(company_id: str = _V59_DEFAULT_COMPANY_ID, active_only: bool = True):
    store = _load_store(); _v59_bootstrap_store(store); _save_store(store)
    rows = [a for a in store.setdefault("chart_of_accounts", []) if a.get("company_id") == company_id]
    if active_only:
        rows = [a for a in rows if a.get("is_active", True)]
    return _api_ok({"company_id": company_id, "count": len(rows), "accounts": sorted(rows, key=lambda x: x.get("code", ""))})


@router.post("/accounting/accounts")
def v60_create_account(req: V60AccountCreate):
    store = _load_store(); _v59_bootstrap_store(store)
    if not _v59_get_company(store, req.company_id):
        _api_error("Company không tồn tại.", [{"field": "company_id", "message": req.company_id}], 404)
    if _v60_account_exists(store, req.company_id, req.code):
        _api_error("Mã tài khoản đã tồn tại trong company này.", [{"field": "code", "message": req.code}])
    account = req.dict()
    account.update({"created_at": _now(), "updated_at": _now(), "source": "manual"})
    store.setdefault("chart_of_accounts", []).append(account)
    _v64_audit_event(store, "account.create", "account", req.code, account, company_id=req.company_id)
    _save_store(store)
    return _api_ok(account, "Đã tạo tài khoản kế toán.")


@router.put("/accounting/accounts/{account_code}")
def v60_update_account(account_code: str, req: V60AccountUpdate, company_id: str = _V59_DEFAULT_COMPANY_ID):
    store = _load_store(); _v59_bootstrap_store(store)
    account = next((a for a in store.setdefault("chart_of_accounts", []) if a.get("company_id") == company_id and a.get("code") == account_code), None)
    if not account:
        _api_error("Không tìm thấy tài khoản.", [{"field": "account_code", "message": account_code}], 404)
    for key, value in req.dict(exclude_unset=True).items():
        account[key] = value
    account["updated_at"] = _now()
    _v64_audit_event(store, "account.update", "account", account_code, req.dict(exclude_unset=True), company_id=company_id)
    _save_store(store)
    return _api_ok(account, "Đã cập nhật tài khoản.")


@router.get("/accounting/accounts/search")
def v60_search_accounts(q: str = "", company_id: str = _V59_DEFAULT_COMPANY_ID, limit: int = 20):
    store = _load_store(); rows = _v60_accounts_for_company(store, company_id)
    q_norm = _normalize_v43(q or "")
    if q_norm:
        rows = [a for a in rows if q_norm in _normalize_v43(a.get("code", "") + " " + a.get("name", "") + " " + a.get("type", ""))]
    return _api_ok({"company_id": company_id, "count": len(rows[:limit]), "accounts": sorted(rows, key=lambda x: x.get("code", ""))[:limit]})


@router.post("/journal-entries")
def v61_create_journal_entry(req: V61JournalEntryCreate):
    store = _load_store(); _v59_bootstrap_store(store)
    if not _v59_get_company(store, req.company_id):
        _api_error("Company không tồn tại.", [{"field": "company_id", "message": req.company_id}], 404)
    lines = _v61_normalize_lines(req.lines)
    errors = _v61_validate_journal(store, req.company_id, lines)
    if errors:
        _api_error("Validation failed", errors)
    entry = {
        "id": _next_id(store, "JE"),
        "company_id": req.company_id,
        "date": req.date,
        "description": req.description,
        "source_document_id": req.source_document_id,
        "lines": lines,
        "total_debit": round(sum(l["debit"] for l in lines), 2),
        "total_credit": round(sum(l["credit"] for l in lines), 2),
        "status": req.status if req.status in {"draft", "approved", "posted"} else "draft",
        "metadata": req.metadata,
        "created_at": _now(),
        "updated_at": _now(),
        "posted_at": _now() if req.status == "posted" else None,
    }
    store.setdefault("v61_journal_entries", []).append(entry)
    _v64_audit_event(store, "journal.create", "journal_entry", entry["id"], {"status": entry["status"], "total": entry["total_debit"]}, actor=req.actor, company_id=req.company_id)
    _save_store(store)
    return _api_ok(entry, "Đã tạo bút toán. Tổng Nợ = Tổng Có.")


@router.get("/journal-entries")
def v61_list_journal_entries(company_id: str = _V59_DEFAULT_COMPANY_ID, status: Optional[str] = None, date_from: Optional[str] = None, date_to: Optional[str] = None, limit: int = 200):
    store = _load_store(); _v59_bootstrap_store(store)
    rows = [e for e in store.setdefault("v61_journal_entries", []) if e.get("company_id") == company_id]
    if status:
        rows = [e for e in rows if e.get("status") == status]
    if date_from:
        rows = [e for e in rows if str(e.get("date", "")) >= date_from]
    if date_to:
        rows = [e for e in rows if str(e.get("date", "")) <= date_to]
    rows = sorted(rows, key=lambda x: (x.get("date", ""), x.get("id", "")), reverse=True)[:limit]
    return _api_ok({"company_id": company_id, "count": len(rows), "journal_entries": rows})


@router.get("/journal-entries/{entry_id}")
def v61_get_journal_entry(entry_id: str):
    store = _load_store(); entry = next((e for e in store.setdefault("v61_journal_entries", []) if e.get("id") == entry_id), None)
    if not entry:
        _api_error("Không tìm thấy bút toán.", [{"field": "entry_id", "message": entry_id}], 404)
    return _api_ok(entry)


@router.put("/journal-entries/{entry_id}")
def v61_update_journal_entry(entry_id: str, req: V61JournalEntryUpdate):
    store = _load_store(); _v59_bootstrap_store(store)
    entry = next((e for e in store.setdefault("v61_journal_entries", []) if e.get("id") == entry_id), None)
    if not entry:
        _api_error("Không tìm thấy bút toán.", [{"field": "entry_id", "message": entry_id}], 404)
    if entry.get("status") == "posted":
        _api_error("Không sửa trực tiếp bút toán đã posted. Hãy tạo bút toán điều chỉnh.", [{"field": "status", "message": "posted"}])
    updates = req.dict(exclude_unset=True)
    if "lines" in updates and updates["lines"] is not None:
        lines = _v61_normalize_lines(updates["lines"])
        errors = _v61_validate_journal(store, entry.get("company_id", _V59_DEFAULT_COMPANY_ID), lines)
        if errors:
            _api_error("Validation failed", errors)
        entry["lines"] = lines
        entry["total_debit"] = round(sum(l["debit"] for l in lines), 2)
        entry["total_credit"] = round(sum(l["credit"] for l in lines), 2)
    for key in ["date", "description", "source_document_id", "status", "metadata"]:
        if key in updates and updates[key] is not None:
            entry[key] = updates[key]
    entry["updated_at"] = _now()
    _v64_audit_event(store, "journal.update", "journal_entry", entry_id, updates, actor=req.actor, company_id=entry.get("company_id"))
    _save_store(store)
    return _api_ok(entry, "Đã cập nhật bút toán.")


@router.delete("/journal-entries/{entry_id}")
def v61_delete_journal_entry(entry_id: str, actor: str = "api"):
    store = _load_store(); rows = store.setdefault("v61_journal_entries", [])
    entry = next((e for e in rows if e.get("id") == entry_id), None)
    if not entry:
        _api_error("Không tìm thấy bút toán.", [{"field": "entry_id", "message": entry_id}], 404)
    if entry.get("status") == "posted":
        _api_error("Không xóa bút toán đã posted. Hãy tạo bút toán điều chỉnh.", [{"field": "status", "message": "posted"}])
    rows.remove(entry)
    _v64_audit_event(store, "journal.delete", "journal_entry", entry_id, {"description": entry.get("description")}, actor=actor, company_id=entry.get("company_id"))
    _save_store(store)
    return _api_ok({"deleted_id": entry_id}, "Đã xóa bút toán nháp/chưa posted.")


@router.post("/documents/{document_id}/submit-review")
def v62_submit_document_review(document_id: str, req: V62WorkflowAction):
    store = _load_store(); _v59_bootstrap_store(store)
    wf = _v62_find_workflow(store, document_id)
    if not wf:
        wf = {"id": _next_id(store, "WF"), "document_id": document_id, "company_id": req.company_id, "status": "draft", "history": [], "created_at": _now(), "updated_at": _now(), "journal_entry_id": None}
        store.setdefault("document_workflows", []).append(wf)
    wf["status"] = "pending_approval"
    wf["updated_at"] = _now()
    wf.setdefault("history", []).append({"at": _now(), "actor": req.actor, "action": "submit-review", "note": req.note})
    _v64_audit_event(store, "document.submit_review", "document", document_id, {"note": req.note}, actor=req.actor, company_id=req.company_id)
    _save_store(store)
    return _api_ok(wf, "Chứng từ đã chuyển sang trạng thái chờ duyệt.")


@router.post("/documents/{document_id}/approve")
def v62_approve_document(document_id: str, req: V62WorkflowAction):
    store = _load_store(); _v59_bootstrap_store(store)
    wf = _v62_find_workflow(store, document_id)
    if not wf:
        wf = {"id": _next_id(store, "WF"), "document_id": document_id, "company_id": req.company_id, "status": "pending_approval", "history": [], "created_at": _now(), "updated_at": _now(), "journal_entry_id": None}
        store.setdefault("document_workflows", []).append(wf)
    wf["status"] = "approved"
    if req.journal_entry_id:
        wf["journal_entry_id"] = req.journal_entry_id
    wf["updated_at"] = _now()
    wf.setdefault("history", []).append({"at": _now(), "actor": req.actor, "action": "approve", "note": req.note, "journal_entry_id": req.journal_entry_id})
    _v64_audit_event(store, "document.approve", "document", document_id, {"journal_entry_id": req.journal_entry_id, "note": req.note}, actor=req.actor, company_id=wf.get("company_id"))
    _save_store(store)
    return _api_ok(wf, "Chứng từ đã được duyệt.")


@router.post("/documents/{document_id}/reject")
def v62_reject_document(document_id: str, req: V62WorkflowAction):
    store = _load_store(); _v59_bootstrap_store(store)
    wf = _v62_find_workflow(store, document_id)
    if not wf:
        wf = {"id": _next_id(store, "WF"), "document_id": document_id, "company_id": req.company_id, "status": "pending_approval", "history": [], "created_at": _now(), "updated_at": _now(), "journal_entry_id": None}
        store.setdefault("document_workflows", []).append(wf)
    wf["status"] = "rejected"
    wf["updated_at"] = _now()
    wf.setdefault("history", []).append({"at": _now(), "actor": req.actor, "action": "reject", "reason": req.rejection_reason or req.note})
    _v64_audit_event(store, "document.reject", "document", document_id, {"reason": req.rejection_reason or req.note}, actor=req.actor, company_id=wf.get("company_id"))
    _save_store(store)
    return _api_ok(wf, "Chứng từ đã bị từ chối.")


@router.post("/documents/{document_id}/post-to-journal")
def v62_post_document_to_journal(document_id: str, req: V62WorkflowAction):
    store = _load_store(); _v59_bootstrap_store(store)
    wf = _v62_find_workflow(store, document_id)
    if not wf or wf.get("status") != "approved":
        _api_error("Chỉ chứng từ đã approved mới được ghi sổ.", [{"field": "status", "message": (wf or {}).get("status", "missing_workflow")}])
    journal_id = req.journal_entry_id or wf.get("journal_entry_id")
    entry = next((e for e in store.setdefault("v61_journal_entries", []) if e.get("id") == journal_id), None)
    if not entry:
        _api_error("Không tìm thấy journal_entry_id để ghi sổ.", [{"field": "journal_entry_id", "message": str(journal_id)}], 404)
    errors = _v61_validate_journal(store, entry.get("company_id", req.company_id), entry.get("lines") or [])
    if errors:
        _api_error("Journal entry không hợp lệ, chưa thể posted.", errors)
    entry["status"] = "posted"
    entry["posted_at"] = _now()
    entry["updated_at"] = _now()
    wf["status"] = "posted"
    wf["journal_entry_id"] = journal_id
    wf["updated_at"] = _now()
    wf.setdefault("history", []).append({"at": _now(), "actor": req.actor, "action": "post-to-journal", "journal_entry_id": journal_id, "note": req.note})
    _v64_audit_event(store, "document.post_to_journal", "document", document_id, {"journal_entry_id": journal_id}, actor=req.actor, company_id=entry.get("company_id"))
    _v64_audit_event(store, "journal.post", "journal_entry", journal_id, {"source_document_id": document_id}, actor=req.actor, company_id=entry.get("company_id"))
    _save_store(store)
    return _api_ok({"workflow": wf, "journal_entry": entry}, "Đã ghi sổ chứng từ vào journal.")


@router.get("/reports/trial-balance")
def v63_trial_balance(company_id: str = _V59_DEFAULT_COMPANY_ID, date_from: Optional[str] = None, date_to: Optional[str] = None, posted_only: bool = True):
    store = _load_store(); _v59_bootstrap_store(store)
    type_map = _account_type_map(store, company_id)
    accounts = {a.get("code"): a for a in _v60_accounts_for_company(store, company_id)}
    bucket: Dict[str, Dict[str, Any]] = {}
    entries = [e for e in store.setdefault("v61_journal_entries", []) if e.get("company_id") == company_id]
    if posted_only:
        entries = [e for e in entries if e.get("status") == "posted"]
    if date_from:
        entries = [e for e in entries if str(e.get("date", "")) >= date_from]
    if date_to:
        entries = [e for e in entries if str(e.get("date", "")) <= date_to]
    for e in entries:
        for line in e.get("lines") or []:
            code = line.get("account")
            rec = bucket.setdefault(code, {"account": code, "account_name": accounts.get(code, {}).get("name", ""), "type": type_map.get(code, "other"), "debit": 0.0, "credit": 0.0})
            rec["debit"] += float(line.get("debit") or 0)
            rec["credit"] += float(line.get("credit") or 0)
    rows = []
    for rec in bucket.values():
        rec["debit"] = round(rec["debit"], 2); rec["credit"] = round(rec["credit"], 2)
        rec["ending_balance"] = round(_v63_signed_delta(rec["type"], rec["debit"], rec["credit"]), 2)
        rows.append(rec)
    totals = {"debit": round(sum(r["debit"] for r in rows), 2), "credit": round(sum(r["credit"] for r in rows), 2)}
    return _api_ok({"company_id": company_id, "posted_only": posted_only, "date_from": date_from, "date_to": date_to, "totals": totals, "rows": sorted(rows, key=lambda x: x.get("account", ""))})


@router.get("/reports/general-ledger")
def v63_general_ledger(company_id: str = _V59_DEFAULT_COMPANY_ID, account: Optional[str] = None, date_from: Optional[str] = None, date_to: Optional[str] = None, posted_only: bool = True):
    store = _load_store(); _v59_bootstrap_store(store)
    type_map = _account_type_map(store, company_id)
    rows = []
    entries = [e for e in store.setdefault("v61_journal_entries", []) if e.get("company_id") == company_id]
    if posted_only:
        entries = [e for e in entries if e.get("status") == "posted"]
    if date_from:
        entries = [e for e in entries if str(e.get("date", "")) >= date_from]
    if date_to:
        entries = [e for e in entries if str(e.get("date", "")) <= date_to]
    for e in sorted(entries, key=lambda x: (x.get("date", ""), x.get("id", ""))):
        for line in e.get("lines") or []:
            code = line.get("account")
            if account and code != account:
                continue
            rows.append({"date": e.get("date"), "entry_id": e.get("id"), "description": line.get("description") or e.get("description"), "account": code, "debit": line.get("debit", 0), "credit": line.get("credit", 0), "status": e.get("status")})
    running = 0.0
    if account:
        acc_type = type_map.get(account, "other")
        for r in rows:
            running += _v63_signed_delta(acc_type, float(r["debit"] or 0), float(r["credit"] or 0))
            r["running_balance"] = round(running, 2)
    return _api_ok({"company_id": company_id, "account": account, "count": len(rows), "rows": rows})


@router.get("/reports/income-statement")
def v63_income_statement(company_id: str = _V59_DEFAULT_COMPANY_ID, date_from: Optional[str] = None, date_to: Optional[str] = None):
    tb = v63_trial_balance(company_id=company_id, date_from=date_from, date_to=date_to, posted_only=True)["data"]["rows"]
    revenue = sum(r["ending_balance"] for r in tb if str(r.get("account", "")).startswith(("511", "515", "711")))
    expense = sum(abs(r["ending_balance"]) for r in tb if str(r.get("account", "")).startswith(("632", "635", "641", "642", "811")))
    return _api_ok({"company_id": company_id, "date_from": date_from, "date_to": date_to, "revenue": round(revenue, 2), "expense": round(expense, 2), "profit_before_tax": round(revenue - expense, 2), "note": "MVP theo tài khoản posted; chưa xử lý bút toán kết chuyển/thuế TNDN đầy đủ."})


@router.get("/reports/balance-sheet")
def v63_balance_sheet(company_id: str = _V59_DEFAULT_COMPANY_ID, date_to: Optional[str] = None):
    tb = v63_trial_balance(company_id=company_id, date_to=date_to, posted_only=True)["data"]["rows"]
    assets = sum(r["ending_balance"] for r in tb if r.get("type") in {"asset", "contra_asset"})
    liabilities = sum(r["ending_balance"] for r in tb if r.get("type") == "liability")
    equity = sum(r["ending_balance"] for r in tb if r.get("type") == "equity")
    return _api_ok({"company_id": company_id, "date_to": date_to, "assets": round(assets, 2), "liabilities": round(liabilities, 2), "equity": round(equity, 2), "check_difference": round(assets - liabilities - equity, 2), "note": "MVP balance sheet theo số dư tài khoản; cần mapping báo cáo chi tiết cho production."})


@router.get("/audit-logs")
def v64_list_audit_logs(company_id: Optional[str] = None, document_id: Optional[str] = None, entry_id: Optional[str] = None, action: Optional[str] = None, limit: int = 300):
    store = _load_store(); rows = store.setdefault("audit_logs", [])
    if company_id:
        rows = [r for r in rows if r.get("company_id") == company_id]
    if document_id:
        rows = [r for r in rows if r.get("entity_id") == document_id or (r.get("detail") or {}).get("source_document_id") == document_id]
    if entry_id:
        rows = [r for r in rows if r.get("entity_id") == entry_id or (r.get("detail") or {}).get("journal_entry_id") == entry_id]
    if action:
        rows = [r for r in rows if r.get("action") == action]
    rows = sorted(rows, key=lambda x: x.get("created_at", ""), reverse=True)[:limit]
    return _api_ok({"count": len(rows), "audit_logs": rows})


@router.get("/ai/v65/response-format")
def v65_response_format():
    return _api_ok({
        "success_example": {"success": True, "data": {"id": "..."}, "message": "OK", "errors": []},
        "error_example": {"success": False, "data": None, "message": "Validation failed", "errors": [{"field": "lines", "message": "Tổng Nợ phải bằng tổng Có"}]},
        "frontend_note": "Các API V59-V65 trả JSON chuẩn để frontend render form, table, workflow, báo cáo và audit log.",
    }, "V65 response format chuẩn cho frontend.")


@router.get("/ai/v59-v65/upgrade-status")
def v59_v65_upgrade_status():
    store = _load_store(); _v59_bootstrap_store(store); _save_store(store)
    return _api_ok({
        "stage": "Finiip V59-V65 backend-only accounting workflow",
        "frontend_policy": "Không thêm UI. Frontend riêng gọi API JSON.",
        "completed": [
            "V59: Company / tenant management",
            "V60: Chart of Accounts TT200 seed + CRUD/search",
            "V61: General journal entries with debit=credit validation",
            "V62: Document approval workflow before posting",
            "V63: Trial balance, general ledger, income statement, balance sheet MVP",
            "V64: Audit logs for company/account/journal/document actions",
            "V65: Standard API response format for frontend",
        ],
        "counts": {
            "companies": len(store.setdefault("companies", [])),
            "accounts": len(store.setdefault("chart_of_accounts", [])),
            "journal_entries": len(store.setdefault("v61_journal_entries", [])),
            "document_workflows": len(store.setdefault("document_workflows", [])),
            "audit_logs": len(store.setdefault("audit_logs", [])),
        },
        "main_apis": [
            "POST /companies", "GET /companies", "GET /companies/{company_id}", "PUT /companies/{company_id}",
            "GET /accounting/accounts", "POST /accounting/accounts", "PUT /accounting/accounts/{account_code}", "GET /accounting/accounts/search",
            "POST /journal-entries", "GET /journal-entries", "GET /journal-entries/{entry_id}", "PUT /journal-entries/{entry_id}", "DELETE /journal-entries/{entry_id}",
            "POST /documents/{document_id}/submit-review", "POST /documents/{document_id}/approve", "POST /documents/{document_id}/reject", "POST /documents/{document_id}/post-to-journal",
            "GET /reports/trial-balance", "GET /reports/general-ledger", "GET /reports/income-statement", "GET /reports/balance-sheet",
            "GET /audit-logs", "GET /ai/v65/response-format", "GET /ai/v59-v65/upgrade-status",
        ],
    }, "V59-V65 đã sẵn sàng cho frontend tích hợp.")

# ---------------------------------------------------------------------------
# V53.1-V63.1: Quality upgrade for existing backend functions only.
# Keeps old endpoints compatible, enriches JSON responses for frontend clients.
# ---------------------------------------------------------------------------

_V53_1_CHECKLISTS = {
    "purchase_invoice": [
        ("invoice_number", "Có số hóa đơn/ký hiệu hóa đơn", ["so hoa don", "số hóa đơn", "ký hiệu", "ky hieu", "invoice no"]),
        ("invoice_date", "Có ngày hóa đơn/ngày phát sinh", ["ngay", "ngày", "/", "thang", "tháng"]),
        ("seller_tax_code", "Có MST/thông tin người bán", ["ma so thue", "mã số thuế", "mst", "nguoi ban", "người bán", "seller"]),
        ("buyer_info", "Có thông tin người mua/đơn vị mua", ["nguoi mua", "người mua", "don vi mua", "đơn vị mua", "buyer"]),
        ("amount", "Có tiền trước thuế/tổng tiền", ["tong tien", "tổng tiền", "thanh toán", "thanh toan", "thành tiền"]),
        ("vat", "Có thông tin VAT/GTGT nếu là hóa đơn VAT", ["vat", "gtgt", "thuế", "thue"]),
        ("payment", "Có phương thức thanh toán", ["chuyen khoan", "chuyển khoản", "ngan hang", "ngân hàng", "tien mat", "tiền mặt", "bank"]),
    ],
    "sale_invoice": [
        ("invoice_number", "Có số hóa đơn/ký hiệu hóa đơn", ["so hoa don", "số hóa đơn", "ký hiệu", "ky hieu"]),
        ("invoice_date", "Có ngày hóa đơn", ["ngay", "ngày", "/", "thang", "tháng"]),
        ("customer", "Có thông tin khách hàng/người mua", ["nguoi mua", "người mua", "khach hang", "khách hàng", "customer"]),
        ("revenue_amount", "Có doanh thu/tổng tiền", ["doanh thu", "tong tien", "tổng tiền", "thanh toán", "thanh toan"]),
        ("output_vat", "Có VAT/GTGT đầu ra nếu áp dụng", ["vat", "gtgt", "thuế", "thue"]),
    ],
    "contract": [
        ("parties", "Có bên A/bên B hoặc thông tin các bên", ["ben a", "bên a", "ben b", "bên b", "party"]),
        ("value", "Có giá trị hợp đồng", ["gia tri", "giá trị", "tong gia tri", "tổng giá trị", "thanh toán"]),
        ("term", "Có thời hạn/hiệu lực", ["hieu luc", "hiệu lực", "thoi han", "thời hạn", "ngày ký"]),
        ("acceptance", "Có điều kiện nghiệm thu/bàn giao nếu là dịch vụ", ["nghiem thu", "nghiệm thu", "ban giao", "bàn giao"]),
    ],
    "cash_payment": [
        ("voucher_no", "Có số phiếu chi", ["phieu chi", "phiếu chi", "so phieu", "số phiếu"]),
        ("receiver", "Có người nhận tiền", ["nguoi nhan", "người nhận", "receiver"]),
        ("reason", "Có lý do chi/diễn giải", ["ly do", "lý do", "dien giai", "diễn giải", "noi dung", "nội dung"]),
        ("amount", "Có số tiền chi", ["so tien", "số tiền", "tong tien", "tổng tiền"]),
        ("signatures", "Có chữ ký/phê duyệt", ["ky", "ký", "chu ky", "chữ ký", "duyet", "duyệt"]),
    ],
    "cash_receipt": [
        ("voucher_no", "Có số phiếu thu", ["phieu thu", "phiếu thu", "so phieu", "số phiếu"]),
        ("payer", "Có người nộp tiền", ["nguoi nop", "người nộp", "payer"]),
        ("reason", "Có lý do thu/diễn giải", ["ly do", "lý do", "dien giai", "diễn giải", "noi dung", "nội dung"]),
        ("amount", "Có số tiền thu", ["so tien", "số tiền", "tong tien", "tổng tiền"]),
    ],
    "bank_statement": [
        ("bank_account", "Có tài khoản ngân hàng", ["tai khoan", "tài khoản", "account", "stk"]),
        ("transaction_date", "Có ngày giao dịch", ["ngay giao dich", "ngày giao dịch", "date", "/"]),
        ("amount", "Có số tiền giao dịch", ["so tien", "số tiền", "amount", "ghi no", "ghi có", "ghi co"]),
        ("counterparty", "Có đối tượng/nội dung giao dịch", ["noi dung", "nội dung", "dien giai", "diễn giải", "beneficiary", "sender"]),
    ],
    "payroll": [
        ("employee", "Có nhân viên/người lao động", ["nhan vien", "nhân viên", "nguoi lao dong", "người lao động"]),
        ("gross_salary", "Có lương gross/tổng thu nhập", ["tong thu nhap", "tổng thu nhập", "gross", "luong", "lương"]),
        ("insurance", "Có BHXH/BHYT/BHTN nếu áp dụng", ["bhxh", "bhyt", "bhtn", "bao hiem", "bảo hiểm"]),
        ("pit", "Có thuế TNCN nếu áp dụng", ["tncn", "thuế thu nhập cá nhân", "thue thu nhap ca nhan"]),
    ],
    "unknown": [
        ("date", "Có ngày chứng từ", ["ngay", "ngày", "/"]),
        ("amount", "Có số tiền", ["so tien", "số tiền", "tong tien", "tổng tiền", "amount"]),
        ("counterparty", "Có đối tượng liên quan", ["nguoi ban", "người bán", "nguoi mua", "người mua", "khach hang", "nhà cung cấp"]),
        ("description", "Có diễn giải/nội dung nghiệp vụ", ["dien giai", "diễn giải", "noi dung", "nội dung"]),
    ],
}


def _v53_1_keyword_present(norm_text: str, raw_text: str, keywords: List[str]) -> bool:
    raw_lower = (raw_text or "").lower()
    for kw in keywords:
        nkw = _normalize_v43(kw)
        if (nkw and nkw in norm_text) or kw.lower() in raw_lower:
            return True
    return False


def _v53_1_document_checklist(doc_type: str, text: str, detected: Dict[str, Any]) -> Dict[str, Any]:
    norm = _normalize_v43(text or "")
    template = _V53_1_CHECKLISTS.get(doc_type) or _V53_1_CHECKLISTS["unknown"]
    items = []
    passed = warnings = failed = 0
    for code, label, keywords in template:
        status = "pass" if _v53_1_keyword_present(norm, text, keywords) else "missing"
        if code in {"invoice_date", "transaction_date", "date"} and detected.get("dates"):
            status = "pass"
        if code in {"amount", "revenue_amount", "value", "gross_salary"} and detected.get("amount_candidates"):
            status = "pass"
        if code in {"vat", "output_vat"} and not detected.get("has_vat_keyword"):
            status = "warning"
        if code == "payment" and not (detected.get("has_bank_payment_keyword") or detected.get("has_cash_payment_keyword")):
            status = "warning"
        if status == "pass":
            passed += 1
        elif status == "warning":
            warnings += 1
        else:
            failed += 1
        items.append({"code": code, "label": label, "status": status})
    total = max(len(items), 1)
    score = round((passed + warnings * 0.5) / total * 100, 1)
    if failed >= 3 or score < 50:
        overall = "need_more_info"
    elif warnings or failed:
        overall = "need_review"
    else:
        overall = "ready_for_accountant_review"
    return {"overall_status": overall, "quality_score": score, "passed": passed, "warnings": warnings, "missing": failed, "items": items}


def _v53_basic_document_review(title: str, text: str, source: str = "manual", tags: Optional[List[str]] = None) -> Dict[str, Any]:
    classification = _v55_classify_accounting_document(text, title, tags)
    norm = _normalize_v43(text)
    amounts = _v53_extract_amount_candidates(text)
    dates = _v53_extract_dates(text)
    has_vat = any(k in norm for k in ["vat", "gtgt", "thue gia tri gia tang", "thuế giá trị gia tăng"])
    has_invoice = any(k in norm for k in ["hoa don", "hóa đơn", "invoice"])
    has_contract = any(k in norm for k in ["hop dong", "hợp đồng", "contract"])
    has_bank_payment = any(k in norm for k in ["chuyen khoan", "chuyển khoản", "ngan hang", "ngân hàng", "bank"])
    has_cash_payment = any(k in norm for k in ["tien mat", "tiền mặt", "cash"])
    total_amount = amounts[0] if amounts else None
    vat_guess = None
    amount_before_vat = None
    if total_amount and has_vat:
        amount_before_vat = round(total_amount / 1.1, 2)
        vat_guess = round(total_amount - amount_before_vat, 2)
    detected = {
        "dates": dates,
        "amount_candidates": amounts,
        "total_amount_guess": total_amount,
        "amount_before_vat_guess": amount_before_vat,
        "vat_amount_guess": vat_guess,
        "has_vat_keyword": has_vat,
        "has_invoice_keyword": has_invoice,
        "has_contract_keyword": has_contract,
        "has_bank_payment_keyword": has_bank_payment,
        "has_cash_payment_keyword": has_cash_payment,
    }
    checklist = _v53_1_document_checklist(classification["document_type"], text, detected)
    missing = []
    for item in checklist["items"]:
        if item["status"] in {"missing", "warning"}:
            missing.append(item["label"])
    if classification["document_type"] in {"purchase_invoice", "sale_invoice"} and not has_invoice:
        missing.append("Chưa thấy dấu hiệu hóa đơn rõ ràng")
    if total_amount and total_amount >= 5_000_000 and not has_bank_payment:
        missing.append("Giao dịch từ 5 triệu đồng trở lên: cần kiểm tra chứng từ thanh toán không dùng tiền mặt")
    risk_level = "high" if checklist["overall_status"] == "need_more_info" else "medium" if checklist["overall_status"] == "need_review" else "low"
    confidence = "high" if classification.get("confidence") == "high" and checklist["quality_score"] >= 75 else "medium" if checklist["quality_score"] >= 50 else "low"
    return {
        "classification": classification,
        "detected_fields": detected,
        "document_checklist": checklist,
        "overall_status": checklist["overall_status"],
        "quality_score": checklist["quality_score"],
        "confidence": confidence,
        "risk_level": risk_level,
        "missing_info": list(dict.fromkeys(missing)),
        "review_summary": " ".join([
            f"Loại tài liệu dự đoán: {classification['label']} ({classification['confidence']}).",
            f"Điểm checklist: {checklist['quality_score']}/100, trạng thái: {checklist['overall_status']}.",
            f"Số tiền lớn nhất phát hiện: {total_amount:,.0f}đ." if total_amount else "Chưa phát hiện số tiền rõ ràng.",
        ]),
        "next_actions": [
            "Frontend có thể render document_checklist.items thành bảng kiểm tra chứng từ.",
            "Gọi /ai/v54/journal-suggestion để lấy nhiều phương án bút toán có kiểm tra cân đối.",
            "Gọi /ai/v56/tax-risk-checklist để lấy checklist rủi ro thuế/chứng từ.",
        ],
        "source": source,
        "tags": tags or [],
    }


def _v54_1_line_sum(lines: List[Dict[str, Any]]) -> Dict[str, float]:
    debit = credit = 0.0
    for line in lines:
        if "debit_account" in line:
            debit += float(line.get("amount") or 0)
        else:
            debit += float(line.get("debit") or 0)
            credit += float(line.get("credit") or 0)
        if "credit_account" in line:
            credit += float(line.get("amount") or 0)
    return {"debit": round(debit, 2), "credit": round(credit, 2), "difference": round(debit - credit, 2)}


def _v54_1_accounts_valid(lines: List[Dict[str, Any]], valid_accounts: Optional[set] = None) -> List[Dict[str, str]]:
    # If no company chart is supplied, validate against the TT200 defaults that this prototype seeds.
    default_codes = {"111", "112", "131", "1331", "152", "153", "156", "211", "214", "242", "331", "3331", "334", "338", "411", "421", "511", "515", "632", "635", "641", "642", "711", "811", "911"}
    valid = valid_accounts or default_codes
    errors = []
    for idx, line in enumerate(lines):
        for key in ["debit_account", "credit_account"]:
            code = str(line.get(key) or "").strip()
            if code and code not in valid:
                errors.append({"field": f"options.lines[{idx}].{key}", "message": f"Tài khoản {code} chưa có trong danh mục tài khoản."})
    return errors


def _v54_1_make_option(name: str, confidence: str, lines: List[Dict[str, Any]], explanation: str, assumptions: List[str], missing: List[str]) -> Dict[str, Any]:
    balance = _v54_1_line_sum(lines)
    validation = []
    if abs(balance["difference"]) > 0.01:
        validation.append({"field": "journal_lines", "message": "Tổng Nợ phải bằng tổng Có."})
    validation.extend(_v54_1_accounts_valid(lines))
    return {
        "name": name,
        "confidence": confidence if not validation else "low",
        "journal_lines": lines,
        "balance_check": balance,
        "is_balanced": abs(balance["difference"]) <= 0.01,
        "account_validation_errors": validation,
        "explanation": explanation,
        "assumptions": assumptions,
        "missing_info": missing,
    }


def _v54_suggest_journal_payload(payload: Dict[str, Any]) -> Dict[str, Any]:
    description = payload.get("description") or ""
    content = payload.get("content") or ""
    doc_type = payload.get("document_type") or _v55_classify_accounting_document(content, description).get("document_type")
    standard = payload.get("accounting_standard") or payload.get("standard") or "TT200"
    norm = _normalize_v43("\n".join([description, content, str(payload.get("purpose") or "")]))
    total = payload.get("total_amount")
    before_vat = payload.get("amount_before_vat")
    vat = payload.get("vat_amount")
    if total is None:
        amounts = _v53_extract_amount_candidates(content + "\n" + description)
        total = amounts[0] if amounts else 0.0
    if before_vat is None and vat is None and total:
        if any(k in norm for k in ["vat", "gtgt", "10%"]):
            before_vat = round(float(total) / 1.1, 2)
            vat = round(float(total) - before_vat, 2)
        else:
            before_vat = float(total)
            vat = 0.0
    before_vat = float(before_vat or 0.0)
    vat = float(vat or 0.0)
    total = float(total or before_vat + vat)
    payment = _normalize_v43(payload.get("payment_method") or "")
    paid_by_bank = any(k in payment or k in norm for k in ["chuyen khoan", "chuyển khoản", "ngan hang", "ngân hàng", "bank", "112"])
    paid_by_cash = any(k in payment or k in norm for k in ["tien mat", "tiền mặt", "cash", "111"])
    counter_account = "112" if paid_by_bank else "111" if paid_by_cash else "331"
    options: List[Dict[str, Any]] = []
    warnings = []
    missing = []
    if not total:
        missing.append("Chưa xác định chắc chắn số tiền chứng từ")
    if total >= 5_000_000 and not paid_by_bank:
        warnings.append("Từ 5 triệu đồng trở lên cần kiểm tra điều kiện thanh toán không dùng tiền mặt để khấu trừ VAT/chi phí thuế.")
    if doc_type == "sale_invoice" or any(k in norm for k in ["ban hang", "bán hàng", "doanh thu", "dau ra"]):
        debit = "112" if paid_by_bank else "111" if paid_by_cash else "131"
        lines = [{"debit_account": debit, "credit_account": "511", "amount": before_vat, "description": "Ghi nhận doanh thu chưa VAT"}]
        if vat:
            lines.append({"debit_account": debit, "credit_account": "3331", "amount": vat, "description": "Thuế GTGT đầu ra phải nộp"})
        options.append(_v54_1_make_option("Ghi nhận doanh thu bán hàng/dịch vụ", "high" if total else "low", lines, "Dùng 511 cho doanh thu, 3331 cho VAT đầu ra; Nợ tiền/khách hàng tùy tình trạng thanh toán.", ["Mặc định là doanh thu chịu VAT nếu chứng từ có VAT."], missing))
    else:
        base_assumptions = ["Mặc định là chứng từ đầu vào hoặc chi phí mua ngoài."]
        if vat == 0 and any(k in norm for k in ["vat", "gtgt"]):
            missing.append("Có nhắc VAT nhưng chưa tách được số VAT")
        if any(k in norm for k in ["laptop", "may tinh", "máy tính", "tai san co dinh", "tài sản cố định", "khau hao", "khấu hao"]):
            lines_ccdc = [{"debit_account": "242", "credit_account": counter_account, "amount": before_vat, "description": "Ghi nhận CCDC/chi phí trả trước"}]
            if vat:
                lines_ccdc.append({"debit_account": "1331", "credit_account": counter_account, "amount": vat, "description": "VAT đầu vào được khấu trừ nếu đủ điều kiện"})
            options.append(_v54_1_make_option("Phương án CCDC/chi phí trả trước", "high" if before_vat < 30000000 and total else "medium", lines_ccdc, "Dùng 242 khi tài sản/công cụ cần phân bổ nhiều kỳ, đặc biệt khi chưa đủ điều kiện ghi TSCĐ.", base_assumptions + ["Cần chính sách ghi nhận CCDC/TSCĐ của công ty."], missing + ["Chưa rõ thời gian sử dụng dự kiến và ngưỡng TSCĐ nội bộ."]))
            lines_fa = [{"debit_account": "211", "credit_account": counter_account, "amount": before_vat, "description": "Ghi nhận tài sản cố định hữu hình"}]
            if vat:
                lines_fa.append({"debit_account": "1331", "credit_account": counter_account, "amount": vat, "description": "VAT đầu vào được khấu trừ nếu đủ điều kiện"})
            options.append(_v54_1_make_option("Phương án tài sản cố định", "medium" if before_vat >= 30000000 else "low", lines_fa, "Dùng 211 nếu đủ điều kiện ghi nhận TSCĐ theo chính sách và quy định hiện hành.", base_assumptions, missing + ["Cần quyết định đưa vào sử dụng/biên bản bàn giao và thời gian sử dụng."]))
        else:
            debit = "642"
            label = "Chi phí quản lý doanh nghiệp"
            if any(k in norm for k in ["quang cao", "quảng cáo", "marketing", "facebook", "google ads"]):
                debit, label = "641", "Chi phí bán hàng/quảng cáo"
            elif any(k in norm for k in ["hang hoa", "hàng hóa", "nhap kho", "nhập kho", "vat tu", "vật tư"]):
                debit, label = "156", "Mua hàng hóa nhập kho"
            elif any(k in norm for k in ["nguyen lieu", "nguyên liệu", "vat lieu", "vật liệu"]):
                debit, label = "152", "Mua nguyên vật liệu"
            elif any(k in norm for k in ["luong", "lương", "nhan vien", "nhân viên"]):
                debit, label = "334", "Ghi nhận phải trả người lao động"
            lines = [{"debit_account": debit, "credit_account": counter_account, "amount": before_vat, "description": label}]
            if vat and debit not in {"334"}:
                lines.append({"debit_account": "1331", "credit_account": counter_account, "amount": vat, "description": "VAT đầu vào được khấu trừ nếu đủ điều kiện"})
            options.append(_v54_1_make_option(f"Phương án mặc định: {label}", "medium" if total else "low", lines, "Tài khoản được chọn theo từ khóa nghiệp vụ trong mô tả/chứng từ; kế toán cần kiểm tra lại trước khi ghi sổ.", base_assumptions, missing))
    primary = options[0] if options else None
    return {
        "journal_type": "sale" if primary and any(l.get("credit_account") == "511" for l in primary["journal_lines"]) else "purchase_or_expense",
        "document_type": doc_type,
        "accounting_standard": standard,
        "currency": "VND",
        "amount_before_vat": before_vat,
        "vat_amount": vat,
        "total_amount": total,
        "options": options,
        "suggested_lines": (primary or {}).get("journal_lines", []),
        "primary_option": primary,
        "warnings": warnings + ["Đây là đề xuất backend/AI; cần kế toán kiểm tra chứng từ gốc và quy định hiện hành trước khi ghi sổ."],
        "missing_info": list(dict.fromkeys(missing)),
        "confidence": (primary or {}).get("confidence", "low"),
    }


def _v58_1_build_answer(query: str, results: List[Dict[str, Any]]) -> Dict[str, Any]:
    usable = [r for r in results if float(r.get("score") or 0) > 0]
    if not usable:
        return {
            "conclusion": "Chưa đủ căn cứ trong kho tài liệu đã upload để trả lời chắc chắn.",
            "answer": "Backend không tìm thấy nguồn phù hợp trong RAG local. Hãy upload thông tư/quy định/tài liệu nội bộ liên quan rồi hỏi lại.",
            "confidence": "low",
            "risk_level": "medium",
            "sources": [],
            "missing_info": ["Thiếu nguồn văn bản trong knowledge base", "Chưa có trích đoạn đủ liên quan"],
        }
    top_score = max(float(r.get("score") or 0) for r in usable)
    confidence = "high" if top_score >= 0.75 and len(usable) >= 2 else "medium" if top_score >= 0.35 else "low"
    srcs = []
    for r in usable[:5]:
        srcs.append({
            "document_id": r.get("document_id"),
            "title": r.get("title"),
            "score": r.get("score"),
            "chunk": (r.get("content") or r.get("chunk") or "")[:800],
            "tags": r.get("tags") or [],
        })
    return {
        "conclusion": "Có căn cứ tham khảo trong kho RAG, nhưng vẫn cần kiểm tra văn bản gốc trước khi áp dụng.",
        "answer": "Các trích đoạn liên quan đã được trả về trong sources. Frontend có thể hiển thị nguồn trước, sau đó cho kế toán/AI tổng hợp theo nghiệp vụ cụ thể.",
        "confidence": confidence,
        "risk_level": "low" if confidence == "high" else "medium",
        "sources": srcs,
        "missing_info": [] if confidence != "low" else ["Điểm liên quan của nguồn còn thấp; nên upload thêm nguồn đúng chủ đề."],
    }


class V58RagAnswerRequest(BaseModel):
    query: str
    category: Optional[str] = None
    limit: int = Field(5, ge=1, le=20)
    min_score: float = 0.0


@router.post("/ai/v58/rag-answer")
def v58_1_rag_answer(req: V58RagAnswerRequest):
    store = _load_store()
    query = req.query if not req.category else f"{req.category} {req.query}"
    results = _v46_vector_search(store, query, req.limit, req.min_score)
    if req.category:
        results = [r for r in results if req.category in (r.get("tags") or []) or req.category in _normalize_v43(r.get("title") or "")]
    answer = _v58_1_build_answer(req.query, results)
    return _api_ok({"query": req.query, "category": req.category, **answer}, "V58.1 RAG answer có sources/confidence/missing_info.")



# =========================
# V64/V65 - Long Exam Legal Solver
# =========================

def _v64_strip_accents(text: str) -> str:
    text = unicodedata.normalize("NFD", text or "")
    return "".join(ch for ch in text if unicodedata.category(ch) != "Mn")


def _v64_norm(text: str) -> str:
    return re.sub(r"\s+", " ", _v64_strip_accents(text).lower()).strip()


def _v64_detect_intent(question: str) -> str:
    q = _v64_norm(question)
    legal_terms = [
        "thong tu", "nghi dinh", "luat", "quy dinh", "dieu kien", "can cu",
        "co duoc", "co phai", "bat buoc", "muc phat", "thoi han", "duoc tru",
        "khong duoc tru", "hoa don", "chung tu", "quyet toan", "thue tndn", "thue gtgt",
    ]
    entry_terms = ["dinh khoan", "hach toan", "but toan", "no tk", "co tk", "no tai khoan", "co tai khoan"]
    calc_terms = [
        "tinh", "bao nhieu", "vat", "gtgt", "khau hao", "phan bo", "ccdc",
        "gia xuat kho", "binh quan", "fifo", "loi nhuan", "gia von", "tndn", "thue phai nop",
    ]
    has_legal = any(t in q for t in legal_terms)
    has_entry = any(t in q for t in entry_terms)
    has_calc = any(t in q for t in calc_terms) or bool(_money_values_v43(question))
    is_long = len(question or "") >= 260 or len(re.findall(r"[\.\?\!;\n]", question or "")) >= 2
    if has_legal and (has_calc or has_entry or is_long):
        return "hybrid_legal_exam"
    if has_legal:
        return "legal_rag"
    if has_entry and has_calc:
        return "calculation_entry"
    if has_entry:
        return "accounting_entry"
    if has_calc:
        return "calculation"
    if is_long:
        return "long_question"
    return "general_accounting"


def _v64_split_question(question: str) -> List[Dict[str, Any]]:
    raw = re.sub(r"\s+", " ", question or "").strip()
    if not raw:
        return []
    # Split by numbered clauses, semicolons, line breaks, and sentence endings, but keep meaningful chunks.
    parts = re.split(r"(?:\n+)|(?:;)+|(?:\s+(?=\d+[\)\.]\s))|(?<=[\?\.\!])\s+", raw)
    cleaned: List[str] = []
    for p in parts:
        p = re.sub(r"^\s*\d+[\)\.]\s*", "", p).strip(" .;,-")
        if len(p) >= 12:
            cleaned.append(p)
    if not cleaned:
        cleaned = [raw]
    # Merge tiny fragments into the previous item.
    merged: List[str] = []
    for p in cleaned:
        if merged and len(p) < 15:
            merged[-1] = (merged[-1] + "; " + p).strip()
        else:
            merged.append(p)
    return [{"order": i + 1, "question": p, "intent": _v64_detect_intent(p)} for i, p in enumerate(merged[:12])]


def _v64_format_money(amount: float) -> str:
    return f"{amount:,.0f}".replace(",", ".") + "đ"


def _v64_extract_facts(question: str) -> Dict[str, Any]:
    amounts = _money_values_v43(question)
    return {
        "amounts": amounts,
        "amounts_formatted": [_v64_format_money(a) for a in amounts],
        "vat_rate": _vat_rate_v43(question),
        "detected_intent": _v64_detect_intent(question),
        "has_payment_method": any(k in _v64_norm(question) for k in ["chua thanh toan", "chuyen khoan", "tien mat", "ngan hang", "phai thu", "phai tra"]),
    }


def _v64_solve_depreciation_or_allocation(question: str) -> Optional[Dict[str, Any]]:
    q = _v64_norm(question)
    amounts = _money_values_v43(question)
    if not amounts:
        return None
    months_match = re.search(r"(\d+)\s*(thang|month)", q)
    years_match = re.search(r"(\d+)\s*(nam|year)", q)
    months = None
    if months_match:
        months = int(months_match.group(1))
    elif years_match:
        months = int(years_match.group(1)) * 12
    if not months:
        return None
    cost = amounts[0]
    monthly = round(cost / months, 2)
    yearly = round(monthly * 12, 2)
    is_allocation = any(k in q for k in ["phan bo", "ccdc", "cong cu dung cu", "242"])
    title = "Phân bổ chi phí trả trước/CCDC" if is_allocation else "Khấu hao theo đường thẳng"
    debit = "642" if any(k in q for k in ["quan ly", "van phong", "bo phan quan ly"]) else "641" if any(k in q for k in ["ban hang", "sales"]) else "642"
    credit = "242" if is_allocation else "214"
    return {
        "type": "allocation" if is_allocation else "depreciation",
        "steps": [{
            "step": title,
            "journal_entry": [
                {"debit": debit, "account_name": ACCOUNTING_ACCOUNTS_V43.get(debit), "amount": monthly},
                {"credit": credit, "account_name": ACCOUNTING_ACCOUNTS_V43.get(credit), "amount": monthly},
            ],
            "explanation": f"Nguyên giá/giá trị phân bổ {_v64_format_money(cost)} trong {months} tháng: mỗi tháng {_v64_format_money(monthly)}.",
        }],
        "calculations": {"cost": cost, "months": months, "monthly_expense": monthly, "yearly_expense": yearly},
        "answer": f"{title}: {_v64_format_money(cost)} / {months} tháng = {_v64_format_money(monthly)}/tháng. Bút toán tháng: Nợ {debit} / Có {credit}: {_v64_format_money(monthly)}.",
        "assumptions": ["Mặc định phân bổ/khấu hao đều theo tháng và dùng bộ phận quản lý nếu đề không nêu rõ bộ phận sử dụng."],
        "warnings": [],
        "check": {"balanced": True, "total_debit": monthly, "total_credit": monthly},
    }


def _v64_solve_average_cost(question: str) -> Optional[Dict[str, Any]]:
    q = _v64_norm(question)
    if not any(k in q for k in ["binh quan", "gia xuat kho", "don gia binh quan"]):
        return None
    nums = [float(x.replace(",", ".")) for x in re.findall(r"\d+(?:[\.,]\d+)?", q)]
    amounts = _money_values_v43(question)
    qtys = []
    for m in re.finditer(r"(\d+(?:[\.,]\d+)?)\s*(cai|sp|san pham|hang|don vi)", q):
        qtys.append(float(m.group(1).replace(",", ".")))
    if len(amounts) < 2 or len(qtys) < 2:
        return None
    begin_value, import_value = amounts[0], amounts[1]
    begin_qty, import_qty = qtys[0], qtys[1]
    total_qty = begin_qty + import_qty
    unit_cost = round((begin_value + import_value) / total_qty, 2)
    sold_qty = qtys[2] if len(qtys) >= 3 else None
    cogs = round(unit_cost * sold_qty, 2) if sold_qty is not None else None
    answer = f"Đơn giá bình quân = ({_v64_format_money(begin_value)} + {_v64_format_money(import_value)}) / ({begin_qty:g} + {import_qty:g}) = {_v64_format_money(unit_cost)}/đơn vị."
    if cogs is not None:
        answer += f" Giá xuất kho cho {sold_qty:g} đơn vị = {_v64_format_money(cogs)}."
    return {
        "type": "average_cost",
        "steps": [],
        "calculations": {"begin_value": begin_value, "begin_qty": begin_qty, "import_value": import_value, "import_qty": import_qty, "unit_cost": unit_cost, "sold_qty": sold_qty, "cogs": cogs},
        "answer": answer,
        "assumptions": ["Dùng phương pháp bình quân sau nhập; đề cần nêu rõ nếu dùng bình quân cuối kỳ hay sau mỗi lần nhập."],
        "warnings": [] if sold_qty is not None else ["Chưa thấy số lượng xuất/bán nên mới tính được đơn giá bình quân."],
        "check": {"balanced": None},
    }


def _v64_solve_calculation(question: str) -> Dict[str, Any]:
    special = _v64_solve_depreciation_or_allocation(question) or _v64_solve_average_cost(question)
    if special:
        return special
    return _solve_accounting_problem_v43(question)


def _v64_render_exam_answer(question: str, facts: Dict[str, Any], sub_results: List[Dict[str, Any]], legal_sources: List[Dict[str, Any]], confidence: str, missing_info: List[str]) -> str:
    lines = [
        "## 1. Tóm tắt đề",
        question.strip(),
        "",
        "## 2. Dữ kiện đã phát hiện",
    ]
    if facts.get("amounts_formatted"):
        lines.append("- Số tiền: " + ", ".join(facts["amounts_formatted"]))
    lines.append(f"- Thuế suất VAT/GTGT nhận diện: {facts.get('vat_rate', 0) * 100:.0f}%")
    lines.append(f"- Loại câu hỏi: {facts.get('detected_intent')}")
    lines += ["", "## 3. Tách yêu cầu và cách xử lý"]
    for item in sub_results:
        lines.append(f"### Ý {item['order']}: {item['question']}")
        lines.append(f"- Nhóm xử lý: {item['intent']}")
        if item.get("answer"):
            lines.append(item["answer"])
        if item.get("assumptions"):
            lines.append("Giả định: " + " ".join(item["assumptions"]))
        if item.get("warnings"):
            lines.append("Cần kiểm tra: " + " ".join(item["warnings"]))
    lines += ["", "## 4. Căn cứ thông tư/nghị định / nguồn RAG"]
    if legal_sources:
        for idx, src in enumerate(legal_sources[:5], 1):
            title = src.get("title") or src.get("document_id") or src.get("doc_id") or "Nguồn chưa đặt tên"
            snippet = src.get("chunk") or src.get("snippet") or src.get("content") or ""
            lines.append(f"{idx}. {title} — điểm liên quan: {src.get('score')}. Trích đoạn: {snippet[:350]}")
    else:
        lines.append("Chưa tìm thấy nguồn RAG đủ rõ. Không nên kết luận pháp lý chắc chắn nếu chưa upload văn bản gốc.")
    lines += ["", "## 5. Kết luận", f"- Mức tin cậy: {confidence}"]
    if missing_info:
        lines.append("- Còn thiếu: " + "; ".join(dict.fromkeys(missing_info)))
    lines.append("- Với câu hỏi pháp lý/kê khai thuế, cần đối chiếu văn bản gốc đang có hiệu lực và chứng từ thực tế trước khi áp dụng.")
    return "\n".join(lines)


class V64LongExamLegalSolverRequest(BaseModel):
    question: str = Field(..., min_length=10)
    standard: str = "TT200"
    category: Optional[str] = None
    use_rag: bool = True
    require_sources: bool = True
    save_learning: bool = True
    limit: int = Field(5, ge=1, le=20)


@router.post("/ai/v64/long-exam-legal-solver")
def v64_long_exam_legal_solver(req: V64LongExamLegalSolverRequest):
    store = _load_store()
    facts = _v64_extract_facts(req.question)
    sub_questions = _v64_split_question(req.question)
    sub_results: List[Dict[str, Any]] = []
    all_sources: List[Dict[str, Any]] = []
    missing_info: List[str] = []
    full_calc_result: Optional[Dict[str, Any]] = None
    calc_inserted = False
    if facts["detected_intent"] in {"hybrid_legal_exam", "long_exam_calculation", "calculation_entry", "calculation", "accounting_entry"}:
        full_calc_result = _v64_solve_calculation(req.question)

    for sub in sub_questions:
        intent = sub["intent"]
        item: Dict[str, Any] = {**sub}
        if intent in {"legal_rag", "hybrid_legal_exam"} and req.use_rag:
            query = sub["question"] if not req.category else f"{req.category} {sub['question']}"
            results = _v67_supabase_vector_search(query, req.limit) or _v46_vector_search(store, query, req.limit, 0.0)
            if not results:
                # fallback to older keyword RAG store
                results = _rag_search_v43(store, query, req.limit)
            rag_answer = _v58_1_build_answer(sub["question"], results)
            item["answer"] = rag_answer.get("answer")
            item["sources"] = rag_answer.get("sources") or results[:5]
            item["confidence"] = rag_answer.get("confidence")
            item["missing_info"] = rag_answer.get("missing_info") or []
            all_sources.extend(item["sources"])
            missing_info.extend(item["missing_info"])
            if "calc" in intent or "exam" in intent:
                calc = full_calc_result or _v64_solve_calculation(sub["question"])
                item["calculation"] = calc.get("calculations")
                item["answer"] = (item.get("answer") or "") + "\n\nPhần tính toán/định khoản tổng hợp cho toàn đề:\n" + calc.get("answer", "")
                item["assumptions"] = calc.get("assumptions", [])
                item["warnings"] = calc.get("warnings", [])
                calc_inserted = True
        elif intent in {"calculation", "calculation_entry", "accounting_entry"}:
            if full_calc_result is not None and not calc_inserted:
                calc = full_calc_result
                item["answer"] = "Phần tính toán/định khoản tổng hợp cho toàn đề:\n" + calc.get("answer", "")
                item["steps"] = calc.get("steps", [])
                item["calculation"] = calc.get("calculations", {})
                item["assumptions"] = calc.get("assumptions", [])
                item["warnings"] = calc.get("warnings", [])
                calc_inserted = True
            else:
                item["answer"] = "Ý này thuộc phần tính toán/định khoản và đã được gộp vào lời giải tổng hợp phía trên để tránh mất dữ kiện giữa các câu."
                item["steps"] = []
                item["calculation"] = {}
                item["assumptions"] = []
                item["warnings"] = []
        else:
            item["answer"] = "Đây là ý mô tả/tổng quát. Backend đã giữ lại để AI/Frontend tổng hợp cùng các ý tính toán và căn cứ pháp lý."
        sub_results.append(item)

    # Also search legal basis once for the full long question if legal/hybrid was detected.
    if req.use_rag and facts["detected_intent"] in {"legal_rag", "hybrid_legal_exam"}:
        full_results = _v67_supabase_vector_search(req.question, req.limit) or _v46_vector_search(store, req.question, req.limit, 0.0) or _rag_search_v43(store, req.question, req.limit)
        full_rag = _v58_1_build_answer(req.question, full_results)
        all_sources.extend(full_rag.get("sources") or full_results[:5])
        missing_info.extend(full_rag.get("missing_info") or [])

    # Deduplicate sources.
    dedup_sources: List[Dict[str, Any]] = []
    seen = set()
    for src in all_sources:
        key = str(src.get("chunk_id") or src.get("document_id") or src.get("doc_id") or src.get("title") or src.get("snippet") or src.get("chunk"))[:180]
        if key and key not in seen:
            seen.add(key)
            dedup_sources.append(src)

    if req.require_sources and facts["detected_intent"] in {"legal_rag", "hybrid_legal_exam"} and not dedup_sources:
        missing_info.append("Câu hỏi có yếu tố thông tư/nghị định nhưng kho RAG chưa có nguồn phù hợp.")
    confidence = "high" if dedup_sources and not missing_info else "medium" if dedup_sources else "low"
    answer = _v64_render_exam_answer(req.question, facts, sub_results, dedup_sources, confidence, missing_info)

    if req.save_learning:
        _append_qa_learning(store, req.question, answer, "v64_long_exam_legal_solver", dedup_sources)
        _save_store(store)
    return _api_ok({
        "version": "V64/V65",
        "standard": req.standard,
        "intent": facts["detected_intent"],
        "facts": facts,
        "sub_questions": sub_results,
        "sources": dedup_sources[:10],
        "confidence": confidence,
        "missing_info": list(dict.fromkeys(missing_info)),
        "answer": answer,
        "frontend_sections": ["Tóm tắt đề", "Dữ kiện", "Tách yêu cầu", "Căn cứ pháp lý", "Tính toán/định khoản", "Kết luận/cảnh báo"],
    }, "V64/V65 xử lý câu hỏi dài kiểu đề thi + tính toán + RAG thông tư/nghị định.")


@router.get("/ai/v64/long-exam-legal-solver/status")
def v64_long_exam_legal_solver_status():
    return _api_ok({
        "version": "V64/V65",
        "ready_for": [
            "Câu hỏi dài nhiều ý",
            "Bài tính toán kế toán cơ bản: VAT, định khoản mua/bán, lợi nhuận gộp",
            "Khấu hao/phân bổ CCDC theo tháng",
            "Đơn giá bình quân xuất kho mức cơ bản",
            "Câu hỏi thông tư/nghị định qua RAG local có sources/confidence/missing_info",
        ],
        "limitations": [
            "Chưa thay thế kế toán/luật sư; bắt buộc kiểm tra chứng từ và văn bản gốc",
            "FIFO, thuế TNDN nhiều điều kiện, bảng dữ liệu phức tạp cần bổ sung solver riêng",
            "Nếu chưa upload thông tư/nghị định thì phần pháp lý sẽ trả low confidence",
        ],
        "endpoint": "POST /ai/v64/long-exam-legal-solver",
    }, "V64/V65 status")




# ---------------------------------------------------------------------------
# V66: File Upload Router
# Separate long-term RAG knowledge uploads from one-time question/exam files.
# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# V67 - Supabase RAG persistence/search
# ---------------------------------------------------------------------------
# V66 saved uploads into the local JSON RAG store. V67 keeps that local behavior,
# and also writes documents/chunks into Supabase PostgreSQL + pgvector when
# DATABASE_URL is configured. This is the production-safe RAG path.


def _v67_database_url() -> str:
    return (os.environ.get("DATABASE_URL") or "").strip()


def _v67_pgvector_enabled() -> bool:
    return bool(_v67_database_url())


def _v67_connect_pg():
    try:
        import psycopg2  # type: ignore
    except Exception as exc:  # pragma: no cover
        raise RuntimeError("Thiếu psycopg2-binary trong requirements.txt") from exc
    url = _v67_database_url()
    if not url:
        raise RuntimeError("DATABASE_URL chưa được cấu hình trong .env")
    return psycopg2.connect(url)


def _v67_vector_literal(values: List[float]) -> str:
    # pgvector accepts text literals such as '[0.1,0.2,...]'. Keep 6 decimals to
    # avoid huge SQL payloads while preserving stable similarity behavior.
    return "[" + ",".join(f"{float(v):.6f}" for v in values) + "]"


def _v67_hash_embedding(text: str, dims: int = 1536) -> List[float]:
    """Dependency-free embedding fallback.

    This is not as smart as OpenAI embeddings, but it is deterministic and works
    immediately for production plumbing tests. Later you can replace this with
    OpenAI/text-embedding and keep the same rag_chunks table.
    """
    vec = [0.0] * dims
    tokens = _v46_tokens(text or "")
    if not tokens:
        return vec
    for tok in tokens:
        digest = hashlib.sha256(tok.encode("utf-8")).digest()
        idx = int.from_bytes(digest[:4], "big") % dims
        sign = 1.0 if digest[4] % 2 == 0 else -1.0
        vec[idx] += sign
    norm = math.sqrt(sum(v * v for v in vec)) or 1.0
    return [v / norm for v in vec]


def _v67_insert_supabase_rag_document(
    title: str,
    content: str,
    source: str,
    category: str,
    document_type: str,
    tags: List[str],
    file_type: str,
    extraction: str,
    chunks: List[str],
    storage_path: Optional[str] = None,
) -> Dict[str, Any]:
    """Insert one RAG document and its chunks into Supabase/Postgres.

    Returns a status dict instead of raising to the caller, so V66 upload still
    works locally even if Supabase is misconfigured.
    """
    if not _v67_pgvector_enabled():
        return {"enabled": False, "saved": False, "reason": "DATABASE_URL chưa cấu hình; chỉ lưu RAG local."}
    conn = None
    try:
        from psycopg2.extras import Json  # type: ignore
        conn = _v67_connect_pg()
        conn.autocommit = False
        with conn.cursor() as cur:
            cur.execute(
                """
                insert into rag_documents (title, document_type, category, source, storage_path, uploaded_by)
                values (%s, %s, %s, %s, %s, %s)
                returning id
                """,
                (title, document_type, category, source, storage_path, "backend_v67"),
            )
            document_id = str(cur.fetchone()[0])
            inserted_chunks = 0
            for idx, chunk in enumerate(chunks, start=1):
                emb = _v67_vector_literal(_v67_hash_embedding(f"{title}\n{chunk}"))
                metadata = {
                    "title": title,
                    "source": source,
                    "category": category,
                    "document_type": document_type,
                    "tags": tags,
                    "file_type": file_type,
                    "extraction": extraction,
                    "embedding_provider": "v67_hash_fallback",
                }
                cur.execute(
                    """
                    insert into rag_chunks (document_id, chunk_index, content, metadata, embedding)
                    values (%s, %s, %s, %s, %s::vector)
                    """,
                    (document_id, idx, chunk, Json(metadata), emb),
                )
                inserted_chunks += 1
        conn.commit()
        return {"enabled": True, "saved": True, "document_id": document_id, "chunks": inserted_chunks}
    except Exception as exc:
        if conn:
            conn.rollback()
        return {"enabled": True, "saved": False, "error": str(exc)}
    finally:
        if conn:
            conn.close()


def _v67_supabase_vector_search(query: str, limit: int = 5) -> List[Dict[str, Any]]:
    if not _v67_pgvector_enabled():
        return []
    conn = None
    try:
        conn = _v67_connect_pg()
        qvec = _v67_vector_literal(_v67_hash_embedding(query))
        with conn.cursor() as cur:
            cur.execute(
                """
                select
                  c.id,
                  c.document_id,
                  c.chunk_index,
                  c.content,
                  c.metadata,
                  d.title,
                  d.source,
                  1 - (c.embedding <=> %s::vector) as score
                from rag_chunks c
                left join rag_documents d on d.id = c.document_id
                where c.embedding is not null
                order by c.embedding <=> %s::vector
                limit %s
                """,
                (qvec, qvec, int(limit)),
            )
            rows = cur.fetchall()
        results: List[Dict[str, Any]] = []
        for row in rows:
            chunk_id, doc_id, chunk_index, content, metadata, title, source, score = row
            content = content or ""
            results.append({
                "chunk_id": str(chunk_id),
                "document_id": str(doc_id),
                "doc_id": str(doc_id),
                "chunk_index": chunk_index,
                "title": title or (metadata or {}).get("title") or "Supabase RAG document",
                "source": source or (metadata or {}).get("source") or "supabase",
                "score": round(float(score or 0), 4),
                "snippet": content[:700] + ("..." if len(content) > 700 else ""),
                "content": content,
                "provider": "supabase_pgvector_v67",
            })
        return results
    except Exception:
        return []
    finally:
        if conn:
            conn.close()


@router.get("/ai/v67/supabase-rag/status")
def v67_supabase_rag_status():
    status: Dict[str, Any] = {
        "version": "V67",
        "database_url_configured": bool(_v67_database_url()),
        "tables_expected": ["rag_documents", "rag_chunks"],
        "note": "V67 ghi tài liệu RAG vào Supabase PostgreSQL + pgvector nếu DATABASE_URL có trong .env.",
    }
    if not _v67_pgvector_enabled():
        return _api_ok(status, "Supabase RAG chưa bật vì thiếu DATABASE_URL.")
    conn = None
    try:
        conn = _v67_connect_pg()
        with conn.cursor() as cur:
            cur.execute("select count(*) from rag_documents")
            status["rag_documents"] = int(cur.fetchone()[0])
            cur.execute("select count(*) from rag_chunks")
            status["rag_chunks"] = int(cur.fetchone()[0])
        return _api_ok(status, "Supabase RAG đã kết nối được.")
    except Exception as exc:
        status["error"] = str(exc)
        return _api_ok(status, "Supabase RAG chưa kết nối được; kiểm tra DATABASE_URL/key/bảng.")
    finally:
        if conn:
            conn.close()


class V67SupabaseSearchRequest(BaseModel):
    question: str = Field(..., min_length=2)
    limit: int = Field(5, ge=1, le=20)


@router.post("/ai/v67/supabase-rag/search")
def v67_supabase_rag_search(req: V67SupabaseSearchRequest):
    results = _v67_supabase_vector_search(req.question, req.limit)
    return _api_ok({
        "version": "V67",
        "count": len(results),
        "results": results,
    }, "Đã tìm trong Supabase pgvector RAG.")

_V66_RAG_DOCUMENT_TYPES = {
    "thong_tu", "nghi_dinh", "luat", "quy_dinh", "quy_trinh", "policy", "manual",
    "legal_policy", "accounting_policy", "knowledge", "training", "sop",
}

_V66_SOLVE_DOCUMENT_TYPES = {
    "de_thi", "bai_tap", "question_file", "exam", "worksheet", "invoice_to_review", "one_time",
}

_V66_RAG_KEYWORDS = [
    "thông tư", "thong tu", "nghị định", "nghi dinh", "luật", "luat", "quy định", "quy dinh",
    "chuẩn mực", "chuan muc", "chế độ kế toán", "che do ke toan", "quy trình", "quy trinh",
    "sổ tay", "so tay", "nghiệp vụ", "nghiep vu", "hướng dẫn", "huong dan", "chính sách", "chinh sach",
]

_V66_SOLVE_KEYWORDS = [
    "đề thi", "de thi", "bài tập", "bai tap", "yêu cầu", "yeu cau", "hãy tính", "hay tinh",
    "định khoản", "dinh khoan", "tính vat", "tinh vat", "lợi nhuận", "loi nhuan", "giải bài", "giai bai",
]


def _v66_clean_tags(*items: Any) -> List[str]:
    tags: List[str] = []
    for item in items:
        if item is None:
            continue
        if isinstance(item, str):
            parts = item.split(",")
        elif isinstance(item, list):
            parts = item
        else:
            parts = [str(item)]
        for part in parts:
            value = str(part).strip()
            if value and value not in tags:
                tags.append(value)
    return tags


def _v66_detect_upload_target(filename: str, text: str, declared_target: str = "auto", document_type: Optional[str] = None, category: Optional[str] = None) -> Dict[str, Any]:
    target = (declared_target or "auto").strip().lower()
    doc_type = (document_type or "").strip().lower()
    norm = _normalize_v43("\n".join([filename or "", text[:4000] or "", doc_type, category or ""]))
    if target in {"rag", "knowledge", "long_term"}:
        return {"target": "rag", "reason": "Frontend/admin chọn đưa file vào kho tri thức RAG."}
    if target in {"solve", "question", "exam", "one_time"}:
        return {"target": "solve", "reason": "Frontend chọn xử lý file như đề bài/câu hỏi một lần."}
    if target in {"temp", "parse", "temporary"}:
        return {"target": "temp", "reason": "Frontend chọn chỉ parse file tạm, không lưu RAG."}

    if doc_type in _V66_RAG_DOCUMENT_TYPES or (category or "") in _V58_KNOWLEDGE_CATEGORIES:
        return {"target": "rag", "reason": "document_type/category thuộc nhóm tài liệu nền cần tra cứu lâu dài."}
    if doc_type in _V66_SOLVE_DOCUMENT_TYPES:
        return {"target": "solve", "reason": "document_type thuộc nhóm đề bài/câu hỏi xử lý một lần."}
    if any(k in norm for k in _V66_RAG_KEYWORDS):
        return {"target": "rag", "reason": "Tên/nội dung có dấu hiệu thông tư, nghị định, luật, quy trình hoặc tài liệu nghiệp vụ."}
    if any(k in norm for k in _V66_SOLVE_KEYWORDS):
        return {"target": "solve", "reason": "Tên/nội dung có dấu hiệu đề thi, bài tập, định khoản hoặc yêu cầu tính toán."}
    return {"target": "temp", "reason": "Chưa đủ dấu hiệu để lưu lâu dài; mặc định chỉ parse tạm để tránh đưa nhầm dữ liệu riêng tư vào RAG."}


def _v66_extract_upload(file_name: str, raw: bytes) -> Dict[str, Any]:
    extracted = _v47_extract_text(file_name or "uploaded", raw)
    text = extracted.get("text") or ""
    return {
        "filename": file_name or "uploaded",
        "file_type": extracted.get("file_type"),
        "extraction": extracted.get("extraction"),
        "text": text,
        "text_length": len(text),
        "preview": text[:1000],
    }


def _v66_store_rag_document_from_text(
    title: str,
    content: str,
    source: str,
    category: str,
    document_type: str,
    tags: List[str],
    file_type: str,
    extraction: str,
    auto_chunk: bool = True,
) -> Dict[str, Any]:
    category = category if category in _V58_KNOWLEDGE_CATEGORIES else "general"
    final_tags = _v66_clean_tags(category, document_type, "rag_knowledge", "v66_upload", tags)
    store = _load_store()
    doc = _v47_store_document(
        store,
        title=title,
        content=content,
        source=source,
        tags=final_tags,
        file_type=file_type,
        extraction=extraction,
        auto_chunk=auto_chunk,
    )
    chunks = _chunk_text_v47(content) if auto_chunk else [content]
    supabase_result = _v67_insert_supabase_rag_document(
        title=title,
        content=content,
        source=source,
        category=category,
        document_type=document_type,
        tags=final_tags,
        file_type=file_type,
        extraction=extraction,
        chunks=chunks,
    )
    record = {
        "id": _next_id(store, "RUP"),
        "document_id": doc.get("id"),
        "title": title,
        "category": category,
        "document_type": document_type,
        "tags": final_tags,
        "created_at": _now(),
        "text_length": len(content),
        "chunk_count": doc.get("chunk_count", 0),
        "source": source,
        "file_type": file_type,
        "supabase": supabase_result,
    }
    store.setdefault("v66_upload_router_logs", []).append({**record, "route": "rag"})
    _save_store(store)
    return {"document": doc, "upload_record": record, "category": category, "tags": final_tags, "supabase": supabase_result}


@router.post("/ai/v66/rag/upload-file")
async def v66_rag_upload_file(
    file: UploadFile = File(...),
    title: Optional[str] = Form(None),
    category: str = Form("general"),
    document_type: str = Form("legal_policy"),
    source: str = Form("v66_rag_upload"),
    tags: str = Form(""),
    auto_chunk: bool = Form(True),
):
    """Upload thông tư/nghị định/quy trình/tài liệu nền vào RAG lâu dài."""
    raw = await file.read()
    extracted = _v66_extract_upload(file.filename or "uploaded", raw)
    saved = _v66_store_rag_document_from_text(
        title=title or extracted["filename"],
        content=extracted["text"],
        source=source,
        category=category,
        document_type=document_type,
        tags=_v66_clean_tags(tags),
        file_type=extracted.get("file_type") or "unknown",
        extraction=extracted.get("extraction") or "unknown",
        auto_chunk=auto_chunk,
    )
    return _api_ok({
        "version": "V66",
        "route": "rag",
        "file": {k: v for k, v in extracted.items() if k != "text"},
        **saved,
        "next_step": "Frontend có thể gọi POST /ai/v64/long-exam-legal-solver với use_rag=true để hỏi theo tài liệu này.",
    }, "Đã upload file vào kho tri thức RAG lâu dài.")


@router.post("/ai/v66/solve/upload-question-file")
async def v66_solve_upload_question_file(
    file: UploadFile = File(...),
    question: str = Form("Giải giúp tôi nội dung trong file này."),
    standard: str = Form("TT200"),
    category: Optional[str] = Form(None),
    use_rag: bool = Form(True),
    require_sources: bool = Form(True),
    save_learning: bool = Form(False),
):
    """Upload đề thi/bài tập/bảng số liệu để giải một lần, không đưa file vào RAG."""
    raw = await file.read()
    extracted = _v66_extract_upload(file.filename or "uploaded", raw)
    combined_question = (
        f"{question.strip()}\n\n"
        f"--- NỘI DUNG FILE {extracted['filename']} ---\n"
        f"{extracted['text']}"
    )
    solution = v64_long_exam_legal_solver(V64LongExamLegalSolverRequest(
        question=combined_question,
        standard=standard,
        category=category,
        use_rag=use_rag,
        require_sources=require_sources,
        save_learning=save_learning,
        limit=8,
    ))
    return _api_ok({
        "version": "V66",
        "route": "solve",
        "stored_in_rag": False,
        "file": {k: v for k, v in extracted.items() if k != "text"},
        "combined_question_length": len(combined_question),
        "solution": solution,
        "note": "File này chỉ dùng cho lần hỏi hiện tại; không lưu vào kho RAG để tránh đưa nhầm đề thi/chứng từ riêng tư vào kiến thức lâu dài.",
    }, "Đã đọc file và gọi V64/V65 solver để giải một lần.")


@router.post("/ai/v66/file-upload-router")
async def v66_file_upload_router(
    file: UploadFile = File(...),
    target: str = Form("auto"),
    question: str = Form("Giải thích/xử lý file này giúp tôi."),
    title: Optional[str] = Form(None),
    category: str = Form("general"),
    document_type: Optional[str] = Form(None),
    source: str = Form("v66_router"),
    tags: str = Form(""),
    standard: str = Form("TT200"),
    use_rag: bool = Form(True),
    require_sources: bool = Form(True),
    save_learning: bool = Form(False),
    auto_chunk: bool = Form(True),
):
    """Một cổng upload chung: auto phân loại RAG / solve một lần / parse tạm."""
    raw = await file.read()
    extracted = _v66_extract_upload(file.filename or "uploaded", raw)
    decision = _v66_detect_upload_target(extracted["filename"], extracted["text"], target, document_type, category)
    route = decision["target"]
    if route == "rag":
        saved = _v66_store_rag_document_from_text(
            title=title or extracted["filename"],
            content=extracted["text"],
            source=source,
            category=category,
            document_type=document_type or "legal_policy",
            tags=_v66_clean_tags(tags),
            file_type=extracted.get("file_type") or "unknown",
            extraction=extracted.get("extraction") or "unknown",
            auto_chunk=auto_chunk,
        )
        payload = {"stored_in_rag": True, **saved}
    elif route == "solve":
        combined_question = f"{question.strip()}\n\n--- NỘI DUNG FILE {extracted['filename']} ---\n{extracted['text']}"
        payload = {
            "stored_in_rag": False,
            "combined_question_length": len(combined_question),
            "solution": v64_long_exam_legal_solver(V64LongExamLegalSolverRequest(
                question=combined_question,
                standard=standard,
                category=category,
                use_rag=use_rag,
                require_sources=require_sources,
                save_learning=save_learning,
                limit=8,
            )),
        }
    else:
        payload = {
            "stored_in_rag": False,
            "parsed_only": True,
            "warning": "File chỉ được parse tạm. Nếu đây là tài liệu nền, hãy gọi lại với target=rag.",
        }
    return _api_ok({
        "version": "V66",
        "route": route,
        "decision": decision,
        "file": {k: v for k, v in extracted.items() if k != "text"},
        **payload,
    }, "V66 đã định tuyến file upload.")


@router.get("/ai/v66/file-upload-router/status")
def v66_file_upload_router_status():
    store = _load_store()
    return _api_ok({
        "version": "V66",
        "purpose": "Tách 2 luồng upload: tài liệu nền vào RAG và file câu hỏi xử lý một lần.",
        "endpoints": {
            "rag_upload": "POST /ai/v66/rag/upload-file",
            "one_time_solve": "POST /ai/v66/solve/upload-question-file",
            "auto_router": "POST /ai/v66/file-upload-router",
            "ask_after_rag": "POST /ai/v64/long-exam-legal-solver",
            "supabase_status": "GET /ai/v67/supabase-rag/status",
            "supabase_search": "POST /ai/v67/supabase-rag/search",
        },
        "targets": ["auto", "rag", "solve", "temp"],
        "rag_examples": ["Thông tư", "Nghị định", "Luật", "Quy trình nội bộ", "Sổ tay nghiệp vụ"],
        "solve_examples": ["Đề thi", "Bài tập kế toán", "Bảng số liệu Excel", "PDF/Word chứa câu hỏi"],
        "counts": {
            "local_rag_documents": len(store.setdefault("v47_documents", [])),
            "local_vector_chunks": len(store.setdefault("v46_vector_chunks", [])),
            "v66_upload_logs": len(store.setdefault("v66_upload_router_logs", [])),
            "supabase_enabled": bool(_v67_database_url()),
        },
        "frontend_rule": "Frontend chỉ chọn file và gửi target; backend mới đọc file, lưu RAG hoặc giải tạm.",
    }, "V66 File Upload Router sẵn sàng.")

@router.get("/documents/{document_id}/workflow")
def v62_1_get_document_workflow(document_id: str):
    store = _load_store(); _v59_bootstrap_store(store)
    wf = _v62_find_workflow(store, document_id)
    if not wf:
        return _api_ok({
            "document_id": document_id,
            "status": "uploaded",
            "allowed_next_actions": ["submit-review"],
            "workflow": None,
            "frontend_steps": ["uploaded", "ai_processing", "ai_reviewed", "need_more_info", "pending_approval", "approved", "posted"],
        }, "Chưa có workflow, frontend có thể gọi submit-review.")
    status = wf.get("status")
    allowed = []
    if status in {"draft", "uploaded", "ai_reviewed", "need_more_info"}:
        allowed.append("submit-review")
    if status == "pending_approval":
        allowed.extend(["approve", "reject"])
    if status == "approved":
        allowed.append("post-to-journal")
    return _api_ok({"document_id": document_id, "status": status, "allowed_next_actions": allowed, "workflow": wf})


def _v57_1_stream_excel(rows: List[Dict[str, Any]], filename_prefix: str, sheet_title: str):
    bio = _v57_make_excel(rows, sheet_title)
    filename = f"{filename_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(bio, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/ai/v57/journal-entries/export-excel")
def v57_1_journal_entries_export_excel(company_id: str = _V59_DEFAULT_COMPANY_ID, status: Optional[str] = None, date_from: Optional[str] = None, date_to: Optional[str] = None, limit: int = 1000):
    store = _load_store(); _v59_bootstrap_store(store)
    rows = [e for e in store.setdefault("v61_journal_entries", []) if e.get("company_id") == company_id]
    if status:
        rows = [e for e in rows if e.get("status") == status]
    if date_from:
        rows = [e for e in rows if str(e.get("date", "")) >= date_from]
    if date_to:
        rows = [e for e in rows if str(e.get("date", "")) <= date_to]
    flat = []
    for e in rows[:limit]:
        for line in e.get("lines") or []:
            flat.append({
                "entry_id": e.get("id"), "date": e.get("date"), "status": e.get("status"), "description": e.get("description"),
                "account": line.get("account"), "debit": line.get("debit"), "credit": line.get("credit"), "line_description": line.get("description"),
                "source_document_id": e.get("source_document_id"), "created_at": e.get("created_at"),
            })
    return _v57_1_stream_excel(flat, "ai_ke_toan_journal_entries", "journal_entries")


@router.get("/ai/v57/reports/trial-balance/export-excel")
def v57_1_trial_balance_export_excel(company_id: str = _V59_DEFAULT_COMPANY_ID, date_from: Optional[str] = None, date_to: Optional[str] = None):
    data = v63_trial_balance(company_id=company_id, date_from=date_from, date_to=date_to, posted_only=True)["data"]
    return _v57_1_stream_excel(data.get("rows") or [], "ai_ke_toan_trial_balance", "trial_balance")


@router.get("/ai/v57/reports/general-ledger/export-excel")
def v57_1_general_ledger_export_excel(company_id: str = _V59_DEFAULT_COMPANY_ID, account: Optional[str] = None, date_from: Optional[str] = None, date_to: Optional[str] = None):
    data = v63_general_ledger(company_id=company_id, account=account, date_from=date_from, date_to=date_to, posted_only=True)["data"]
    return _v57_1_stream_excel(data.get("rows") or [], "ai_ke_toan_general_ledger", "general_ledger")


@router.get("/ai/v53-v63/quality-upgrade-status")
def v53_v63_quality_upgrade_status():
    return _api_ok({
        "stage": "V53.1-V63.1 backend quality upgrade",
        "policy": "Không thêm frontend; chỉ làm giàu API JSON/download Excel để frontend riêng gọi.",
        "upgraded_existing_behavior": [
            "V53.1: document review có checklist theo loại chứng từ, quality_score, overall_status, risk_level.",
            "V54.1: journal suggestion có nhiều phương án, balance_check Nợ/Có, validation tài khoản, missing_info.",
            "V58.1: RAG answer có sources, confidence, missing_info và cơ chế không đủ căn cứ.",
            "V62.1: workflow có endpoint đọc trạng thái + allowed_next_actions cho frontend.",
            "V63.1: báo cáo mặc định lấy posted journal entries; thêm export Excel qua V57.1.",
            "V57.1: export Excel cho journal entries, trial balance, general ledger.",
        ],
        "new_apis": [
            "POST /ai/v58/rag-answer",
            "GET  /documents/{document_id}/workflow",
            "GET  /ai/v57/journal-entries/export-excel",
            "GET  /ai/v57/reports/trial-balance/export-excel",
            "GET  /ai/v57/reports/general-ledger/export-excel",
            "GET  /ai/v53-v63/quality-upgrade-status",
        ],
    }, "V53.1-V63.1 đã sẵn sàng.")


# ============================================================================
# V64.1 / V53.2 / V54.2 / V58.2 / V57.2 / VDEV.1
# Backend integration quality upgrade: frontend contracts, extraction,
# validation engine, corrections, enhanced exports and demo seed data.
# ============================================================================

_DOCUMENT_TYPES_V64_1 = [
    "invoice_input", "invoice_output", "contract", "payment_voucher", "receipt_voucher",
    "bank_statement", "payroll", "fixed_asset", "tool_equipment", "tax_declaration",
    "financial_statement", "legal_document", "unknown",
]
_DOCUMENT_STATUSES_V64_1 = [
    "uploaded", "ai_processing", "ai_reviewed", "need_more_info", "pending_approval",
    "approved", "posted", "rejected", "cancelled",
]
_RISK_LEVELS_V64_1 = ["low", "medium", "high"]
_CORRECTION_STATUSES_V58_2 = ["pending", "applied", "ignored"]
_VAT_RATES_V54_2 = [0, 5, 8, 10]


class V53ExtractFieldsRequest(BaseModel):
    content: str = ""
    document_type: Optional[str] = None
    filename: Optional[str] = None
    company_id: str = _V59_DEFAULT_COMPANY_ID


class V54ValidateJournalRequest(BaseModel):
    company_id: str = _V59_DEFAULT_COMPANY_ID
    date: Optional[str] = None
    description: str = ""
    lines: List[Dict[str, Any]] = Field(default_factory=list)
    status: str = "draft"


class V54ValidateDocumentRequest(BaseModel):
    company_id: str = _V59_DEFAULT_COMPANY_ID
    document_type: str = "unknown"
    fields: Dict[str, Any] = Field(default_factory=dict)
    content: str = ""


class V54ValidateTaxRequest(BaseModel):
    company_id: str = _V59_DEFAULT_COMPANY_ID
    document_type: str = "unknown"
    amount_before_tax: Optional[float] = None
    vat_amount: Optional[float] = None
    total_amount: Optional[float] = None
    payment_method: Optional[str] = None
    fields: Dict[str, Any] = Field(default_factory=dict)


class V58CorrectionRequest(BaseModel):
    company_id: str = _V59_DEFAULT_COMPANY_ID
    document_id: Optional[str] = None
    entity_type: str = "document_field"
    entity_id: Optional[str] = None
    field: str
    ai_value: Any = None
    correct_value: Any = None
    reason: Optional[str] = None
    actor: str = "api"


class V58ApplyCorrectionRequest(BaseModel):
    correction_id: str
    actor: str = "api"
    apply_to: str = "stored_correction_only"
    note: Optional[str] = None


class VDevSeedRequest(BaseModel):
    company_id: str = _V59_DEFAULT_COMPANY_ID
    reset_first: bool = False
    include_posted_entries: bool = True
    actor: str = "dev-seed"


def _safe_decode_bytes(data: bytes) -> str:
    for enc in ("utf-8", "utf-8-sig", "cp1258", "latin-1"):
        try:
            return data.decode(enc)
        except Exception:
            continue
    return data.decode("utf-8", errors="ignore")


def _extract_docx_text_from_bytes(data: bytes) -> str:
    try:
        with zipfile.ZipFile(BytesIO(data)) as zf:
            chunks = []
            for name in zf.namelist():
                if name.startswith("word/") and name.endswith(".xml"):
                    raw = zf.read(name).decode("utf-8", errors="ignore")
                    raw = re.sub(r"<[^>]+>", " ", raw)
                    chunks.append(raw)
            return re.sub(r"\s+", " ", "\n".join(chunks)).strip()
    except Exception:
        return ""


def _extract_xlsx_text_from_bytes(data: bytes, max_rows_per_sheet: int = 80) -> str:
    if Workbook is None:
        return ""
    try:
        from openpyxl import load_workbook  # type: ignore
        wb = load_workbook(BytesIO(data), data_only=True, read_only=True)
        out = []
        for ws in wb.worksheets:
            out.append(f"[SHEET] {ws.title}")
            for ridx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if ridx > max_rows_per_sheet:
                    out.append("...")
                    break
                vals = [str(v) for v in row if v is not None and str(v).strip()]
                if vals:
                    out.append(" | ".join(vals))
        return "\n".join(out).strip()
    except Exception:
        return ""


def _parse_uploaded_bytes(filename: str, data: bytes) -> Dict[str, Any]:
    lower = (filename or "").lower()
    file_type = lower.rsplit(".", 1)[-1] if "." in lower else "unknown"
    text = ""
    parser = "basic-text"
    warnings: List[str] = []
    if lower.endswith(".docx"):
        text = _extract_docx_text_from_bytes(data)
        parser = "docx-zip-xml"
    elif lower.endswith(".xlsx") or lower.endswith(".xlsm"):
        text = _extract_xlsx_text_from_bytes(data)
        parser = "openpyxl"
    elif lower.endswith(".csv"):
        text = _safe_decode_bytes(data)
        parser = "csv-text"
    elif lower.endswith(".pdf"):
        text = _safe_decode_bytes(data)
        parser = "pdf-basic-decode"
        warnings.append("PDF scan/ảnh cần OCR chuyên dụng; parser cơ bản chỉ đọc được PDF có text layer hoặc metadata text.")
    else:
        text = _safe_decode_bytes(data)
    if not text:
        warnings.append("Không trích xuất được text đáng tin cậy; frontend nên yêu cầu OCR hoặc nhập nội dung thủ công.")
    return {
        "filename": filename,
        "file_type": file_type,
        "size_bytes": len(data),
        "parser": parser,
        "extracted_text": text[:200000],
        "text_length": len(text),
        "warnings": warnings,
    }


def _regex_first(patterns: List[str], text: str, flags: int = re.IGNORECASE) -> Optional[str]:
    for pat in patterns:
        m = re.search(pat, text, flags)
        if m:
            val = (m.group(1) if m.groups() else m.group(0)).strip(" :\t\n\r")
            if val:
                return val
    return None


def _parse_money_value(value: Any) -> Optional[float]:
    if value is None:
        return None
    s = str(value)
    s = re.sub(r"[^0-9,.-]", "", s)
    if not s:
        return None
    # Vietnamese format usually uses dot as thousands separator and comma as decimal.
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s and s.count(",") == 1 and len(s.split(",")[-1]) <= 2:
        s = s.replace(",", ".")
    else:
        s = s.replace(",", "").replace(".", "") if re.search(r"[.,]\d{3}(\D|$)", s) else s.replace(",", "")
    try:
        return round(float(s), 2)
    except Exception:
        return None


def _money_after_labels(labels: List[str], text: str) -> Optional[float]:
    for label in labels:
        pat = rf"{label}\s*[:：]?\s*([0-9][0-9\.,\s]*)(?:\s*VND|\s*đ|\s*dong|\s*đồng)?"
        raw = _regex_first([pat], text, flags=re.IGNORECASE)
        val = _parse_money_value(raw)
        if val is not None:
            return val
    return None


def _normalize_date_value(raw: Optional[str]) -> Optional[str]:
    if not raw:
        return None
    raw = raw.strip()
    m = re.search(r"(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})", raw)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100:
            y += 2000
        try:
            return datetime(y, mo, d).date().isoformat()
        except Exception:
            return raw
    m = re.search(r"ngày\s*(\d{1,2})\s*tháng\s*(\d{1,2})\s*năm\s*(\d{4})", raw, re.IGNORECASE)
    if m:
        try:
            return datetime(int(m.group(3)), int(m.group(2)), int(m.group(1))).date().isoformat()
        except Exception:
            return raw
    return raw


def _extract_document_fields_from_text(content: str, document_type: Optional[str] = None, filename: Optional[str] = None) -> Dict[str, Any]:
    text = content or ""
    norm = _normalize_v43(text + " " + (filename or ""))
    if not document_type or document_type == "unknown":
        classified = _v55_classify_accounting_document(text, filename or "")
        document_type = classified.get("document_type", "unknown")
    tax_codes = re.findall(r"(?:mã\s*số\s*thuế|mst|tax\s*code)\s*[:：]?\s*([0-9\-\s]{8,20})", text, re.IGNORECASE)
    invoice_no = _regex_first([
        r"(?:số\s*hoá\s*đơn|số\s*hóa\s*đơn|invoice\s*no\.?|số)\s*[:：]?\s*([A-Z0-9\-/]{3,30})",
        r"(?:No\.?|Số)\s*[:：]?\s*([A-Z0-9\-/]{3,30})",
    ], text)
    invoice_date_raw = _regex_first([
        r"(ngày\s*\d{1,2}\s*tháng\s*\d{1,2}\s*năm\s*\d{4})",
        r"(?:ngày|date)\s*[:：]?\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})",
        r"(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})",
    ], text)
    supplier_name = _regex_first([
        r"(?:người\s*bán|đơn\s*vị\s*bán|seller|supplier)\s*[:：]?\s*(.+)",
        r"(?:tên\s*đơn\s*vị\s*bán)\s*[:：]?\s*(.+)",
    ], text)
    buyer_name = _regex_first([
        r"(?:người\s*mua|đơn\s*vị\s*mua|buyer|customer)\s*[:：]?\s*(.+)",
        r"(?:tên\s*đơn\s*vị\s*mua)\s*[:：]?\s*(.+)",
    ], text)
    amount_before_tax = _money_after_labels([r"cộng\s*tiền\s*hàng", r"tiền\s*hàng", r"amount\s*before\s*tax", r"subtotal"], text)
    vat_amount = _money_after_labels([r"tiền\s*thuế\s*gtgt", r"thuế\s*gtgt", r"vat\s*amount", r"vat"], text)
    total_amount = _money_after_labels([r"tổng\s*cộng\s*tiền\s*thanh\s*toán", r"tổng\s*thanh\s*toán", r"total\s*amount", r"grand\s*total"], text)
    amounts = _v53_extract_amount_candidates(text)
    if total_amount is None and amounts:
        total_amount = max(amounts)
    if amount_before_tax is None and total_amount and vat_amount:
        amount_before_tax = round(total_amount - vat_amount, 2)
    if vat_amount is None and total_amount and amount_before_tax:
        vat_amount = round(total_amount - amount_before_tax, 2)
    payment_method = None
    if any(k in norm for k in ["chuyen khoan", "chuyển khoản", "ngan hang", "ngân hàng", "bank", "uy nhiem chi", "ủy nhiệm chi"]):
        payment_method = "bank_transfer"
    elif any(k in norm for k in ["tien mat", "tiền mặt", "cash"]):
        payment_method = "cash"
    vat_rate = None
    mrate = re.search(r"(0|5|8|10)\s*%", text)
    if mrate:
        vat_rate = int(mrate.group(1))
    elif amount_before_tax and vat_amount:
        calc = round(vat_amount / amount_before_tax * 100)
        if calc in _VAT_RATES_V54_2:
            vat_rate = calc
    fields = {
        "invoice_no": invoice_no,
        "invoice_date": _normalize_date_value(invoice_date_raw),
        "supplier_name": supplier_name[:180] if supplier_name else None,
        "buyer_name": buyer_name[:180] if buyer_name else None,
        "supplier_tax_code": tax_codes[0].replace(" ", "") if len(tax_codes) >= 1 else None,
        "buyer_tax_code": tax_codes[1].replace(" ", "") if len(tax_codes) >= 2 else None,
        "amount_before_tax": amount_before_tax,
        "vat_rate": vat_rate,
        "vat_amount": vat_amount,
        "total_amount": total_amount,
        "payment_method": payment_method,
        "currency": "VND" if any(k in norm for k in ["vnd", "đồng", "dong"]) else None,
    }
    important = ["invoice_no", "invoice_date", "supplier_name", "total_amount"]
    if document_type in {"invoice_input", "invoice_output", "purchase_invoice", "sale_invoice"}:
        important += ["supplier_tax_code", "amount_before_tax", "vat_amount"]
    missing = [k for k in important if fields.get(k) in (None, "")]
    present_count = len([k for k in important if fields.get(k) not in (None, "")])
    confidence = "high" if important and present_count / len(important) >= 0.75 else "medium" if present_count >= 2 else "low"
    return {
        "document_type": document_type,
        "fields": fields,
        "missing_fields": missing,
        "confidence": confidence,
        "raw_amount_candidates": amounts[:10],
        "warnings": ["Kết quả bóc tách bằng rule/heuristic; frontend nên cho người dùng kiểm tra và sửa trước khi ghi sổ."],
    }


def _split_documents_from_text(text: str) -> List[Dict[str, Any]]:
    content = text or ""
    if not content.strip():
        return []
    markers = list(re.finditer(r"(?i)(h[oó]a\s*đ[oơ]n|invoice|hợp\s*đồng|contract|sao\s*kê|bank\s*statement)", content))
    if len(markers) <= 1:
        classified = _v55_classify_accounting_document(content, "")
        return [{"index": 1, "document_type": classified.get("document_type", "unknown"), "char_from": 0, "char_to": len(content), "confidence": classified.get("confidence", "medium"), "preview": content[:500]}]
    docs = []
    positions = [m.start() for m in markers] + [len(content)]
    # De-duplicate very close markers.
    compact = []
    for p in positions:
        if not compact or p - compact[-1] > 300:
            compact.append(p)
    if compact[-1] != len(content):
        compact.append(len(content))
    for idx in range(len(compact) - 1):
        start, end = compact[idx], compact[idx + 1]
        part = content[start:end].strip()
        if len(part) < 20:
            continue
        classified = _v55_classify_accounting_document(part, "")
        docs.append({"index": len(docs) + 1, "document_type": classified.get("document_type", "unknown"), "char_from": start, "char_to": end, "confidence": classified.get("confidence", "medium"), "preview": part[:500]})
    return docs


def _validate_tax_payload(document_type: str, amount_before_tax: Optional[float], vat_amount: Optional[float], total_amount: Optional[float], payment_method: Optional[str], fields: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    fields = fields or {}
    before = amount_before_tax if amount_before_tax is not None else _parse_money_value(fields.get("amount_before_tax"))
    vat = vat_amount if vat_amount is not None else _parse_money_value(fields.get("vat_amount"))
    total = total_amount if total_amount is not None else _parse_money_value(fields.get("total_amount"))
    payment = payment_method or fields.get("payment_method")
    errors: List[Dict[str, Any]] = []
    warnings: List[Dict[str, Any]] = []
    info: List[Dict[str, Any]] = []
    if before is not None and vat is not None and total is not None and abs((before + vat) - total) > 1:
        errors.append({"field": "total_amount", "message": "Tổng thanh toán không khớp tiền trước thuế + VAT."})
    if before and vat is not None:
        rate = round(vat / before * 100, 2) if before else 0
        if round(rate) not in _VAT_RATES_V54_2:
            warnings.append({"field": "vat_amount", "message": f"VAT rate tính ra khoảng {rate}%, không thuộc nhóm phổ biến 0/5/8/10%."})
    if total and total >= 5_000_000 and payment != "bank_transfer":
        warnings.append({"field": "payment_method", "message": "Chứng từ từ 5 triệu đồng trở lên cần kiểm tra thanh toán không dùng tiền mặt để giảm rủi ro thuế."})
    if document_type in {"invoice_input", "invoice_output", "purchase_invoice", "sale_invoice"}:
        for key in ["invoice_no", "invoice_date", "supplier_tax_code", "total_amount"]:
            if not fields.get(key) and key in {"invoice_no", "invoice_date", "supplier_tax_code"}:
                warnings.append({"field": key, "message": "Thiếu thông tin hóa đơn quan trọng."})
    if not errors and not warnings:
        info.append({"field": "tax", "message": "Chưa phát hiện rủi ro thuế cơ bản theo rule engine."})
    risk_level = "high" if errors else "medium" if warnings else "low"
    return {"valid": not errors, "risk_level": risk_level, "errors": errors, "warnings": warnings, "info": info}


@router.post("/ai/file/parse")
async def v64_1_parse_file(file: UploadFile = File(...)):
    data = await file.read()
    parsed = _parse_uploaded_bytes(file.filename or "uploaded", data)
    detected = _split_documents_from_text(parsed.get("extracted_text") or "")
    parsed["document_count"] = len(detected)
    parsed["detected_documents"] = detected[:20]
    return _api_ok(parsed, "Đã parse file ở mức backend cơ bản.")


@router.post("/ai/file/extract-fields")
async def v64_1_file_extract_fields(file: UploadFile = File(...), document_type: Optional[str] = None):
    data = await file.read()
    parsed = _parse_uploaded_bytes(file.filename or "uploaded", data)
    extracted = _extract_document_fields_from_text(parsed.get("extracted_text") or "", document_type=document_type, filename=file.filename)
    return _api_ok({"file": {k: v for k, v in parsed.items() if k != "extracted_text"}, **extracted}, "Đã bóc tách field từ file.")


@router.post("/ai/file/split-documents")
async def v64_1_file_split_documents(file: UploadFile = File(...)):
    data = await file.read()
    parsed = _parse_uploaded_bytes(file.filename or "uploaded", data)
    docs = _split_documents_from_text(parsed.get("extracted_text") or "")
    return _api_ok({"filename": file.filename, "document_count": len(docs), "documents": docs, "warnings": parsed.get("warnings", [])}, "Đã tách chứng từ ở mức heuristic.")


@router.post("/ai/v53/extract-document-fields")
def v53_2_extract_document_fields(req: V53ExtractFieldsRequest):
    extracted = _extract_document_fields_from_text(req.content, document_type=req.document_type, filename=req.filename)
    tax_check = _validate_tax_payload(extracted.get("document_type", "unknown"), extracted["fields"].get("amount_before_tax"), extracted["fields"].get("vat_amount"), extracted["fields"].get("total_amount"), extracted["fields"].get("payment_method"), extracted["fields"])
    return _api_ok({**extracted, "tax_validation_preview": tax_check}, "V53.2 đã bóc tách field chứng từ.")


@router.post("/accounting/validate/journal-entry")
def v54_2_validate_journal_entry(req: V54ValidateJournalRequest):
    store = _load_store(); _v59_bootstrap_store(store)
    lines = _v61_normalize_lines(req.lines)
    errors = _v61_validate_journal(store, req.company_id, lines)
    warnings: List[Dict[str, Any]] = []
    if req.date:
        try:
            d = datetime.fromisoformat(req.date[:10]).date()
            if d > datetime.now().date():
                warnings.append({"field": "date", "message": "Ngày chứng từ đang nằm trong tương lai."})
        except Exception:
            errors.append({"field": "date", "message": "Ngày không đúng định dạng ISO YYYY-MM-DD."})
    totals = {"debit": round(sum(l["debit"] for l in lines), 2), "credit": round(sum(l["credit"] for l in lines), 2)}
    return _api_ok({"valid": not errors, "company_id": req.company_id, "totals": totals, "normalized_lines": lines, "errors": errors, "warnings": warnings}, "Đã validate journal entry bằng rule engine.")


@router.post("/accounting/validate/document")
def v54_2_validate_document(req: V54ValidateDocumentRequest):
    fields = dict(req.fields or {})
    if req.content and not fields:
        fields = _extract_document_fields_from_text(req.content, req.document_type).get("fields", {})
    required_by_type = {
        "invoice_input": ["invoice_no", "invoice_date", "supplier_tax_code", "total_amount"],
        "invoice_output": ["invoice_no", "invoice_date", "buyer_tax_code", "total_amount"],
        "contract": ["contract_no", "contract_date", "partner_name"],
        "bank_statement": ["transaction_date", "amount"],
    }
    required = required_by_type.get(req.document_type, [])
    errors = [{"field": k, "message": "Thiếu trường bắt buộc theo loại chứng từ."} for k in required if not fields.get(k)]
    warnings = []
    if req.document_type == "unknown":
        warnings.append({"field": "document_type", "message": "Chưa phân loại được chứng từ."})
    completeness = 1.0 if not required else round((len(required) - len(errors)) / len(required), 4)
    return _api_ok({"valid": not errors, "document_type": req.document_type, "completeness": completeness, "errors": errors, "warnings": warnings, "fields": fields}, "Đã validate document fields.")


@router.post("/accounting/validate/tax")
def v54_2_validate_tax(req: V54ValidateTaxRequest):
    result = _validate_tax_payload(req.document_type, req.amount_before_tax, req.vat_amount, req.total_amount, req.payment_method, req.fields)
    return _api_ok(result, "Đã validate tax bằng rule engine.")


@router.post("/ai/corrections")
def v58_2_create_correction(req: V58CorrectionRequest):
    store = _load_store(); _v59_bootstrap_store(store)
    row = req.dict()
    row.update({"id": _next_id(store, "CORR"), "status": "pending", "created_at": _now(), "updated_at": _now(), "version": "V58.2"})
    store.setdefault("ai_corrections_v58_2", []).append(row)
    _v64_audit_event(store, "ai.correction.create", req.entity_type, req.entity_id or req.document_id or row["id"], {"field": req.field, "reason": req.reason}, actor=req.actor, company_id=req.company_id)
    _save_store(store)
    return _api_ok(row, "Đã lưu correction của người dùng.")


@router.get("/ai/corrections")
def v58_2_list_corrections(company_id: str = _V59_DEFAULT_COMPANY_ID, document_id: Optional[str] = None, status: Optional[str] = None, limit: int = 200):
    store = _load_store(); _v59_bootstrap_store(store)
    rows = [r for r in store.setdefault("ai_corrections_v58_2", []) if r.get("company_id") == company_id]
    if document_id:
        rows = [r for r in rows if r.get("document_id") == document_id]
    if status:
        rows = [r for r in rows if r.get("status") == status]
    rows = sorted(rows, key=lambda x: x.get("created_at", ""), reverse=True)[:limit]
    return _api_ok({"company_id": company_id, "count": len(rows), "corrections": rows})


@router.post("/ai/corrections/apply")
def v58_2_apply_correction(req: V58ApplyCorrectionRequest):
    store = _load_store(); _v59_bootstrap_store(store)
    row = next((r for r in store.setdefault("ai_corrections_v58_2", []) if r.get("id") == req.correction_id), None)
    if not row:
        _api_error("Không tìm thấy correction.", [{"field": "correction_id", "message": req.correction_id}], 404)
    row["status"] = "applied"
    row["applied_at"] = _now()
    row["applied_by"] = req.actor
    row["apply_to"] = req.apply_to
    row["note"] = req.note
    row["updated_at"] = _now()
    # MVP: store correction as learning signal. Direct mutation of arbitrary documents is intentionally not automatic.
    _v64_audit_event(store, "ai.correction.apply", row.get("entity_type", "correction"), row.get("entity_id") or row.get("document_id") or row["id"], {"field": row.get("field"), "apply_to": req.apply_to}, actor=req.actor, company_id=row.get("company_id"))
    _save_store(store)
    return _api_ok(row, "Đã đánh dấu correction là applied; backend không tự sửa chứng từ gốc nếu chưa có mapping an toàn.")


@router.get("/api/enums")
def v64_1_api_enums():
    return _api_ok({
        "document_statuses": _DOCUMENT_STATUSES_V64_1,
        "document_types": _DOCUMENT_TYPES_V64_1,
        "risk_levels": _RISK_LEVELS_V64_1,
        "correction_statuses": _CORRECTION_STATUSES_V58_2,
        "journal_statuses": ["draft", "pending_approval", "approved", "posted", "rejected", "cancelled"],
        "accounting_modes": ["TT200", "TT133", "custom"],
        "vat_rates": _VAT_RATES_V54_2,
        "workflow_actions": ["submit-review", "approve", "reject", "post-to-journal"],
        "standard_response": {"success": True, "data": {}, "message": "OK", "errors": []},
    }, "Enums cho frontend.")


@router.get("/api/schema/frontend")
def v64_1_frontend_schema():
    return _api_ok({
        "version": "V64.1-VDEV.1 integration quality",
        "base_response": {"success": "boolean", "data": "object|array|null", "message": "string", "errors": "array"},
        "recommended_frontend_flow": [
            "POST /ai/file/parse hoặc /ai/file/extract-fields",
            "POST /ai/v53/document-review/text",
            "POST /ai/v54/journal-suggestion",
            "POST /accounting/validate/journal-entry",
            "POST /journal-entries",
            "POST /documents/{document_id}/submit-review",
            "POST /documents/{document_id}/approve",
            "POST /documents/{document_id}/post-to-journal",
            "GET /reports/trial-balance",
        ],
        "forms": {
            "document_field_editor": ["document_type", "invoice_no", "invoice_date", "supplier_name", "supplier_tax_code", "amount_before_tax", "vat_amount", "total_amount", "payment_method"],
            "journal_entry": ["date", "description", "source_document_id", "lines[].account", "lines[].debit", "lines[].credit", "lines[].description"],
            "correction": ["document_id", "field", "ai_value", "correct_value", "reason"],
        },
        "download_endpoints": [
            "GET /ai/v57/journal-entries/export-excel",
            "GET /ai/v57/reports/trial-balance/export-excel",
            "GET /ai/v57/reports/general-ledger/export-excel",
            "GET /ai/v57/export-all-excel",
        ],
    }, "API contract tóm tắt cho frontend.")


@router.get("/api/health/deep")
def v64_1_deep_health():
    checks = []
    try:
        store = _load_store(); _v59_bootstrap_store(store)
        checks.append({"name": "json_store", "ok": True, "path": str(STORE_PATH), "keys": list(store.keys())[:20]})
    except Exception as exc:
        checks.append({"name": "json_store", "ok": False, "error": str(exc)})
        store = {}
    checks.append({"name": "openpyxl", "ok": Workbook is not None})
    checks.append({"name": "companies", "ok": bool(store.get("companies")), "count": len(store.get("companies", []))})
    checks.append({"name": "chart_of_accounts", "ok": bool(store.get("chart_of_accounts")), "count": len(store.get("chart_of_accounts", []))})
    checks.append({"name": "journal_entries", "ok": True, "count": len(store.get("v61_journal_entries", []))})
    ok = all(c.get("ok") for c in checks)
    return _api_ok({"ok": ok, "checks": checks, "time": _now()}, "Deep health check hoàn tất.")


@router.get("/ai/v57/export-all-excel")
def v57_2_export_all_excel(company_id: str = _V59_DEFAULT_COMPANY_ID):
    if Workbook is None:
        _api_error("openpyxl chưa khả dụng, không thể export Excel.", status_code=500)
    store = _load_store(); _v59_bootstrap_store(store)
    wb = Workbook()
    def write_sheet(title: str, rows: List[Dict[str, Any]]):
        ws = wb.create_sheet(title[:31])
        if not rows:
            ws.append(["empty"]); return
        headers = sorted({k for r in rows for k in r.keys() if not isinstance(r.get(k), (dict, list))})
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True) if Font else cell.font
        for r in rows:
            ws.append([r.get(h) for h in headers])
    # remove default sheet after creating real sheets
    default = wb.active
    wb.remove(default)
    write_sheet("companies", [c for c in store.get("companies", []) if c.get("id") == company_id])
    write_sheet("accounts", [a for a in store.get("chart_of_accounts", []) if a.get("company_id") == company_id])
    flat_entries = []
    for e in store.get("v61_journal_entries", []):
        if e.get("company_id") != company_id:
            continue
        for line in e.get("lines") or []:
            flat_entries.append({"entry_id": e.get("id"), "date": e.get("date"), "status": e.get("status"), "description": e.get("description"), **line})
    write_sheet("journal_entries", flat_entries)
    tb = v63_trial_balance(company_id=company_id, posted_only=True)["data"].get("rows", [])
    write_sheet("trial_balance", tb)
    write_sheet("corrections", [r for r in store.get("ai_corrections_v58_2", []) if r.get("company_id") == company_id])
    bio = BytesIO(); wb.save(bio); bio.seek(0)
    filename = f"ai_ke_toan_backend_export_all_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(bio, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})


def _dev_seed_demo_entries(store: Dict[str, Any], company_id: str) -> List[Dict[str, Any]]:
    samples = [
        {
            "date": datetime.now().date().isoformat(),
            "description": "Bán hàng hóa thu tiền chuyển khoản",
            "source_document_id": "DEMO-DOC-SALE-001",
            "lines": [
                {"account": "112", "debit": 11000000, "credit": 0, "description": "Thu tiền khách hàng"},
                {"account": "511", "debit": 0, "credit": 10000000, "description": "Doanh thu bán hàng"},
                {"account": "3331", "debit": 0, "credit": 1000000, "description": "VAT đầu ra"},
            ],
        },
        {
            "date": datetime.now().date().isoformat(),
            "description": "Mua laptop phục vụ văn phòng chưa thanh toán",
            "source_document_id": "DEMO-DOC-PUR-001",
            "lines": [
                {"account": "242", "debit": 18000000, "credit": 0, "description": "Ghi nhận CCDC/chi phí trả trước"},
                {"account": "1331", "debit": 1800000, "credit": 0, "description": "VAT đầu vào"},
                {"account": "331", "debit": 0, "credit": 19800000, "description": "Phải trả nhà cung cấp"},
            ],
        },
        {
            "date": datetime.now().date().isoformat(),
            "description": "Chi phí quảng cáo Facebook thanh toán ngân hàng",
            "source_document_id": "DEMO-DOC-ADS-001",
            "lines": [
                {"account": "641", "debit": 5000000, "credit": 0, "description": "Chi phí quảng cáo"},
                {"account": "1331", "debit": 500000, "credit": 0, "description": "VAT đầu vào"},
                {"account": "112", "debit": 0, "credit": 5500000, "description": "Thanh toán ngân hàng"},
            ],
        },
    ]
    created = []
    existing_ids = {e.get("source_document_id") for e in store.setdefault("v61_journal_entries", [])}
    for s in samples:
        if s["source_document_id"] in existing_ids:
            continue
        lines = _v61_normalize_lines(s["lines"])
        entry = {"id": _next_id(store, "JE"), "company_id": company_id, "date": s["date"], "description": s["description"], "source_document_id": s["source_document_id"], "lines": lines, "status": "posted", "total_debit": round(sum(l["debit"] for l in lines), 2), "total_credit": round(sum(l["credit"] for l in lines), 2), "created_at": _now(), "updated_at": _now(), "posted_at": _now(), "metadata": {"demo": True}}
        store.setdefault("v61_journal_entries", []).append(entry)
        created.append(entry)
    return created


@router.post("/dev/seed-demo-data")
def vdev_1_seed_demo_data(req: VDevSeedRequest):
    store = _load_store(); _v59_bootstrap_store(store)
    if req.reset_first:
        store["v61_journal_entries"] = [e for e in store.get("v61_journal_entries", []) if not e.get("metadata", {}).get("demo")]
        store["document_workflows"] = [w for w in store.get("document_workflows", []) if not str(w.get("document_id", "")).startswith("DEMO-")]
        store["ai_corrections_v58_2"] = [c for c in store.get("ai_corrections_v58_2", []) if not str(c.get("document_id", "")).startswith("DEMO-")]
    if not _v59_get_company(store, req.company_id):
        store.setdefault("companies", []).append({"id": req.company_id, "company_name": "Demo Company", "tax_code": "010DEMO", "accounting_mode": "TT200", "currency": "VND", "fiscal_year": datetime.now().year, "is_active": True, "created_at": _now(), "updated_at": _now()})
    created_entries = _dev_seed_demo_entries(store, req.company_id) if req.include_posted_entries else []
    demo_workflows = []
    for doc_id, status in [("DEMO-DOC-SALE-001", "posted"), ("DEMO-DOC-PUR-001", "pending_approval"), ("DEMO-DOC-RISK-001", "need_more_info")]:
        if not _v62_find_workflow(store, doc_id):
            wf = {"id": _next_id(store, "WF"), "document_id": doc_id, "company_id": req.company_id, "status": status, "history": [{"at": _now(), "actor": req.actor, "action": "seed", "note": "Demo workflow"}], "created_at": _now(), "updated_at": _now(), "journal_entry_id": None}
            store.setdefault("document_workflows", []).append(wf); demo_workflows.append(wf)
    _v64_audit_event(store, "dev.seed_demo_data", "company", req.company_id, {"created_entries": len(created_entries), "created_workflows": len(demo_workflows)}, actor=req.actor, company_id=req.company_id)
    _save_store(store)
    return _api_ok({"company_id": req.company_id, "created_entries": created_entries, "created_workflows": demo_workflows}, "Đã seed demo data cho frontend test.")


@router.delete("/dev/clear-demo-data")
def vdev_1_clear_demo_data(company_id: str = _V59_DEFAULT_COMPANY_ID, actor: str = "dev-clear"):
    store = _load_store(); _v59_bootstrap_store(store)
    before = {"entries": len(store.get("v61_journal_entries", [])), "workflows": len(store.get("document_workflows", [])), "corrections": len(store.get("ai_corrections_v58_2", []))}
    store["v61_journal_entries"] = [e for e in store.get("v61_journal_entries", []) if not e.get("metadata", {}).get("demo")]
    store["document_workflows"] = [w for w in store.get("document_workflows", []) if not str(w.get("document_id", "")).startswith("DEMO-")]
    store["ai_corrections_v58_2"] = [c for c in store.get("ai_corrections_v58_2", []) if not str(c.get("document_id", "")).startswith("DEMO-")]
    after = {"entries": len(store.get("v61_journal_entries", [])), "workflows": len(store.get("document_workflows", [])), "corrections": len(store.get("ai_corrections_v58_2", []))}
    _v64_audit_event(store, "dev.clear_demo_data", "company", company_id, {"before": before, "after": after}, actor=actor, company_id=company_id)
    _save_store(store)
    return _api_ok({"before": before, "after": after}, "Đã xóa demo data.")


@router.get("/dev/demo-scenarios")
def vdev_1_demo_scenarios():
    return _api_ok({
        "scenarios": [
            {"name": "Upload hóa đơn mua laptop", "steps": ["POST /ai/v53/extract-document-fields", "POST /ai/v53/document-review/text", "POST /ai/v54/journal-suggestion", "POST /accounting/validate/journal-entry"]},
            {"name": "Duyệt và ghi sổ", "steps": ["POST /journal-entries", "POST /documents/{document_id}/submit-review", "POST /documents/{document_id}/approve", "POST /documents/{document_id}/post-to-journal"]},
            {"name": "Xem báo cáo", "steps": ["GET /reports/trial-balance", "GET /reports/general-ledger", "GET /ai/v57/export-all-excel"]},
            {"name": "Người dùng sửa AI", "steps": ["POST /ai/corrections", "POST /ai/corrections/apply", "GET /ai/corrections"]},
        ],
        "sample_invoice_text": "HÓA ĐƠN GTGT số 0000123 ngày 30/05/2026 Người bán: ABC Co MST: 0101234567 Cộng tiền hàng: 18.000.000 Thuế GTGT 10%: 1.800.000 Tổng thanh toán: 19.800.000 Thanh toán chuyển khoản",
    }, "Demo scenarios cho frontend.")


@router.get("/ai/v53-2-vdev1/integration-quality-status")
def v53_2_vdev1_integration_quality_status():
    return _api_ok({
        "stage": "Backend integration quality upgrade V64.1 + V53.2 + V54.2 + V58.2 + V57.2 + VDEV.1",
        "policy": "Không thêm frontend; chỉ nâng API để frontend riêng dễ gọi, validate chắc hơn và có demo data.",
        "new_groups": {
            "file_processing": ["POST /ai/file/parse", "POST /ai/file/extract-fields", "POST /ai/file/split-documents"],
            "field_extraction": ["POST /ai/v53/extract-document-fields"],
            "validation_engine": ["POST /accounting/validate/journal-entry", "POST /accounting/validate/document", "POST /accounting/validate/tax"],
            "human_corrections": ["POST /ai/corrections", "GET /ai/corrections", "POST /ai/corrections/apply"],
            "frontend_contract": ["GET /api/enums", "GET /api/schema/frontend", "GET /api/health/deep"],
            "export": ["GET /ai/v57/export-all-excel"],
            "dev_seed": ["POST /dev/seed-demo-data", "DELETE /dev/clear-demo-data", "GET /dev/demo-scenarios"],
        },
    }, "Bản nâng backend integration quality đã sẵn sàng.")

# ---------------------------------------------------------------------------
# V48-V55: MVP product flow upgrade
# Upload Excel -> OCR invoice -> RAG knowledge -> smart reports -> dashboard -> audit -> feedback-to-rules -> demo UI
# ---------------------------------------------------------------------------

class V48BatchConfirmRequest(BaseModel):
    batch_id: str
    row_ids: Optional[List[str]] = None
    post_immediately: bool = False

class V50KnowledgeSearchRequest(BaseModel):
    query: str = Field(..., min_length=2)
    limit: int = Field(5, ge=1, le=20)

class V51AskReportRequest(BaseModel):
    question: str = Field(..., min_length=2)
    company_id: Optional[str] = None

class V53AuditRequest(BaseModel):
    batch_id: Optional[str] = None
    include_drafts: bool = True
    include_journal_entries: bool = True

class V54ApplyFeedbackRequest(BaseModel):
    dry_run: bool = True
    min_rating: int = Field(4, ge=1, le=5)


def _v48_cell_value(row: Dict[str, Any], names: List[str]) -> Any:
    normalized = {_normalize_v43(str(k)): v for k, v in row.items()}
    for name in names:
        key = _normalize_v43(name)
        if key in normalized and normalized[key] not in (None, ""):
            return normalized[key]
    for k, v in normalized.items():
        if any(_normalize_v43(name) in k for name in names) and v not in (None, ""):
            return v
    return None


def _v48_read_xlsx_rows(raw: bytes) -> List[Dict[str, Any]]:
    if Workbook is None:
        raise HTTPException(status_code=500, detail="openpyxl chưa khả dụng")
    from openpyxl import load_workbook  # type: ignore
    wb = load_workbook(BytesIO(raw), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(c or "").strip() or f"col_{i+1}" for i, c in enumerate(rows[0])]
    out = []
    for vals in rows[1:]:
        if not vals or not any(v not in (None, "") for v in vals):
            continue
        out.append({headers[i]: vals[i] if i < len(vals) else None for i in range(len(headers))})
    return out


def _v48_parse_amount(value: Any, fallback_text: str = "") -> Optional[float]:
    if isinstance(value, (int, float)):
        return float(value)
    if value is not None and str(value).strip():
        extracted = _extract_amount_from_text(str(value))
        if extracted is not None:
            return extracted
        try:
            cleaned = re.sub(r"[^0-9,.-]", "", str(value)).replace(".", "").replace(",", ".")
            return float(cleaned) if cleaned else None
        except Exception:
            pass
    return _extract_amount_from_text(fallback_text)


def _v48_make_row_proposal(raw_row: Dict[str, Any], idx: int) -> Dict[str, Any]:
    desc = _v48_cell_value(raw_row, ["description", "mo ta", "mô tả", "noi dung", "nội dung", "dien giai", "diễn giải", "transaction"])
    amount_raw = _v48_cell_value(raw_row, ["amount", "so tien", "số tiền", "tong tien", "tổng tiền", "debit", "credit"])
    payment = _v48_cell_value(raw_row, ["payment_method", "phuong thuc", "phương thức", "thanh toán", "payment"])
    date_val = _v48_cell_value(raw_row, ["date", "ngay", "ngày", "transaction_date"])
    description = str(desc or "").strip()
    amount = _v48_parse_amount(amount_raw, description)
    message = description
    if amount is not None and str(amount) not in message:
        message = f"{description} {amount:,.0f}"
    if payment:
        message = f"{message} {payment}"
    proposal = _proposal_from_message_v42(message)
    if amount is not None:
        proposal["amount"] = amount
    if payment:
        proposal["payment_method"] = _detect_payment_method_from_text(str(payment))
    proposal["date"] = str(date_val) if date_val is not None else None
    proposal["raw_row"] = raw_row
    proposal["row_index"] = idx
    proposal["row_id"] = f"ROW-{idx:04d}"
    proposal["status"] = "ready" if description and amount is not None else "need_review"
    if not description:
        proposal["risk_note"] = "Thiếu mô tả giao dịch. Cần bổ sung trước khi lưu bút toán."
    elif amount is None:
        proposal["risk_note"] = "Thiếu số tiền. Cần bổ sung trước khi lưu bút toán."
    return proposal


@router.post("/ai/v48/upload-transactions")
async def v48_upload_transactions(file: UploadFile = File(...)):
    raw = await file.read()
    filename = file.filename or "uploaded.xlsx"
    if not filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="V48 MVP hiện hỗ trợ Excel .xlsx/.xlsm. Với CSV hãy chuyển sang Excel trước.")
    rows = _v48_read_xlsx_rows(raw)
    proposals = [_v48_make_row_proposal(r, i + 1) for i, r in enumerate(rows)]
    store = _load_store()
    batch_id = _next_id(store, "BATCH")
    batch = {"id": batch_id, "version": "V48", "filename": filename, "created_at": _now(), "count": len(proposals), "ready_count": len([p for p in proposals if p.get("status") == "ready"]), "rows": proposals}
    store.setdefault("v48_transaction_batches", []).append(batch)
    _audit(store, "upload_transactions", "v48_batch", batch_id, {"filename": filename, "count": len(proposals)})
    _save_store(store)
    return {"version": "V48", "batch_id": batch_id, "count": len(proposals), "ready_count": batch["ready_count"], "rows": proposals[:200], "next_steps": ["GET /ai/v48/transactions/review?batch_id=...", "POST /ai/v48/transactions/confirm-batch"]}


@router.get("/ai/v48/transactions/review")
def v48_transactions_review(batch_id: Optional[str] = None, limit: int = 200):
    store = _load_store()
    batches = store.setdefault("v48_transaction_batches", [])
    if batch_id:
        batch = next((b for b in batches if b.get("id") == batch_id), None)
    else:
        batch = batches[-1] if batches else None
    if not batch:
        raise HTTPException(status_code=404, detail="Chưa có batch giao dịch V48")
    rows = batch.get("rows", [])[:max(1, min(limit, 1000))]
    return {"version": "V48", "batch_id": batch.get("id"), "filename": batch.get("filename"), "count": batch.get("count"), "ready_count": batch.get("ready_count"), "rows": rows}


@router.post("/ai/v48/transactions/confirm-batch")
def v48_confirm_batch(req: V48BatchConfirmRequest):
    store = _load_store()
    batch = next((b for b in store.setdefault("v48_transaction_batches", []) if b.get("id") == req.batch_id), None)
    if not batch:
        raise HTTPException(status_code=404, detail="Không tìm thấy batch")
    selected = set(req.row_ids or [])
    created = []
    skipped = []
    for p in batch.get("rows", []):
        if selected and p.get("row_id") not in selected:
            continue
        if p.get("status") != "ready" or p.get("amount") is None:
            skipped.append({"row_id": p.get("row_id"), "reason": p.get("risk_note") or "not_ready"})
            continue
        draft_req = JournalDraftCreate(description=p.get("description") or "", amount=float(p.get("amount") or 0), vat_rate=float(p.get("vat_rate") or 0.1), payment_method=p.get("payment_method") or "bank", category=p.get("category"), debit_account=p.get("debit_account"), credit_account=p.get("credit_account"), source="v48_confirm_batch")
        lines = _build_lines(draft_req)
        draft_id = _next_id(store, "DRAFT")
        draft = {"id": draft_id, "version": "V48", "created_at": _now(), "updated_at": _now(), "status": "approved" if req.post_immediately else "draft", "source": "v48_batch", "batch_id": req.batch_id, "row_id": p.get("row_id"), "description": p.get("description"), "amount": float(p.get("amount") or 0), "vat_rate": float(p.get("vat_rate") or 0.1), "payment_method": p.get("payment_method"), "category": p.get("category"), "lines": lines, "debit_total": _line_total(lines, "debit"), "credit_total": _line_total(lines, "credit"), "balanced": _line_total(lines, "debit") == _line_total(lines, "credit"), "risk_flags": _detect_errors(lines, p.get("description") or ""), "risk_note": p.get("risk_note")}
        store.setdefault("journal_drafts", []).append(draft)
        created.append(draft)
    batch["confirmed_at"] = _now()
    batch["confirmed_count"] = len(created)
    _audit(store, "confirm_batch", "v48_batch", req.batch_id, {"created": len(created), "skipped": len(skipped)})
    _save_store(store)
    return {"version": "V48", "status": "saved", "created_count": len(created), "skipped_count": len(skipped), "created_drafts": created[:200], "skipped": skipped[:200]}


@router.post("/ai/v49/ocr-invoice")
async def v49_ocr_invoice(file: UploadFile = File(...)):
    raw = await file.read()
    filename = file.filename or "invoice"
    try:
        from invoice_ocr import read_text_from_upload, parse_invoice_text  # type: ignore
        text_info = read_text_from_upload(filename, raw)
        parsed = parse_invoice_text(text_info.get("text") or "")
    except Exception as exc:
        extracted = _v47_extract_text(filename, raw)
        text_info = {"text": extracted.get("text", ""), "method": extracted.get("extraction", "fallback"), "warning": str(exc)}
        parsed = {"description": "Hóa đơn/Chứng từ", "total_amount": _extract_amount_from_text(text_info.get("text") or ""), "raw_text": text_info.get("text") or "", "confidence": 0.25}
    store = _load_store()
    invoice_id = _next_id(store, "INV")
    item = {"id": invoice_id, "version": "V49", "filename": filename, "created_at": _now(), "text_method": text_info.get("method"), "warning": text_info.get("warning"), "parsed": parsed}
    store.setdefault("v49_invoices", []).append(item)
    _audit(store, "ocr_invoice", "v49_invoice", invoice_id, {"filename": filename, "confidence": parsed.get("confidence")})
    _save_store(store)
    return {"version": "V49", "invoice_id": invoice_id, "extraction": text_info, "invoice": parsed, "next_step": "POST /ai/v49/invoice-to-journal?invoice_id=..."}


@router.post("/ai/v49/invoice-to-journal")
def v49_invoice_to_journal(invoice_id: str):
    store = _load_store()
    inv = next((i for i in store.setdefault("v49_invoices", []) if i.get("id") == invoice_id), None)
    if not inv:
        raise HTTPException(status_code=404, detail="Không tìm thấy invoice_id")
    parsed = inv.get("parsed") or {}
    description = parsed.get("description") or "Hóa đơn/Chứng từ"
    total = parsed.get("amount_before_tax") or parsed.get("subtotal") or parsed.get("total_amount") or _extract_amount_from_text(parsed.get("raw_text") or "") or 0
    vat = parsed.get("vat_amount") or 0
    vat_rate = 0.1 if vat else 0.0
    message = f"{description} {float(total):,.0f} chuyển khoản"
    proposal = _proposal_from_message_v42(message)
    proposal.update({"amount": float(total), "vat_rate": vat_rate, "source_invoice_id": invoice_id, "invoice_fields": parsed})
    return {"version": "V49", "invoice_id": invoice_id, "proposal": proposal, "journal_suggestion": _v54_suggest_journal_payload({"description": description, "content": parsed.get("raw_text") or "", "total_amount": float(total) + float(vat or 0), "amount_before_vat": float(total), "vat_amount": float(vat or 0), "payment_method": "bank"})}


def _v50_search_knowledge(query: str, limit: int = 5) -> List[Dict[str, Any]]:
    q_words = {w for w in re.findall(r"[a-zA-Z0-9À-ỹ]+", _normalize_v43(query)) if len(w) >= 2}
    results: List[Dict[str, Any]] = []
    if KNOWLEDGE_BASE_DIR.exists():
        for file in KNOWLEDGE_BASE_DIR.rglob("*.md"):
            text = file.read_text(encoding="utf-8", errors="ignore")
            chunks = _chunk_text_v47(text, max_chars=900, overlap=80) if "_chunk_text_v47" in globals() else [text]
            for idx, chunk in enumerate(chunks):
                words = {w for w in re.findall(r"[a-zA-Z0-9À-ỹ]+", _normalize_v43(chunk)) if len(w) >= 2}
                score = len(q_words & words) / max(1, len(q_words))
                if score > 0:
                    results.append({"source": str(file.relative_to(BASE_DIR)), "chunk_index": idx, "score": round(score, 3), "content": chunk[:1200]})
    store = _load_store()
    for d in store.setdefault("v47_documents", []):
        for idx, ch in enumerate(d.get("chunks") or []):
            content = ch.get("content") if isinstance(ch, dict) else str(ch)
            words = {w for w in re.findall(r"[a-zA-Z0-9À-ỹ]+", _normalize_v43(content)) if len(w) >= 2}
            score = len(q_words & words) / max(1, len(q_words))
            if score > 0:
                results.append({"source": d.get("title"), "document_id": d.get("id"), "chunk_index": idx, "score": round(score, 3), "content": content[:1200]})
    return sorted(results, key=lambda x: x.get("score", 0), reverse=True)[:limit]


@router.post("/ai/v50/knowledge-search")
def v50_knowledge_search(req: V50KnowledgeSearchRequest):
    return {"version": "V50", "query": req.query, "results": _v50_search_knowledge(req.query, req.limit)}


@router.post("/ai/v50/rag-chat")
def v50_rag_chat(req: V50KnowledgeSearchRequest):
    results = _v50_search_knowledge(req.query, req.limit)
    if not results:
        answer = "Chưa tìm thấy nguồn phù hợp trong knowledge_base/RAG. Hãy bổ sung tài liệu kế toán, thuế hoặc quy trình nội bộ rồi hỏi lại."
        confidence = "low"
    else:
        top = results[0]
        answer = "Dựa trên nguồn phù hợp nhất, hệ thống tìm thấy nội dung sau để kế toán tham khảo:\n" + top.get("content", "")[:900] + "\n\nLưu ý: đây là hỗ trợ tra cứu, cần kiểm tra văn bản gốc trước khi áp dụng chính thức."
        confidence = "high" if top.get("score", 0) >= 0.65 else "medium"
    return {"version": "V50", "query": req.query, "answer": answer, "confidence": confidence, "sources": results}


def _v51_report_payload(store: Dict[str, Any]) -> Dict[str, Any]:
    summary = _chat_summary(store)
    drafts = store.get("journal_drafts", [])
    risky = []
    for d in drafts:
        if d.get("risk_flags") or d.get("risk_note"):
            risky.append({"id": d.get("id"), "description": d.get("description"), "risk_flags": d.get("risk_flags"), "risk_note": d.get("risk_note"), "status": d.get("status")})
    return {"summary": summary, "drafts_pending": len([d for d in drafts if d.get("status") == "draft"]), "risk_count": len(risky), "risky_items": risky[:50]}


@router.get("/ai/v51/report/vat")
def v51_report_vat():
    store = _load_store()
    data = dict(_chat_summary(store)["vat"]); data.pop("version", None); return {"version": "V51", "report": "vat", **data}


@router.get("/ai/v51/report/profit")
def v51_report_profit():
    store = _load_store()
    data = dict(_chat_summary(store)["income"]); data.pop("version", None); return {"version": "V51", "report": "profit", **data}


@router.get("/ai/v51/report/cashflow")
def v51_report_cashflow():
    store = _load_store()
    entries = store.get("journal_entries", [])
    cash_in = cash_out = 0.0
    for e in entries:
        for l in e.get("lines", []):
            acc = str(l.get("account") or "")
            if acc.startswith(("111", "112")):
                cash_in += float(l.get("debit") or 0)
                cash_out += float(l.get("credit") or 0)
    return {"version": "V51", "report": "cashflow", "cash_in": round(cash_in, 2), "cash_out": round(cash_out, 2), "net_cashflow": round(cash_in - cash_out, 2), "note": "MVP tính theo phát sinh Nợ/Có tài khoản 111/112 trong journal_entries."}


@router.post("/ai/v51/ask-report")
def v51_ask_report(req: V51AskReportRequest):
    q = _normalize_v43(req.question)
    store = _load_store()
    if any(k in q for k in ["vat", "thue", "thuế"]):
        data = _chat_summary(store)["vat"]
        answer = f"VAT đầu ra: {data['vat_output']:,.0f}đ\nVAT đầu vào: {data['vat_input']:,.0f}đ\nVAT phải nộp: {data['vat_payable']:,.0f}đ"
        intent = "vat_report"
    elif any(k in q for k in ["dong tien", "dòng tiền", "cash", "tien mat", "tiền mặt"]):
        data = v51_report_cashflow()
        answer = f"Tiền vào: {data['cash_in']:,.0f}đ\nTiền ra: {data['cash_out']:,.0f}đ\nDòng tiền thuần: {data['net_cashflow']:,.0f}đ"
        intent = "cashflow_report"
    else:
        data = _chat_summary(store)["income"]
        answer = f"Doanh thu: {data['revenue']:,.0f}đ\nChi phí: {data['expenses']:,.0f}đ\nLợi nhuận trước thuế: {data['profit_before_tax']:,.0f}đ"
        intent = "profit_report"
    return {"version": "V51", "intent": intent, "answer": answer, "data": data}


@router.get("/ai/v52/dashboard-data")
def v52_dashboard_data():
    store = _load_store()
    payload = _v51_report_payload(store)
    summary = payload["summary"]
    cards = summary["dashboard"]["cards"]
    drafts = store.get("journal_drafts", [])
    by_category: Dict[str, float] = defaultdict(float)
    for d in drafts:
        by_category[str(d.get("category") or "Chưa phân loại")] += float(d.get("amount") or 0)
    return {"version": "V52", "cards": {**cards, "risk_count": payload["risk_count"], "drafts_pending": payload["drafts_pending"]}, "top_categories": sorted([{"category": k, "amount": round(v, 2)} for k, v in by_category.items()], key=lambda x: x["amount"], reverse=True)[:10], "risky_items": payload["risky_items"]}


@router.get("/v52/dashboard")
def v52_dashboard_page():
    path = FRONTEND_DIR / "v52_dashboard.html"
    if not path.exists():
        raise HTTPException(status_code=404, detail="Không tìm thấy frontend/v52_dashboard.html")
    return FileResponse(path, media_type="text/html")


@router.post("/ai/v53/audit-transactions")
def v53_audit_transactions(req: V53AuditRequest):
    store = _load_store()
    alerts = []
    if req.batch_id:
        batch = next((b for b in store.setdefault("v48_transaction_batches", []) if b.get("id") == req.batch_id), None)
        if batch:
            for r in batch.get("rows", []):
                if r.get("status") != "ready":
                    alerts.append({"level": "medium", "source": "v48_batch", "id": r.get("row_id"), "message": r.get("risk_note") or "Dòng cần kiểm tra"})
                if (r.get("payment_method") == "cash") and float(r.get("amount") or 0) >= 5_000_000:
                    alerts.append({"level": "high", "source": "v48_batch", "id": r.get("row_id"), "message": "Giao dịch tiền mặt từ 5 triệu đồng trở lên cần kiểm tra chứng từ và điều kiện thanh toán không dùng tiền mặt."})
    if req.include_drafts:
        for d in store.get("journal_drafts", []):
            if d.get("risk_flags"):
                alerts.append({"level": "high", "source": "journal_draft", "id": d.get("id"), "message": "Draft có risk_flags", "detail": d.get("risk_flags")})
            if not d.get("balanced"):
                alerts.append({"level": "high", "source": "journal_draft", "id": d.get("id"), "message": "Bút toán chưa cân Nợ/Có"})
            if not d.get("description") or not d.get("amount"):
                alerts.append({"level": "medium", "source": "journal_draft", "id": d.get("id"), "message": "Thiếu mô tả hoặc số tiền"})
    seen = set()
    if req.include_journal_entries:
        for e in store.get("journal_entries", []):
            sig = (e.get("description"), round(float(e.get("debit_total") or 0), 2), round(float(e.get("credit_total") or 0), 2))
            if sig in seen:
                alerts.append({"level": "medium", "source": "journal_entry", "id": e.get("id"), "message": "Có dấu hiệu trùng bút toán"})
            seen.add(sig)
    store["v53_risk_alerts"] = [{"created_at": _now(), **a} for a in alerts]
    _save_store(store)
    return {"version": "V53", "alert_count": len(alerts), "alerts": alerts}


@router.get("/ai/v53/risk-alerts")
def v53_risk_alerts(limit: int = 100):
    store = _load_store()
    return {"version": "V53", "count": len(store.get("v53_risk_alerts", [])), "alerts": store.get("v53_risk_alerts", [])[:max(1, min(limit, 500))]}


def _v54_extract_rule_suggestion(feedback: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    correction = feedback.get("user_correction")
    prediction = feedback.get("ai_prediction") or {}
    if isinstance(correction, str):
        text = correction
        debit = re.search(r"(?:nợ|no|debit)\s*([0-9]{3,4})", text, re.IGNORECASE)
        credit = re.search(r"(?:có|co|credit)\s*([0-9]{3,4})", text, re.IGNORECASE)
        if debit or credit:
            return {"feedback_id": feedback.get("id"), "description": feedback.get("user_message"), "suggested_debit_account": debit.group(1) if debit else prediction.get("debit_account"), "suggested_credit_account": credit.group(1) if credit else prediction.get("credit_account"), "reason": text, "status": "suggested"}
    if isinstance(correction, dict):
        if correction.get("debit_account") or correction.get("credit_account") or correction.get("category"):
            return {"feedback_id": feedback.get("id"), "description": feedback.get("user_message"), "suggested_category": correction.get("category"), "suggested_debit_account": correction.get("debit_account"), "suggested_credit_account": correction.get("credit_account"), "reason": correction, "status": "suggested"}
    return None


@router.get("/ai/v54/rule-suggestions")
def v54_rule_suggestions():
    store = _load_store()
    suggestions = store.get("v54_rule_suggestions")
    if suggestions is None:
        suggestions = [s for s in (_v54_extract_rule_suggestion(f) for f in store.get("ai_feedback", [])) if s]
        store["v54_rule_suggestions"] = suggestions
        _save_store(store)
    return {"version": "V54", "count": len(suggestions), "suggestions": suggestions}


@router.post("/ai/v54/apply-feedback-to-rules")
def v54_apply_feedback_to_rules(req: V54ApplyFeedbackRequest):
    store = _load_store()
    feedbacks = [f for f in store.get("ai_feedback", []) if (f.get("rating") or 0) >= req.min_rating or f.get("user_correction")]
    suggestions = [s for s in (_v54_extract_rule_suggestion(f) for f in feedbacks) if s]
    store["v54_rule_suggestions"] = suggestions
    applied = []
    if not req.dry_run and ACCOUNTING_RULES_PATH.exists():
        try:
            rules = json.loads(ACCOUNTING_RULES_PATH.read_text(encoding="utf-8"))
            if isinstance(rules, dict):
                rules_list = rules.get("rules", [])
            else:
                rules_list = rules
            for s in suggestions:
                desc = s.get("description") or ""
                keywords = [w for w in re.findall(r"[A-Za-zÀ-ỹ0-9]+", desc.lower()) if len(w) >= 4][:5]
                new_rule = {"rule_id": f"feedback_{s.get('feedback_id')}", "keywords": keywords, "category": s.get("suggested_category") or "Theo feedback người dùng", "debit_account": s.get("suggested_debit_account") or "642", "credit_account_options": [s.get("suggested_credit_account") or "112"], "risk": "Rule được tạo từ feedback, cần chief accountant duyệt trước khi dùng production.", "confidence": 0.6}
                if not any(r.get("rule_id") == new_rule["rule_id"] for r in rules_list):
                    rules_list.append(new_rule); applied.append(new_rule)
            ACCOUNTING_RULES_PATH.write_text(json.dumps(rules_list, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"Không ghi được accounting_rules.json: {exc}")
    store["v54_last_apply"] = {"created_at": _now(), "dry_run": req.dry_run, "suggestion_count": len(suggestions), "applied_count": len(applied)}
    _save_store(store)
    return {"version": "V54", "dry_run": req.dry_run, "suggestion_count": len(suggestions), "applied_count": len(applied), "suggestions": suggestions, "applied_rules": applied}


@router.get("/v55/mvp-demo")
def v55_mvp_demo_page():
    path = FRONTEND_DIR / "v55_mvp_demo.html"
    if not path.exists():
        raise HTTPException(status_code=404, detail="Không tìm thấy frontend/v55_mvp_demo.html")
    return FileResponse(path, media_type="text/html")


@router.get("/ai/v48-v55/upgrade-status")
def v48_v55_upgrade_status():
    store = _load_store()
    return {"version": "V48-V55", "name": "Finiip AI Accounting Assistant MVP", "completed": ["V48 upload Excel giao dịch + batch confirm", "V49 OCR hóa đơn/chứng từ + invoice-to-journal", "V50 knowledge search/RAG chat", "V51 báo cáo thông minh VAT/lợi nhuận/dòng tiền", "V52 dashboard AI", "V53 audit sai sót/rủi ro", "V54 học từ feedback -> rule suggestions", "V55 demo UI tổng hợp"], "main_urls": ["POST /ai/v48/upload-transactions", "GET /ai/v48/transactions/review", "POST /ai/v48/transactions/confirm-batch", "POST /ai/v49/ocr-invoice", "POST /ai/v49/invoice-to-journal", "POST /ai/v50/rag-chat", "POST /ai/v51/ask-report", "GET /v52/dashboard", "POST /ai/v53/audit-transactions", "GET /ai/v54/rule-suggestions", "GET /v55/mvp-demo"], "counts": {"v48_batches": len(store.get("v48_transaction_batches", [])), "v49_invoices": len(store.get("v49_invoices", [])), "risk_alerts": len(store.get("v53_risk_alerts", [])), "rule_suggestions": len(store.get("v54_rule_suggestions", []))}}
