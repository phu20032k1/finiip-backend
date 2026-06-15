from datetime import date, datetime, timedelta
import json
import os
import re
import tempfile
import glob
from typing import Any, Dict, List, Optional, Tuple
from difflib import SequenceMatcher

from fastapi import BackgroundTasks, Depends, FastAPI, File, Form, HTTPException, Query, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy import func, text
from sqlalchemy.orm import Session

from ai_engine import ACCOUNTING_RULES, add_custom_rule, benchmark_ai_cases, classify_transaction, demo_ai_cases, suggest_journal_entry
from ai_ml import (
    build_training_examples_from_corrections,
    model_status,
    predict_with_model,
    save_model,
    train_naive_bayes,
)
from database import Base, engine, get_db
from models import (
    Account,
    AccountingPeriod,
    AuditLog,
    AICorrection,
    AILog,
    AIReviewItem,
    Customer,
    JournalEntry,
    PurchaseInvoice,
    SalesInvoice,
    Supplier,
    Transaction,
)
from schemas import (
    AccountCreate,
    AccountResponse,
    AIAnalyzeRequest,
    AIBatchAnalyzeRequest,
    AICorrectionCreate,
    AICustomRuleCreate,
    AICorrectionResponse,
    AICreateTransactionRequest,
    AILogResponse,
    AccountingPeriodCreate,
    AccountingPeriodResponse,
    AuditLogResponse,
    ClosePeriodRequest,
    ImportExcelResponse,
    CustomerCreate,
    CustomerResponse,
    JournalEntryCreate,
    JournalEntryResponse,
    PurchaseInvoiceCreate,
    PurchaseInvoiceResponse,
    SalesInvoiceCreate,
    SalesInvoiceResponse,
    SupplierCreate,
    SupplierResponse,
    TransactionCreate,
    TransactionResponse,
    TransactionUpdate,
    AITeachExampleCreate,
    AITeachBatchRequest,
    AITrainingExampleUpdate,
    AIFeedbackRequest,
    AIMLEvaluateRequest,
    AIMLTrainRequest,
    AIMLPredictRequest,
    InvoiceOCRTextRequest,
    VATFormulaRequest,
    DepreciationFormulaRequest,
    PrepaidAllocationRequest,
    GrossProfitRequest,
    NetProfitRequest,
    CorporateIncomeTaxRequest,
    JournalCheckRequest,
    FinancialRatiosRequest,
    BreakEvenRequest,
    FIFOInventoryRequest,
    WeightedAverageInventoryRequest,
    PayrollBasicRequest,
    AccountsAgingRequest,
    PeriodClosingRequest,
    BasicFinancialStatementsRequest,
    FrontendTransactionPreviewRequest,
    FrontendInvoiceTextPreviewRequest,
    BulkReanalyzeTransactionsRequest,
    AIV18FeedbackLearningRequest,
    AIV19ReviewDecisionRequest,
    AIV20RetrainFeedbackRequest,
    AIV21OCRImproveTextRequest,
    AIV22DoubleEntryRequest,
)
from seed_data import DEFAULT_ACCOUNTS
from invoice_ocr import parse_invoice_text, read_text_from_upload
from ai_autopilot import autopilot_response
from finiip_v25_v40 import router as v25_v40_router
from ai_quality import AI_V3_DEMO_CASES, enhance_ai_result, run_ai_v3_test_suite
from accounting_formulas import (
    calculate_break_even,
    calculate_corporate_income_tax,
    calculate_financial_ratios,
    calculate_gross_profit,
    calculate_net_profit,
    calculate_prepaid_allocation,
    calculate_straight_line_depreciation,
    calculate_vat,
    check_journal_balance,
    formula_catalog,
    calculate_fifo_inventory,
    calculate_weighted_average_inventory,
    calculate_payroll_basic,
    calculate_accounts_aging,
    generate_period_closing_entries,
    build_basic_financial_statements,
)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover
    Workbook = None
    Font = PatternFill = Alignment = get_column_letter = None

Base.metadata.create_all(bind=engine)


def migrate_sqlite_schema():
    """Bổ sung cột còn thiếu cho database SQLite cũ mà không xóa dữ liệu."""
    if not str(engine.url).startswith("sqlite"):
        return
    migrations = {
        "transactions": {
            "debit_account_code": "VARCHAR",
            "credit_account_code": "VARCHAR",
            "ai_confidence": "FLOAT",
            "status": "VARCHAR DEFAULT 'draft'",
            "confirmed_at": "DATETIME",
            "cancelled_at": "DATETIME",
            "accounting_period": "VARCHAR",
            "updated_at": "DATETIME DEFAULT CURRENT_TIMESTAMP",
        },
        "journal_entries": {
            "line_no": "INTEGER DEFAULT 1",
            "status": "VARCHAR DEFAULT 'draft'",
            "accounting_period": "VARCHAR",
            "updated_at": "DATETIME DEFAULT CURRENT_TIMESTAMP",
        },
        "accounts": {
            "updated_at": "DATETIME DEFAULT CURRENT_TIMESTAMP",
        },
        "customers": {
            "updated_at": "DATETIME DEFAULT CURRENT_TIMESTAMP",
        },
        "suppliers": {
            "updated_at": "DATETIME DEFAULT CURRENT_TIMESTAMP",
        },
        "ai_logs": {
            "updated_at": "DATETIME DEFAULT CURRENT_TIMESTAMP",
        },
        "ai_corrections": {
            "updated_at": "DATETIME DEFAULT CURRENT_TIMESTAMP",
        },
        "ai_review_items": {
            "transaction_id": "INTEGER",
            "source": "VARCHAR DEFAULT 'ai'",
            "description": "TEXT",
            "amount": "FLOAT",
            "ai_category": "VARCHAR",
            "ai_type": "VARCHAR",
            "ai_debit_account_code": "VARCHAR",
            "ai_credit_account_code": "VARCHAR",
            "ai_confidence": "FLOAT",
            "ai_result_json": "TEXT",
            "status": "VARCHAR DEFAULT 'pending'",
            "priority": "VARCHAR DEFAULT 'medium'",
            "reason": "TEXT",
            "reviewer_note": "TEXT",
            "reviewed_at": "DATETIME",
            "created_at": "DATETIME DEFAULT CURRENT_TIMESTAMP",
            "updated_at": "DATETIME DEFAULT CURRENT_TIMESTAMP",
        },
    }
    with engine.begin() as conn:
        for table_name, columns in migrations.items():
            existing = {row[1] for row in conn.execute(text(f"PRAGMA table_info({table_name})"))}
            if not existing:
                continue
            for column_name, column_type in columns.items():
                if column_name not in existing:
                    conn.execute(text(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {column_type}"))


migrate_sqlite_schema()

app = FastAPI(
    title="Finiip AI Accounting API",
    description="Finiip V15: Backend API-first cho frontend có sẵn, AI kế toán, OCR, ML feedback và Accounting Engine.",
    version="15.0.0",
)

def _get_cors_origins() -> List[str]:
    """
    Local dev: allow all origins by default.
    Production: set CORS_ORIGINS=https://your-frontend.vercel.app,https://your-domain.com
    """
    raw_origins = os.getenv("CORS_ORIGINS", "*").strip()
    if raw_origins == "*":
        return ["*"]
    return [origin.strip() for origin in raw_origins.split(",") if origin.strip()]


_cors_origins = _get_cors_origins()
app.add_middleware(
    CORSMiddleware,
    allow_origins=_cors_origins,
    # Browsers reject credentialed CORS with a wildcard origin. The chat MVP
    # uses headers, not cookies, so wildcard mode stays credential-free.
    allow_credentials=_cors_origins != ["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.middleware("http")
async def optional_api_key_guard(request: Request, call_next):
    """Bật bảo vệ API bằng FINIIP_API_KEY nếu deploy thật; mặc định tắt để dev frontend dễ gọi."""
    api_key = os.getenv("FINIIP_API_KEY")
    if api_key:
        public_paths = {"/", "/docs", "/redoc", "/openapi.json", "/api/v1/health"}
        path = request.url.path
        # V100 backend-only Admin RAG UI uses normal HTML forms, so it cannot
        # attach X-API-Key headers. Let the route protect itself with
        # FINIIP_ADMIN_KEY/FINIIP_API_KEY via query/form field.
        if path.startswith("/admin/rag-ui"):
            return await call_next(request)
        if path not in public_paths and not path.startswith("/docs") and not path.startswith("/redoc"):
            provided = request.headers.get("X-API-Key")
            if provided != api_key:
                raise HTTPException(status_code=401, detail="Thiếu hoặc sai X-API-Key")
    return await call_next(request)


# Optional legacy IIP Steel module. Bản MVP kế toán có thể chạy độc lập nếu module này không có.
try:
    from iip_steel_platform import router as iip_steel_router
    app.include_router(iip_steel_router)
except ModuleNotFoundError:
    iip_steel_router = None




# =========================
# V5 Accounting Core helpers
# =========================

VALID_TRANSACTION_STATUSES = {"draft", "confirmed", "cancelled"}
VALID_JOURNAL_STATUSES = {"draft", "posted", "cancelled"}


def period_key(d: date) -> str:
    return d.strftime("%Y-%m")


def month_bounds(period: str) -> tuple[date, date]:
    try:
        year, month = map(int, period.split("-"))
        start = date(year, month, 1)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Kỳ kế toán phải có dạng YYYY-MM") from exc
    if month == 12:
        end = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        end = date(year, month + 1, 1) - timedelta(days=1)
    return start, end


def get_or_create_period(db: Session, d: date) -> AccountingPeriod:
    key = period_key(d)
    period = db.query(AccountingPeriod).filter(AccountingPeriod.period == key).first()
    if period:
        return period
    start, end = month_bounds(key)
    period = AccountingPeriod(period=key, start_date=start, end_date=end, status="open")
    db.add(period)
    db.flush()
    return period


def ensure_period_open(db: Session, d: date) -> None:
    period = get_or_create_period(db, d)
    if period.status == "closed":
        raise HTTPException(status_code=409, detail=f"Kỳ kế toán {period.period} đã khóa, không thể ghi/sửa/xóa dữ liệu")


def ensure_account_exists(db: Session, code: Optional[str]) -> None:
    if not code:
        raise HTTPException(status_code=400, detail="Thiếu mã tài khoản kế toán")
    if not db.query(Account).filter(Account.code == code).first():
        raise HTTPException(status_code=400, detail=f"Tài khoản {code} chưa tồn tại. Hãy chạy /setup/default-accounts trước")


def audit_log(db: Session, *, action: str, entity_type: str, entity_id: Optional[int] = None, old: Any = None, new: Any = None, note: Optional[str] = None, actor: str = "system") -> None:
    db.add(AuditLog(
        action=action,
        entity_type=entity_type,
        entity_id=entity_id,
        old_value_json=json.dumps(old, ensure_ascii=False, default=str) if old is not None else None,
        new_value_json=json.dumps(new, ensure_ascii=False, default=str) if new is not None else None,
        note=note,
        actor=actor,
    ))


def account_to_dict(a: Account) -> Dict[str, Any]:
    return {
        "id": a.id,
        "code": a.code,
        "name": a.name,
        "account_type": a.account_type,
        "created_at": str(a.created_at),
        "updated_at": str(getattr(a, "updated_at", "")),
    }


def transaction_to_dict(t: Transaction) -> Dict[str, Any]:
    return {
        "id": t.id,
        "transaction_date": str(t.transaction_date),
        "description": t.description,
        "amount": t.amount,
        "type": t.type,
        "category": t.category,
        "status": getattr(t, "status", "draft"),
        "debit_account_code": t.debit_account_code,
        "credit_account_code": t.credit_account_code,
        "ai_confidence": t.ai_confidence,
        "accounting_period": getattr(t, "accounting_period", None),
    }



def purchase_invoice_to_dict(i: PurchaseInvoice) -> Dict[str, Any]:
    return {
        "id": i.id,
        "invoice_date": str(i.invoice_date),
        "invoice_number": i.invoice_number,
        "supplier_id": i.supplier_id,
        "supplier_name": i.supplier_name,
        "description": i.description,
        "subtotal": i.subtotal,
        "vat_rate": i.vat_rate,
        "vat_amount": i.vat_amount,
        "total_amount": i.total_amount,
        "status": i.status,
    }



def journal_entry_to_dict(e: JournalEntry) -> Dict[str, Any]:
    return {
        "id": e.id,
        "transaction_id": e.transaction_id,
        "entry_date": str(e.entry_date),
        "description": e.description,
        "debit_account_code": e.debit_account_code,
        "debit_account_name": e.debit_account_name,
        "credit_account_code": e.credit_account_code,
        "credit_account_name": e.credit_account_name,
        "amount": e.amount,
        "line_no": e.line_no,
        "status": e.status,
        "accounting_period": e.accounting_period,
    }


def validate_journal_payload(db: Session, entry_date: date, debit_code: str, credit_code: str, amount: float) -> None:
    if amount <= 0:
        raise HTTPException(status_code=400, detail="Số tiền bút toán phải lớn hơn 0")
    if debit_code == credit_code:
        raise HTTPException(status_code=400, detail="Tài khoản Nợ và Có không được trùng nhau")
    ensure_account_exists(db, debit_code)
    ensure_account_exists(db, credit_code)
    ensure_period_open(db, entry_date)


def validate_transaction_journal_balance(db: Session, transaction_id: int) -> Dict[str, Any]:
    entries = db.query(JournalEntry).filter(JournalEntry.transaction_id == transaction_id, JournalEntry.status != "cancelled").all()
    debit_total = sum(float(e.amount or 0) for e in entries)
    credit_total = sum(float(e.amount or 0) for e in entries)
    balanced = round(debit_total - credit_total, 2) == 0 and debit_total > 0
    return {"transaction_id": transaction_id, "debit_total": debit_total, "credit_total": credit_total, "balanced": balanced, "entries": len(entries)}


def get_account_name(db: Session, code: Optional[str]) -> str:
    if not code:
        return "Unknown"
    account = db.query(Account).filter(Account.code == code).first()
    return account.name if account else "Unknown"


def create_ai_log(
    db: Session,
    *,
    action: str,
    description: Optional[str] = None,
    amount: Optional[float] = None,
    result: Optional[Dict[str, Any]] = None,
    transaction_id: Optional[int] = None,
) -> AILog:
    result = result or {}
    log = AILog(
        transaction_id=transaction_id,
        action=action,
        input_description=description,
        input_amount=amount,
        predicted_category=result.get("category"),
        predicted_type=result.get("transaction_type"),
        debit_account_code=result.get("debit_account_code") or result.get("debit_account"),
        credit_account_code=result.get("credit_account_code") or result.get("credit_account"),
        ai_confidence=result.get("confidence"),
        ai_result_json=json.dumps(result, ensure_ascii=False, default=str),
    )
    db.add(log)
    db.commit()
    db.refresh(log)
    return log



# =========================
# V6 - AI Learning Lite helpers
# =========================

VIETNAMESE_STOPWORDS = {
    "thang", "nam", "ngay", "cho", "cua", "va", "voi", "bang", "tu", "den", "da",
    "theo", "mot", "cac", "khoan", "tien", "thanh", "toan", "chi", "thu", "nhan",
}


def normalize_vi_text(value: Optional[str]) -> str:
    """Chuẩn hóa tiếng Việt có dấu/không dấu để so khớp correction memory."""
    import unicodedata
    value = (value or "").lower().strip()
    value = unicodedata.normalize("NFD", value)
    value = "".join(ch for ch in value if unicodedata.category(ch) != "Mn")
    value = value.replace("đ", "d")
    value = re.sub(r"[^a-z0-9\s]", " ", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value


def text_tokens(value: Optional[str]) -> set:
    normalized = normalize_vi_text(value)
    return {token for token in normalized.split() if len(token) >= 2 and token not in VIETNAMESE_STOPWORDS}


def jaccard_similarity(a: Optional[str], b: Optional[str]) -> float:
    left = text_tokens(a)
    right = text_tokens(b)
    if not left or not right:
        return 0.0
    return round(len(left & right) / len(left | right), 4)


def correction_to_learning_result(correction: AICorrection, amount: float, similarity: float) -> Dict[str, Any]:
    debit_name = "Unknown"
    credit_name = "Unknown"
    # Names are filled later in suggest_journal_entry_with_learning because it has db access.
    return {
        "category": correction.user_category,
        "transaction_type": correction.user_type,
        "debit_account_code": correction.user_debit_account_code,
        "debit_account": correction.user_debit_account_code,
        "debit_account_name": debit_name,
        "credit_account_code": correction.user_credit_account_code,
        "credit_account": correction.user_credit_account_code,
        "credit_account_name": credit_name,
        "amount": amount,
        "confidence": min(0.98, max(0.78, similarity + 0.18)),
        "source": "learning_memory",
        "learning_correction_id": correction.id,
        "learning_similarity": similarity,
        "matched_description": correction.original_description,
        "warnings": ["Kết quả ưu tiên từ correction memory - nên kiểm tra lại trước khi confirmed"],
        "journal_lines": [
            {"side": "debit", "account_code": correction.user_debit_account_code, "account_name": debit_name, "amount": amount},
            {"side": "credit", "account_code": correction.user_credit_account_code, "account_name": credit_name, "amount": amount},
        ],
    }


def find_best_correction_match(db: Session, description: str, min_similarity: float = 0.42) -> Tuple[Optional[AICorrection], float]:
    corrections = db.query(AICorrection).order_by(AICorrection.id.desc()).limit(300).all()
    best = None
    best_score = 0.0
    for correction in corrections:
        score = jaccard_similarity(description, correction.original_description)
        if score > best_score:
            best = correction
            best_score = score
    if best and best_score >= min_similarity:
        return best, best_score
    return None, best_score


def _fill_account_names_for_ai_result(db: Session, result: Dict[str, Any]) -> Dict[str, Any]:
    """Điền tên tài khoản cho mọi kết quả AI có debit/credit code."""
    if result.get("debit_account_code"):
        result["debit_account_name"] = get_account_name(db, result.get("debit_account_code"))
    if result.get("credit_account_code"):
        result["credit_account_name"] = get_account_name(db, result.get("credit_account_code"))
    for line in result.get("journal_lines") or []:
        if line.get("account_code"):
            line["account_name"] = get_account_name(db, line.get("account_code"))
    return result


def suggest_journal_entry_with_learning(db: Session, description: str, amount: float, min_similarity: float = 0.42, ml_min_confidence: float = 0.55) -> Dict[str, Any]:
    """Kết hợp 3 tầng AI: correction memory -> ML model -> rule-based fallback."""
    rule_result = suggest_journal_entry(description, amount)

    correction, similarity = find_best_correction_match(db, description, min_similarity=min_similarity)
    if correction:
        learned = correction_to_learning_result(correction, amount, similarity)
        _fill_account_names_for_ai_result(db, learned)
        learned["rule_based_result"] = rule_result
        return learned

    ml_result = predict_with_model(description, amount)
    if ml_result and float(ml_result.get("confidence") or 0) >= ml_min_confidence:
        _fill_account_names_for_ai_result(db, ml_result)
        ml_result["rule_based_result"] = rule_result
        ml_result["best_learning_similarity"] = similarity
        return ml_result

    rule_result["source"] = "rule_based"
    rule_result["best_learning_similarity"] = similarity
    if ml_result:
        rule_result["ml_candidate"] = {
            "category": ml_result.get("category"),
            "transaction_type": ml_result.get("transaction_type"),
            "debit_account_code": ml_result.get("debit_account_code"),
            "credit_account_code": ml_result.get("credit_account_code"),
            "confidence": ml_result.get("confidence"),
            "reason_not_used": f"ML confidence thấp hơn ngưỡng {ml_min_confidence}",
        }
    return rule_result

def create_journal_from_ai_result(
    db: Session,
    *,
    transaction_id: Optional[int],
    description: str,
    result: Dict[str, Any],
    commit: bool = True,
    status: str = "draft",
    entry_date: Optional[date] = None,
) -> List[JournalEntry]:
    lines = result.get("journal_lines") or []
    entries: List[JournalEntry] = []
    entry_date = entry_date or date.today()
    ensure_period_open(db, entry_date)

    if lines:
        debit_lines = [line for line in lines if line.get("side") == "debit"]
        credit_lines = [line for line in lines if line.get("side") == "credit"]
        line_no = 1
        for debit in debit_lines:
            for credit in credit_lines:
                amount = min(float(debit.get("amount") or 0), float(credit.get("amount") or 0))
                if amount <= 0:
                    continue
                entry = JournalEntry(
                    transaction_id=transaction_id,
                    entry_date=entry_date,
                    description=description,
                    debit_account_code=str(debit.get("account_code")),
                    debit_account_name=str(debit.get("account_name")),
                    credit_account_code=str(credit.get("account_code")),
                    credit_account_name=str(credit.get("account_name")),
                    amount=amount,
                    line_no=line_no,
                    status=status,
                    accounting_period=period_key(entry_date),
                )
                db.add(entry)
                entries.append(entry)
                line_no += 1
    else:
        debit_code = result.get("debit_account_code")
        credit_code = result.get("credit_account_code")
        amount = result.get("amount")
        if debit_code and credit_code and amount:
            entry = JournalEntry(
                transaction_id=transaction_id,
                entry_date=entry_date,
                description=description,
                debit_account_code=debit_code,
                debit_account_name=get_account_name(db, debit_code),
                credit_account_code=credit_code,
                credit_account_name=get_account_name(db, credit_code),
                amount=amount,
                status=status,
                accounting_period=period_key(entry_date),
            )
            db.add(entry)
            entries.append(entry)

    if commit:
        db.commit()
        for entry in entries:
            db.refresh(entry)
    return entries


app.include_router(v25_v40_router)

@app.get("/")
def home():
    return {
        "name": "Finiip AI Accounting API",
        "version": "6.0.0",
        "stage": "V6 AI Learning Lite - correction memory/similarity/accuracy/rule suggestions",
        "next_step": "AI đã có thể ưu tiên học từ correction; frontend có thể làm sau",
        "docs": "/docs",
    }


@app.get("/app")
def finiip_demo_app():
    """Backend-only build: frontend được tách sang repository/project riêng."""
    return {
        "message": "Finiip backend-only API is running. Frontend should call this backend via HTTP API.",
        "docs": "/docs",
        "openapi": "/openapi.json",
        "health": "/api/v1/health",
        "frontend_contract": "/api/v1/frontend/contract",
        "ai_v3": {
            "single_analyze": "POST /ai/v3/analyze",
            "batch_analyze": "POST /ai/v3/batch-analyze",
            "demo_cases": "GET /ai/v3/demo-cases",
            "test_suite": "GET /ai/v3/test-suite"
        }
    }


@app.get("/frontend/index.html")
def finiip_demo_index_html():
    """Alias cũ, giữ để không phá link; bản này không kèm frontend HTML."""
    return finiip_demo_app()


@app.get("/system/status")
def system_status(db: Session = Depends(get_db)):
    return {
        "status": "ok",
        "version": "5.0.0",
        "database": "connected",
        "counts": {
            "accounts": db.query(Account).count(),
            "transactions": db.query(Transaction).count(),
            "journal_entries": db.query(JournalEntry).count(),
            "ai_logs": db.query(AILog).count(),
            "ai_corrections": db.query(AICorrection).count(),
            "ai_review_items": db.query(AIReviewItem).count(),
            "accounting_periods": db.query(AccountingPeriod).count(),
            "audit_logs": db.query(AuditLog).count(),
        },
    }


@app.post("/setup/default-accounts")
def setup_default_accounts(db: Session = Depends(get_db)):
    created = 0
    skipped = 0
    for item in DEFAULT_ACCOUNTS:
        exists = db.query(Account).filter(Account.code == item["code"]).first()
        if exists:
            skipped += 1
            continue
        db.add(Account(**item))
        created += 1
    db.commit()
    return {"message": "Đã thiết lập hệ thống tài khoản", "created": created, "skipped": skipped}


@app.get("/accounts", response_model=List[AccountResponse])
def list_accounts(db: Session = Depends(get_db)):
    return db.query(Account).order_by(Account.code).all()


@app.post("/accounts", response_model=AccountResponse)
def create_account(payload: AccountCreate, db: Session = Depends(get_db)):
    if db.query(Account).filter(Account.code == payload.code).first():
        raise HTTPException(status_code=400, detail="Mã tài khoản đã tồn tại")
    account = Account(**payload.model_dump())
    db.add(account)
    db.commit()
    db.refresh(account)
    return account


@app.get("/accounts/{code}", response_model=AccountResponse)
def get_account(code: str, db: Session = Depends(get_db)):
    account = db.query(Account).filter(Account.code == code).first()
    if not account:
        raise HTTPException(status_code=404, detail="Không tìm thấy tài khoản")
    return account


@app.post("/customers", response_model=CustomerResponse)
def create_customer(payload: CustomerCreate, db: Session = Depends(get_db)):
    obj = Customer(**payload.model_dump())
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return obj


@app.get("/customers", response_model=List[CustomerResponse])
def list_customers(db: Session = Depends(get_db)):
    return db.query(Customer).order_by(Customer.id.desc()).all()


@app.post("/suppliers", response_model=SupplierResponse)
def create_supplier(payload: SupplierCreate, db: Session = Depends(get_db)):
    obj = Supplier(**payload.model_dump())
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return obj


@app.get("/suppliers", response_model=List[SupplierResponse])
def list_suppliers(db: Session = Depends(get_db)):
    return db.query(Supplier).order_by(Supplier.id.desc()).all()


@app.post("/transactions", response_model=TransactionResponse)
def create_transaction(payload: TransactionCreate, db: Session = Depends(get_db)):
    data = payload.model_dump()
    if data["transaction_date"] is None:
        data["transaction_date"] = date.today()
    ensure_period_open(db, data["transaction_date"])
    data["accounting_period"] = period_key(data["transaction_date"])
    transaction = Transaction(**data)
    db.add(transaction)
    db.flush()
    audit_log(db, action="create", entity_type="transaction", entity_id=transaction.id, new=transaction_to_dict(transaction))
    db.commit()
    db.refresh(transaction)
    return transaction


@app.get("/transactions", response_model=List[TransactionResponse])
def list_transactions(
    db: Session = Depends(get_db),
    skip: int = 0,
    limit: int = Query(100, le=500),
    type: Optional[str] = None,
    category: Optional[str] = None,
    status: Optional[str] = None,
    period: Optional[str] = None,
    keyword: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
):
    query = db.query(Transaction)
    if type:
        query = query.filter(Transaction.type == type)
    if category:
        query = query.filter(Transaction.category == category)
    if status:
        query = query.filter(Transaction.status == status)
    if period:
        query = query.filter(Transaction.accounting_period == period)
    if keyword:
        query = query.filter(Transaction.description.ilike(f"%{keyword}%"))
    if date_from:
        query = query.filter(Transaction.transaction_date >= date_from)
    if date_to:
        query = query.filter(Transaction.transaction_date <= date_to)
    return query.order_by(Transaction.transaction_date.desc(), Transaction.id.desc()).offset(skip).limit(limit).all()


@app.get("/transactions/{transaction_id}", response_model=TransactionResponse)
def get_transaction(transaction_id: int, db: Session = Depends(get_db)):
    transaction = db.get(Transaction, transaction_id)
    if not transaction:
        raise HTTPException(status_code=404, detail="Không tìm thấy giao dịch")
    return transaction


@app.put("/transactions/{transaction_id}", response_model=TransactionResponse)
def update_transaction(transaction_id: int, payload: TransactionUpdate, db: Session = Depends(get_db)):
    transaction = db.get(Transaction, transaction_id)
    if not transaction:
        raise HTTPException(status_code=404, detail="Không tìm thấy giao dịch")
    ensure_period_open(db, transaction.transaction_date)
    old_value = transaction_to_dict(transaction)
    for key, value in payload.model_dump(exclude_unset=True).items():
        setattr(transaction, key, value)
    if payload.transaction_date:
        ensure_period_open(db, payload.transaction_date)
        transaction.accounting_period = period_key(payload.transaction_date)
    audit_log(db, action="update", entity_type="transaction", entity_id=transaction.id, old=old_value, new=transaction_to_dict(transaction))
    db.commit()
    db.refresh(transaction)
    return transaction


@app.delete("/transactions/{transaction_id}")
def delete_transaction(transaction_id: int, db: Session = Depends(get_db)):
    transaction = db.get(Transaction, transaction_id)
    if not transaction:
        raise HTTPException(status_code=404, detail="Không tìm thấy giao dịch")
    ensure_period_open(db, transaction.transaction_date)
    old_value = transaction_to_dict(transaction)
    db.query(JournalEntry).filter(JournalEntry.transaction_id == transaction_id).delete()
    db.delete(transaction)
    audit_log(db, action="delete", entity_type="transaction", entity_id=transaction_id, old=old_value)
    db.commit()
    return {"message": "Đã xóa giao dịch và bút toán liên quan"}


@app.post("/transactions/{transaction_id}/reanalyze")
def reanalyze_transaction(transaction_id: int, db: Session = Depends(get_db)):
    transaction = db.get(Transaction, transaction_id)
    if not transaction:
        raise HTTPException(status_code=404, detail="Không tìm thấy giao dịch")
    result = suggest_journal_entry_with_learning(db, transaction.description, transaction.amount)
    transaction.category = result.get("category")
    transaction.type = result.get("transaction_type") or transaction.type
    transaction.debit_account_code = result.get("debit_account_code")
    transaction.credit_account_code = result.get("credit_account_code")
    transaction.ai_confidence = result.get("confidence")
    db.commit()
    db.refresh(transaction)
    log = create_ai_log(db, action="reanalyze_transaction", description=transaction.description, amount=transaction.amount, result=result, transaction_id=transaction.id)
    return {"transaction": transaction, "ai_result": result, "ai_log_id": log.id}


@app.post("/journal-entries", response_model=JournalEntryResponse)
def create_journal_entry(payload: JournalEntryCreate, db: Session = Depends(get_db)):
    debit_name = get_account_name(db, payload.debit_account_code)
    credit_name = get_account_name(db, payload.credit_account_code)
    entry = JournalEntry(
        transaction_id=payload.transaction_id,
        entry_date=payload.entry_date or date.today(),
        description=payload.description,
        debit_account_code=payload.debit_account_code,
        debit_account_name=debit_name,
        credit_account_code=payload.credit_account_code,
        credit_account_name=credit_name,
        amount=payload.amount,
    )
    db.add(entry)
    db.commit()
    db.refresh(entry)
    return entry


@app.get("/journal-entries", response_model=List[JournalEntryResponse])
def list_journal_entries(db: Session = Depends(get_db), transaction_id: Optional[int] = None):
    query = db.query(JournalEntry)
    if transaction_id:
        query = query.filter(JournalEntry.transaction_id == transaction_id)
    return query.order_by(JournalEntry.entry_date.desc(), JournalEntry.id.desc()).all()


@app.delete("/journal-entries/{entry_id}")
def delete_journal_entry(entry_id: int, db: Session = Depends(get_db)):
    entry = db.get(JournalEntry, entry_id)
    if not entry:
        raise HTTPException(status_code=404, detail="Không tìm thấy bút toán")
    db.delete(entry)
    db.commit()
    return {"message": "Đã xóa bút toán"}


@app.post("/sales-invoices", response_model=SalesInvoiceResponse)
def create_sales_invoice(payload: SalesInvoiceCreate, db: Session = Depends(get_db)):
    vat_amount = payload.subtotal * payload.vat_rate / 100
    invoice = SalesInvoice(**payload.model_dump(exclude={"invoice_date"}), invoice_date=payload.invoice_date or date.today(), vat_amount=vat_amount, total_amount=payload.subtotal + vat_amount)
    db.add(invoice)
    db.commit()
    db.refresh(invoice)
    return invoice


@app.get("/sales-invoices", response_model=List[SalesInvoiceResponse])
def list_sales_invoices(db: Session = Depends(get_db)):
    return db.query(SalesInvoice).order_by(SalesInvoice.invoice_date.desc(), SalesInvoice.id.desc()).all()


@app.post("/purchase-invoices", response_model=PurchaseInvoiceResponse)
def create_purchase_invoice(payload: PurchaseInvoiceCreate, db: Session = Depends(get_db)):
    vat_amount = payload.subtotal * payload.vat_rate / 100
    invoice = PurchaseInvoice(**payload.model_dump(exclude={"invoice_date"}), invoice_date=payload.invoice_date or date.today(), vat_amount=vat_amount, total_amount=payload.subtotal + vat_amount)
    db.add(invoice)
    db.commit()
    db.refresh(invoice)
    return invoice


@app.get("/purchase-invoices", response_model=List[PurchaseInvoiceResponse])
def list_purchase_invoices(db: Session = Depends(get_db)):
    return db.query(PurchaseInvoice).order_by(PurchaseInvoice.invoice_date.desc(), PurchaseInvoice.id.desc()).all()




# =========================
# V12 - OCR đọc hóa đơn
# =========================

def _date_from_ocr(value: Optional[str]) -> date:
    if not value:
        return date.today()
    try:
        return date.fromisoformat(value)
    except ValueError:
        return date.today()


def _invoice_ai_payload(parsed: Dict[str, Any], db: Session) -> Dict[str, Any]:
    amount = parsed.get("total_amount") or parsed.get("subtotal") or 0
    description = parsed.get("description") or "Hóa đơn mua hàng"
    ai_result = None
    if amount and amount > 0:
        ai_result = suggest_journal_entry_with_learning(db, description, float(amount))
    return {
        "description_for_ai": description,
        "amount_for_ai": amount,
        "ai_result": ai_result,
    }


def _create_purchase_invoice_from_ocr(parsed: Dict[str, Any], db: Session) -> Optional[PurchaseInvoice]:
    subtotal = parsed.get("subtotal")
    total = parsed.get("total_amount")
    vat_amount = parsed.get("vat_amount")
    vat_rate = parsed.get("vat_rate") or 0
    if subtotal is None and total is not None:
        if vat_amount is not None:
            subtotal = max(float(total) - float(vat_amount), 0)
        elif vat_rate:
            subtotal = float(total) / (1 + float(vat_rate) / 100)
        else:
            subtotal = float(total)
    if not subtotal or subtotal <= 0:
        return None
    vat_amount_calc = float(subtotal) * float(vat_rate) / 100
    invoice = PurchaseInvoice(
        invoice_date=_date_from_ocr(parsed.get("invoice_date")),
        invoice_number=parsed.get("invoice_number") or f"OCR-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}",
        supplier_name=parsed.get("supplier_name") or "Nhà cung cấp chưa xác định",
        description=parsed.get("description") or "Hóa đơn OCR",
        subtotal=float(subtotal),
        vat_rate=float(vat_rate or 0),
        vat_amount=vat_amount if vat_amount is not None else vat_amount_calc,
        total_amount=total if total is not None else float(subtotal) + vat_amount_calc,
        status="unpaid",
    )
    db.add(invoice)
    db.flush()
    audit_log(db, action="ocr_create_purchase_invoice", entity_type="purchase_invoice", entity_id=invoice.id, new={
        "invoice_number": invoice.invoice_number,
        "supplier_name": invoice.supplier_name,
        "total_amount": invoice.total_amount,
    })
    return invoice


def _create_transaction_from_ocr(parsed: Dict[str, Any], ai_result: Dict[str, Any], db: Session, auto_create_journal: bool = False) -> Optional[Transaction]:
    amount = parsed.get("total_amount") or parsed.get("subtotal")
    if not amount or amount <= 0:
        return None
    tx_date = _date_from_ocr(parsed.get("invoice_date"))
    ensure_period_open(db, tx_date)
    transaction = Transaction(
        transaction_date=tx_date,
        description=parsed.get("description") or "Hóa đơn OCR",
        amount=float(amount),
        type=ai_result.get("transaction_type") or "expense",
        category=ai_result.get("category"),
        note=f"Tạo từ OCR hóa đơn số {parsed.get('invoice_number') or 'chưa rõ'}",
        debit_account_code=ai_result.get("debit_account_code"),
        credit_account_code=ai_result.get("credit_account_code"),
        ai_confidence=ai_result.get("confidence"),
        status="draft",
        accounting_period=period_key(tx_date),
    )
    db.add(transaction)
    db.flush()
    if auto_create_journal:
        create_journal_from_ai_result(db, transaction_id=transaction.id, description=transaction.description, result=ai_result, commit=False, status="draft", entry_date=tx_date)
    audit_log(db, action="ocr_create_transaction", entity_type="transaction", entity_id=transaction.id, new=transaction_to_dict(transaction))
    return transaction


def _build_ocr_response(parsed: Dict[str, Any], db: Session, *, create_purchase_invoice: bool = False, create_transaction: bool = False, auto_create_journal: bool = False, source: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    ai_payload = _invoice_ai_payload(parsed, db)
    created_invoice = None
    created_transaction = None
    if create_purchase_invoice:
        created_invoice = _create_purchase_invoice_from_ocr(parsed, db)
    if create_transaction and ai_payload["ai_result"]:
        created_transaction = _create_transaction_from_ocr(parsed, ai_payload["ai_result"], db, auto_create_journal=auto_create_journal)
    create_ai_log(db, action="ocr_invoice_extract", description=parsed.get("description"), amount=parsed.get("total_amount") or parsed.get("subtotal"), result={
        "source": source or {},
        "parsed": {k: v for k, v in parsed.items() if k != "raw_text"},
        "ai_result": ai_payload["ai_result"],
    }, transaction_id=created_transaction.id if created_transaction else None)
    db.commit()
    if created_invoice:
        db.refresh(created_invoice)
    if created_transaction:
        db.refresh(created_transaction)
    return {
        "message": "Đã đọc hóa đơn bằng OCR-lite",
        "source": source or {},
        "extracted": parsed,
        "ai_suggestion": ai_payload,
        "created_purchase_invoice": purchase_invoice_to_dict(created_invoice) if created_invoice else None,
        "created_transaction": transaction_to_dict(created_transaction) if created_transaction else None,
        "next_step": "Nếu kết quả sai, sửa bằng /ai/feedback để AI học lại; nếu ảnh chưa đọc được, cài Tesseract OCR hoặc upload PDF có text.",
    }


@app.post("/ocr/invoice/text")
def ocr_invoice_from_text(payload: InvoiceOCRTextRequest, db: Session = Depends(get_db)):
    """Đọc hóa đơn từ text đã OCR sẵn. Dùng tốt cho test backend và copy text từ PDF/hóa đơn điện tử."""
    parsed = parse_invoice_text(payload.raw_text)
    return _build_ocr_response(
        parsed,
        db,
        create_purchase_invoice=payload.create_purchase_invoice,
        create_transaction=payload.create_transaction,
        auto_create_journal=payload.auto_create_journal,
        source={"method": "raw_text"},
    )


@app.post("/ocr/invoice/upload")
async def ocr_invoice_upload(
    file: UploadFile = File(...),
    create_purchase_invoice: bool = False,
    create_transaction: bool = False,
    auto_create_journal: bool = False,
    db: Session = Depends(get_db),
):
    """Upload TXT/PDF/ảnh hóa đơn để trích xuất số hóa đơn, ngày, nhà cung cấp, VAT, tổng tiền và gợi ý bút toán."""
    content = await file.read()
    read_result = read_text_from_upload(file.filename or "invoice", content)
    raw_text = read_result.get("text") or ""
    if not raw_text.strip():
        return {
            "message": "Chưa trích xuất được text từ file hóa đơn",
            "source": {"filename": file.filename, **{k: v for k, v in read_result.items() if k != "text"}},
            "extracted": None,
            "ai_suggestion": None,
            "created_purchase_invoice": None,
            "created_transaction": None,
            "next_step": "Upload file PDF có text-layer hoặc ảnh rõ nét sau khi cài Tesseract OCR. Bạn cũng có thể dùng /ocr/invoice/text để gửi text hóa đơn đã copy.",
        }
    parsed = parse_invoice_text(raw_text)
    return _build_ocr_response(
        parsed,
        db,
        create_purchase_invoice=create_purchase_invoice,
        create_transaction=create_transaction,
        auto_create_journal=auto_create_journal,
        source={"filename": file.filename, **{k: v for k, v in read_result.items() if k != "text"}},
    )


@app.get("/ocr/invoice/demo")
def ocr_invoice_demo(db: Session = Depends(get_db)):
    """Demo nhanh OCR hóa đơn bằng text mẫu."""
    sample = """
HÓA ĐƠN GIÁ TRỊ GIA TĂNG
Số hóa đơn: HD001234
Ngày 15/05/2026
Đơn vị bán hàng: Công ty Điện lực EVN Hà Nội
Mã số thuế: 0100100417
Đơn vị mua hàng: Công ty Finiip
Cộng tiền hàng: 2.000.000
Thuế suất GTGT: 10%
Tiền thuế GTGT: 200.000
Tổng cộng thanh toán: 2.200.000
"""
    parsed = parse_invoice_text(sample)
    return _build_ocr_response(parsed, db, source={"method": "demo_text"})


@app.post("/ai/classify-transaction")
def ai_classify_transaction(payload: AIAnalyzeRequest, db: Session = Depends(get_db)):
    result = classify_transaction(payload.description)
    log = create_ai_log(db, action="classify_transaction", description=payload.description, amount=payload.amount, result=result)
    return {"ai_result": result, "ai_log_id": log.id}


@app.post("/ai/suggest-journal-entry")
def ai_suggest_journal_entry(payload: AIAnalyzeRequest, db: Session = Depends(get_db)):
    result = suggest_journal_entry(payload.description, payload.amount)
    log = create_ai_log(db, action="suggest_journal_entry", description=payload.description, amount=payload.amount, result=result)
    return {"ai_result": result, "ai_log_id": log.id}


@app.post("/ai/analyze")
def ai_analyze(payload: AIAnalyzeRequest, db: Session = Depends(get_db)):
    result = suggest_journal_entry_with_learning(db, payload.description, payload.amount)
    log = create_ai_log(db, action="analyze_with_learning_default", description=payload.description, amount=payload.amount, result=result)
    return {"ai_result": result, "ai_log_id": log.id}


@app.post("/ai/v3/analyze")
def ai_v3_analyze(payload: AIAnalyzeRequest, db: Session = Depends(get_db)):
    """AI V3: phân tích + giải thích + confidence calibration + risk gate."""
    base_result = suggest_journal_entry_with_learning(db, payload.description, payload.amount)
    result = enhance_ai_result(payload.description, payload.amount, base_result)
    log = create_ai_log(db, action="ai_v3_analyze", description=payload.description, amount=payload.amount, result=result)
    return {
        "ai_result": result,
        "ai_log_id": log.id,
        "workflow_hint": result.get("quality_gate"),
        "message": "AI V3 đã trả về giải thích, rủi ro, confidence hiệu chỉnh và đề xuất quy trình duyệt.",
    }


@app.post("/ai/v3/batch-analyze")
def ai_v3_batch_analyze(payload: AIBatchAnalyzeRequest, db: Session = Depends(get_db)):
    items = []
    for item in payload.items:
        base_result = suggest_journal_entry_with_learning(db, item.description, item.amount)
        items.append({
            "description": item.description,
            "amount": item.amount,
            "ai_result": enhance_ai_result(item.description, item.amount, base_result),
        })
    return {"total": len(items), "items": items}


@app.get("/ai/v3/demo-cases")
def ai_v3_demo_cases():
    return {"items": AI_V3_DEMO_CASES}


@app.get("/ai/v3/test-suite")
def ai_v3_test_suite(db: Session = Depends(get_db)):
    def analyzer(description: str, amount: float) -> Dict[str, Any]:
        base_result = suggest_journal_entry_with_learning(db, description, amount)
        return enhance_ai_result(description, amount, base_result)
    return run_ai_v3_test_suite(analyzer)


@app.post("/ai/analyze-with-learning")
def ai_analyze_with_learning(payload: AIAnalyzeRequest, db: Session = Depends(get_db), min_similarity: float = Query(0.42, ge=0, le=1)):
    result = suggest_journal_entry_with_learning(db, payload.description, payload.amount, min_similarity=min_similarity)
    log = create_ai_log(db, action="analyze_with_learning", description=payload.description, amount=payload.amount, result=result)
    return {"ai_result": result, "ai_log_id": log.id}


@app.post("/ai/batch-analyze")
def ai_batch_analyze(payload: AIBatchAnalyzeRequest, db: Session = Depends(get_db)):
    """Phân tích nhiều giao dịch một lần, dùng để test AI trước khi làm frontend."""
    cases = [item.model_dump() for item in payload.items]
    report = benchmark_ai_cases(cases)
    log = create_ai_log(
        db,
        action="batch_analyze",
        description=f"Batch analyze {len(cases)} items",
        amount=None,
        result={"summary": {k: report[k] for k in ["total", "recognized", "unknown", "recognized_rate", "high_confidence_rate", "warnings_count"]}},
    )
    return {"benchmark": report, "ai_log_id": log.id}


@app.post("/ai/rules/custom")
def create_custom_ai_rule(payload: AICustomRuleCreate):
    """Thêm rule tạm thời trong RAM để test nhanh. Muốn lưu lâu dài thì thêm vào ai_engine.py."""
    return add_custom_rule(**payload.model_dump())


@app.get("/ai/test-suite")
def ai_test_suite():
    """Bộ test mẫu 20 giao dịch nên dùng trước khi làm frontend."""
    cases = [
        {"description":"Thanh toán tiền điện EVN tháng 5 bằng chuyển khoản", "amount":2500000},
        {"description":"Thanh toán tiền nước tháng 5", "amount":900000},
        {"description":"Trả lương nhân viên tháng 5 qua ngân hàng", "amount":35000000},
        {"description":"Mua máy tính văn phòng 22 triệu có VAT 10% chuyển khoản", "amount":22000000},
        {"description":"Chi phí quảng cáo Facebook Ads có VAT 10%", "amount":5500000},
        {"description":"Chi phí quảng cáo TikTok Ads", "amount":8000000},
        {"description":"Thu tiền bán hàng từ khách hàng A", "amount":12000000},
        {"description":"Nhận chuyển khoản doanh thu bán hàng có VAT 10%", "amount":11000000},
        {"description":"Mua văn phòng phẩm", "amount":1200000},
        {"description":"Trả tiền thuê văn phòng", "amount":15000000},
        {"description":"Phí ngân hàng Vietcombank", "amount":55000},
        {"description":"Rút tiền mặt từ ngân hàng", "amount":10000000},
        {"description":"Nộp tiền mặt vào tài khoản ngân hàng", "amount":15000000},
        {"description":"Mua hàng hóa nhập kho", "amount":40000000},
        {"description":"Thanh toán tiền internet FPT", "amount":700000},
        {"description":"Chi phí tiếp khách", "amount":3000000},
        {"description":"Tạm ứng cho nhân viên", "amount":3000000},
        {"description":"Hoàn ứng nhân viên", "amount":500000},
        {"description":"Trả tiền nhà cung cấp bằng chuyển khoản", "amount":45000000},
        {"description":"Nhận vốn góp của chủ sở hữu", "amount":200000000},
    ]
    return benchmark_ai_cases(cases)


@app.post("/ai/create-transaction")
def ai_create_transaction(payload: AICreateTransactionRequest, db: Session = Depends(get_db)):
    result = suggest_journal_entry_with_learning(db, payload.description, payload.amount)
    tx_date = payload.transaction_date or date.today()
    ensure_period_open(db, tx_date)
    transaction = Transaction(
        transaction_date=tx_date,
        description=payload.description,
        amount=payload.amount,
        type=result.get("transaction_type") or "unknown",
        category=result.get("category"),
        note=payload.note,
        debit_account_code=result.get("debit_account_code"),
        credit_account_code=result.get("credit_account_code"),
        ai_confidence=result.get("confidence"),
        status="draft",
        accounting_period=period_key(tx_date),
    )
    db.add(transaction)
    db.flush()
    entries: List[JournalEntry] = []
    if payload.auto_create_journal:
        entries = create_journal_from_ai_result(db, transaction_id=transaction.id, description=transaction.description, result=result, commit=False, status="draft", entry_date=tx_date)
    audit_log(db, action="ai_create_transaction", entity_type="transaction", entity_id=transaction.id, new=transaction_to_dict(transaction))
    db.commit()
    db.refresh(transaction)
    for e in entries:
        db.refresh(e)
    log = create_ai_log(db, action="create_transaction", description=payload.description, amount=payload.amount, result=result, transaction_id=transaction.id)
    return {"transaction": transaction_to_dict(transaction), "ai_result": result, "journal_entries": [journal_entry_to_dict(e) for e in entries], "ai_log_id": log.id}


@app.post("/ai/transactions/{transaction_id}/confirm-journal-entry")
def confirm_ai_journal_entry(transaction_id: int, db: Session = Depends(get_db)):
    transaction = db.get(Transaction, transaction_id)
    if not transaction:
        raise HTTPException(status_code=404, detail="Không tìm thấy giao dịch")
    result = suggest_journal_entry(transaction.description, transaction.amount)
    entries = create_journal_from_ai_result(db, transaction_id=transaction.id, description=transaction.description, result=result)
    log = create_ai_log(db, action="confirm_journal_entry", description=transaction.description, amount=transaction.amount, result=result, transaction_id=transaction.id)
    return {"message": "Đã tạo bút toán từ gợi ý AI", "journal_entries": [journal_entry_to_dict(e) for e in entries], "ai_log_id": log.id}


@app.post("/ai/transactions/{transaction_id}/correct", response_model=AICorrectionResponse)
def correct_ai_transaction(transaction_id: int, payload: AICorrectionCreate, db: Session = Depends(get_db)):
    transaction = db.get(Transaction, transaction_id)
    if not transaction:
        raise HTTPException(status_code=404, detail="Không tìm thấy giao dịch")
    ensure_account_exists(db, payload.user_debit_account_code)
    ensure_account_exists(db, payload.user_credit_account_code)
    old_value = transaction_to_dict(transaction)
    correction = AICorrection(
        transaction_id=transaction.id,
        original_description=transaction.description,
        original_amount=transaction.amount,
        ai_category=transaction.category,
        ai_type=transaction.type,
        ai_debit_account_code=transaction.debit_account_code,
        ai_credit_account_code=transaction.credit_account_code,
        ai_confidence=transaction.ai_confidence,
        user_category=payload.user_category,
        user_type=payload.user_type,
        user_debit_account_code=payload.user_debit_account_code,
        user_credit_account_code=payload.user_credit_account_code,
        note=payload.note,
    )
    transaction.category = payload.user_category
    transaction.type = payload.user_type
    transaction.debit_account_code = payload.user_debit_account_code
    transaction.credit_account_code = payload.user_credit_account_code
    db.add(correction)
    db.flush()
    audit_log(db, action="ai_correct", entity_type="transaction", entity_id=transaction.id, old=old_value, new=transaction_to_dict(transaction), note=payload.note)
    create_ai_log(db, action="correction_saved", description=transaction.description, amount=transaction.amount, result={
        "category": payload.user_category,
        "transaction_type": payload.user_type,
        "debit_account_code": payload.user_debit_account_code,
        "credit_account_code": payload.user_credit_account_code,
        "confidence": 1.0,
        "source": "user_correction",
    }, transaction_id=transaction.id)
    db.commit()
    db.refresh(correction)
    return correction


@app.get("/ai/logs", response_model=List[AILogResponse])
def list_ai_logs(db: Session = Depends(get_db), limit: int = Query(100, le=500)):
    return db.query(AILog).order_by(AILog.id.desc()).limit(limit).all()


@app.get("/ai/corrections", response_model=List[AICorrectionResponse])
def list_ai_corrections(db: Session = Depends(get_db), limit: int = Query(100, le=500)):
    return db.query(AICorrection).order_by(AICorrection.id.desc()).limit(limit).all()


@app.get("/ai/corrections/stats")
def ai_correction_stats(db: Session = Depends(get_db)):
    rows = db.query(AICorrection.user_category, func.count(AICorrection.id)).group_by(AICorrection.user_category).all()
    return {
        "total_corrections": db.query(AICorrection).count(),
        "by_user_category": [{"category": r[0], "count": r[1]} for r in rows],
        "message": "Dữ liệu này là nền tảng để Finiip học từ phản hồi người dùng ở cấp 3 nhẹ.",
    }



@app.get("/ai/learning-memory")
def ai_learning_memory(db: Session = Depends(get_db), limit: int = Query(100, le=500)):
    corrections = db.query(AICorrection).order_by(AICorrection.id.desc()).limit(limit).all()
    return {
        "total": db.query(AICorrection).count(),
        "items": [
            {
                "correction_id": c.id,
                "original_description": c.original_description,
                "tokens": sorted(text_tokens(c.original_description)),
                "user_category": c.user_category,
                "user_type": c.user_type,
                "user_debit_account_code": c.user_debit_account_code,
                "user_credit_account_code": c.user_credit_account_code,
                "created_at": c.created_at,
            }
            for c in corrections
        ],
    }


@app.get("/ai/learning-memory/search")
def search_learning_memory(description: str, db: Session = Depends(get_db), min_similarity: float = Query(0.0, ge=0, le=1), limit: int = Query(10, le=50)):
    corrections = db.query(AICorrection).order_by(AICorrection.id.desc()).limit(300).all()
    matches = []
    for c in corrections:
        score = jaccard_similarity(description, c.original_description)
        if score >= min_similarity:
            matches.append({
                "correction_id": c.id,
                "similarity": score,
                "original_description": c.original_description,
                "user_category": c.user_category,
                "user_type": c.user_type,
                "user_debit_account_code": c.user_debit_account_code,
                "user_credit_account_code": c.user_credit_account_code,
            })
    matches.sort(key=lambda x: x["similarity"], reverse=True)
    return {"description": description, "matches": matches[:limit]}


@app.get("/ai/accuracy")
def ai_accuracy_report(db: Session = Depends(get_db)):
    total_ai = db.query(AILog).filter(AILog.action.in_(["analyze", "analyze_with_learning", "analyze_with_learning_default", "create_transaction", "reanalyze_transaction"])).count()
    total_transactions = db.query(Transaction).count()
    corrections = db.query(AICorrection).count()
    confirmed = db.query(Transaction).filter(Transaction.status == "confirmed").count()
    low_confidence = db.query(Transaction).filter(Transaction.ai_confidence.isnot(None), Transaction.ai_confidence < 0.7).count()
    unknown = db.query(Transaction).filter(Transaction.type == "unknown").count()
    avg_confidence = db.query(func.avg(Transaction.ai_confidence)).filter(Transaction.ai_confidence.isnot(None)).scalar() or 0
    estimated_accuracy = None
    if total_transactions:
        estimated_accuracy = round(max(0, (total_transactions - corrections) / total_transactions * 100), 2)
    return {
        "total_ai_actions": total_ai,
        "total_transactions": total_transactions,
        "confirmed_transactions": confirmed,
        "corrections": corrections,
        "low_confidence_transactions": low_confidence,
        "unknown_transactions": unknown,
        "average_confidence": round(float(avg_confidence), 3),
        "estimated_accuracy_percent": estimated_accuracy,
        "note": "Độ chính xác là ước tính theo số correction; muốn chính xác hơn cần người dùng xác nhận đúng/sai rõ ràng.",
    }


@app.get("/ai/rule-suggestions")
def ai_rule_suggestions(db: Session = Depends(get_db), limit: int = Query(20, le=100)):
    corrections = db.query(AICorrection).order_by(AICorrection.id.desc()).limit(500).all()
    buckets: Dict[str, Dict[str, Any]] = {}
    for c in corrections:
        key = f"{c.user_category}|{c.user_type}|{c.user_debit_account_code}|{c.user_credit_account_code}"
        bucket = buckets.setdefault(key, {
            "category": c.user_category,
            "transaction_type": c.user_type,
            "debit_account_code": c.user_debit_account_code,
            "credit_account_code": c.user_credit_account_code,
            "count": 0,
            "keywords": {},
            "examples": [],
        })
        bucket["count"] += 1
        if len(bucket["examples"]) < 3:
            bucket["examples"].append(c.original_description)
        for token in text_tokens(c.original_description):
            bucket["keywords"][token] = bucket["keywords"].get(token, 0) + 1
    suggestions = []
    for bucket in buckets.values():
        keywords = sorted(bucket["keywords"].items(), key=lambda item: item[1], reverse=True)[:8]
        suggestions.append({
            "category": bucket["category"],
            "transaction_type": bucket["transaction_type"],
            "debit_account_code": bucket["debit_account_code"],
            "credit_account_code": bucket["credit_account_code"],
            "correction_count": bucket["count"],
            "suggested_keywords": [k for k, _ in keywords],
            "examples": bucket["examples"],
            "rule_template": {
                "keywords": [k for k, _ in keywords],
                "category": bucket["category"],
                "transaction_type": bucket["transaction_type"],
                "debit_account": bucket["debit_account_code"],
                "credit_account": bucket["credit_account_code"],
                "confidence": 0.86,
            },
        })
    suggestions.sort(key=lambda x: x["correction_count"], reverse=True)
    return {"total_suggestions": len(suggestions), "items": suggestions[:limit]}


@app.get("/ai/demo")
def get_ai_demo():
    return {"examples": demo_ai_cases()}


@app.get("/ai/knowledge")
def get_ai_knowledge():
    categories = sorted({rule.get("category") for rule in ACCOUNTING_RULES if rule.get("category")})
    return {"rule_count": len(ACCOUNTING_RULES), "categories": categories, "sample_rules": ACCOUNTING_RULES[:10]}


@app.get("/ai/dashboard")
def ai_dashboard(db: Session = Depends(get_db)):
    total_transactions = db.query(Transaction).count()
    ai_transactions = db.query(Transaction).filter(Transaction.ai_confidence.isnot(None)).count()
    low_confidence = db.query(Transaction).filter(Transaction.ai_confidence.isnot(None), Transaction.ai_confidence < 0.7).count()
    unknown = db.query(Transaction).filter(Transaction.type == "unknown").count()
    avg_confidence = db.query(func.avg(Transaction.ai_confidence)).filter(Transaction.ai_confidence.isnot(None)).scalar() or 0
    return {
        "total_transactions": total_transactions,
        "ai_transactions": ai_transactions,
        "ai_coverage_rate": round(ai_transactions / total_transactions * 100, 2) if total_transactions else 0,
        "average_confidence": round(float(avg_confidence), 3),
        "low_confidence_transactions": low_confidence,
        "unknown_transactions": unknown,
        "ai_logs": db.query(AILog).count(),
        "ai_corrections": db.query(AICorrection).count(),
        "stage": "Cấp 3 nhẹ: rule-based + learning from correction memory",
    }




# =========================
# V5 status, periods, audit, import
# =========================

@app.post("/transactions/{transaction_id}/confirm", response_model=TransactionResponse)
def confirm_transaction(transaction_id: int, db: Session = Depends(get_db)):
    transaction = db.query(Transaction).filter(Transaction.id == transaction_id).first()
    if not transaction:
        raise HTTPException(status_code=404, detail="Không tìm thấy giao dịch")
    ensure_period_open(db, transaction.transaction_date)
    if transaction.status == "cancelled":
        raise HTTPException(status_code=409, detail="Giao dịch đã hủy")
    balance = validate_transaction_journal_balance(db, transaction_id)
    if not balance["balanced"]:
        raise HTTPException(status_code=400, detail={"message": "Giao dịch chưa có bút toán hợp lệ để confirmed", "balance": balance})
    old = transaction_to_dict(transaction)
    transaction.status = "confirmed"
    transaction.confirmed_at = datetime.utcnow()
    transaction.accounting_period = period_key(transaction.transaction_date)
    db.query(JournalEntry).filter(JournalEntry.transaction_id == transaction_id, JournalEntry.status != "cancelled").update({"status": "posted", "accounting_period": transaction.accounting_period})
    audit_log(db, action="confirm", entity_type="transaction", entity_id=transaction.id, old=old, new=transaction_to_dict(transaction))
    db.commit()
    db.refresh(transaction)
    return transaction


@app.post("/transactions/{transaction_id}/cancel", response_model=TransactionResponse)
def cancel_transaction(transaction_id: int, note: Optional[str] = None, db: Session = Depends(get_db)):
    transaction = db.query(Transaction).filter(Transaction.id == transaction_id).first()
    if not transaction:
        raise HTTPException(status_code=404, detail="Không tìm thấy giao dịch")
    ensure_period_open(db, transaction.transaction_date)
    old = transaction_to_dict(transaction)
    transaction.status = "cancelled"
    transaction.cancelled_at = datetime.utcnow()
    db.query(JournalEntry).filter(JournalEntry.transaction_id == transaction_id).update({"status": "cancelled"})
    audit_log(db, action="cancel", entity_type="transaction", entity_id=transaction.id, old=old, new=transaction_to_dict(transaction), note=note)
    db.commit()
    db.refresh(transaction)
    return transaction


@app.get("/transactions/{transaction_id}/validation")
def validate_transaction(transaction_id: int, db: Session = Depends(get_db)):
    transaction = db.query(Transaction).filter(Transaction.id == transaction_id).first()
    if not transaction:
        raise HTTPException(status_code=404, detail="Không tìm thấy giao dịch")
    return {
        "transaction": transaction_to_dict(transaction),
        "journal_balance": validate_transaction_journal_balance(db, transaction_id),
        "period_locked": get_or_create_period(db, transaction.transaction_date).status == "closed",
    }


@app.get("/accounting-periods", response_model=List[AccountingPeriodResponse])
def list_accounting_periods(db: Session = Depends(get_db)):
    return db.query(AccountingPeriod).order_by(AccountingPeriod.period.desc()).all()


@app.post("/accounting-periods", response_model=AccountingPeriodResponse)
def create_accounting_period(payload: AccountingPeriodCreate, db: Session = Depends(get_db)):
    if db.query(AccountingPeriod).filter(AccountingPeriod.period == payload.period).first():
        raise HTTPException(status_code=400, detail="Kỳ kế toán đã tồn tại")
    start, end = month_bounds(payload.period)
    period = AccountingPeriod(period=payload.period, start_date=start, end_date=end, status="open", note=payload.note)
    db.add(period)
    db.flush()
    audit_log(db, action="create", entity_type="accounting_period", entity_id=period.id, new={"period": period.period, "status": period.status})
    db.commit()
    db.refresh(period)
    return period


@app.post("/accounting-periods/close", response_model=AccountingPeriodResponse)
def close_accounting_period(payload: ClosePeriodRequest, db: Session = Depends(get_db)):
    start, end = month_bounds(payload.period)
    period = db.query(AccountingPeriod).filter(AccountingPeriod.period == payload.period).first()
    if not period:
        period = AccountingPeriod(period=payload.period, start_date=start, end_date=end, status="open")
        db.add(period)
        db.flush()
    unconfirmed = db.query(Transaction).filter(Transaction.accounting_period == payload.period, Transaction.status == "draft").count()
    if unconfirmed:
        raise HTTPException(status_code=409, detail=f"Còn {unconfirmed} giao dịch draft trong kỳ {payload.period}. Hãy confirmed hoặc cancel trước khi khóa sổ")
    old = {"status": period.status}
    period.status = "closed"
    period.closed_at = datetime.utcnow()
    period.note = payload.note
    audit_log(db, action="close", entity_type="accounting_period", entity_id=period.id, old=old, new={"status": period.status}, note=payload.note)
    db.commit()
    db.refresh(period)
    return period


@app.post("/accounting-periods/reopen", response_model=AccountingPeriodResponse)
def reopen_accounting_period(payload: ClosePeriodRequest, db: Session = Depends(get_db)):
    period = db.query(AccountingPeriod).filter(AccountingPeriod.period == payload.period).first()
    if not period:
        raise HTTPException(status_code=404, detail="Không tìm thấy kỳ kế toán")
    old = {"status": period.status}
    period.status = "open"
    period.reopened_at = datetime.utcnow()
    period.note = payload.note
    audit_log(db, action="reopen", entity_type="accounting_period", entity_id=period.id, old=old, new={"status": period.status}, note=payload.note)
    db.commit()
    db.refresh(period)
    return period


@app.get("/audit-logs", response_model=List[AuditLogResponse])
def list_audit_logs(entity_type: Optional[str] = None, entity_id: Optional[int] = None, limit: int = Query(default=100, le=500), db: Session = Depends(get_db)):
    query = db.query(AuditLog)
    if entity_type:
        query = query.filter(AuditLog.entity_type == entity_type)
    if entity_id:
        query = query.filter(AuditLog.entity_id == entity_id)
    return query.order_by(AuditLog.id.desc()).limit(limit).all()


def _make_workbook(rows: List[List[Any]], headers: List[str], title: str):
    if Workbook is None:
        raise HTTPException(status_code=500, detail="openpyxl chưa được cài đặt")
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(row)
    for col_idx, _ in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22
    return wb


@app.get("/import/templates/transactions")
def download_transactions_import_template():
    rows = [[date.today().isoformat(), "Thanh toán tiền điện EVN tháng 5 bằng chuyển khoản", 2500000, "Ghi chú tùy chọn"]]
    wb = _make_workbook(rows, ["transaction_date", "description", "amount", "note"], "Import transactions")
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    return FileResponse(tmp.name, filename="finiip_transactions_import_template.xlsx", media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.post("/import/transactions/excel", response_model=ImportExcelResponse)
async def import_transactions_excel(file: UploadFile = File(...), preview: bool = False, auto_confirm: bool = False, db: Session = Depends(get_db)):
    if Workbook is None:
        raise HTTPException(status_code=500, detail="openpyxl chưa được cài đặt")
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Chỉ hỗ trợ file Excel .xlsx/.xlsm")
    from openpyxl import load_workbook
    content = await file.read()
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.write(content)
    tmp.close()
    wb = load_workbook(tmp.name, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    required = {"transaction_date", "description", "amount"}
    if not required.issubset(set(headers)):
        raise HTTPException(status_code=400, detail="File cần có các cột transaction_date, description, amount")
    idx = {h: headers.index(h) for h in headers if h}
    items = []
    imported = 0
    for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[idx["description"]]:
            continue
        raw_date = row[idx["transaction_date"]]
        if isinstance(raw_date, datetime):
            tx_date = raw_date.date()
        elif isinstance(raw_date, date):
            tx_date = raw_date
        else:
            tx_date = date.fromisoformat(str(raw_date)[:10])
        description = str(row[idx["description"]]).strip()
        amount = float(row[idx["amount"]])
        note = str(row[idx.get("note")]).strip() if "note" in idx and row[idx.get("note")] is not None else None
        ai = suggest_journal_entry(description, amount)
        item = {"row": row_no, "transaction_date": str(tx_date), "description": description, "amount": amount, "ai_result": ai}
        items.append(item)
        if preview:
            continue
        ensure_period_open(db, tx_date)
        tx = Transaction(transaction_date=tx_date, description=description, amount=amount, type=ai.get("transaction_type", "unknown"), category=ai.get("category"), note=note, debit_account_code=ai.get("debit_account_code"), credit_account_code=ai.get("credit_account_code"), ai_confidence=ai.get("confidence"), status="draft", accounting_period=period_key(tx_date))
        db.add(tx)
        db.flush()
        create_journal_from_ai_result(db, transaction_id=tx.id, description=description, result=ai, commit=False, status="draft", entry_date=tx_date)
        if auto_confirm:
            balance = validate_transaction_journal_balance(db, tx.id)
            if balance["balanced"]:
                tx.status = "confirmed"
                tx.confirmed_at = datetime.utcnow()
                db.query(JournalEntry).filter(JournalEntry.transaction_id == tx.id).update({"status": "posted"})
        audit_log(db, action="import_excel", entity_type="transaction", entity_id=tx.id, new=transaction_to_dict(tx), note=f"row={row_no}; file={file.filename}")
        imported += 1
    if not preview:
        db.commit()
    return {"imported": imported, "status": "preview" if preview else "imported", "items": items}


@app.get("/reports/summary")
def summary_report(db: Session = Depends(get_db)):
    income = db.query(func.sum(Transaction.amount)).filter(Transaction.status == "confirmed", Transaction.type == "income").scalar() or 0
    expense = db.query(func.sum(Transaction.amount)).filter(Transaction.status == "confirmed", Transaction.type == "expense").scalar() or 0
    return {
        "income": income,
        "expense": expense,
        "profit": income - expense,
        "transaction_count": db.query(Transaction).filter(Transaction.status == "confirmed").count(),
        "journal_entry_count": db.query(JournalEntry).filter(JournalEntry.status == "posted").count(),
    }


@app.get("/reports/categories")
def category_report(db: Session = Depends(get_db)):
    rows = (
        db.query(Transaction.category, Transaction.type, func.count(Transaction.id), func.sum(Transaction.amount))
        .filter(Transaction.status == "confirmed")
        .group_by(Transaction.category, Transaction.type)
        .all()
    )
    return [
        {"category": row[0] or "Chưa phân loại", "type": row[1], "count": row[2], "amount": row[3] or 0}
        for row in rows
    ]


@app.get("/reports/account-balances")
def account_balances(db: Session = Depends(get_db)):
    debit_rows = db.query(JournalEntry.debit_account_code, JournalEntry.debit_account_name, func.sum(JournalEntry.amount)).filter(JournalEntry.status == "posted").group_by(JournalEntry.debit_account_code, JournalEntry.debit_account_name).all()
    credit_rows = db.query(JournalEntry.credit_account_code, JournalEntry.credit_account_name, func.sum(JournalEntry.amount)).filter(JournalEntry.status == "posted").group_by(JournalEntry.credit_account_code, JournalEntry.credit_account_name).all()
    balances: Dict[str, Dict[str, Any]] = {}
    for code, name, amount in debit_rows:
        balances.setdefault(code, {"account_code": code, "account_name": name, "debit": 0, "credit": 0, "balance": 0})["debit"] += amount or 0
    for code, name, amount in credit_rows:
        balances.setdefault(code, {"account_code": code, "account_name": name, "debit": 0, "credit": 0, "balance": 0})["credit"] += amount or 0
    for item in balances.values():
        item["balance"] = item["debit"] - item["credit"]
    return sorted(balances.values(), key=lambda x: x["account_code"])


@app.get("/reports/overview")
def overview_report(db: Session = Depends(get_db)):
    summary = summary_report(db)
    return {
        "summary": summary,
        "ai_dashboard": ai_dashboard(db),
        "categories": category_report(db),
        "recent_transactions": list_transactions(db=db, limit=10),
    }


@app.post("/demo/seed-transactions")
def seed_demo_transactions(db: Session = Depends(get_db)):
    examples = [
        ("Thanh toán tiền điện EVN tháng 5 bằng chuyển khoản", 2500000),
        ("Thanh toán tiền nước văn phòng", 900000),
        ("Trả lương nhân viên qua ngân hàng tháng 5", 35000000),
        ("Mua máy tính văn phòng 22 triệu có VAT 10% chuyển khoản", 22000000),
        ("Doanh thu bán hàng cho khách chuyển khoản có VAT 10%", 11000000),
        ("Chi phí quảng cáo Facebook Ads tháng 5", 8000000),
        ("Chi phí quảng cáo TikTok Ads", 6500000),
        ("Phí ngân hàng Vietcombank", 55000),
        ("Rút tiền mặt từ ngân hàng", 10000000),
        ("Tạm ứng nhân viên đi công tác", 3000000),
    ]
    created = []
    for description, amount in examples:
        result = suggest_journal_entry(description, amount)
        transaction = Transaction(
            transaction_date=date.today(),
            description=description,
            amount=amount,
            type=result.get("transaction_type") or "unknown",
            category=result.get("category"),
            debit_account_code=result.get("debit_account_code"),
            credit_account_code=result.get("credit_account_code"),
            ai_confidence=result.get("confidence"),
            note="Dữ liệu demo",
        )
        db.add(transaction)
        db.commit()
        db.refresh(transaction)
        create_ai_log(db, action="seed_demo_transaction", description=description, amount=amount, result=result, transaction_id=transaction.id)
        created.append(transaction)
    return {"message": "Đã tạo dữ liệu demo", "created": len(created), "transactions": created}


# =========================
# V5 - Accounting reports + Excel export
# =========================


def _parse_date(value: Optional[date]) -> Optional[date]:
    return value


def _filter_transactions_by_date(query, start_date: Optional[date], end_date: Optional[date]):
    if start_date:
        query = query.filter(Transaction.transaction_date >= start_date)
    if end_date:
        query = query.filter(Transaction.transaction_date <= end_date)
    return query


def _filter_journal_entries_by_date(query, start_date: Optional[date], end_date: Optional[date]):
    if start_date:
        query = query.filter(JournalEntry.entry_date >= start_date)
    if end_date:
        query = query.filter(JournalEntry.entry_date <= end_date)
    return query


def _money(value: Optional[float]) -> float:
    return round(float(value or 0), 2)


def _sum_transactions(db: Session, tx_type: Optional[str] = None, category_like: Optional[str] = None, start_date: Optional[date] = None, end_date: Optional[date] = None) -> float:
    query = db.query(func.sum(Transaction.amount)).filter(Transaction.status == "confirmed")
    if tx_type:
        query = query.filter(Transaction.type == tx_type)
    if category_like:
        query = query.filter(Transaction.category.ilike(f"%{category_like}%"))
    query = _filter_transactions_by_date(query, start_date, end_date)
    return _money(query.scalar())


def _journal_rows(db: Session, start_date: Optional[date] = None, end_date: Optional[date] = None):
    query = db.query(JournalEntry).filter(JournalEntry.status == "posted")
    query = _filter_journal_entries_by_date(query, start_date, end_date)
    return query.order_by(JournalEntry.entry_date.asc(), JournalEntry.id.asc()).all()


def _build_workbook(title: str, headers: List[str], rows: List[List[Any]], summary: Optional[List[List[Any]]] = None):
    if Workbook is None:
        raise HTTPException(status_code=500, detail="Thiếu thư viện openpyxl. Hãy chạy: pip install openpyxl")
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(headers)))
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    row_index = 3
    if summary:
        for item in summary:
            ws.append(item)
            row_index += 1
        row_index += 1
        ws.append([])
    ws.append(headers)
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(row)
    for col_idx, _ in enumerate(headers, start=1):
        max_len = 12
        for cell in ws[get_column_letter(col_idx)]:
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)
    return wb


def _save_workbook_response(wb, filename: str):
    temp_dir = tempfile.mkdtemp(prefix="finiip_export_")
    path = os.path.join(temp_dir, filename)
    wb.save(path)
    return FileResponse(
        path,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/reports/revenue")
def revenue_report(db: Session = Depends(get_db), start_date: Optional[date] = None, end_date: Optional[date] = None):
    query = db.query(Transaction.category, func.count(Transaction.id), func.sum(Transaction.amount)).filter(Transaction.status == "confirmed", Transaction.type == "income")
    query = _filter_transactions_by_date(query, start_date, end_date)
    rows = query.group_by(Transaction.category).all()
    items = [{"category": r[0] or "Chưa phân loại", "count": r[1], "amount": _money(r[2])} for r in rows]
    return {"start_date": start_date, "end_date": end_date, "total_revenue": _money(sum(i["amount"] for i in items)), "items": items}


@app.get("/reports/expenses")
def expenses_report(db: Session = Depends(get_db), start_date: Optional[date] = None, end_date: Optional[date] = None):
    query = db.query(Transaction.category, func.count(Transaction.id), func.sum(Transaction.amount)).filter(Transaction.status == "confirmed", Transaction.type == "expense")
    query = _filter_transactions_by_date(query, start_date, end_date)
    rows = query.group_by(Transaction.category).all()
    items = [{"category": r[0] or "Chưa phân loại", "count": r[1], "amount": _money(r[2])} for r in rows]
    return {"start_date": start_date, "end_date": end_date, "total_expense": _money(sum(i["amount"] for i in items)), "items": items}


@app.get("/reports/profit-loss")
def profit_loss_report(db: Session = Depends(get_db), start_date: Optional[date] = None, end_date: Optional[date] = None):
    revenue = _sum_transactions(db, tx_type="income", start_date=start_date, end_date=end_date)
    expense = _sum_transactions(db, tx_type="expense", start_date=start_date, end_date=end_date)
    unknown = _sum_transactions(db, tx_type="unknown", start_date=start_date, end_date=end_date)
    gross_profit = revenue - expense
    margin = round((gross_profit / revenue * 100), 2) if revenue else 0
    return {
        "start_date": start_date,
        "end_date": end_date,
        "revenue": revenue,
        "expense": expense,
        "profit": _money(gross_profit),
        "profit_margin_percent": margin,
        "unknown_amount": unknown,
        "note": "Báo cáo dựa trên transaction.type. Sau này có thể nâng cấp sang chuẩn báo cáo theo hệ tài khoản.",
    }


@app.get("/reports/vat")
def vat_report(db: Session = Depends(get_db), start_date: Optional[date] = None, end_date: Optional[date] = None):
    sales_query = db.query(func.sum(SalesInvoice.vat_amount), func.sum(SalesInvoice.total_amount), func.count(SalesInvoice.id))
    purchase_query = db.query(func.sum(PurchaseInvoice.vat_amount), func.sum(PurchaseInvoice.total_amount), func.count(PurchaseInvoice.id))
    if start_date:
        sales_query = sales_query.filter(SalesInvoice.invoice_date >= start_date)
        purchase_query = purchase_query.filter(PurchaseInvoice.invoice_date >= start_date)
    if end_date:
        sales_query = sales_query.filter(SalesInvoice.invoice_date <= end_date)
        purchase_query = purchase_query.filter(PurchaseInvoice.invoice_date <= end_date)
    sales_vat, sales_total, sales_count = sales_query.one()
    purchase_vat, purchase_total, purchase_count = purchase_query.one()
    output_vat = _money(sales_vat)
    input_vat = _money(purchase_vat)
    return {
        "start_date": start_date,
        "end_date": end_date,
        "output_vat": output_vat,
        "input_vat": input_vat,
        "vat_payable": _money(output_vat - input_vat),
        "sales_invoice_count": sales_count or 0,
        "purchase_invoice_count": purchase_count or 0,
        "sales_total_amount": _money(sales_total),
        "purchase_total_amount": _money(purchase_total),
    }


@app.get("/reports/cashflow")
def cashflow_report(db: Session = Depends(get_db), start_date: Optional[date] = None, end_date: Optional[date] = None):
    rows = _journal_rows(db, start_date, end_date)
    cash_accounts = {"111", "112"}
    inflow = 0.0
    outflow = 0.0
    details = []
    for e in rows:
        debit_root = str(e.debit_account_code)[:3]
        credit_root = str(e.credit_account_code)[:3]
        if debit_root in cash_accounts and credit_root not in cash_accounts:
            inflow += e.amount
            details.append({"date": e.entry_date, "description": e.description, "direction": "inflow", "amount": e.amount, "account": e.debit_account_code})
        elif credit_root in cash_accounts and debit_root not in cash_accounts:
            outflow += e.amount
            details.append({"date": e.entry_date, "description": e.description, "direction": "outflow", "amount": e.amount, "account": e.credit_account_code})
    return {"start_date": start_date, "end_date": end_date, "cash_inflow": _money(inflow), "cash_outflow": _money(outflow), "net_cashflow": _money(inflow - outflow), "items": details}


@app.get("/reports/journal-entries")
def journal_entries_report(db: Session = Depends(get_db), start_date: Optional[date] = None, end_date: Optional[date] = None):
    rows = _journal_rows(db, start_date, end_date)
    total_debit = _money(sum(e.amount for e in rows))
    total_credit = _money(sum(e.amount for e in rows))
    return {
        "start_date": start_date,
        "end_date": end_date,
        "total_entries": len(rows),
        "total_debit": total_debit,
        "total_credit": total_credit,
        "balanced": total_debit == total_credit,
        "items": rows,
    }


@app.get("/reports/trial-balance")
def trial_balance_report(db: Session = Depends(get_db), start_date: Optional[date] = None, end_date: Optional[date] = None):
    rows = _journal_rows(db, start_date, end_date)
    balances: Dict[str, Dict[str, Any]] = {}
    for e in rows:
        balances.setdefault(e.debit_account_code, {"account_code": e.debit_account_code, "account_name": e.debit_account_name, "debit": 0.0, "credit": 0.0})["debit"] += e.amount
        balances.setdefault(e.credit_account_code, {"account_code": e.credit_account_code, "account_name": e.credit_account_name, "debit": 0.0, "credit": 0.0})["credit"] += e.amount
    items = []
    for item in balances.values():
        debit = _money(item["debit"])
        credit = _money(item["credit"])
        items.append({**item, "debit": debit, "credit": credit, "balance": _money(debit - credit)})
    items.sort(key=lambda x: x["account_code"])
    return {"start_date": start_date, "end_date": end_date, "total_debit": _money(sum(i["debit"] for i in items)), "total_credit": _money(sum(i["credit"] for i in items)), "balanced": _money(sum(i["debit"] for i in items)) == _money(sum(i["credit"] for i in items)), "items": items}


@app.get("/export/transactions/excel")
def export_transactions_excel(db: Session = Depends(get_db), start_date: Optional[date] = None, end_date: Optional[date] = None):
    query = db.query(Transaction).filter(Transaction.status == "confirmed")
    query = _filter_transactions_by_date(query, start_date, end_date)
    txs = query.order_by(Transaction.transaction_date.asc(), Transaction.id.asc()).all()
    rows = [[t.id, t.transaction_date.isoformat(), t.description, t.type, t.category, t.debit_account_code, t.credit_account_code, t.amount, t.ai_confidence, t.note] for t in txs]
    wb = _build_workbook("Finiip Transactions", ["ID", "Ngày", "Mô tả", "Loại", "Danh mục", "TK Nợ", "TK Có", "Số tiền", "AI confidence", "Ghi chú"], rows, [["Tổng giao dịch", len(rows)], ["Tổng tiền", _money(sum(t.amount for t in txs))]])
    return _save_workbook_response(wb, "finiip_transactions.xlsx")


@app.get("/export/journal-entries/excel")
def export_journal_entries_excel(db: Session = Depends(get_db), start_date: Optional[date] = None, end_date: Optional[date] = None):
    entries = _journal_rows(db, start_date, end_date)
    rows = [[e.id, e.transaction_id, e.entry_date.isoformat(), e.description, e.debit_account_code, e.debit_account_name, e.credit_account_code, e.credit_account_name, e.amount, e.line_no] for e in entries]
    wb = _build_workbook("Finiip Journal Entries", ["ID", "Transaction ID", "Ngày", "Diễn giải", "TK Nợ", "Tên TK Nợ", "TK Có", "Tên TK Có", "Số tiền", "Dòng"], rows, [["Tổng bút toán", len(rows)], ["Tổng phát sinh Nợ/Có", _money(sum(e.amount for e in entries))]])
    return _save_workbook_response(wb, "finiip_journal_entries.xlsx")


@app.get("/export/reports/profit-loss/excel")
def export_profit_loss_excel(db: Session = Depends(get_db), start_date: Optional[date] = None, end_date: Optional[date] = None):
    report = profit_loss_report(db, start_date, end_date)
    rows = [
        ["Doanh thu", report["revenue"]],
        ["Chi phí", report["expense"]],
        ["Lợi nhuận", report["profit"]],
        ["Biên lợi nhuận %", report["profit_margin_percent"]],
        ["Số tiền chưa phân loại", report["unknown_amount"]],
    ]
    wb = _build_workbook("Finiip Profit Loss", ["Chỉ tiêu", "Giá trị"], rows, [["Từ ngày", str(start_date or "")], ["Đến ngày", str(end_date or "")]])
    return _save_workbook_response(wb, "finiip_profit_loss.xlsx")


@app.get("/export/reports/vat/excel")
def export_vat_excel(db: Session = Depends(get_db), start_date: Optional[date] = None, end_date: Optional[date] = None):
    report = vat_report(db, start_date, end_date)
    rows = [
        ["VAT đầu ra", report["output_vat"]],
        ["VAT đầu vào", report["input_vat"]],
        ["VAT phải nộp", report["vat_payable"]],
        ["Số hóa đơn bán", report["sales_invoice_count"]],
        ["Số hóa đơn mua", report["purchase_invoice_count"]],
    ]
    wb = _build_workbook("Finiip VAT Report", ["Chỉ tiêu", "Giá trị"], rows, [["Từ ngày", str(start_date or "")], ["Đến ngày", str(end_date or "")]])
    return _save_workbook_response(wb, "finiip_vat_report.xlsx")


@app.post("/demo/seed-full")
def seed_full_demo(db: Session = Depends(get_db), reset: bool = False):
    """Tạo bộ dữ liệu demo đầy đủ: accounts, transactions, journal entries, invoices."""
    if reset:
        db.query(JournalEntry).delete()
        db.query(Transaction).delete()
        db.query(SalesInvoice).delete()
        db.query(PurchaseInvoice).delete()
        db.query(AILog).delete()
        db.query(AuditLog).delete()
        db.query(AccountingPeriod).delete()
        db.commit()
    setup_default_accounts(db)
    demo_transactions = [
        ("Thu tiền bán hàng từ khách hàng A có VAT 10%", 11000000),
        ("Nhận chuyển khoản doanh thu bán hàng tháng 5", 18000000),
        ("Thanh toán tiền điện EVN tháng 5 bằng chuyển khoản", 2500000),
        ("Thanh toán tiền nước văn phòng", 900000),
        ("Thanh toán tiền internet FPT", 700000),
        ("Trả lương nhân viên qua ngân hàng tháng 5", 35000000),
        ("Chi phí quảng cáo Facebook Ads có VAT 10%", 5500000),
        ("Chi phí quảng cáo TikTok Ads", 6500000),
        ("Mua máy tính văn phòng 22 triệu có VAT 10% chuyển khoản", 22000000),
        ("Mua văn phòng phẩm", 1200000),
        ("Trả tiền thuê văn phòng", 15000000),
        ("Phí ngân hàng Vietcombank", 55000),
        ("Mua hàng hóa nhập kho", 40000000),
        ("Tạm ứng cho nhân viên đi công tác", 3000000),
        ("Hoàn ứng nhân viên", 500000),
        ("Nhận vốn góp của chủ sở hữu", 200000000),
    ]
    created_transactions = []
    for description, amount in demo_transactions:
        result = suggest_journal_entry(description, amount)
        tx_date = date.today()
        transaction = Transaction(
            transaction_date=tx_date,
            description=description,
            amount=amount,
            type=result.get("transaction_type") or "unknown",
            category=result.get("category"),
            debit_account_code=result.get("debit_account_code"),
            credit_account_code=result.get("credit_account_code"),
            ai_confidence=result.get("confidence"),
            note="Dữ liệu demo V5",
            status="confirmed",
            confirmed_at=datetime.utcnow(),
            accounting_period=period_key(tx_date),
        )
        db.add(transaction)
        db.flush()
        create_journal_from_ai_result(db, transaction_id=transaction.id, description=description, result=result, commit=False, status="posted", entry_date=tx_date)
        audit_log(db, action="seed_full_demo", entity_type="transaction", entity_id=transaction.id, new=transaction_to_dict(transaction))
        db.commit()
        db.refresh(transaction)
        create_ai_log(db, action="seed_full_demo", description=description, amount=amount, result=result, transaction_id=transaction.id)
        created_transactions.append(transaction)
    sales_invoice = SalesInvoice(invoice_date=date.today(), invoice_number="HD-BAN-001", customer_name="Khách hàng A", description="Bán hàng demo", subtotal=10000000, vat_rate=10, vat_amount=1000000, total_amount=11000000, status="paid")
    purchase_invoice = PurchaseInvoice(invoice_date=date.today(), invoice_number="HD-MUA-001", supplier_name="Nhà cung cấp Demo", description="Mua hàng hóa nhập kho", subtotal=40000000, vat_rate=10, vat_amount=4000000, total_amount=44000000, status="paid")
    db.add_all([sales_invoice, purchase_invoice])
    db.commit()
    return {
        "message": "Đã tạo dữ liệu demo đầy đủ cho báo cáo và xuất Excel",
        "reset": reset,
        "created_transactions": len(created_transactions),
        "journal_entries": db.query(JournalEntry).count(),
        "sales_invoices": db.query(SalesInvoice).count(),
        "purchase_invoices": db.query(PurchaseInvoice).count(),
        "next_test_endpoints": [
            "/reports/profit-loss",
            "/reports/vat",
            "/reports/cashflow",
            "/reports/trial-balance",
            "/export/transactions/excel",
            "/export/journal-entries/excel",
        ],
    }


@app.get("/docs/backend-checklist")
def backend_checklist():
    return {
        "stage": "Backend V5 - chưa cần frontend",
        "done": [
            "AI phân tích giao dịch",
            "Tạo giao dịch từ AI",
            "Tạo bút toán từ AI",
            "Sổ nhật ký chung",
            "Báo cáo doanh thu",
            "Báo cáo chi phí",
            "Báo cáo lãi/lỗ",
            "Báo cáo VAT",
            "Báo cáo dòng tiền đơn giản",
            "Bảng cân đối phát sinh thử",
            "Xuất Excel giao dịch / bút toán / P&L / VAT",
            "Dữ liệu demo đầy đủ",
            "Validate bút toán Nợ/Có",
            "Trạng thái draft/confirmed/cancelled",
            "Kỳ kế toán và khóa sổ",
            "Audit log",
            "Import Excel",
        ],
        "next_backend_steps_before_frontend": [
            "Frontend sau khi backend ổn",
            "Đăng nhập JWT khi cần nhiều người dùng",
            "Phân quyền admin/accountant/viewer khi cần",
            "Đổi SQLite sang PostgreSQL khi deploy",
            "Thêm unit test tự động",
            "Chuẩn hóa folder routers/services nếu dự án lớn hơn",
        ],
    }


# =========================
# V7 - Knowledge Teaching helpers & endpoints
# =========================

def hybrid_similarity(a: Optional[str], b: Optional[str]) -> float:
    """So khớp tốt hơn Jaccard: kết hợp token overlap + thứ tự chữ."""
    norm_a = normalize_vi_text(a)
    norm_b = normalize_vi_text(b)
    if not norm_a or not norm_b:
        return 0.0
    jaccard = jaccard_similarity(norm_a, norm_b)
    seq = SequenceMatcher(None, norm_a, norm_b).ratio()
    left = text_tokens(norm_a)
    right = text_tokens(norm_b)
    containment = len(left & right) / max(1, min(len(left), len(right))) if left and right else 0
    return round((jaccard * 0.45) + (seq * 0.25) + (containment * 0.30), 4)


def find_best_correction_match(db: Session, description: str, min_similarity: float = 0.42) -> Tuple[Optional[AICorrection], float]:
    """Tìm memory gần nhất bằng hybrid similarity, ưu tiên correction mới."""
    corrections = db.query(AICorrection).order_by(AICorrection.id.desc()).limit(700).all()
    best = None
    best_score = 0.0
    for correction in corrections:
        score = hybrid_similarity(description, correction.original_description)
        if score > best_score:
            best = correction
            best_score = score
    if best and best_score >= min_similarity:
        return best, best_score
    return None, best_score


def _save_ai_teaching_example(db: Session, payload: AITeachExampleCreate) -> AICorrection:
    ensure_account_exists(db, payload.user_debit_account_code)
    ensure_account_exists(db, payload.user_credit_account_code)
    ai_result = suggest_journal_entry_with_learning(db, payload.description, payload.amount, min_similarity=0.99)
    correction = AICorrection(
        transaction_id=None,
        original_description=payload.description,
        original_amount=payload.amount,
        ai_category=ai_result.get("category"),
        ai_type=ai_result.get("transaction_type"),
        ai_debit_account_code=ai_result.get("debit_account_code"),
        ai_credit_account_code=ai_result.get("credit_account_code"),
        ai_confidence=ai_result.get("confidence"),
        user_category=payload.user_category,
        user_type=payload.user_type,
        user_debit_account_code=payload.user_debit_account_code,
        user_credit_account_code=payload.user_credit_account_code,
        note=payload.note or "Dạy AI trực tiếp qua /ai/teach",
    )
    db.add(correction)
    db.flush()
    create_ai_log(db, action="knowledge_taught", description=payload.description, amount=payload.amount, result={
        "source": "manual_teaching",
        "category": payload.user_category,
        "transaction_type": payload.user_type,
        "debit_account_code": payload.user_debit_account_code,
        "credit_account_code": payload.user_credit_account_code,
    })
    db.refresh(correction)
    return correction


@app.post("/ai/teach")
def teach_ai(payload: AITeachExampleCreate, db: Session = Depends(get_db)):
    """Dạy AI một ví dụ đúng mà không cần tạo giao dịch thật."""
    correction = _save_ai_teaching_example(db, payload)
    db.commit()
    return {
        "message": "Đã thêm kiến thức mới vào AI learning memory",
        "correction_id": correction.id,
        "next_test": "Gọi POST /ai/analyze-with-learning với mô tả tương tự để kiểm tra AI đã học chưa",
    }


@app.post("/ai/teach-batch")
def teach_ai_batch(payload: AITeachBatchRequest, db: Session = Depends(get_db)):
    """Dạy AI nhiều ví dụ cùng lúc, phù hợp seed knowledge kế toán."""
    created = []
    for item in payload.items:
        correction = _save_ai_teaching_example(db, item)
        created.append(correction.id)
    db.commit()
    return {
        "message": "Đã thêm bộ kiến thức mới vào AI",
        "created": len(created),
        "correction_ids": created,
    }




# ============================================================
# V9 / Level 3 - Trainable ML classifier endpoints
# ============================================================

@app.get("/ai/ml/dataset")
def ai_ml_dataset(db: Session = Depends(get_db)):
    """Xem dataset học máy được tạo từ correction và /ai/teach."""
    corrections = db.query(AICorrection).order_by(AICorrection.id.asc()).all()
    examples = build_training_examples_from_corrections(corrections)
    label_summary: Dict[str, int] = {}
    for item in examples:
        label = f"{item['category']} | {item['transaction_type']} | Nợ {item['debit_account_code']} / Có {item['credit_account_code']}"
        label_summary[label] = label_summary.get(label, 0) + 1
    return {
        "example_count": len(examples),
        "label_count": len(label_summary),
        "label_summary": label_summary,
        "items": examples[-100:],
        "next_step": "Gọi POST /ai/ml/train để train model phân loại giao dịch Cấp 3.",
    }


@app.get("/ai/ml/status")
def ai_ml_status(db: Session = Depends(get_db)):
    """Kiểm tra model ML đã train chưa và dataset hiện có bao nhiêu ví dụ."""
    examples = build_training_examples_from_corrections(db.query(AICorrection).all())
    status = model_status()
    status["current_dataset_examples"] = len(examples)
    status["current_stage"] = "Cấp 3 thật nếu model đã train và /ai/analyze dùng source=ml_model" if status.get("trained") else "Cấp 2.5/3 nhẹ - chưa train model"
    return status


@app.post("/ai/ml/train")
def ai_ml_train(payload: AIMLTrainRequest = AIMLTrainRequest(), db: Session = Depends(get_db)):
    """Train model phân loại giao dịch từ dữ liệu đã được người dùng sửa/dạy."""
    if not payload.include_corrections:
        raise HTTPException(status_code=400, detail="Hiện tại nguồn train chính là corrections/teach examples. Hãy bật include_corrections=true.")
    examples = build_training_examples_from_corrections(db.query(AICorrection).order_by(AICorrection.id.asc()).all())
    if len(examples) < payload.min_examples:
        raise HTTPException(
            status_code=400,
            detail=f"Chưa đủ dữ liệu train: có {len(examples)} ví dụ, cần ít nhất {payload.min_examples}. Hãy dùng /ai/teach hoặc /ai/transactions/{{id}}/correct.",
        )
    model = train_naive_bayes(examples)
    save_model(model)
    return {
        "message": "Đã train xong model ML phân loại giao dịch",
        "stage": "Cấp 3 - trainable transaction classifier",
        "example_count": model.get("example_count"),
        "label_count": model.get("label_count"),
        "trained_at": model.get("trained_at"),
        "next_test": "Gọi POST /ai/ml/predict hoặc POST /ai/analyze với mô tả mới.",
    }




@app.post("/ai/ml/seed-and-train")
def ai_ml_seed_and_train(db: Session = Depends(get_db)):
    """Nạp toàn bộ dữ liệu học mẫu đi kèm dự án và train model ML ngay lập tức."""
    data_dir = os.path.join(os.path.dirname(__file__), "data")
    dataset_paths = sorted(glob.glob(os.path.join(data_dir, "ai_training_examples*.json")))
    if not dataset_paths:
        raise HTTPException(status_code=500, detail="Không tìm thấy data/ai_training_examples*.json")

    # Tạo hệ thống tài khoản mặc định nếu database đang trống/chưa setup.
    created_accounts = 0
    skipped_accounts = 0
    for item in DEFAULT_ACCOUNTS:
        exists = db.query(Account).filter(Account.code == item["code"]).first()
        if exists:
            skipped_accounts += 1
            continue
        db.add(Account(**item))
        created_accounts += 1
    db.flush()

    items = []
    dataset_summary = []
    for dataset_path in dataset_paths:
        with open(dataset_path, "r", encoding="utf-8") as f:
            payload = json.load(f)
        file_items = payload.get("items", [])
        dataset_summary.append({"file": os.path.basename(dataset_path), "count": len(file_items)})
        items.extend(file_items)

    created_examples = 0
    skipped_examples = 0
    for item in items:
        description = str(item.get("description") or "").strip()
        category = str(item.get("user_category") or "").strip()
        user_type = str(item.get("user_type") or "expense").strip()
        debit = str(item.get("user_debit_account_code") or "").strip()
        credit = str(item.get("user_credit_account_code") or "").strip()
        amount = float(item.get("amount") or 1)
        if not all([description, category, user_type, debit, credit]):
            continue
        ensure_account_exists(db, debit)
        ensure_account_exists(db, credit)
        exists = (
            db.query(AICorrection)
            .filter(AICorrection.original_description == description)
            .filter(AICorrection.user_category == category)
            .filter(AICorrection.user_debit_account_code == debit)
            .filter(AICorrection.user_credit_account_code == credit)
            .first()
        )
        if exists:
            skipped_examples += 1
            continue
        db.add(AICorrection(
            transaction_id=None,
            original_description=description,
            original_amount=amount,
            ai_category=None,
            ai_type=None,
            ai_debit_account_code=None,
            ai_credit_account_code=None,
            ai_confidence=None,
            user_category=category,
            user_type=user_type,
            user_debit_account_code=debit,
            user_credit_account_code=credit,
            note=item.get("note") or "Seed training example",
        ))
        created_examples += 1

    db.commit()

    examples = build_training_examples_from_corrections(db.query(AICorrection).order_by(AICorrection.id.asc()).all())
    if not examples:
        raise HTTPException(status_code=400, detail="Không có dữ liệu để train model")
    model = train_naive_bayes(examples)
    save_model(model)
    return {
        "message": "Đã nạp dữ liệu học mẫu và train xong AI ML",
        "stage": "Cấp 3 - trainable transaction classifier",
        "accounts": {"created": created_accounts, "skipped": skipped_accounts},
        "datasets": dataset_summary,
        "training_examples": {"created": created_examples, "skipped": skipped_examples, "total_for_training": model.get("example_count")},
        "label_count": model.get("label_count"),
        "trained_at": model.get("trained_at"),
        "next_test": "Gọi POST /ai/ml/predict hoặc POST /ai/analyze. Nếu source=ml_model nghĩa là AI đang dùng model đã học.",
    }


@app.post("/ai/ml/predict")
def ai_ml_predict(payload: AIMLPredictRequest, db: Session = Depends(get_db)):
    """Dự đoán trực tiếp bằng model ML, không dùng rule fallback."""
    result = predict_with_model(payload.description, payload.amount)
    if not result:
        raise HTTPException(status_code=400, detail="Chưa có model ML. Hãy gọi /ai/ml/train trước.")
    _fill_account_names_for_ai_result(db, result)
    result["accepted_by_threshold"] = float(result.get("confidence") or 0) >= payload.min_confidence
    result["min_confidence"] = payload.min_confidence
    log = create_ai_log(db, action="ml_predict", description=payload.description, amount=payload.amount, result=result)
    return {"ai_result": result, "ai_log_id": log.id}



# ============================================================
# V11 - AI Feedback, Training Dataset Management & Evaluation
# ============================================================

def _correction_to_training_item(correction: AICorrection) -> Dict[str, Any]:
    return {
        "id": correction.id,
        "transaction_id": correction.transaction_id,
        "description": correction.original_description,
        "amount": correction.original_amount,
        "ai_category": correction.ai_category,
        "ai_type": correction.ai_type,
        "ai_debit_account_code": correction.ai_debit_account_code,
        "ai_credit_account_code": correction.ai_credit_account_code,
        "ai_confidence": correction.ai_confidence,
        "user_category": correction.user_category,
        "user_type": correction.user_type,
        "user_debit_account_code": correction.user_debit_account_code,
        "user_credit_account_code": correction.user_credit_account_code,
        "note": correction.note,
        "created_at": correction.created_at,
        "updated_at": correction.updated_at,
    }


def _train_model_from_current_dataset(db: Session, min_examples: int = 1) -> Dict[str, Any]:
    examples = build_training_examples_from_corrections(db.query(AICorrection).order_by(AICorrection.id.asc()).all())
    if len(examples) < min_examples:
        raise HTTPException(status_code=400, detail=f"Chưa đủ dữ liệu train: có {len(examples)} ví dụ, cần ít nhất {min_examples}.")
    model = train_naive_bayes(examples)
    save_model(model)
    return model


@app.get("/ai/training-examples")
def list_ai_training_examples(
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    category: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """Quản lý dataset AI đã học: xem các ví dụ correction/teach đang dùng để train."""
    query = db.query(AICorrection)
    if category:
        query = query.filter(AICorrection.user_category == category)
    total = query.count()
    rows = query.order_by(AICorrection.id.desc()).offset(offset).limit(limit).all()
    return {
        "total": total,
        "limit": limit,
        "offset": offset,
        "items": [_correction_to_training_item(row) for row in rows],
        "next_actions": [
            "POST /ai/training-examples để thêm ví dụ học mới",
            "PUT /ai/training-examples/{id} để sửa ví dụ sai",
            "DELETE /ai/training-examples/{id} để xóa dữ liệu rác",
            "POST /ai/ml/train để train lại model sau khi chỉnh dataset",
        ],
    }


@app.post("/ai/training-examples")
def create_ai_training_example(payload: AITeachExampleCreate, db: Session = Depends(get_db)):
    """Thêm một ví dụ học mới vào dataset AI, tương đương /ai/teach nhưng trả về dữ liệu chi tiết hơn."""
    correction = _save_ai_teaching_example(db, payload)
    db.commit()
    db.refresh(correction)
    return {
        "message": "Đã thêm ví dụ học mới vào dataset AI",
        "item": _correction_to_training_item(correction),
        "next_step": "Gọi POST /ai/ml/train để model học lại dữ liệu mới.",
    }


@app.put("/ai/training-examples/{example_id}")
def update_ai_training_example(example_id: int, payload: AITrainingExampleUpdate, db: Session = Depends(get_db)):
    """Sửa một ví dụ học nếu nhãn/category/tài khoản bị sai."""
    correction = db.get(AICorrection, example_id)
    if not correction:
        raise HTTPException(status_code=404, detail="Không tìm thấy training example")
    data = payload.model_dump(exclude_unset=True)
    if "description" in data and data["description"] is not None:
        correction.original_description = data["description"]
    if "amount" in data and data["amount"] is not None:
        correction.original_amount = data["amount"]
    if "user_category" in data and data["user_category"] is not None:
        correction.user_category = data["user_category"]
    if "user_type" in data and data["user_type"] is not None:
        correction.user_type = data["user_type"]
    if "user_debit_account_code" in data and data["user_debit_account_code"] is not None:
        ensure_account_exists(db, data["user_debit_account_code"])
        correction.user_debit_account_code = data["user_debit_account_code"]
    if "user_credit_account_code" in data and data["user_credit_account_code"] is not None:
        ensure_account_exists(db, data["user_credit_account_code"])
        correction.user_credit_account_code = data["user_credit_account_code"]
    if "note" in data:
        correction.note = data["note"]
    audit_log(db, action="ai_training_example_updated", entity_type="ai_correction", entity_id=correction.id, new=_correction_to_training_item(correction))
    db.commit()
    db.refresh(correction)
    return {
        "message": "Đã cập nhật training example",
        "item": _correction_to_training_item(correction),
        "next_step": "Gọi POST /ai/ml/train để model học lại sau khi sửa nhãn.",
    }


@app.delete("/ai/training-examples/{example_id}")
def delete_ai_training_example(example_id: int, db: Session = Depends(get_db)):
    """Xóa một ví dụ học rác khỏi dataset AI."""
    correction = db.get(AICorrection, example_id)
    if not correction:
        raise HTTPException(status_code=404, detail="Không tìm thấy training example")
    old = _correction_to_training_item(correction)
    db.delete(correction)
    audit_log(db, action="ai_training_example_deleted", entity_type="ai_correction", entity_id=example_id, old=old)
    db.commit()
    return {
        "message": "Đã xóa training example",
        "deleted_id": example_id,
        "next_step": "Gọi POST /ai/ml/train để model bỏ học dữ liệu đã xóa.",
    }


@app.post("/ai/feedback")
def save_ai_feedback(payload: AIFeedbackRequest, db: Session = Depends(get_db)):
    """Lưu phản hồi của người dùng khi AI đoán sai; đây là vòng lặp để AI càng dùng càng học."""
    ensure_account_exists(db, payload.correct_debit_account_code)
    ensure_account_exists(db, payload.correct_credit_account_code)
    correction = AICorrection(
        transaction_id=None,
        original_description=payload.description,
        original_amount=payload.amount,
        ai_category=payload.ai_category,
        ai_type=payload.ai_type,
        ai_debit_account_code=payload.ai_debit_account_code,
        ai_credit_account_code=payload.ai_credit_account_code,
        ai_confidence=payload.ai_confidence,
        user_category=payload.correct_category,
        user_type=payload.correct_type,
        user_debit_account_code=payload.correct_debit_account_code,
        user_credit_account_code=payload.correct_credit_account_code,
        note=payload.note or "Feedback từ người dùng qua /ai/feedback",
    )
    db.add(correction)
    db.flush()
    audit_log(db, action="ai_feedback_saved", entity_type="ai_correction", entity_id=correction.id, new=_correction_to_training_item(correction))
    create_ai_log(db, action="feedback_saved", description=payload.description, amount=payload.amount, result={
        "source": "user_feedback",
        "category": payload.correct_category,
        "transaction_type": payload.correct_type,
        "debit_account_code": payload.correct_debit_account_code,
        "credit_account_code": payload.correct_credit_account_code,
        "confidence": 1.0,
    })
    db.commit()
    db.refresh(correction)
    response: Dict[str, Any] = {
        "message": "Đã lưu feedback. AI sẽ dùng ví dụ này trong lần train tiếp theo.",
        "item": _correction_to_training_item(correction),
        "trained": False,
        "next_step": "Gọi POST /ai/ml/train hoặc gửi train_after=true để train ngay.",
    }
    if payload.train_after:
        model = _train_model_from_current_dataset(db, min_examples=1)
        response["trained"] = True
        response["model"] = {
            "example_count": model.get("example_count"),
            "label_count": model.get("label_count"),
            "trained_at": model.get("trained_at"),
        }
    return response


def _evaluate_examples(examples: List[Dict[str, Any]], test_ratio: float) -> Dict[str, Any]:
    total = len(examples)
    if total < 2:
        raise HTTPException(status_code=400, detail="Cần ít nhất 2 ví dụ để đánh giá model")
    test_size = max(1, int(round(total * test_ratio)))
    if test_size >= total:
        test_size = total - 1
    # Chia deterministic: lấy mỗi bước đều nhau làm test để tránh phụ thuộc random.
    step = max(2, total // test_size)
    test_indexes = set(range(step - 1, total, step))
    if len(test_indexes) < test_size:
        for idx in range(total - 1, -1, -1):
            test_indexes.add(idx)
            if len(test_indexes) >= test_size:
                break
    train_set = [item for idx, item in enumerate(examples) if idx not in test_indexes]
    test_set = [item for idx, item in enumerate(examples) if idx in test_indexes]
    model = train_naive_bayes(train_set)
    rows = []
    correct = 0
    category_stats: Dict[str, Dict[str, int]] = {}
    confidence_sum = 0.0
    for item in test_set:
        prediction = predict_with_model(item["description"], item.get("amount") or 0, model=model)
        expected = {
            "category": item["category"],
            "transaction_type": item["transaction_type"],
            "debit_account_code": item["debit_account_code"],
            "credit_account_code": item["credit_account_code"],
        }
        actual = {
            "category": prediction.get("category") if prediction else None,
            "transaction_type": prediction.get("transaction_type") if prediction else None,
            "debit_account_code": prediction.get("debit_account_code") if prediction else None,
            "credit_account_code": prediction.get("credit_account_code") if prediction else None,
        }
        is_correct = expected == actual
        correct += 1 if is_correct else 0
        conf = float((prediction or {}).get("confidence") or 0)
        confidence_sum += conf
        category = item["category"]
        category_stats.setdefault(category, {"total": 0, "correct": 0, "wrong": 0})
        category_stats[category]["total"] += 1
        category_stats[category]["correct"] += 1 if is_correct else 0
        category_stats[category]["wrong"] += 0 if is_correct else 1
        rows.append({
            "description": item["description"],
            "amount": item.get("amount"),
            "expected": expected,
            "predicted": actual,
            "confidence": round(conf, 4),
            "correct": is_correct,
        })
    accuracy = correct / len(test_set) if test_set else 0
    weak_categories = [
        {"category": cat, **stats, "accuracy": round(stats["correct"] / stats["total"], 4)}
        for cat, stats in category_stats.items()
        if stats["wrong"] > 0
    ]
    weak_categories.sort(key=lambda x: (x["accuracy"], -x["total"]))
    return {
        "stage": "V11 - AI Feedback & Evaluation",
        "total_examples": total,
        "train_samples": len(train_set),
        "test_samples": len(test_set),
        "correct": correct,
        "wrong": len(test_set) - correct,
        "accuracy": round(accuracy, 4),
        "accuracy_percent": round(accuracy * 100, 2),
        "average_confidence": round(confidence_sum / len(test_set), 4) if test_set else 0,
        "weak_categories": weak_categories[:10],
        "results": rows[:100],
        "recommendation": "Dạy thêm ví dụ cho weak_categories, xóa/sửa nhãn sai trong /ai/training-examples, rồi POST /ai/ml/train.",
    }


@app.get("/ai/ml/evaluate")
def ai_ml_evaluate_get(
    test_ratio: float = Query(0.2, gt=0, lt=0.8),
    min_examples: int = Query(10, ge=2),
    db: Session = Depends(get_db),
):
    """Đánh giá nhanh độ đúng của model bằng holdout set từ dataset hiện tại."""
    examples = build_training_examples_from_corrections(db.query(AICorrection).order_by(AICorrection.id.asc()).all())
    if len(examples) < min_examples:
        raise HTTPException(status_code=400, detail=f"Chưa đủ dữ liệu evaluate: có {len(examples)} ví dụ, cần ít nhất {min_examples}.")
    return _evaluate_examples(examples, test_ratio)


@app.post("/ai/ml/evaluate")
def ai_ml_evaluate_post(payload: AIMLEvaluateRequest = AIMLEvaluateRequest(), db: Session = Depends(get_db)):
    """Đánh giá model bằng body JSON, tiện cho frontend sau này."""
    examples = build_training_examples_from_corrections(db.query(AICorrection).order_by(AICorrection.id.asc()).all())
    if len(examples) < payload.min_examples:
        raise HTTPException(status_code=400, detail=f"Chưa đủ dữ liệu evaluate: có {len(examples)} ví dụ, cần ít nhất {payload.min_examples}.")
    return _evaluate_examples(examples, payload.test_ratio)


@app.get("/ai/knowledge-health")
def ai_knowledge_health(db: Session = Depends(get_db)):
    """Kiểm tra sức khỏe kiến thức AI để biết cần dạy thêm phần nào."""
    total_rules = len(ACCOUNTING_RULES)
    total_corrections = db.query(AICorrection).count()
    low_conf = db.query(Transaction).filter(Transaction.ai_confidence.isnot(None), Transaction.ai_confidence < 0.7).count()
    unknown = db.query(Transaction).filter(Transaction.type == "unknown").count()
    suggestions = ai_rule_suggestions(db, limit=10).get("items", [])
    return {
        "stage": "Cấp 3 nhẹ+ / AI học từ rule + correction memory + manual teaching",
        "rule_count": total_rules,
        "learning_examples": total_corrections,
        "low_confidence_transactions": low_conf,
        "unknown_transactions": unknown,
        "priority": [
            "Dạy thêm các giao dịch AI chưa nhận diện qua /ai/teach",
            "Sau 3-5 correction cùng loại, dùng /ai/rule-suggestions để chuyển thành rule cố định",
            "Chỉ confirm giao dịch khi đã kiểm tra bút toán Nợ/Có",
        ],
        "top_rule_suggestions": suggestions,
    }

# ============================================================
# FINIIP BACKEND - V8 AI KNOWLEDGE / INTELLIGENCE ENDPOINTS
# Các API mới không phá API cũ. Dùng để nâng AI trước khi làm frontend.
# ============================================================

@app.get("/ai/v8/knowledge-store")
def ai_v8_knowledge_store():
    from ai_engine import V8_KNOWLEDGE_DOMAINS, ACCOUNTING_RULES
    category_count: Dict[str, int] = {}
    for rule in ACCOUNTING_RULES:
        category = rule.get("category") or "Chưa phân loại"
        category_count[category] = category_count.get(category, 0) + 1
    return {
        "stage": "V8 Knowledge Store",
        "domains": V8_KNOWLEDGE_DOMAINS,
        "rule_count": len(ACCOUNTING_RULES),
        "category_count": category_count,
        "how_to_improve": [
            "Dùng POST /ai/teach để dạy ví dụ mới.",
            "Dùng POST /ai/v8/accuracy-test để đo độ đúng theo bộ test có expected answer.",
            "Dùng POST /ai/v8/analyze-deep để xem explanation, confidence gate, risk score và journal_entry.",
        ],
    }



@app.post("/ai/v17/autopilot-analyze")
def ai_v17_autopilot_analyze(payload: AIAnalyzeRequest, db: Session = Depends(get_db)):
    """V17: Self-made AI autopilot, không dùng OpenAI/Ollama/LLM ngoài.

    API này dùng kết quả AI hiện có, rồi thêm lớp quyết định:
    auto_approve / needs_review / reject_or_teach.
    """
    result = suggest_journal_entry_with_learning(db, payload.description, payload.amount)
    wrapped = autopilot_response(payload.description, payload.amount, result)
    log = create_ai_log(
        db,
        action="v17_autopilot_analyze",
        description=payload.description,
        amount=payload.amount,
        result={
            "ai_result": result,
            "autopilot": wrapped.get("autopilot"),
            "no_external_llm": True,
        },
    )
    wrapped["ai_log_id"] = log.id
    return wrapped


@app.get("/ai/v17/upgrade-status")
def ai_v17_upgrade_status():
    """Cho biết bản V17 đã nâng Finiip lên đâu trong roadmap AI kế toán."""
    return {
        "stage": "V17 - Self-made AI Accounting Autopilot",
        "no_openai": True,
        "no_ollama": True,
        "no_external_llm": True,
        "current_level": "Cấp 3+",
        "why": [
            "Có rule-based accounting engine",
            "Có ML classifier tự viết bằng Python/Naive Bayes",
            "Có feedback loop và training examples",
            "Có OCR invoice parser",
            "Có autopilot safety layer để quyết định auto_approve/needs_review/reject_or_teach",
        ],
        "not_yet": [
            "Chưa phải ChatGPT kế toán Cấp 6",
            "Chưa có Transformer/LLM tự train",
            "Chưa nên tự động ghi sổ mọi giao dịch nếu chưa qua review",
        ],
        "next_recommended_steps": [
            "Làm màn hình AI Review Queue cho frontend",
            "Import dữ liệu thật và dùng /ai/feedback để sửa kết quả sai",
            "Đo auto_approve_rate, review_rate, wrong_rate",
            "Nâng OCR từ parser text lên pipeline ảnh hóa đơn thực tế",
        ],
    }

@app.post("/ai/v8/analyze-deep")
def ai_v8_analyze_deep(payload: AIAnalyzeRequest, db: Session = Depends(get_db)):
    """Phân tích sâu: phân loại + explanation + confidence gate + journal_entry + risk score."""
    result = suggest_journal_entry_with_learning(db, payload.description, payload.amount)
    from ai_engine import v8_enrich_ai_result
    enriched = v8_enrich_ai_result(payload.description, payload.amount, result)
    log = create_ai_log(db, action="v8_analyze_deep", description=payload.description, amount=payload.amount, result=enriched)
    return {"ai_result": enriched, "ai_log_id": log.id}


@app.post("/ai/v8/journal-entry")
def ai_v8_journal_entry(payload: AIAnalyzeRequest, db: Session = Depends(get_db)):
    """Chỉ lấy bút toán AI gợi ý, phù hợp để test trước khi tạo bút toán thật."""
    result = suggest_journal_entry_with_learning(db, payload.description, payload.amount)
    from ai_engine import v8_enrich_ai_result
    enriched = v8_enrich_ai_result(payload.description, payload.amount, result)
    return {
        "description": payload.description,
        "amount": payload.amount,
        "category": enriched.get("category"),
        "confidence": enriched.get("confidence"),
        "needs_review": enriched.get("needs_review"),
        "journal_entry": enriched.get("journal_entry"),
        "explanation": enriched.get("explanation"),
    }


@app.post("/ai/v8/risk-check")
def ai_v8_risk_check(payload: AIAnalyzeRequest, db: Session = Depends(get_db)):
    """Kiểm tra rủi ro thuế/chứng từ cho một giao dịch."""
    result = suggest_journal_entry_with_learning(db, payload.description, payload.amount)
    from ai_engine import v8_tax_risk_check, v8_explain_result
    risk = v8_tax_risk_check(payload.description, payload.amount, result)
    return {
        "description": payload.description,
        "amount": payload.amount,
        "category": result.get("category"),
        "confidence": result.get("confidence"),
        "risk": risk,
        "explanation": v8_explain_result(payload.description, payload.amount, result),
    }


@app.post("/ai/v8/accuracy-test")
def ai_v8_accuracy_test(payload: Optional[Dict[str, Any]] = None):
    """Chạy test AI với expected answer. Nếu không truyền cases thì dùng bộ test mẫu V8."""
    from ai_engine import v8_run_expected_accuracy_test
    cases = None
    if payload and isinstance(payload, dict):
        cases = payload.get("cases")
    return v8_run_expected_accuracy_test(cases)


@app.get("/ai/v8/roadmap")
def ai_v8_roadmap():
    return {
        "current_stage": "Cấp 3+ tiến gần Cấp 5: phân loại + học ví dụ + explanation + risk + gợi ý bút toán",
        "done": [
            "Rule-based AI knowledge pack",
            "Learning memory từ correction và /ai/teach",
            "Hybrid similarity cho học ví dụ",
            "Explanation cho kết quả AI",
            "Confidence gate / needs_review",
            "Journal entry chuẩn hóa",
            "Tax risk score",
            "Expected accuracy test",
        ],
        "next_backend_targets": [
            "Lưu knowledge store vào database thay vì chỉ code/RAM",
            "Thêm RAG/vector search cho tài liệu kế toán và dữ liệu công ty",
            "Thêm chatbot đọc báo cáo từ database",
            "Thêm OCR hóa đơn",
            "Thêm workflow duyệt bút toán trước khi post",
        ],
    }


# =========================
# V13/V14 Accounting Formula Engine endpoints
# =========================

def _formula_result(name: str, result: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "engine": "Finiip V14 Advanced Accounting Engine",
        "formula": name,
        "result": result,
    }


@app.get("/formulas/catalog")
def get_formula_catalog():
    """Danh sách công thức kế toán mà AI/backend hiện biết tính."""
    return {
        "engine": "Finiip V14 Advanced Accounting Engine",
        "count": len(formula_catalog()),
        "items": formula_catalog(),
        "note": "Đây là bộ công thức lõi + nâng cao: VAT, khấu hao, tồn kho, giá vốn, lương, công nợ, kết chuyển và BCTC cơ bản.",
    }


@app.post("/formulas/vat")
def formula_vat(payload: VATFormulaRequest):
    try:
        return _formula_result("vat", calculate_vat(**payload.model_dump()))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.post("/formulas/depreciation")
def formula_depreciation(payload: DepreciationFormulaRequest):
    try:
        return _formula_result("straight_line_depreciation", calculate_straight_line_depreciation(**payload.model_dump()))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.post("/formulas/prepaid-allocation")
def formula_prepaid_allocation(payload: PrepaidAllocationRequest):
    try:
        return _formula_result("prepaid_allocation", calculate_prepaid_allocation(**payload.model_dump()))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.post("/formulas/profit/gross")
def formula_profit_gross(payload: GrossProfitRequest):
    return _formula_result("gross_profit", calculate_gross_profit(**payload.model_dump()))


@app.post("/formulas/profit/net")
def formula_profit_net(payload: NetProfitRequest):
    return _formula_result("net_profit", calculate_net_profit(**payload.model_dump()))


@app.post("/formulas/tax/cit")
def formula_tax_cit(payload: CorporateIncomeTaxRequest):
    return _formula_result("corporate_income_tax", calculate_corporate_income_tax(**payload.model_dump()))


@app.post("/formulas/journal/check-balance")
def formula_journal_check_balance(payload: JournalCheckRequest):
    lines = [line.model_dump() for line in payload.lines]
    return _formula_result("journal_balance", check_journal_balance(lines))


@app.post("/formulas/ratios")
def formula_ratios(payload: FinancialRatiosRequest):
    return _formula_result("financial_ratios", calculate_financial_ratios(**payload.model_dump()))


@app.post("/formulas/break-even")
def formula_break_even(payload: BreakEvenRequest):
    try:
        return _formula_result("break_even", calculate_break_even(**payload.model_dump()))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.get("/formulas/ledger/profit-loss")
def formula_ledger_profit_loss(period: Optional[str] = None, db: Session = Depends(get_db)):
    """Tính lợi nhuận nhanh từ bảng transactions, dùng dữ liệu đã có trong backend."""
    query = db.query(Transaction)
    if period:
        query = query.filter(Transaction.accounting_period == period)
    transactions = query.filter(Transaction.status != "cancelled").all()
    revenue = sum(float(t.amount or 0) for t in transactions if t.type == "income")
    expenses = sum(float(t.amount or 0) for t in transactions if t.type == "expense")
    result = calculate_net_profit(revenue=revenue, cogs=0, operating_expenses=expenses)
    return {
        "engine": "Finiip V14 Advanced Accounting Engine",
        "formula": "ledger_profit_loss",
        "period": period,
        "transaction_count": len(transactions),
        "result": result,
    }


@app.get("/formulas/ledger/trial-balance-check")
def formula_ledger_trial_balance_check(period: Optional[str] = None, db: Session = Depends(get_db)):
    """Kiểm tra cân bằng Nợ/Có từ bảng journal_entries."""
    query = db.query(JournalEntry).filter(JournalEntry.status != "cancelled")
    if period:
        query = query.filter(JournalEntry.accounting_period == period)
    entries = query.all()
    lines: List[Dict[str, Any]] = []
    for entry in entries:
        lines.append({"side": "debit", "account_code": entry.debit_account_code, "account_name": entry.debit_account_name, "amount": entry.amount})
        lines.append({"side": "credit", "account_code": entry.credit_account_code, "account_name": entry.credit_account_name, "amount": entry.amount})
    result = check_journal_balance(lines)
    return {
        "engine": "Finiip V14 Advanced Accounting Engine",
        "formula": "ledger_trial_balance_check",
        "period": period,
        "journal_entry_count": len(entries),
        "result": result,
    }


# =========================
# V14 Advanced Accounting Engine endpoints
# =========================

@app.post("/formulas/inventory/fifo")
def formula_inventory_fifo(payload: FIFOInventoryRequest):
    try:
        data = payload.model_dump()
        return _formula_result("inventory_fifo", calculate_fifo_inventory(**data))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.post("/formulas/inventory/weighted-average")
def formula_inventory_weighted_average(payload: WeightedAverageInventoryRequest):
    try:
        data = payload.model_dump()
        return _formula_result("inventory_weighted_average", calculate_weighted_average_inventory(**data))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.post("/formulas/payroll/basic")
def formula_payroll_basic(payload: PayrollBasicRequest):
    try:
        return _formula_result("payroll_basic", calculate_payroll_basic(**payload.model_dump()))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.post("/formulas/accounts/aging")
def formula_accounts_aging(payload: AccountsAgingRequest):
    return _formula_result("accounts_aging", calculate_accounts_aging(**payload.model_dump()))


@app.post("/formulas/closing/period")
def formula_closing_period(payload: PeriodClosingRequest):
    return _formula_result("period_closing", generate_period_closing_entries(**payload.model_dump()))


@app.post("/formulas/statements/basic")
def formula_statements_basic(payload: BasicFinancialStatementsRequest):
    return _formula_result("financial_statements_basic", build_basic_financial_statements(**payload.model_dump()))


# =========================
# V15 Backend API-first Pack
# =========================

API_VERSION = "16.0.0"
API_STAGE = "V16 - Large AI Training Pack"


def _route_summary() -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for route in app.routes:
        methods = sorted(getattr(route, "methods", []) or [])
        path = getattr(route, "path", "")
        if not path or path.startswith("/openapi"):
            continue
        rows.append({
            "path": path,
            "methods": [m for m in methods if m not in {"HEAD", "OPTIONS"}],
            "name": getattr(route, "name", None),
        })
    rows.sort(key=lambda r: (r["path"], ",".join(r["methods"])))
    return rows


def _count_table(db: Session, model: Any) -> int:
    return int(db.query(model).count())


@app.get("/api/v1/health")
def api_v1_health(db: Session = Depends(get_db)):
    """Health-check ổn định cho frontend/deploy monitor."""
    try:
        db.execute(text("SELECT 1"))
        database_status = "ok"
    except Exception as exc:  # pragma: no cover
        database_status = f"error: {exc}"
    return {
        "ok": database_status == "ok",
        "service": "Finiip AI Accounting Backend",
        "version": API_VERSION,
        "stage": API_STAGE,
        "database": database_status,
        "time": datetime.utcnow().isoformat() + "Z",
        "docs": {"swagger": "/docs", "redoc": "/redoc", "openapi": "/openapi.json"},
    }


@app.get("/api/v1/meta")
def api_v1_meta():
    """Thông tin contract cho frontend có sẵn."""
    return {
        "service": "Finiip AI Accounting API",
        "version": API_VERSION,
        "stage": API_STAGE,
        "api_prefix": "/api/v1",
        "auth": {
            "mode": "optional_api_key",
            "header": "X-API-Key",
            "enabled_when_env_FINIIP_API_KEY_is_set": bool(os.getenv("FINIIP_API_KEY")),
        },
        "cors": "enabled_for_dev_all_origins",
        "recommended_frontend_flow": [
            "GET /api/v1/health",
            "GET /api/v1/frontend/bootstrap",
            "POST /api/v1/ai/transaction-preview",
            "POST /ai/create-transaction",
            "POST /api/v1/ocr/invoice-preview",
            "POST /api/v1/frontend/train-ai",
        ],
        "openapi": "/openapi.json",
    }


@app.get("/api/v1/routes")
def api_v1_routes(group: Optional[str] = None):
    """Danh sách route để frontend map API nhanh hơn."""
    routes = _route_summary()
    if group:
        routes = [r for r in routes if r["path"].startswith(group)]
    return {"total": len(routes), "routes": routes}


@app.get("/api/v1/frontend/bootstrap")
def api_v1_frontend_bootstrap(db: Session = Depends(get_db)):
    """Một endpoint gom dữ liệu nền cho frontend: tài khoản, kỳ, AI, dashboard."""
    accounts = db.query(Account).order_by(Account.code.asc()).all()
    periods = db.query(AccountingPeriod).order_by(AccountingPeriod.period.desc()).limit(24).all()
    transactions_count = _count_table(db, Transaction)
    journal_count = _count_table(db, JournalEntry)
    correction_count = _count_table(db, AICorrection)
    ml_status = model_status()
    latest_transactions = db.query(Transaction).order_by(Transaction.transaction_date.desc(), Transaction.id.desc()).limit(10).all()
    return {
        "stage": API_STAGE,
        "version": API_VERSION,
        "counts": {
            "accounts": len(accounts),
            "periods": len(periods),
            "transactions": transactions_count,
            "journal_entries": journal_count,
            "ai_training_examples": correction_count,
        },
        "accounts": [account_to_dict(a) for a in accounts],
        "periods": [{
            "id": p.id,
            "period": p.period,
            "start_date": p.start_date.isoformat(),
            "end_date": p.end_date.isoformat(),
            "status": p.status,
        } for p in periods],
        "ai": {
            "ml_model": ml_status,
            "has_model": bool(ml_status.get("exists") or ml_status.get("is_trained")),
            "training_examples": correction_count,
        },
        "formula_catalog": formula_catalog(),
        "latest_transactions": [transaction_to_dict(t) for t in latest_transactions],
        "frontend_notes": {
            "money_unit": "VND",
            "date_format": "YYYY-MM-DD",
            "transaction_statuses": sorted(VALID_TRANSACTION_STATUSES),
            "journal_statuses": sorted(VALID_JOURNAL_STATUSES),
        },
    }


@app.post("/api/v1/ai/transaction-preview")
def api_v1_transaction_preview(payload: FrontendTransactionPreviewRequest, db: Session = Depends(get_db)):
    """Preview đầy đủ cho form nhập giao dịch: AI classify, bút toán, kiểm tra Nợ/Có, gợi ý frontend action."""
    ai_result = suggest_journal_entry_with_learning(db, payload.description, payload.amount)
    _fill_account_names_for_ai_result(db, ai_result)
    amount = float(payload.amount)
    lines = [
        {
            "side": "debit",
            "account_code": ai_result.get("debit_account_code"),
            "account_name": ai_result.get("debit_account_name"),
            "amount": amount,
        },
        {
            "side": "credit",
            "account_code": ai_result.get("credit_account_code"),
            "account_name": ai_result.get("credit_account_name"),
            "amount": amount,
        },
    ]
    balance = check_journal_balance(lines)
    confidence = float(ai_result.get("confidence") or 0)
    log = create_ai_log(db, action="api_v1_transaction_preview", description=payload.description, amount=payload.amount, result=ai_result)
    return {
        "stage": API_STAGE,
        "preview": {
            "description": payload.description,
            "amount": amount,
            "transaction_date": (payload.transaction_date or date.today()).isoformat(),
            "period": period_key(payload.transaction_date or date.today()),
        },
        "ai_result": ai_result,
        "journal_lines": lines,
        "journal_balance": balance,
        "frontend_decision": {
            "can_auto_create": (balance.get("is_balanced") or balance.get("balanced")) and confidence >= payload.min_confidence,
            "needs_review": confidence < payload.min_confidence or ai_result.get("category") == "Chưa phân loại",
            "min_confidence": payload.min_confidence,
            "next_api_to_create": "/ai/create-transaction",
            "next_api_to_correct": "/ai/feedback",
        },
        "ai_log_id": log.id,
    }


@app.post("/api/v1/ocr/invoice-preview")
def api_v1_invoice_text_preview(payload: FrontendInvoiceTextPreviewRequest, db: Session = Depends(get_db)):
    """Preview đọc hóa đơn cho frontend. Mặc định không ghi DB, chỉ trả data để người dùng xác nhận."""
    parsed = parse_invoice_text(payload.raw_text)
    result = _build_ocr_response(
        parsed,
        db,
        create_purchase_invoice=payload.create_drafts,
        create_transaction=payload.create_drafts,
        auto_create_journal=False,
        source={"method": "api_v1_raw_text_preview"},
    )
    result["stage"] = API_STAGE
    result["frontend_decision"] = {
        "can_show_confirmation_screen": bool(result.get("extracted")),
        "create_drafts_was_enabled": payload.create_drafts,
        "next_api_to_persist": "/ocr/invoice/text",
        "next_api_to_correct_ai": "/ai/feedback",
    }
    return result


@app.post("/api/v1/bulk/transactions/reanalyze")
def api_v1_bulk_reanalyze_transactions(payload: BulkReanalyzeTransactionsRequest, db: Session = Depends(get_db)):
    """Re-analyze nhiều giao dịch cho màn hình admin/frontend. Có thể chỉ preview hoặc update vào DB."""
    if len(payload.transaction_ids) > 200:
        raise HTTPException(status_code=400, detail="Tối đa 200 giao dịch mỗi lần")
    rows = []
    updated = 0
    for transaction_id in payload.transaction_ids:
        transaction = db.get(Transaction, transaction_id)
        if not transaction:
            rows.append({"transaction_id": transaction_id, "ok": False, "error": "not_found"})
            continue
        result = suggest_journal_entry_with_learning(db, transaction.description, transaction.amount)
        _fill_account_names_for_ai_result(db, result)
        confidence = float(result.get("confidence") or 0)
        if payload.update_transactions and confidence >= payload.min_confidence:
            ensure_period_open(db, transaction.transaction_date)
            transaction.category = result.get("category")
            transaction.type = result.get("transaction_type") or transaction.type
            transaction.debit_account_code = result.get("debit_account_code")
            transaction.credit_account_code = result.get("credit_account_code")
            transaction.ai_confidence = confidence
            updated += 1
        rows.append({
            "transaction_id": transaction_id,
            "ok": True,
            "updated": payload.update_transactions and confidence >= payload.min_confidence,
            "confidence": confidence,
            "ai_result": result,
        })
    if payload.update_transactions:
        db.commit()
    return {
        "stage": API_STAGE,
        "total": len(payload.transaction_ids),
        "updated": updated,
        "min_confidence": payload.min_confidence,
        "items": rows,
    }


@app.post("/api/v1/frontend/train-ai")
def api_v1_frontend_train_ai(payload: AIMLTrainRequest = AIMLTrainRequest(), db: Session = Depends(get_db)):
    """Frontend-friendly wrapper để train AI và trả trạng thái mới."""
    model = _train_model_from_current_dataset(db, min_examples=payload.min_examples)
    return {
        "stage": API_STAGE,
        "message": "Đã train lại model AI",
        "model": {
            "example_count": model.get("example_count"),
            "label_count": model.get("label_count"),
            "trained_at": model.get("trained_at"),
        },
        "status": model_status(),
        "next_step": "Gọi POST /api/v1/ai/transaction-preview hoặc POST /ai/analyze để dùng model mới.",
    }


@app.get("/api/v1/openapi-summary")
def api_v1_openapi_summary():
    """Tóm tắt contract cho frontend mà không cần parse toàn bộ OpenAPI."""
    groups: Dict[str, int] = {}
    for r in _route_summary():
        prefix = "/" + r["path"].strip("/").split("/")[0] if r["path"].strip("/") else "/"
        groups[prefix] = groups.get(prefix, 0) + 1
    return {
        "stage": API_STAGE,
        "version": API_VERSION,
        "total_routes": len(_route_summary()),
        "groups": groups,
        "important_for_frontend": [
            "/api/v1/health",
            "/api/v1/frontend/bootstrap",
            "/api/v1/ai/transaction-preview",
            "/api/v1/ocr/invoice-preview",
            "/transactions",
            "/journal-entries",
            "/reports/overview",
            "/ai/feedback",
            "/api/v1/frontend/train-ai",
        ],
    }


# ============================================================
# FINIIP V18-V22 - Feedback learning -> review queue -> retrain
# -> better OCR flow -> double-entry autoposting
# ============================================================

def _ai_result_payload(result: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "category": result.get("category"),
        "transaction_type": result.get("transaction_type"),
        "debit_account_code": result.get("debit_account_code") or result.get("debit_account"),
        "credit_account_code": result.get("credit_account_code") or result.get("credit_account"),
        "confidence": result.get("confidence"),
        "source": result.get("source"),
        "journal_lines": result.get("journal_lines") or [],
        "warnings": result.get("warnings") or [],
    }


def _priority_from_ai_result(result: Dict[str, Any]) -> str:
    confidence = float(result.get("confidence") or 0)
    if confidence < 0.55 or result.get("category") == "Chưa phân loại":
        return "high"
    if confidence < 0.78:
        return "medium"
    return "low"


def _review_reason(result: Dict[str, Any]) -> str:
    confidence = float(result.get("confidence") or 0)
    reasons = []
    if confidence < 0.55:
        reasons.append("confidence thấp")
    elif confidence < 0.78:
        reasons.append("confidence trung bình")
    if result.get("category") == "Chưa phân loại":
        reasons.append("AI chưa phân loại được")
    if result.get("source") == "rule_based":
        reasons.append("kết quả từ rule, nên kiểm tra trước khi ghi sổ")
    if not reasons:
        reasons.append("đưa vào queue để kế toán xác nhận")
    return ", ".join(reasons)


def _create_review_item(
    db: Session,
    *,
    description: str,
    amount: float,
    result: Dict[str, Any],
    transaction_id: Optional[int] = None,
    source: str = "ai",
    reason: Optional[str] = None,
) -> AIReviewItem:
    item = AIReviewItem(
        transaction_id=transaction_id,
        source=source,
        description=description,
        amount=amount,
        ai_category=result.get("category"),
        ai_type=result.get("transaction_type"),
        ai_debit_account_code=result.get("debit_account_code") or result.get("debit_account"),
        ai_credit_account_code=result.get("credit_account_code") or result.get("credit_account"),
        ai_confidence=result.get("confidence"),
        ai_result_json=json.dumps(result, ensure_ascii=False, default=str),
        status="pending",
        priority=_priority_from_ai_result(result),
        reason=reason or _review_reason(result),
    )
    db.add(item)
    db.flush()
    return item


def _review_item_to_dict(item: AIReviewItem) -> Dict[str, Any]:
    ai_result = {}
    if item.ai_result_json:
        try:
            ai_result = json.loads(item.ai_result_json)
        except Exception:
            ai_result = {}
    return {
        "id": item.id,
        "transaction_id": item.transaction_id,
        "source": item.source,
        "description": item.description,
        "amount": item.amount,
        "ai_category": item.ai_category,
        "ai_type": item.ai_type,
        "ai_debit_account_code": item.ai_debit_account_code,
        "ai_credit_account_code": item.ai_credit_account_code,
        "ai_confidence": item.ai_confidence,
        "priority": item.priority,
        "reason": item.reason,
        "status": item.status,
        "reviewer_note": item.reviewer_note,
        "created_at": item.created_at,
        "reviewed_at": item.reviewed_at,
        "ai_result": ai_result,
    }


def _save_feedback_learning(
    db: Session,
    *,
    description: str,
    amount: float,
    correct_category: str,
    correct_type: str,
    correct_debit_account_code: str,
    correct_credit_account_code: str,
    ai_category: Optional[str] = None,
    ai_type: Optional[str] = None,
    ai_debit_account_code: Optional[str] = None,
    ai_credit_account_code: Optional[str] = None,
    ai_confidence: Optional[float] = None,
    note: Optional[str] = None,
) -> AICorrection:
    ensure_account_exists(db, correct_debit_account_code)
    ensure_account_exists(db, correct_credit_account_code)
    correction = AICorrection(
        transaction_id=None,
        original_description=description,
        original_amount=amount,
        ai_category=ai_category,
        ai_type=ai_type,
        ai_debit_account_code=ai_debit_account_code,
        ai_credit_account_code=ai_credit_account_code,
        ai_confidence=ai_confidence,
        user_category=correct_category,
        user_type=correct_type,
        user_debit_account_code=correct_debit_account_code,
        user_credit_account_code=correct_credit_account_code,
        note=note or "V18 feedback learning",
    )
    db.add(correction)
    db.flush()
    return correction


def _train_feedback_model(db: Session, min_examples: int = 1) -> Dict[str, Any]:
    examples = build_training_examples_from_corrections(db.query(AICorrection).order_by(AICorrection.id.asc()).all())
    if len(examples) < min_examples:
        raise HTTPException(status_code=400, detail=f"Chưa đủ dữ liệu train: có {len(examples)} ví dụ, cần ít nhất {min_examples}.")
    model = train_naive_bayes(examples)
    save_model(model)
    return model


def _build_double_entry_lines(db: Session, payload: AIV22DoubleEntryRequest) -> Dict[str, Any]:
    base_result = suggest_journal_entry_with_learning(db, payload.description, payload.amount)
    mode = payload.mode
    if mode == "auto":
        desc = normalize_vi_text(payload.description)
        if any(k in desc for k in ["ban hang", "doanh thu", "xuat hoa don", "khach hang"]):
            mode = "sales"
        elif any(k in desc for k in ["may tinh", "laptop", "tai san", "thiet bi"]):
            mode = "asset"
        elif any(k in desc for k in ["mua", "hoa don", "nha cung cap", "vat", "gtgt"]):
            mode = "purchase"
        else:
            mode = "expense"

    subtotal = float(payload.subtotal or 0)
    vat_amount = float(payload.vat_amount or 0)
    if subtotal <= 0:
        if vat_amount > 0 and payload.amount > vat_amount:
            subtotal = payload.amount - vat_amount
        elif payload.vat_rate > 0 and mode in {"purchase", "sales", "asset"}:
            subtotal = round(payload.amount / (1 + payload.vat_rate / 100), 2)
            vat_amount = round(payload.amount - subtotal, 2)
        else:
            subtotal = payload.amount
    elif vat_amount <= 0 and payload.vat_rate > 0 and mode in {"purchase", "sales", "asset"}:
        vat_amount = round(subtotal * payload.vat_rate / 100, 2)

    total = round(subtotal + vat_amount, 2) if vat_amount else float(payload.amount)
    lines: List[Dict[str, Any]] = []
    if mode == "sales":
        lines = [
            {"side": "debit", "account_code": payload.receivable_account_code, "account_name": get_account_name(db, payload.receivable_account_code), "amount": total},
            {"side": "credit", "account_code": "511", "account_name": get_account_name(db, "511"), "amount": subtotal},
        ]
        if vat_amount > 0:
            lines.append({"side": "credit", "account_code": "3331", "account_name": get_account_name(db, "3331"), "amount": vat_amount})
    elif mode == "asset":
        asset_code = base_result.get("debit_account_code") or "211"
        lines = [
            {"side": "debit", "account_code": asset_code, "account_name": get_account_name(db, asset_code), "amount": subtotal},
        ]
        if vat_amount > 0:
            lines.append({"side": "debit", "account_code": "1331", "account_name": get_account_name(db, "1331"), "amount": vat_amount})
        lines.append({"side": "credit", "account_code": payload.payable_account_code, "account_name": get_account_name(db, payload.payable_account_code), "amount": total})
    elif mode == "purchase":
        purchase_code = base_result.get("debit_account_code") or "156"
        lines = [
            {"side": "debit", "account_code": purchase_code, "account_name": get_account_name(db, purchase_code), "amount": subtotal},
        ]
        if vat_amount > 0:
            lines.append({"side": "debit", "account_code": "1331", "account_name": get_account_name(db, "1331"), "amount": vat_amount})
        lines.append({"side": "credit", "account_code": payload.payable_account_code, "account_name": get_account_name(db, payload.payable_account_code), "amount": total})
    else:
        expense_code = base_result.get("debit_account_code") or "642"
        credit_code = base_result.get("credit_account_code") or payload.cash_account_code
        lines = [
            {"side": "debit", "account_code": expense_code, "account_name": get_account_name(db, expense_code), "amount": subtotal},
        ]
        if vat_amount > 0:
            lines.append({"side": "debit", "account_code": "1331", "account_name": get_account_name(db, "1331"), "amount": vat_amount})
            credit_amount = total
        else:
            credit_amount = payload.amount
        lines.append({"side": "credit", "account_code": credit_code, "account_name": get_account_name(db, credit_code), "amount": credit_amount})

    debit_total = round(sum(float(x.get("amount") or 0) for x in lines if x.get("side") == "debit"), 2)
    credit_total = round(sum(float(x.get("amount") or 0) for x in lines if x.get("side") == "credit"), 2)
    return {
        "mode": mode,
        "base_ai_result": _ai_result_payload(base_result),
        "journal_lines": lines,
        "balance_check": {
            "debit_total": debit_total,
            "credit_total": credit_total,
            "balanced": debit_total == credit_total and debit_total > 0,
            "difference": round(debit_total - credit_total, 2),
        },
        "amount_breakdown": {"subtotal": subtotal, "vat_rate": payload.vat_rate, "vat_amount": vat_amount, "total": total},
    }


@app.post("/ai/v18/feedback-learning")
def ai_v18_feedback_learning(payload: AIV18FeedbackLearningRequest, db: Session = Depends(get_db)):
    correction = _save_feedback_learning(
        db,
        description=payload.description,
        amount=payload.amount,
        correct_category=payload.correct_category,
        correct_type=payload.correct_type,
        correct_debit_account_code=payload.correct_debit_account_code,
        correct_credit_account_code=payload.correct_credit_account_code,
        ai_category=payload.ai_category,
        ai_type=payload.ai_type,
        ai_debit_account_code=payload.ai_debit_account_code,
        ai_credit_account_code=payload.ai_credit_account_code,
        ai_confidence=payload.ai_confidence,
        note=payload.note,
    )
    review_item = None
    ai_result = {
        "category": payload.ai_category,
        "transaction_type": payload.ai_type,
        "debit_account_code": payload.ai_debit_account_code,
        "credit_account_code": payload.ai_credit_account_code,
        "confidence": payload.ai_confidence,
        "source": "user_feedback",
    }
    if payload.create_review_item:
        review_item = _create_review_item(db, description=payload.description, amount=payload.amount, result=ai_result, source="v18_feedback", reason="Feedback đã lưu, dùng để audit quá trình AI học")
        review_item.status = "corrected"
        review_item.reviewer_note = payload.note
        review_item.reviewed_at = datetime.utcnow()
    create_ai_log(db, action="v18_feedback_learning", description=payload.description, amount=payload.amount, result={
        "source": "v18_feedback_learning",
        "category": payload.correct_category,
        "transaction_type": payload.correct_type,
        "debit_account_code": payload.correct_debit_account_code,
        "credit_account_code": payload.correct_credit_account_code,
        "confidence": 1.0,
    })
    model_info = None
    if payload.train_after:
        model = _train_feedback_model(db, min_examples=1)
        model_info = {"example_count": model.get("example_count"), "label_count": model.get("label_count"), "trained_at": model.get("trained_at")}
    db.commit()
    db.refresh(correction)
    return {
        "stage": "V18 - Feedback Learning System",
        "message": "Đã lưu feedback để AI học từ sửa lỗi của kế toán",
        "correction_id": correction.id,
        "learning_item": _correction_to_training_item(correction),
        "review_item": _review_item_to_dict(review_item) if review_item else None,
        "trained_after_save": bool(model_info),
        "model": model_info,
    }


@app.get("/ai/v19/review-queue")
def ai_v19_review_queue(status: str = "pending", limit: int = Query(50, ge=1, le=200), db: Session = Depends(get_db)):
    query = db.query(AIReviewItem)
    if status != "all":
        query = query.filter(AIReviewItem.status == status)
    priority_rank = {"high": 0, "medium": 1, "low": 2}
    items = query.order_by(AIReviewItem.created_at.desc()).limit(limit).all()
    items = sorted(items, key=lambda x: (priority_rank.get(x.priority, 9), -x.id))
    return {
        "stage": "V19 - AI Review Queue",
        "status_filter": status,
        "count": len(items),
        "items": [_review_item_to_dict(item) for item in items],
    }


@app.post("/ai/v19/review-queue/from-analyze")
def ai_v19_push_analyze_to_queue(payload: AIAnalyzeRequest, db: Session = Depends(get_db)):
    result = suggest_journal_entry_with_learning(db, payload.description, payload.amount)
    item = _create_review_item(db, description=payload.description, amount=payload.amount, result=result, source="v19_analyze")
    create_ai_log(db, action="v19_review_queue_created", description=payload.description, amount=payload.amount, result=result)
    db.commit()
    db.refresh(item)
    return {"stage": "V19 - AI Review Queue", "message": "Đã đưa giao dịch vào hàng chờ kế toán duyệt", "item": _review_item_to_dict(item)}


@app.post("/ai/v19/review-queue/{review_item_id}/decision")
def ai_v19_review_decision(review_item_id: int, payload: AIV19ReviewDecisionRequest, db: Session = Depends(get_db)):
    item = db.get(AIReviewItem, review_item_id)
    if not item:
        raise HTTPException(status_code=404, detail="Không tìm thấy review item")
    if item.status != "pending":
        raise HTTPException(status_code=400, detail=f"Review item đã xử lý: {item.status}")

    created_correction = None
    created_journal_entries: List[JournalEntry] = []
    if payload.action == "approve":
        item.status = "approved"
        if payload.create_journal:
            result = json.loads(item.ai_result_json or "{}")
            created_journal_entries = create_journal_from_ai_result(db, transaction_id=item.transaction_id, description=item.description, result=result, commit=False, status="draft")
    elif payload.action == "reject":
        item.status = "rejected"
    else:
        required = [payload.correct_category, payload.correct_type, payload.correct_debit_account_code, payload.correct_credit_account_code]
        if not all(required):
            raise HTTPException(status_code=400, detail="Khi action=correct cần đủ correct_category/correct_type/correct_debit_account_code/correct_credit_account_code")
        item.status = "corrected"
        created_correction = _save_feedback_learning(
            db,
            description=item.description,
            amount=item.amount,
            correct_category=str(payload.correct_category),
            correct_type=str(payload.correct_type),
            correct_debit_account_code=str(payload.correct_debit_account_code),
            correct_credit_account_code=str(payload.correct_credit_account_code),
            ai_category=item.ai_category,
            ai_type=item.ai_type,
            ai_debit_account_code=item.ai_debit_account_code,
            ai_credit_account_code=item.ai_credit_account_code,
            ai_confidence=item.ai_confidence,
            note=payload.note or "Sửa từ V19 review queue",
        )
        if payload.create_journal:
            result = {
                "category": payload.correct_category,
                "transaction_type": payload.correct_type,
                "debit_account_code": payload.correct_debit_account_code,
                "credit_account_code": payload.correct_credit_account_code,
                "amount": item.amount,
                "confidence": 1.0,
            }
            created_journal_entries = create_journal_from_ai_result(db, transaction_id=item.transaction_id, description=item.description, result=result, commit=False, status="draft")
        if payload.train_after_correction:
            _train_feedback_model(db, min_examples=1)
    item.reviewer_note = payload.note
    item.reviewed_at = datetime.utcnow()
    create_ai_log(db, action=f"v19_review_{payload.action}", description=item.description, amount=item.amount, result={
        "category": payload.correct_category or item.ai_category,
        "transaction_type": payload.correct_type or item.ai_type,
        "debit_account_code": payload.correct_debit_account_code or item.ai_debit_account_code,
        "credit_account_code": payload.correct_credit_account_code or item.ai_credit_account_code,
        "confidence": 1.0 if payload.action == "correct" else item.ai_confidence,
    })
    db.commit()
    db.refresh(item)
    return {
        "stage": "V19 - AI Review Queue",
        "message": f"Đã {payload.action} review item",
        "item": _review_item_to_dict(item),
        "correction_id": created_correction.id if created_correction else None,
        "created_journal_entry_ids": [entry.id for entry in created_journal_entries],
    }


@app.post("/ai/v20/retrain-from-feedback")
def ai_v20_retrain_from_feedback(payload: AIV20RetrainFeedbackRequest = AIV20RetrainFeedbackRequest(), db: Session = Depends(get_db)):
    model = _train_feedback_model(db, min_examples=payload.min_examples)
    evaluation = None
    examples = build_training_examples_from_corrections(db.query(AICorrection).order_by(AICorrection.id.asc()).all())
    if payload.evaluate_after and len(examples) >= 2:
        evaluation = _evaluate_examples(examples, payload.test_ratio)
    return {
        "stage": "V20 - Retrain model từ feedback",
        "message": "Đã retrain model Naive Bayes tự viết từ dữ liệu feedback",
        "model": {
            "version": model.get("version"),
            "example_count": model.get("example_count"),
            "label_count": model.get("label_count"),
            "trained_at": model.get("trained_at"),
        },
        "evaluation": evaluation,
        "next_step": "Dùng /ai/v19/review-queue để tiếp tục gom lỗi, rồi retrain định kỳ.",
    }


@app.post("/ocr/v21/invoice-improved/text")
def ai_v21_ocr_improved_text(payload: AIV21OCRImproveTextRequest, db: Session = Depends(get_db)):
    parsed = parse_invoice_text(payload.raw_text)
    result = _build_ocr_response(
        parsed,
        db,
        create_purchase_invoice=payload.create_purchase_invoice,
        create_transaction=payload.create_transaction,
        auto_create_journal=payload.auto_create_journal,
        source={"method": "v21_improved_raw_text"},
    )
    review_item = None
    ai_result = (result.get("ai_suggestion") or {}).get("ai_result")
    if payload.auto_push_review_queue and ai_result:
        amount = float(parsed.get("total_amount") or parsed.get("subtotal") or 0)
        if amount > 0:
            review_item = _create_review_item(db, description=parsed.get("description") or "Hóa đơn", amount=amount, result=ai_result, source="v21_ocr", reason="OCR hóa đơn cần kế toán xác nhận trước khi ghi sổ")
            db.commit()
            db.refresh(review_item)
    result["stage"] = "V21 - OCR hóa đơn tốt hơn + review queue"
    result["review_item"] = _review_item_to_dict(review_item) if review_item else None
    result["ocr_quality"] = {
        "confidence": parsed.get("confidence"),
        "has_invoice_number": bool(parsed.get("invoice_number")),
        "has_date": bool(parsed.get("invoice_date")),
        "has_supplier": bool(parsed.get("supplier_name")),
        "has_total": bool(parsed.get("total_amount")),
        "recommendation": "Nếu thiếu số hóa đơn/ngày/MST/tổng tiền, đưa vào review queue và sửa bằng feedback.",
    }
    return result


@app.post("/ai/v22/double-entry/generate")
def ai_v22_double_entry_generate(payload: AIV22DoubleEntryRequest, db: Session = Depends(get_db)):
    generated = _build_double_entry_lines(db, payload)
    if not generated["balance_check"]["balanced"]:
        raise HTTPException(status_code=400, detail={"message": "Bút toán chưa cân", "generated": generated})
    created_entries: List[JournalEntry] = []
    if payload.auto_create_journal:
        ai_result = {
            "category": generated["base_ai_result"].get("category"),
            "transaction_type": generated["base_ai_result"].get("transaction_type"),
            "debit_account_code": next((x["account_code"] for x in generated["journal_lines"] if x["side"] == "debit"), None),
            "credit_account_code": next((x["account_code"] for x in generated["journal_lines"] if x["side"] == "credit"), None),
            "amount": generated["balance_check"]["debit_total"],
            "confidence": generated["base_ai_result"].get("confidence"),
            "journal_lines": generated["journal_lines"],
        }
        created_entries = create_journal_from_ai_result(
            db,
            transaction_id=payload.transaction_id,
            description=payload.description,
            result=ai_result,
            commit=False,
            status=payload.status,
            entry_date=payload.transaction_date or date.today(),
        )
    create_ai_log(db, action="v22_double_entry_generate", description=payload.description, amount=payload.amount, result={
        "source": "v22_double_entry",
        "category": generated["base_ai_result"].get("category"),
        "transaction_type": generated["base_ai_result"].get("transaction_type"),
        "debit_account_code": next((x["account_code"] for x in generated["journal_lines"] if x["side"] == "debit"), None),
        "credit_account_code": next((x["account_code"] for x in generated["journal_lines"] if x["side"] == "credit"), None),
        "confidence": generated["base_ai_result"].get("confidence"),
        "journal_lines": generated["journal_lines"],
    }, transaction_id=payload.transaction_id)
    db.commit()
    for entry in created_entries:
        db.refresh(entry)
    return {
        "stage": "V22 - Tự sinh bút toán kép",
        "message": "Đã sinh bút toán Nợ/Có cân bằng" + (" và tạo journal_entries" if created_entries else " ở chế độ preview"),
        **generated,
        "created_journal_entry_ids": [entry.id for entry in created_entries],
        "safety_note": "Mặc định nên để status=draft và duyệt qua V19 trước khi posted.",
    }


# ============================================================
# FINIIP V23 - Frontend AI Review Queue UI
# ============================================================

@app.get("/v23/review-queue-ui")
def v23_review_queue_ui():
    """Backend-only build: UI review queue đã tách riêng; dùng API bên dưới."""
    return {
        "message": "Frontend UI is not bundled in backend-only build.",
        "api_status": "/ai/v23/review-ui/status",
        "queue_api": "/ai/v19/review-queue",
        "decision_api": "/ai/v19/review-queue/{item_id}/decision",
        "docs": "/docs"
    }


@app.get("/ai/v23/review-ui/status")
def ai_v23_review_ui_status(db: Session = Depends(get_db)):
    """Tóm tắt trạng thái V23 để frontend hoặc tester kiểm tra nhanh."""
    pending = db.query(AIReviewItem).filter(AIReviewItem.status == "pending").count()
    corrected = db.query(AIReviewItem).filter(AIReviewItem.status == "corrected").count()
    approved = db.query(AIReviewItem).filter(AIReviewItem.status == "approved").count()
    rejected = db.query(AIReviewItem).filter(AIReviewItem.status == "rejected").count()
    return {
        "stage": "V23 - Frontend AI Review Queue",
        "message": "Đã có màn hình duyệt AI Review Queue cho kế toán",
        "ui_url": "/v23/review-queue-ui",
        "safe_posting_flow": "AI chỉ đề xuất; kế toán approve/correct/reject; journal entry tạo từ UI vẫn là draft.",
        "counts": {
            "review_pending": pending,
            "review_corrected": corrected,
            "review_approved": approved,
            "review_rejected": rejected,
            "feedback_examples": db.query(AICorrection).count(),
            "model": model_status(),
        },
        "actions_supported": [
            "create review item from AI analyze",
            "filter queue by status",
            "approve AI result",
            "correct AI result and save feedback",
            "reject AI result",
            "create draft journal entry after approve/correct",
            "retrain model from feedback",
        ],
        "next_step": "V24 - Dashboard chất lượng AI kế toán",
    }


# ============================================================
# FINIIP V24 - AI Quality Dashboard
# ============================================================

def _safe_rate(part: int, total: int) -> float:
    if not total:
        return 0.0
    return round((part / total) * 100, 2)


def _confidence_bucket(confidence: Optional[float]) -> str:
    value = float(confidence or 0)
    if value < 0.55:
        return "low"
    if value < 0.78:
        return "medium"
    return "high"


@app.get("/ai/v24/quality-dashboard")
def ai_v24_quality_dashboard(db: Session = Depends(get_db)):
    """Dashboard số liệu chất lượng AI: queue, feedback, confidence, draft journal và khuyến nghị vận hành."""
    total_review = db.query(AIReviewItem).count()
    pending = db.query(AIReviewItem).filter(AIReviewItem.status == "pending").count()
    approved = db.query(AIReviewItem).filter(AIReviewItem.status == "approved").count()
    corrected = db.query(AIReviewItem).filter(AIReviewItem.status == "corrected").count()
    rejected = db.query(AIReviewItem).filter(AIReviewItem.status == "rejected").count()

    avg_confidence_raw = db.query(func.avg(AIReviewItem.ai_confidence)).scalar()
    avg_confidence = round(float(avg_confidence_raw or 0), 4)

    priority_counts = {"high": 0, "medium": 0, "low": 0}
    status_counts = {"pending": pending, "approved": approved, "corrected": corrected, "rejected": rejected}
    confidence_buckets = {"low": 0, "medium": 0, "high": 0}

    review_items = db.query(AIReviewItem).all()
    for item in review_items:
        if item.priority in priority_counts:
            priority_counts[item.priority] += 1
        confidence_buckets[_confidence_bucket(item.ai_confidence)] += 1

    feedback_examples = db.query(AICorrection).count()
    ai_log_total = db.query(AILog).count()
    journal_drafts = db.query(JournalEntry).filter(JournalEntry.status == "draft").count()
    journal_posted = db.query(JournalEntry).filter(JournalEntry.status == "posted").count()

    quality_score = 0
    if total_review:
        quality_score = max(0, 100 - (_safe_rate(corrected + rejected, total_review) * 0.7) - (_safe_rate(pending, total_review) * 0.3))
        quality_score = round(quality_score, 2)

    recommendations = []
    if pending > 0:
        recommendations.append("Ưu tiên xử lý các item pending, đặc biệt priority=high.")
    if total_review and _safe_rate(corrected, total_review) > 25:
        recommendations.append("Tỷ lệ corrected cao: nên retrain model bằng /ai/v20/retrain-from-feedback.")
    if avg_confidence and avg_confidence < 0.70:
        recommendations.append("Confidence trung bình còn thấp: cần thêm dữ liệu feedback thật.")
    if feedback_examples < 20:
        recommendations.append("Nên tích lũy ít nhất 20-50 feedback examples trước khi tin model hơn.")
    if not recommendations:
        recommendations.append("Dashboard ổn định. Có thể tiếp tục sang V25 - Auto Journal Draft Flow.")

    return {
        "stage": "V24 - Dashboard chất lượng AI kế toán",
        "message": "Đã có dashboard đo chất lượng AI, review queue và feedback learning.",
        "ui_url": "/v24/quality-dashboard-ui",
        "summary": {
            "total_review_items": total_review,
            "pending_review_items": pending,
            "feedback_examples": feedback_examples,
            "ai_logs": ai_log_total,
            "avg_confidence": avg_confidence,
            "quality_score": quality_score,
        },
        "rates": {
            "pending_rate": _safe_rate(pending, total_review),
            "approved_rate": _safe_rate(approved, total_review),
            "corrected_rate": _safe_rate(corrected, total_review),
            "rejected_rate": _safe_rate(rejected, total_review),
            "human_review_rate": _safe_rate(pending + corrected + rejected, total_review),
            "ai_accept_rate": _safe_rate(approved, total_review),
        },
        "status_counts": status_counts,
        "priority_counts": priority_counts,
        "confidence_buckets": confidence_buckets,
        "journal_counts": {
            "draft": journal_drafts,
            "posted": journal_posted,
        },
        "model": model_status(),
        "recommendations": recommendations,
        "next_step": "V25 - Auto Journal Draft: AI tạo bút toán nháp an toàn, chờ kế toán posted.",
    }


@app.get("/v24/quality-dashboard-ui")
def v24_quality_dashboard_ui():
    """Backend-only build: dashboard UI đã tách riêng; dùng API JSON bên dưới."""
    return {
        "message": "Frontend UI is not bundled in backend-only build.",
        "quality_dashboard_api": "/ai/v24/quality-dashboard",
        "docs": "/docs"
    }


@app.get("/ai/v18-v22/upgrade-status")
def ai_v18_v22_upgrade_status(db: Session = Depends(get_db)):
    pending = db.query(AIReviewItem).filter(AIReviewItem.status == "pending").count()
    corrected = db.query(AIReviewItem).filter(AIReviewItem.status == "corrected").count()
    approved = db.query(AIReviewItem).filter(AIReviewItem.status == "approved").count()
    return {
        "stage": "Finiip V18-V22 - Self-made AI accounting workflow",
        "no_openai": True,
        "no_ollama": True,
        "completed": [
            "V18 Feedback Learning",
            "V19 AI Review Queue",
            "V20 Retrain model từ feedback",
            "V21 OCR hóa đơn tốt hơn + review queue",
            "V22 Tự sinh bút toán kép",
            "V23 Frontend AI Review Queue",
            "V24 Dashboard chất lượng AI kế toán",
        ],
        "current_level": "Cấp 4.5 / Cấp 5 prototype có dashboard kiểm soát chất lượng",
        "why_not_full_level_5_yet": "Đã sinh bút toán kép được, nhưng vẫn cần nhiều dữ liệu thật, kiểm thử kế toán và quy trình duyệt trước khi tự động posted.",
        "counts": {
            "feedback_examples": db.query(AICorrection).count(),
            "review_pending": pending,
            "review_corrected": corrected,
            "review_approved": approved,
            "review_rejected": db.query(AIReviewItem).filter(AIReviewItem.status == "rejected").count(),
            "model": model_status(),
        },
        "main_apis": [
            "POST /ai/v18/feedback-learning",
            "GET  /ai/v19/review-queue",
            "POST /ai/v19/review-queue/from-analyze",
            "POST /ai/v19/review-queue/{id}/decision",
            "POST /ai/v20/retrain-from-feedback",
            "POST /ocr/v21/invoice-improved/text",
            "POST /ai/v22/double-entry/generate",
            "GET  /v23/review-queue-ui",
            "GET  /ai/v23/review-ui/status",
            "GET  /ai/v24/quality-dashboard",
            "GET  /v24/quality-dashboard-ui",
        ],
    }


# ============================================================
# V4 - AI bulk import / bank statement intelligence (backend-only)
# ============================================================
# Mục tiêu: frontend có thể upload CSV/XLSX sao kê hoặc danh sách giao dịch,
# backend tự đọc dòng, chuẩn hoá amount/description/date, phân tích AI V3 hàng loạt,
# phát hiện rủi ro, giao dịch trùng và kiểm tra bút toán cơ bản.

from pydantic import BaseModel, Field
import csv
import io
import hashlib


class AIV4TransactionItem(BaseModel):
    description: str
    amount: float = Field(gt=0)
    transaction_date: Optional[str] = None
    reference: Optional[str] = None
    counterparty: Optional[str] = None
    raw: Optional[Dict[str, Any]] = None


class AIV4BatchAnalyzeItemsRequest(BaseModel):
    items: List[AIV4TransactionItem]
    duplicate_window_days: int = 7


class AIV4JournalValidateRequest(BaseModel):
    description: Optional[str] = None
    amount: Optional[float] = None
    debit_account_code: Optional[str] = None
    credit_account_code: Optional[str] = None
    entries: Optional[List[Dict[str, Any]]] = None


class AIV4DuplicateDetectRequest(BaseModel):
    items: List[AIV4TransactionItem]
    duplicate_window_days: int = 7


_AMOUNT_COLUMNS = [
    "amount", "số tiền", "so tien", "sotien", "money", "value", "gia tri", "giá trị",
    "credit", "debit", "phat sinh co", "phát sinh có", "phat sinh no", "phát sinh nợ",
    "thu", "chi", "deposit", "withdrawal",
]
_DESCRIPTION_COLUMNS = [
    "description", "mô tả", "mo ta", "nội dung", "noi dung", "dien giai", "diễn giải",
    "transaction", "memo", "remark", "ghi chú", "ghi chu", "details", "content",
]
_DATE_COLUMNS = ["date", "ngày", "ngay", "transaction_date", "posting date", "book date", "ngày giao dịch"]
_REFERENCE_COLUMNS = ["reference", "ref", "mã giao dịch", "ma giao dich", "transaction id", "id", "số chứng từ", "so chung tu"]
_COUNTERPARTY_COLUMNS = ["counterparty", "đối tác", "doi tac", "beneficiary", "sender", "receiver", "nhà cung cấp", "khách hàng"]


def _norm_col(value: Any) -> str:
    return str(value or "").strip().lower().replace("_", " ").replace("-", " ")


def _pick_value(row: Dict[str, Any], candidates: List[str]) -> Optional[Any]:
    normalized = {_norm_col(k): v for k, v in row.items()}
    for c in candidates:
        if c in normalized and normalized[c] not in (None, ""):
            return normalized[c]
    for key, value in normalized.items():
        if any(c in key for c in candidates) and value not in (None, ""):
            return value
    return None


def _parse_money(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(abs(value))
    s = str(value).strip()
    if not s:
        return None
    s = s.replace("₫", "").replace("VND", "").replace("vnd", "").replace(" ", "")
    negative = s.startswith("-") or s.endswith("CR-")
    s = s.replace("+", "").replace("-", "")
    # Hỗ trợ cả 1,234,567 và 1.234.567 và 1234567.89
    if s.count(",") > 0 and s.count(".") > 0:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    else:
        if s.count(",") == 1 and len(s.split(",")[-1]) <= 2:
            s = s.replace(",", ".")
        else:
            s = s.replace(",", "").replace(".", "")
    try:
        n = float(s)
        return abs(n) if not negative else abs(n)
    except Exception:
        return None


def _row_to_item(row: Dict[str, Any], index: int) -> Dict[str, Any]:
    description = _pick_value(row, _DESCRIPTION_COLUMNS)
    amount = _pick_value(row, _AMOUNT_COLUMNS)
    # Nếu có cả debit/credit riêng, ưu tiên ô có số tiền khác 0
    if amount in (None, "", 0):
        for col in _AMOUNT_COLUMNS:
            v = _pick_value(row, [col])
            money = _parse_money(v)
            if money and money > 0:
                amount = v
                break
    parsed_amount = _parse_money(amount)
    if not description:
        non_empty = [str(v).strip() for v in row.values() if v not in (None, "")]
        description = " | ".join(non_empty[:3]) if non_empty else f"Giao dịch dòng {index}"
    item = {
        "row_index": index,
        "description": str(description).strip(),
        "amount": parsed_amount or 0,
        "transaction_date": str(_pick_value(row, _DATE_COLUMNS) or "").strip() or None,
        "reference": str(_pick_value(row, _REFERENCE_COLUMNS) or "").strip() or None,
        "counterparty": str(_pick_value(row, _COUNTERPARTY_COLUMNS) or "").strip() or None,
        "raw": row,
    }
    return item


def _read_csv_upload(content: bytes) -> List[Dict[str, Any]]:
    text_content = None
    for enc in ("utf-8-sig", "utf-8", "cp1258", "latin-1"):
        try:
            text_content = content.decode(enc)
            break
        except Exception:
            continue
    if text_content is None:
        raise HTTPException(status_code=400, detail="Không đọc được encoding CSV")
    sample = text_content[:2048]
    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|") if sample.strip() else csv.excel
    reader = csv.DictReader(io.StringIO(text_content), dialect=dialect)
    return [dict(row) for row in reader]


def _read_xlsx_upload(content: bytes) -> List[Dict[str, Any]]:
    if Workbook is None:
        raise HTTPException(status_code=500, detail="Thiếu openpyxl để đọc XLSX")
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h or "").strip() for h in rows[0]]
    data = []
    for values in rows[1:]:
        if not any(v not in (None, "") for v in values):
            continue
        data.append({headers[i] if i < len(headers) and headers[i] else f"column_{i+1}": values[i] if i < len(values) else None for i in range(len(headers))})
    return data


def _read_transaction_upload(file: UploadFile, content: bytes) -> List[Dict[str, Any]]:
    name = (file.filename or "").lower()
    if name.endswith(".csv") or file.content_type in {"text/csv", "application/csv"}:
        return _read_csv_upload(content)
    if name.endswith(".xlsx"):
        return _read_xlsx_upload(content)
    raise HTTPException(status_code=400, detail="Chỉ hỗ trợ .csv hoặc .xlsx")


def _fingerprint_item(item: Dict[str, Any]) -> str:
    date_part = str(item.get("transaction_date") or "")[:10]
    amount_part = round(float(item.get("amount") or 0), 0)
    desc_part = re.sub(r"\s+", " ", str(item.get("description") or "").lower()).strip()[:80]
    key = f"{date_part}|{amount_part}|{desc_part}"
    return hashlib.sha1(key.encode("utf-8")).hexdigest()[:12]


def _detect_duplicates(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    seen: Dict[str, List[int]] = {}
    duplicates = []
    for idx, item in enumerate(items):
        fp = _fingerprint_item(item)
        seen.setdefault(fp, []).append(idx)
    for fp, indexes in seen.items():
        if len(indexes) > 1:
            duplicates.append({
                "fingerprint": fp,
                "row_indexes": [items[i].get("row_index", i + 1) for i in indexes],
                "reason": "Cùng ngày/số tiền/mô tả gần giống nhau trong file import",
            })
    return duplicates


def _validate_ai_journal(ai_result: Dict[str, Any], amount: float) -> Dict[str, Any]:
    debit = ai_result.get("debit_account") or ai_result.get("debit_account_code") or ai_result.get("suggested_debit_account")
    credit = ai_result.get("credit_account") or ai_result.get("credit_account_code") or ai_result.get("suggested_credit_account")
    issues = []
    severity = "ok"
    if not debit or not credit:
        issues.append("Thiếu tài khoản Nợ hoặc Có")
        severity = "review"
    if debit and credit and str(debit) == str(credit):
        issues.append("Tài khoản Nợ và Có đang giống nhau")
        severity = "block"
    if not amount or amount <= 0:
        issues.append("Số tiền không hợp lệ")
        severity = "block"
    if str(credit) == "111" and amount >= 5_000_000:
        issues.append("Giao dịch tiền mặt từ 5 triệu đồng trở lên: cần kiểm tra chứng từ và điều kiện thanh toán không dùng tiền mặt")
        severity = "review" if severity != "block" else severity
    return {
        "is_valid_for_draft": severity in {"ok", "review"},
        "is_safe_to_auto_post": severity == "ok",
        "severity": severity,
        "issues": issues,
        "debit_account_code": debit,
        "credit_account_code": credit,
    }


def _analyze_v4_item(db: Session, item: Dict[str, Any]) -> Dict[str, Any]:
    amount = float(item.get("amount") or 0)
    description = str(item.get("description") or "")
    if not description or amount <= 0:
        return {
            **item,
            "status": "invalid_input",
            "ai_result": None,
            "journal_validation": {"severity": "block", "issues": ["Thiếu mô tả hoặc số tiền không hợp lệ"]},
        }
    base_result = suggest_journal_entry_with_learning(db, description, amount)
    ai_result = enhance_ai_result(description, amount, base_result)
    validation = _validate_ai_journal(ai_result, amount)
    gate = ai_result.get("quality_gate") or {}
    workflow = gate.get("decision") if isinstance(gate, dict) else gate
    if validation["severity"] == "block":
        workflow = "BLOCK_AUTO_POSTING"
    elif validation["severity"] == "review" and workflow == "AUTO_DRAFT_ALLOWED":
        workflow = "REVIEW_REQUIRED"
    return {
        **item,
        "status": "analyzed",
        "fingerprint": _fingerprint_item(item),
        "ai_result": ai_result,
        "journal_validation": validation,
        "recommended_workflow": workflow,
    }


@app.post("/ai/v4/import-preview")
async def ai_v4_import_preview(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Upload CSV/XLSX sao kê hoặc danh sách giao dịch, AI đọc và phân tích hàng loạt, chưa ghi database."""
    content = await file.read()
    rows = _read_transaction_upload(file, content)
    normalized = [_row_to_item(row, idx + 1) for idx, row in enumerate(rows)]
    analyzed = [_analyze_v4_item(db, item) for item in normalized]
    duplicates = _detect_duplicates(analyzed)
    summary = {
        "file_name": file.filename,
        "total_rows": len(rows),
        "valid_amount_rows": sum(1 for i in normalized if (i.get("amount") or 0) > 0),
        "analyzed_rows": sum(1 for i in analyzed if i.get("status") == "analyzed"),
        "review_required": sum(1 for i in analyzed if i.get("recommended_workflow") == "REVIEW_REQUIRED"),
        "blocked": sum(1 for i in analyzed if i.get("recommended_workflow") == "BLOCK_AUTO_POSTING" or i.get("status") == "invalid_input"),
        "duplicate_groups": len(duplicates),
    }
    return {"summary": summary, "duplicates": duplicates, "items": analyzed}


@app.post("/ai/v4/batch-analyze-items")
def ai_v4_batch_analyze_items(payload: AIV4BatchAnalyzeItemsRequest, db: Session = Depends(get_db)):
    items = [i.model_dump() for i in payload.items]
    for idx, item in enumerate(items, start=1):
        item.setdefault("row_index", idx)
    analyzed = [_analyze_v4_item(db, item) for item in items]
    duplicates = _detect_duplicates(analyzed)
    return {
        "summary": {
            "total": len(items),
            "analyzed": sum(1 for i in analyzed if i.get("status") == "analyzed"),
            "review_required": sum(1 for i in analyzed if i.get("recommended_workflow") == "REVIEW_REQUIRED"),
            "blocked": sum(1 for i in analyzed if i.get("recommended_workflow") == "BLOCK_AUTO_POSTING" or i.get("status") == "invalid_input"),
            "duplicate_groups": len(duplicates),
        },
        "duplicates": duplicates,
        "items": analyzed,
    }


@app.post("/ai/v4/detect-duplicates")
def ai_v4_detect_duplicates(payload: AIV4DuplicateDetectRequest):
    items = [i.model_dump() for i in payload.items]
    for idx, item in enumerate(items, start=1):
        item.setdefault("row_index", idx)
    return {"duplicate_groups": _detect_duplicates(items)}


@app.post("/ai/v4/validate-journal")
def ai_v4_validate_journal(payload: AIV4JournalValidateRequest):
    if payload.entries:
        debit_total = sum(float(e.get("amount", 0) or 0) for e in payload.entries if str(e.get("side", "")).lower() in {"debit", "no", "nợ"})
        credit_total = sum(float(e.get("amount", 0) or 0) for e in payload.entries if str(e.get("side", "")).lower() in {"credit", "co", "có"})
        issues = []
        if round(debit_total, 2) != round(credit_total, 2):
            issues.append("Tổng Nợ không bằng Tổng Có")
        if debit_total <= 0 or credit_total <= 0:
            issues.append("Bút toán thiếu dòng Nợ hoặc Có")
        return {
            "debit_total": debit_total,
            "credit_total": credit_total,
            "balanced": round(debit_total, 2) == round(credit_total, 2),
            "is_safe_to_post": not issues,
            "issues": issues,
        }
    ai_result = {
        "debit_account": payload.debit_account_code,
        "credit_account": payload.credit_account_code,
    }
    return _validate_ai_journal(ai_result, float(payload.amount or 0))


@app.get("/ai/v4/demo-bank-statement")
def ai_v4_demo_bank_statement():
    return {
        "items": [
            {"transaction_date": "2026-06-01", "description": "Thu tiền bán hàng từ khách hàng A", "amount": 12000000, "reference": "GD001"},
            {"transaction_date": "2026-06-01", "description": "Thanh toán quảng cáo Facebook bằng chuyển khoản", "amount": 5000000, "reference": "GD002"},
            {"transaction_date": "2026-06-02", "description": "Thanh toán tiền điện EVN", "amount": 2300000, "reference": "GD003"},
            {"transaction_date": "2026-06-02", "description": "Rút tiền mặt 80000000 chi dịch vụ tư vấn", "amount": 80000000, "reference": "GD004"},
            {"transaction_date": "2026-06-02", "description": "Thanh toán tiền điện EVN", "amount": 2300000, "reference": "GD003-DUP"},
        ],
        "how_to_test": "POST body này vào /ai/v4/batch-analyze-items hoặc tạo CSV/XLSX rồi upload vào /ai/v4/import-preview",
    }


@app.get("/ai/v4/capabilities")
def ai_v4_capabilities():
    return {
        "version": "V4 AI bulk accounting intelligence",
        "new_capabilities": [
            "Import preview CSV/XLSX sao kê ngân hàng hoặc danh sách giao dịch",
            "Chuẩn hoá cột ngày/mô tả/số tiền từ nhiều mẫu file khác nhau",
            "AI V3 phân tích hàng loạt: category, bút toán, explanation, confidence, risk flags",
            "Phát hiện giao dịch nghi trùng trong file import",
            "Kiểm tra bút toán cơ bản: Nợ/Có, số tiền, tiền mặt lớn",
            "Đưa ra workflow: AUTO_DRAFT_ALLOWED / REVIEW_REQUIRED / BLOCK_AUTO_POSTING",
        ],
        "main_endpoints": [
            "POST /ai/v4/import-preview",
            "POST /ai/v4/batch-analyze-items",
            "POST /ai/v4/detect-duplicates",
            "POST /ai/v4/validate-journal",
            "GET  /ai/v4/demo-bank-statement",
        ],
    }


# ============================================================
# V5-V9 - AI learning, OCR draft, RAG, anomaly, API security helpers
# ============================================================
# Backend-only extensions for a separate frontend to call.

from fastapi import Header, HTTPException
from ai_v5_v9 import (
    add_knowledge_doc,
    anomaly_score,
    apply_learning,
    delete_learning_rule,
    invoice_to_transaction_draft,
    learning_stats,
    load_feedback_rules,
    save_feedback_event,
    search_knowledge,
    validate_api_key,
    api_key_required,
)


def finiip_api_guard(x_api_key: Optional[str] = Header(default=None)):
    if not validate_api_key(x_api_key):
        raise HTTPException(status_code=401, detail="Invalid or missing X-API-Key")
    return True


class AIV5FeedbackRequest(BaseModel):
    description: str
    amount: float = 0
    ai_result: Dict[str, Any] = Field(default_factory=dict)
    correction: Dict[str, Any] = Field(default_factory=dict)
    user_note: Optional[str] = ""
    user_id: Optional[str] = "anonymous"


class AIV5AnalyzeLearningRequest(BaseModel):
    description: str
    amount: float = Field(default=0, ge=0)


class AIV6InvoiceTextRequest(BaseModel):
    text: str
    auto_analyze: bool = True


class AIV7KnowledgeUploadRequest(BaseModel):
    title: str
    content: str
    source: Optional[str] = "manual"
    tags: List[str] = Field(default_factory=list)


class AIV7AskRequest(BaseModel):
    question: str
    limit: int = 5


class AIV8AnomalyRequest(BaseModel):
    items: List[AIV4TransactionItem]


@app.post("/ai/v5/feedback")
def ai_v5_feedback(payload: AIV5FeedbackRequest, _: bool = Depends(finiip_api_guard)):
    event = save_feedback_event(
        description=payload.description,
        amount=payload.amount,
        ai_result=payload.ai_result,
        correction=payload.correction,
        user_note=payload.user_note or "",
        user_id=payload.user_id or "anonymous",
    )
    return {"saved": True, "event": event, "stats": learning_stats()}


@app.post("/ai/v5/analyze-with-learning")
def ai_v5_analyze_with_learning(payload: AIV5AnalyzeLearningRequest, db: Session = Depends(get_db), _: bool = Depends(finiip_api_guard)):
    base_result = suggest_journal_entry_with_learning(db, payload.description, payload.amount)
    v3_result = enhance_ai_result(payload.description, payload.amount, base_result)
    learned = apply_learning(payload.description, v3_result)
    log = create_ai_log(db, action="ai_v5_analyze_with_learning", description=payload.description, amount=payload.amount, result=learned)
    return {"ai_result": learned, "ai_log_id": log.id, "workflow_hint": learned.get("quality_gate")}


@app.get("/ai/v5/learning-rules")
def ai_v5_learning_rules(_: bool = Depends(finiip_api_guard)):
    return load_feedback_rules()


@app.get("/ai/v5/learning-stats")
def ai_v5_learning_stats(_: bool = Depends(finiip_api_guard)):
    return learning_stats()


@app.delete("/ai/v5/learning-rules/{rule_id}")
def ai_v5_delete_learning_rule(rule_id: str, _: bool = Depends(finiip_api_guard)):
    deleted = delete_learning_rule(rule_id)
    return {"deleted": deleted, "rule_id": rule_id, "stats": learning_stats()}


@app.post("/ai/v6/invoice-to-transaction")
def ai_v6_invoice_to_transaction(payload: AIV6InvoiceTextRequest, db: Session = Depends(get_db), _: bool = Depends(finiip_api_guard)):
    parsed = parse_invoice_text(payload.text)
    draft = invoice_to_transaction_draft(parsed)
    ai_result = None
    if payload.auto_analyze:
        base_result = suggest_journal_entry_with_learning(db, draft.get("description", ""), float(draft.get("amount") or 0))
        ai_result = enhance_ai_result(draft.get("description", ""), float(draft.get("amount") or 0), base_result)
    return {"parsed_invoice": parsed, "transaction_draft": draft, "ai_result": ai_result, "next_step": "Frontend nên hiển thị cho kế toán duyệt trước khi ghi sổ."}


@app.post("/ai/v7/knowledge/upload-text")
def ai_v7_upload_knowledge(payload: AIV7KnowledgeUploadRequest, _: bool = Depends(finiip_api_guard)):
    doc = add_knowledge_doc(payload.title, payload.content, payload.source or "manual", payload.tags)
    return {"saved": True, "document": {"id": doc["id"], "title": doc["title"], "source": doc["source"], "tags": doc["tags"]}}


@app.post("/ai/v7/ask")
def ai_v7_ask(payload: AIV7AskRequest, _: bool = Depends(finiip_api_guard)):
    result = search_knowledge(payload.question, payload.limit)
    return {"question": payload.question, **result, "safety_note": "Câu trả lời dựa trên tài liệu đã upload; kế toán vẫn cần kiểm tra quy định hiện hành."}


@app.post("/ai/v8/anomaly-score")
def ai_v8_anomaly_score(payload: AIV8AnomalyRequest, _: bool = Depends(finiip_api_guard)):
    items = [i.model_dump() for i in payload.items]
    return anomaly_score(items)


@app.get("/ai/v9/security-status")
def ai_v9_security_status():
    return {
        "api_key_required": api_key_required(),
        "how_to_enable": "Set FINIIP_API_KEY in .env/environment. Frontend sends header X-API-Key.",
        "note": "Nếu FINIIP_API_KEY chưa được set, các endpoint V5-V8 vẫn mở để dễ demo local.",
    }




# ============================================================
# V66/V67 - File upload router + Supabase RAG storage
# ============================================================
# Luồng deploy thật:
# - target=rag: upload PDF/DOCX/TXT/XLSX vào kho tri thức lâu dài.
# - target=solve: chỉ đọc file để giải/hỏi một lần, không lưu vào RAG.
# - target=temp: chỉ trích text tạm.

from services.rag_v66_v67 import (
    UPLOAD_DIR,
    chunk_text as v66_chunk_text,
    read_upload_bytes as v66_read_upload_bytes,
    safe_filename as v66_safe_filename,
    save_local_rag_document,
    save_supabase_rag_document,
    search_local_rag,
    search_supabase_rag,
    split_tags as v66_split_tags,
    supabase_status as v67_supabase_status,
)


class V67RAGSearchRequest(BaseModel):
    question: str
    limit: int = Field(default=5, ge=1, le=20)
    category: Optional[str] = None


def _guess_upload_target(filename: str, document_type: Optional[str], question: Optional[str]) -> str:
    name = (filename or "").lower()
    dtype = (document_type or "").lower()
    q = (question or "").strip()
    if q:
        return "solve"
    if any(k in dtype for k in ["thong_tu", "nghi_dinh", "luat", "quy_dinh", "quy_trinh", "so_tay"]):
        return "rag"
    if any(k in name for k in ["thong-tu", "thong_tu", "nghi-dinh", "nghi_dinh", "luat", "quy-trinh", "quy_trinh"]):
        return "rag"
    if any(k in name for k in ["de-thi", "de_thi", "bai-tap", "bai_tap", "question", "cau-hoi", "cau_hoi"]):
        return "solve"
    return "temp"


async def _read_uploaded_file(file: UploadFile) -> Dict[str, Any]:
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="File rỗng")
    filename = file.filename or "upload"
    # V71: giới hạn loại file/dung lượng trước khi parse để tránh spam khi deploy.
    try:
        from services.rag_v66_v67 import validate_rag_file
        validation = validate_rag_file(filename, len(raw))
        if not validation.get("ok"):
            raise HTTPException(status_code=400, detail=validation)
    except HTTPException:
        raise
    except Exception:
        pass
    safe_name = f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{v66_safe_filename(filename)}"
    storage_path = str(UPLOAD_DIR / safe_name)
    with open(storage_path, "wb") as f:
        f.write(raw)
    try:
        text = v66_read_upload_bytes(filename, raw)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    if not text.strip():
        raise HTTPException(status_code=400, detail="Không trích được text từ file. Nếu là ảnh scan, cần OCR riêng.")
    return {"filename": filename, "storage_path": storage_path, "text": text, "size_bytes": len(raw)}


@app.get("/ai/v66/file-upload-router/status")
def ai_v66_file_upload_router_status(_: bool = Depends(finiip_api_guard)):
    local = search_local_rag("kế toán thuế hóa đơn", limit=1)
    store_path = "data/v66_rag_store.json"
    return {
        "ready": True,
        "version": "V66/V67",
        "targets": ["rag", "solve", "temp", "auto"],
        "endpoints": [
            "POST /ai/v66/rag/upload-file",
            "POST /ai/v66/solve/upload-question-file",
            "POST /ai/v66/file-upload-router",
            "GET  /ai/v67/supabase-rag/status",
            "POST /ai/v67/supabase-rag/search",
        ],
        "local_rag_store": store_path,
        "local_sample_matched": local.get("matched", 0),
        "supabase": v67_supabase_status(),
    }


@app.post("/ai/v66/rag/upload-file")
async def ai_v66_rag_upload_file(
    file: UploadFile = File(...),
    title: Optional[str] = Form(default=None),
    category: str = Form(default="general"),
    document_type: str = Form(default="document"),
    source: str = Form(default="admin_upload"),
    tags: Optional[str] = Form(default=""),
    uploaded_by: str = Form(default="admin"),
    document_number: Optional[str] = Form(default=None),
    issued_date: Optional[str] = Form(default=None),
    effective_date: Optional[str] = Form(default=None),
    authority: Optional[str] = Form(default=None),
    status: str = Form(default="active"),
    version: Optional[str] = Form(default=None),
    workspace_id: Optional[str] = Form(default=None),
    user_id: Optional[str] = Form(default=None),
    language: str = Form(default="vi"),
    jurisdiction: Optional[str] = Form(default=None),
    extra_metadata_json: Optional[str] = Form(default=None),
    auto_chunk: bool = Form(default=True),
    _: bool = Depends(finiip_api_guard),
):
    uploaded = await _read_uploaded_file(file)
    doc_title = title or uploaded["filename"]
    tag_list = v66_split_tags(tags)
    chunks = v66_chunk_text(uploaded["text"]) if auto_chunk else [uploaded["text"]]
    if not chunks:
        raise HTTPException(status_code=400, detail="File không có nội dung đủ để chunk")

    local = save_local_rag_document(
        title=doc_title,
        content=uploaded["text"],
        filename=uploaded["filename"],
        category=category,
        document_type=document_type,
        source=source,
        tags=tag_list,
        storage_path=uploaded["storage_path"],
        document_number=document_number,
        issued_date=issued_date,
        effective_date=effective_date,
        authority=authority,
        status=status,
        version=version,
        workspace_id=workspace_id,
        user_id=user_id,
        language=language,
        jurisdiction=jurisdiction,
        extra_metadata=json.loads(extra_metadata_json) if extra_metadata_json else None,
    )
    supabase = save_supabase_rag_document(
        title=doc_title,
        chunks=chunks,
        filename=uploaded["filename"],
        category=category,
        document_type=document_type,
        source=source,
        tags=tag_list,
        storage_path=uploaded["storage_path"],
        uploaded_by=uploaded_by,
        document_number=document_number,
        issued_date=issued_date,
        effective_date=effective_date,
        authority=authority,
        status=status,
        version=version,
        workspace_id=workspace_id,
        user_id=user_id,
        language=language,
        jurisdiction=jurisdiction,
        extra_metadata=json.loads(extra_metadata_json) if extra_metadata_json else None,
    )
    return {
        "saved": True,
        "target": "rag",
        "document": local["document"],
        "file": {"filename": uploaded["filename"], "size_bytes": uploaded["size_bytes"], "text_length": len(uploaded["text"])},
        "chunks": len(chunks),
        "supabase": supabase,
        "next_step": "Hỏi lại bằng POST /ai/v67/supabase-rag/search hoặc /ai/v66/file-upload-router target=solve.",
    }


@app.post("/ai/v66/solve/upload-question-file")
async def ai_v66_solve_upload_question_file(
    file: UploadFile = File(...),
    question: str = Form(default="Giải thích nội dung file này và trả lời chi tiết."),
    standard: str = Form(default=""),
    category: Optional[str] = Form(default=None),
    use_rag: bool = Form(default=True),
    require_sources: bool = Form(default=True),
    save_learning: bool = Form(default=False),
    _: bool = Depends(finiip_api_guard),
):
    uploaded = await _read_uploaded_file(file)
    rag = search_supabase_rag(question + "\n" + uploaded["text"][:2000], limit=5, category=category) if use_rag else {"matched": 0, "results": []}
    answer_parts = [
        "Đã đọc file upload. Đây là bản xử lý một lần, không lưu file vào RAG.",
        f"Câu hỏi: {question}",
    ]
    if standard:
        answer_parts.append(f"Chuẩn/quy định người dùng chọn: {standard}")
    answer_parts.append("\nNội dung file trích được, phần đầu:")
    preview = uploaded["text"][:2500].strip()
    answer_parts.append(preview + ("..." if len(uploaded["text"]) > 2500 else ""))
    if use_rag:
        answer_parts.append("\nTài liệu RAG liên quan:")
        answer_parts.append(rag.get("answer", "Không có kết quả RAG."))
    answer_parts.append("\nGợi ý frontend: gửi text này sang solver/LLM để tạo lời giải đầy đủ nếu cần suy luận dài.")
    return {
        "target": "solve",
        "saved_to_rag": False,
        "file": {"filename": uploaded["filename"], "size_bytes": uploaded["size_bytes"], "text_length": len(uploaded["text"])},
        "question": question,
        "standard": standard,
        "use_rag": use_rag,
        "rag": rag if require_sources else {"matched": rag.get("matched", 0)},
        "answer": "\n".join(answer_parts),
        "save_learning": save_learning,
    }


@app.post("/ai/v66/file-upload-router")
async def ai_v66_file_upload_router(
    file: UploadFile = File(...),
    target: str = Form(default="auto"),
    question: Optional[str] = Form(default=None),
    title: Optional[str] = Form(default=None),
    category: str = Form(default="general"),
    document_type: str = Form(default="document"),
    source: str = Form(default="router_upload"),
    tags: Optional[str] = Form(default=""),
    uploaded_by: str = Form(default="admin"),
    document_number: Optional[str] = Form(default=None),
    issued_date: Optional[str] = Form(default=None),
    effective_date: Optional[str] = Form(default=None),
    authority: Optional[str] = Form(default=None),
    status: str = Form(default="active"),
    version: Optional[str] = Form(default=None),
    workspace_id: Optional[str] = Form(default=None),
    user_id: Optional[str] = Form(default=None),
    language: str = Form(default="vi"),
    jurisdiction: Optional[str] = Form(default=None),
    extra_metadata_json: Optional[str] = Form(default=None),
    _: bool = Depends(finiip_api_guard),
):
    target_clean = (target or "auto").lower().strip()
    if target_clean == "auto":
        target_clean = _guess_upload_target(file.filename or "", document_type, question)
    if target_clean not in {"rag", "solve", "temp"}:
        raise HTTPException(status_code=400, detail="target phải là auto, rag, solve hoặc temp")
    uploaded = await _read_uploaded_file(file)
    if target_clean == "temp":
        return {
            "target": "temp",
            "saved_to_rag": False,
            "file": {"filename": uploaded["filename"], "size_bytes": uploaded["size_bytes"], "text_length": len(uploaded["text"])},
            "text_preview": uploaded["text"][:3000],
        }
    if target_clean == "solve":
        rag = search_supabase_rag((question or "") + "\n" + uploaded["text"][:2000], limit=5, category=category)
        return {
            "target": "solve",
            "saved_to_rag": False,
            "question": question or "Giải thích nội dung file này.",
            "file": {"filename": uploaded["filename"], "size_bytes": uploaded["size_bytes"], "text_length": len(uploaded["text"])},
            "text_preview": uploaded["text"][:2500],
            "rag": rag,
        }
    # target rag
    tag_list = v66_split_tags(tags)
    chunks = v66_chunk_text(uploaded["text"])
    local = save_local_rag_document(
        title=title or uploaded["filename"],
        content=uploaded["text"],
        filename=uploaded["filename"],
        category=category,
        document_type=document_type,
        source=source,
        tags=tag_list,
        storage_path=uploaded["storage_path"],
        document_number=document_number,
        issued_date=issued_date,
        effective_date=effective_date,
        authority=authority,
        status=status,
        version=version,
        workspace_id=workspace_id,
        user_id=user_id,
        language=language,
        jurisdiction=jurisdiction,
        extra_metadata=json.loads(extra_metadata_json) if extra_metadata_json else None,
    )
    supabase = save_supabase_rag_document(
        title=title or uploaded["filename"],
        chunks=chunks,
        filename=uploaded["filename"],
        category=category,
        document_type=document_type,
        source=source,
        tags=tag_list,
        storage_path=uploaded["storage_path"],
        uploaded_by=uploaded_by,
        document_number=document_number,
        issued_date=issued_date,
        effective_date=effective_date,
        authority=authority,
        status=status,
        version=version,
        workspace_id=workspace_id,
        user_id=user_id,
        language=language,
        jurisdiction=jurisdiction,
        extra_metadata=json.loads(extra_metadata_json) if extra_metadata_json else None,
    )
    return {"target": "rag", "saved_to_rag": True, "document": local["document"], "chunks": len(chunks), "supabase": supabase}


@app.get("/ai/v67/supabase-rag/status")
def ai_v67_supabase_rag_status(_: bool = Depends(finiip_api_guard)):
    return v67_supabase_status()


@app.post("/ai/v67/supabase-rag/search")
def ai_v67_supabase_rag_search(payload: V67RAGSearchRequest, _: bool = Depends(finiip_api_guard)):
    return search_supabase_rag(payload.question, limit=payload.limit, category=payload.category)


@app.get("/ai/v5-v9/capabilities")
def ai_v5_v9_capabilities():
    return {
        "version": "V5-V9 AI upgrades",
        "capabilities": [
            "V5 AI học từ correction/feedback của người dùng",
            "V6 OCR invoice text -> transaction draft -> AI analyze",
            "V7 lightweight RAG/knowledge QA cho tài liệu kế toán nội bộ",
            "V8 anomaly scoring cho danh sách giao dịch/import sao kê",
            "V9 optional X-API-Key guard cho frontend riêng",
        ],
        "main_endpoints": [
            "POST /ai/v5/feedback",
            "POST /ai/v5/analyze-with-learning",
            "GET  /ai/v5/learning-rules",
            "POST /ai/v6/invoice-to-transaction",
            "POST /ai/v7/knowledge/upload-text",
            "POST /ai/v7/ask",
            "POST /ai/v8/anomaly-score",
            "GET  /ai/v9/security-status",
        ],
    }


# ============================================================
# V68/V69/V70/V71 - Backend RAG production endpoints, no frontend required
# ============================================================
# V68: embedding/pgvector/schema/health
# V69: hỏi đáp RAG hoàn chỉnh có nguồn
# V70: quản lý tài liệu RAG
# V71: guard/limit/status cho deploy an toàn hơn

import services.rag_v66_v67 as rag68


class V68SchemaInitRequest(BaseModel):
    confirm: bool = Field(default=True, description="true để tạo/nâng schema RAG trong Supabase")


class V69RAGAnswerRequest(BaseModel):
    question: str = Field(..., min_length=3)
    limit: int = Field(default=6, ge=1, le=12)
    category: Optional[str] = None
    workspace_id: Optional[str] = None
    user_id: Optional[str] = None
    status: Optional[str] = Field(default=None, description="Ví dụ: active, draft, archived. Để trống nếu muốn tìm mọi trạng thái.")
    search_mode: str = Field(default="hybrid", description="hybrid | vector | keyword")
    vector_weight: float = Field(default=0.60, ge=0.0, le=1.0)
    keyword_weight: float = Field(default=0.40, ge=0.0, le=1.0)
    style: str = Field(default="detailed")
    use_llm: bool = Field(default=False, description="true nếu đã cấu hình OPENAI_API_KEY")


class V70RAGReindexRequest(BaseModel):
    chunk_size: int = Field(default=1200, ge=500, le=3000)
    overlap: int = Field(default=180, ge=0, le=600)


@app.get("/ai/v68/rag/health")
def ai_v68_rag_health(_: bool = Depends(finiip_api_guard)):
    return rag68.rag_health()


@app.post("/ai/v68/rag/init-schema")
def ai_v68_rag_init_schema(payload: V68SchemaInitRequest, _: bool = Depends(finiip_api_guard)):
    if not payload.confirm:
        raise HTTPException(status_code=400, detail="confirm phải bằng true để init schema")
    return rag68.ensure_supabase_rag_schema()


@app.get("/ai/v68/rag/embedding-status")
def ai_v68_rag_embedding_status(_: bool = Depends(finiip_api_guard)):
    return rag68.embedding_provider()


@app.post("/ai/v69/rag/answer")
def ai_v69_rag_answer(payload: V69RAGAnswerRequest, _: bool = Depends(finiip_api_guard)):
    return rag68.answer_rag_question(
        question=payload.question,
        limit=payload.limit,
        category=payload.category,
        workspace_id=payload.workspace_id,
        user_id=payload.user_id,
        status=payload.status,
        search_mode=payload.search_mode,
        vector_weight=payload.vector_weight,
        keyword_weight=payload.keyword_weight,
        style=payload.style,
        use_llm=payload.use_llm,
    )


@app.get("/ai/v70/rag/documents")
def ai_v70_rag_documents(
    limit: int = Query(default=50, ge=1, le=200),
    offset: int = Query(default=0, ge=0),
    category: Optional[str] = Query(default=None),
    source: Optional[str] = Query(default=None),
    workspace_id: Optional[str] = Query(default=None),
    user_id: Optional[str] = Query(default=None),
    status: Optional[str] = Query(default=None),
    _: bool = Depends(finiip_api_guard),
):
    return rag68.list_rag_documents(
        limit=limit,
        offset=offset,
        category=category,
        source=source,
        workspace_id=workspace_id,
        user_id=user_id,
        status=status,
    )


@app.get("/ai/v70/rag/documents/{document_id}")
def ai_v70_rag_document_detail(
    document_id: str,
    include_chunks: bool = Query(default=True),
    chunk_limit: int = Query(default=100, ge=0, le=500),
    workspace_id: Optional[str] = Query(default=None),
    user_id: Optional[str] = Query(default=None),
    _: bool = Depends(finiip_api_guard),
):
    result = rag68.get_rag_document(document_id, include_chunks=include_chunks, chunk_limit=chunk_limit, workspace_id=workspace_id, user_id=user_id)
    if not result.get("document") and not result.get("error"):
        raise HTTPException(status_code=404, detail="Không tìm thấy document")
    return result


@app.delete("/ai/v70/rag/documents/{document_id}")
def ai_v70_rag_document_delete(
    document_id: str,
    workspace_id: Optional[str] = Query(default=None),
    user_id: Optional[str] = Query(default=None),
    _: bool = Depends(finiip_api_guard),
):
    return rag68.delete_rag_document(document_id, workspace_id=workspace_id, user_id=user_id)


@app.post("/ai/v70/rag/documents/{document_id}/reindex")
def ai_v70_rag_document_reindex(
    document_id: str,
    payload: V70RAGReindexRequest,
    workspace_id: Optional[str] = Query(default=None),
    user_id: Optional[str] = Query(default=None),
    _: bool = Depends(finiip_api_guard),
):
    return rag68.reindex_rag_document(document_id, chunk_size=payload.chunk_size, overlap=payload.overlap, workspace_id=workspace_id, user_id=user_id)


@app.get("/ai/v71/security/deploy-check")
def ai_v71_security_deploy_check(_: bool = Depends(finiip_api_guard)):
    return {
        "api_key_required_when_FINIIP_API_KEY_set": bool(os.getenv("FINIIP_API_KEY")),
        "cors_origins": _get_cors_origins(),
        "cors_currently_allows_all_origins": "*" in _get_cors_origins(),
        "recommendation": "Khi deploy thật, đặt CORS_ORIGINS thành domain frontend riêng, ví dụ https://ten-app.vercel.app.",
        "rag_limits": {
            "max_upload_bytes": rag68.MAX_UPLOAD_BYTES,
            "allowed_extensions": sorted(rag68.ALLOWED_RAG_EXTENSIONS),
        },
        "required_env_for_rag": ["DATABASE_URL"],
        "v82_v83_enabled": {
            "workspace_user_scope": True,
            "hybrid_search": "vector + keyword + rerank",
            "answer_filters": ["category", "workspace_id", "user_id", "status"],
        },
        "optional_env_for_better_answers": ["OPENAI_API_KEY", "OPENAI_EMBEDDING_MODEL", "OPENAI_CHAT_MODEL", "FINIIP_API_KEY"],
    }



# ============================================================
# V84 - Frontend-ready RAG document management API, backend only
# ============================================================
# Stable endpoints for a separate frontend: dashboard/list/detail/upload/update/status/reindex/delete.

class V84DocumentPatchRequest(BaseModel):
    title: Optional[str] = None
    filename: Optional[str] = None
    document_type: Optional[str] = None
    category: Optional[str] = None
    source: Optional[str] = None
    uploaded_by: Optional[str] = None
    document_number: Optional[str] = None
    issued_date: Optional[str] = None
    effective_date: Optional[str] = None
    authority: Optional[str] = None
    status: Optional[str] = None
    version: Optional[str] = None
    workspace_id: Optional[str] = None
    user_id: Optional[str] = None
    language: Optional[str] = None
    jurisdiction: Optional[str] = None
    tags: Optional[List[str]] = None
    metadata: Optional[Dict[str, Any]] = None


class V84DocumentStatusRequest(BaseModel):
    status: str = Field(..., description="active | draft | archived | replaced | inactive")


@app.get("/ai/v84/rag/dashboard/status")
def ai_v84_rag_dashboard_status(
    workspace_id: Optional[str] = Query(default=None),
    user_id: Optional[str] = Query(default=None),
    category: Optional[str] = Query(default=None),
    _: bool = Depends(finiip_api_guard),
):
    return rag68.rag_document_status_summary(workspace_id=workspace_id, user_id=user_id, category=category)


@app.post("/ai/v84/rag/documents/upload")
async def ai_v84_rag_documents_upload(
    file: UploadFile = File(...),
    title: Optional[str] = Form(default=None),
    category: str = Form(default="general"),
    document_type: str = Form(default="document"),
    source: str = Form(default="frontend_upload"),
    tags: Optional[str] = Form(default=""),
    uploaded_by: str = Form(default="admin"),
    document_number: Optional[str] = Form(default=None),
    issued_date: Optional[str] = Form(default=None),
    effective_date: Optional[str] = Form(default=None),
    authority: Optional[str] = Form(default=None),
    status: str = Form(default="active"),
    version: Optional[str] = Form(default=None),
    workspace_id: Optional[str] = Form(default=None),
    user_id: Optional[str] = Form(default=None),
    language: str = Form(default="vi"),
    jurisdiction: Optional[str] = Form(default=None),
    extra_metadata_json: Optional[str] = Form(default=None),
    auto_chunk: bool = Form(default=True),
    _: bool = Depends(finiip_api_guard),
):
    uploaded = await _read_uploaded_file(file)
    doc_title = title or uploaded["filename"]
    tag_list = v66_split_tags(tags)
    try:
        extra_metadata = json.loads(extra_metadata_json) if extra_metadata_json else None
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"extra_metadata_json không phải JSON hợp lệ: {exc}") from exc
    chunks = v66_chunk_text(uploaded["text"]) if auto_chunk else [uploaded["text"]]
    if not chunks:
        raise HTTPException(status_code=400, detail="File không có nội dung đủ để chunk")
    local = save_local_rag_document(
        title=doc_title,
        content=uploaded["text"],
        filename=uploaded["filename"],
        category=category,
        document_type=document_type,
        source=source,
        tags=tag_list,
        storage_path=uploaded["storage_path"],
        document_number=document_number,
        issued_date=issued_date,
        effective_date=effective_date,
        authority=authority,
        status=status,
        version=version,
        workspace_id=workspace_id,
        user_id=user_id,
        language=language,
        jurisdiction=jurisdiction,
        extra_metadata=extra_metadata,
    )
    supabase = save_supabase_rag_document(
        title=doc_title,
        chunks=chunks,
        filename=uploaded["filename"],
        category=category,
        document_type=document_type,
        source=source,
        tags=tag_list,
        storage_path=uploaded["storage_path"],
        uploaded_by=uploaded_by,
        document_number=document_number,
        issued_date=issued_date,
        effective_date=effective_date,
        authority=authority,
        status=status,
        version=version,
        workspace_id=workspace_id,
        user_id=user_id,
        language=language,
        jurisdiction=jurisdiction,
        extra_metadata=extra_metadata,
    )
    frontend_doc = rag68.normalize_document_for_frontend(local["document"])
    return {
        "ok": True,
        "version": "v84_frontend_document_api",
        "saved": True,
        "target": "rag",
        "document": frontend_doc,
        "file": {"filename": uploaded["filename"], "size_bytes": uploaded["size_bytes"], "text_length": len(uploaded["text"])},
        "chunks": len(chunks),
        "supabase": supabase,
        "frontend_next": ["refresh /ai/v84/rag/documents", "show /ai/v84/rag/dashboard/status"],
    }


@app.get("/ai/v84/rag/documents")
def ai_v84_rag_documents(
    limit: int = Query(default=50, ge=1, le=200),
    offset: int = Query(default=0, ge=0),
    q: Optional[str] = Query(default=None),
    category: Optional[str] = Query(default=None),
    source: Optional[str] = Query(default=None),
    workspace_id: Optional[str] = Query(default=None),
    user_id: Optional[str] = Query(default=None),
    status: Optional[str] = Query(default=None),
    _: bool = Depends(finiip_api_guard),
):
    return rag68.list_rag_documents_frontend(
        limit=limit,
        offset=offset,
        q=q,
        category=category,
        source=source,
        workspace_id=workspace_id,
        user_id=user_id,
        status=status,
    )


@app.get("/ai/v84/rag/documents/{document_id}")
def ai_v84_rag_document_detail(
    document_id: str,
    include_chunks: bool = Query(default=True),
    chunk_limit: int = Query(default=100, ge=0, le=500),
    workspace_id: Optional[str] = Query(default=None),
    user_id: Optional[str] = Query(default=None),
    _: bool = Depends(finiip_api_guard),
):
    result = rag68.get_rag_document(document_id, include_chunks=include_chunks, chunk_limit=chunk_limit, workspace_id=workspace_id, user_id=user_id)
    if not result.get("document") and not result.get("error"):
        raise HTTPException(status_code=404, detail="Không tìm thấy document")
    if result.get("document"):
        result["document"] = rag68.normalize_document_for_frontend(result["document"])
    result["version"] = "v84_frontend_document_api"
    return result


@app.patch("/ai/v84/rag/documents/{document_id}")
def ai_v84_rag_document_patch(
    document_id: str,
    payload: V84DocumentPatchRequest,
    workspace_id: Optional[str] = Query(default=None),
    user_id: Optional[str] = Query(default=None),
    _: bool = Depends(finiip_api_guard),
):
    updates = payload.model_dump(exclude_none=True)
    return rag68.update_rag_document_metadata(document_id, updates, workspace_id=workspace_id, user_id=user_id)


@app.post("/ai/v84/rag/documents/{document_id}/status")
def ai_v84_rag_document_status(
    document_id: str,
    payload: V84DocumentStatusRequest,
    workspace_id: Optional[str] = Query(default=None),
    user_id: Optional[str] = Query(default=None),
    _: bool = Depends(finiip_api_guard),
):
    return rag68.set_rag_document_status(document_id, payload.status, workspace_id=workspace_id, user_id=user_id)


@app.post("/ai/v84/rag/documents/{document_id}/reindex")
def ai_v84_rag_document_reindex(
    document_id: str,
    payload: V70RAGReindexRequest,
    workspace_id: Optional[str] = Query(default=None),
    user_id: Optional[str] = Query(default=None),
    _: bool = Depends(finiip_api_guard),
):
    result = rag68.reindex_rag_document(document_id, chunk_size=payload.chunk_size, overlap=payload.overlap, workspace_id=workspace_id, user_id=user_id)
    result["version"] = "v84_frontend_document_api"
    return result


@app.delete("/ai/v84/rag/documents/{document_id}")
def ai_v84_rag_document_delete(
    document_id: str,
    workspace_id: Optional[str] = Query(default=None),
    user_id: Optional[str] = Query(default=None),
    _: bool = Depends(finiip_api_guard),
):
    result = rag68.delete_rag_document(document_id, workspace_id=workspace_id, user_id=user_id)
    result["version"] = "v84_frontend_document_api"
    return result


@app.get("/ai/v84/rag/frontend-contract")
def ai_v84_rag_frontend_contract(_: bool = Depends(finiip_api_guard)):
    return {
        "version": "V84 Frontend Document API",
        "base": "/ai/v84/rag",
        "auth_header": "X-API-Key nếu FINIIP_API_KEY được set",
        "endpoints": [
            "GET    /ai/v84/rag/dashboard/status",
            "POST   /ai/v84/rag/documents/upload",
            "GET    /ai/v84/rag/documents",
            "GET    /ai/v84/rag/documents/{document_id}",
            "PATCH  /ai/v84/rag/documents/{document_id}",
            "POST   /ai/v84/rag/documents/{document_id}/status",
            "POST   /ai/v84/rag/documents/{document_id}/reindex",
            "DELETE /ai/v84/rag/documents/{document_id}",
        ],
        "document_table_columns": ["title", "filename", "category", "status", "document_number", "authority", "chunk_count", "updated_at"],
        "filters": ["q", "workspace_id", "user_id", "category", "source", "status", "limit", "offset"],
        "allowed_statuses": ["active", "draft", "archived", "replaced", "inactive"],
    }

# ============================================================
# V72/V73/V74/V75 - Pro backend RAG upgrades, still no frontend
# ============================================================
# V72: structure-aware chunking, pro schema, batch upload
# V73: hybrid vector + keyword retrieval + rerank
# V74: cited answer + quality check
# V75: audit logs + pro health

import services.rag_pro as ragpro


class V72ProSchemaInitRequest(BaseModel):
    confirm: bool = Field(default=True)


class V73HybridSearchRequest(BaseModel):
    question: str = Field(..., min_length=3)
    limit: int = Field(default=8, ge=1, le=20)
    category: Optional[str] = None
    document_type: Optional[str] = None
    tags: Optional[List[str]] = None


class V74ProAnswerRequest(BaseModel):
    question: str = Field(..., min_length=3)
    limit: int = Field(default=8, ge=1, le=20)
    category: Optional[str] = None
    document_type: Optional[str] = None
    tags: Optional[List[str]] = None
    style: str = Field(default="detailed")
    use_llm: bool = Field(default=False)


class V74QualityCheckRequest(BaseModel):
    questions: List[str] = Field(..., min_length=1, max_length=50)
    category: Optional[str] = None
    limit: int = Field(default=5, ge=1, le=12)


@app.get("/ai/v72/pro-rag/health")
def ai_v72_pro_rag_health(_: bool = Depends(finiip_api_guard)):
    return ragpro.pro_health()


@app.post("/ai/v72/pro-rag/init-schema")
def ai_v72_pro_rag_init_schema(payload: V72ProSchemaInitRequest, _: bool = Depends(finiip_api_guard)):
    if not payload.confirm:
        raise HTTPException(status_code=400, detail="confirm phải bằng true")
    return ragpro.ensure_pro_schema()


@app.post("/ai/v72/pro-rag/upload-file")
async def ai_v72_pro_rag_upload_file(
    file: UploadFile = File(...),
    title: Optional[str] = Form(default=None),
    category: str = Form(default="general"),
    document_type: str = Form(default="document"),
    tags: str = Form(default=""),
    source: str = Form(default="admin_upload"),
    uploaded_by: str = Form(default="admin"),
    chunk_size: int = Form(default=1600),
    overlap: int = Form(default=220),
    _: bool = Depends(finiip_api_guard),
):
    uploaded = await _read_uploaded_file(file)
    tag_list = v66_split_tags(tags)
    result = ragpro.save_pro_document(
        title=title or uploaded["filename"],
        content=uploaded["text"],
        filename=uploaded["filename"],
        category=category,
        document_type=document_type,
        source=source,
        tags=tag_list,
        storage_path=uploaded["storage_path"],
        uploaded_by=uploaded_by,
        chunk_size=chunk_size,
        overlap=overlap,
    )
    return {
        "target": "rag_pro",
        "saved_to_rag": True,
        "filename": uploaded["filename"],
        "size_bytes": uploaded["size_bytes"],
        **result,
    }


@app.post("/ai/v72/pro-rag/upload-batch")
async def ai_v72_pro_rag_upload_batch(
    files: List[UploadFile] = File(...),
    category: str = Form(default="general"),
    document_type: str = Form(default="document"),
    tags: str = Form(default=""),
    source: str = Form(default="admin_batch_upload"),
    uploaded_by: str = Form(default="admin"),
    chunk_size: int = Form(default=1600),
    overlap: int = Form(default=220),
    _: bool = Depends(finiip_api_guard),
):
    if len(files) > ragpro.MAX_BATCH_FILES:
        raise HTTPException(status_code=400, detail=f"Tối đa {ragpro.MAX_BATCH_FILES} file/lần upload")
    tag_list = v66_split_tags(tags)
    items = []
    ok = 0
    failed = 0
    for file in files:
        try:
            uploaded = await _read_uploaded_file(file)
            result = ragpro.save_pro_document(
                title=uploaded["filename"],
                content=uploaded["text"],
                filename=uploaded["filename"],
                category=category,
                document_type=document_type,
                source=source,
                tags=tag_list,
                storage_path=uploaded["storage_path"],
                uploaded_by=uploaded_by,
                chunk_size=chunk_size,
                overlap=overlap,
            )
            ok += 1
            items.append({"filename": uploaded["filename"], "ok": True, **result})
        except Exception as exc:
            failed += 1
            items.append({"filename": getattr(file, "filename", "upload"), "ok": False, "error": str(exc)})
    return {"total": len(files), "ok": ok, "failed": failed, "items": items}


@app.post("/ai/v73/pro-rag/hybrid-search")
def ai_v73_pro_rag_hybrid_search(payload: V73HybridSearchRequest, _: bool = Depends(finiip_api_guard)):
    return ragpro.hybrid_search(
        question=payload.question,
        limit=payload.limit,
        category=payload.category,
        document_type=payload.document_type,
        tags=payload.tags,
    )


@app.post("/ai/v74/pro-rag/answer")
def ai_v74_pro_rag_answer(payload: V74ProAnswerRequest, _: bool = Depends(finiip_api_guard)):
    return ragpro.answer_pro(
        question=payload.question,
        limit=payload.limit,
        category=payload.category,
        document_type=payload.document_type,
        tags=payload.tags,
        style=payload.style,
        use_llm=payload.use_llm,
    )


@app.post("/ai/v74/pro-rag/quality-check")
def ai_v74_pro_rag_quality_check(payload: V74QualityCheckRequest, _: bool = Depends(finiip_api_guard)):
    return ragpro.rag_quality_check(payload.questions, category=payload.category, limit=payload.limit)


@app.get("/ai/v75/pro-rag/audit-logs")
def ai_v75_pro_rag_audit_logs(
    limit: int = Query(default=100, ge=1, le=1000),
    action: Optional[str] = Query(default=None),
    status: Optional[str] = Query(default=None),
    _: bool = Depends(finiip_api_guard),
):
    return ragpro.list_audit_logs(limit=limit, action=action, status=status)


# ============================================================
# V76/V77/V78/V79/V80 - Full Pro backend RAG upgrades, no frontend
# ============================================================
# V76: multi-turn RAG chat sessions
# V77: query cache + metrics
# V78: local/worker-ready job queue
# V79: document versioning + soft delete/restore
# V80: full-pro health + role-aware admin guard


def fullpro_role_guard(request: Request, _: bool = Depends(finiip_api_guard)):
    """Role guard for full-pro endpoints.

    In production set RAG_ADMIN_TOKEN. Then send header:
    X-RAG-Admin-Token: <token>

    Optional role header for internal systems:
    X-User-Role: admin|user|viewer
    """
    admin_token = os.getenv("RAG_ADMIN_TOKEN")
    provided_token = request.headers.get("X-RAG-Admin-Token")
    role = (request.headers.get("X-User-Role") or "admin").lower()
    if admin_token and provided_token != admin_token:
        raise HTTPException(status_code=403, detail="Thiếu hoặc sai X-RAG-Admin-Token")
    if role not in {"admin", "user", "viewer"}:
        raise HTTPException(status_code=403, detail="X-User-Role không hợp lệ")
    return {"role": role, "admin_token_enabled": bool(admin_token)}


class V76ChatAnswerRequest(BaseModel):
    session_id: Optional[str] = Field(default=None)
    question: str = Field(..., min_length=3)
    limit: int = Field(default=8, ge=1, le=20)
    category: Optional[str] = None
    document_type: Optional[str] = None
    tags: Optional[List[str]] = None
    style: str = Field(default="detailed")
    use_llm: bool = Field(default=False)


class V77AdvancedAnswerRequest(BaseModel):
    question: str = Field(..., min_length=3)
    limit: int = Field(default=8, ge=1, le=20)
    category: Optional[str] = None
    document_type: Optional[str] = None
    tags: Optional[List[str]] = None
    style: str = Field(default="detailed")
    bypass_cache: bool = Field(default=False)


class V78JobCreateRequest(BaseModel):
    job_type: str = Field(..., min_length=3)
    payload: Dict[str, Any] = Field(default_factory=dict)


class V79SoftDeleteRequest(BaseModel):
    reason: str = Field(default="manual_delete")


class V79RegisterVersionRequest(BaseModel):
    title: str = Field(..., min_length=1)
    content_sha256: Optional[str] = None
    metadata: Dict[str, Any] = Field(default_factory=dict)


@app.get("/ai/v80/full-pro-rag/health")
def ai_v80_full_pro_rag_health(_: Dict[str, Any] = Depends(fullpro_role_guard)):
    return ragpro.full_pro_health()


@app.post("/ai/v80/full-pro-rag/init-schema")
def ai_v80_full_pro_rag_init_schema(payload: V72ProSchemaInitRequest, _: Dict[str, Any] = Depends(fullpro_role_guard)):
    if not payload.confirm:
        raise HTTPException(status_code=400, detail="confirm phải bằng true")
    return ragpro.ensure_full_pro_schema()


@app.post("/ai/v76/full-pro-rag/chat/answer")
def ai_v76_full_pro_chat_answer(payload: V76ChatAnswerRequest, _: Dict[str, Any] = Depends(fullpro_role_guard)):
    session_id = payload.session_id or f"session_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{os.urandom(4).hex()}"
    return ragpro.chat_answer(
        session_id=session_id,
        question=payload.question,
        limit=payload.limit,
        category=payload.category,
        document_type=payload.document_type,
        tags=payload.tags,
        style=payload.style,
        use_llm=payload.use_llm,
    )


@app.get("/ai/v76/full-pro-rag/chat/sessions")
def ai_v76_full_pro_chat_sessions(limit: int = Query(default=50, ge=1, le=200), _: Dict[str, Any] = Depends(fullpro_role_guard)):
    return ragpro.list_chat_sessions(limit=limit)


@app.get("/ai/v76/full-pro-rag/chat/sessions/{session_id}")
def ai_v76_full_pro_chat_session_detail(session_id: str, _: Dict[str, Any] = Depends(fullpro_role_guard)):
    return ragpro.get_chat_session(session_id)


@app.delete("/ai/v76/full-pro-rag/chat/sessions/{session_id}")
def ai_v76_full_pro_chat_session_reset(session_id: str, _: Dict[str, Any] = Depends(fullpro_role_guard)):
    return ragpro.reset_chat_session(session_id)


@app.post("/ai/v77/full-pro-rag/answer")
def ai_v77_full_pro_advanced_answer(payload: V77AdvancedAnswerRequest, _: Dict[str, Any] = Depends(fullpro_role_guard)):
    return ragpro.advanced_answer(
        question=payload.question,
        limit=payload.limit,
        category=payload.category,
        document_type=payload.document_type,
        tags=payload.tags,
        style=payload.style,
        bypass_cache=payload.bypass_cache,
    )


@app.post("/ai/v77/full-pro-rag/cache/clear")
def ai_v77_full_pro_cache_clear(_: Dict[str, Any] = Depends(fullpro_role_guard)):
    return ragpro.clear_query_cache()


@app.get("/ai/v77/full-pro-rag/metrics")
def ai_v77_full_pro_metrics(limit: int = Query(default=1000, ge=1, le=10000), _: Dict[str, Any] = Depends(fullpro_role_guard)):
    return ragpro.metrics_summary(limit=limit)


@app.post("/ai/v78/full-pro-rag/jobs")
def ai_v78_full_pro_create_job(payload: V78JobCreateRequest, _: Dict[str, Any] = Depends(fullpro_role_guard)):
    return ragpro.create_job(payload.job_type, payload.payload)


@app.get("/ai/v78/full-pro-rag/jobs")
def ai_v78_full_pro_list_jobs(
    limit: int = Query(default=100, ge=1, le=500),
    status: Optional[str] = Query(default=None),
    _: Dict[str, Any] = Depends(fullpro_role_guard),
):
    return ragpro.list_jobs(limit=limit, status=status)


@app.get("/ai/v78/full-pro-rag/jobs/{job_id}")
def ai_v78_full_pro_get_job(job_id: str, _: Dict[str, Any] = Depends(fullpro_role_guard)):
    return ragpro.get_job(job_id)


@app.post("/ai/v78/full-pro-rag/jobs/{job_id}/run")
def ai_v78_full_pro_run_job(job_id: str, _: Dict[str, Any] = Depends(fullpro_role_guard)):
    return ragpro.run_job(job_id)


@app.delete("/ai/v79/full-pro-rag/documents/{document_id}")
def ai_v79_full_pro_soft_delete_document(
    document_id: str,
    payload: V79SoftDeleteRequest = V79SoftDeleteRequest(),
    _: Dict[str, Any] = Depends(fullpro_role_guard),
):
    return ragpro.soft_delete_document(document_id, reason=payload.reason)


@app.post("/ai/v79/full-pro-rag/documents/{document_id}/restore")
def ai_v79_full_pro_restore_document(document_id: str, _: Dict[str, Any] = Depends(fullpro_role_guard)):
    return ragpro.restore_document(document_id)


@app.get("/ai/v79/full-pro-rag/document-versions")
def ai_v79_full_pro_list_versions(
    document_id: Optional[str] = Query(default=None),
    limit: int = Query(default=100, ge=1, le=500),
    _: Dict[str, Any] = Depends(fullpro_role_guard),
):
    return ragpro.list_document_versions(document_id=document_id, limit=limit)


@app.post("/ai/v79/full-pro-rag/documents/{document_id}/versions")
def ai_v79_full_pro_register_version(
    document_id: str,
    payload: V79RegisterVersionRequest,
    _: Dict[str, Any] = Depends(fullpro_role_guard),
):
    return ragpro.register_document_version(document_id=document_id, title=payload.title, content_sha256=payload.content_sha256, metadata=payload.metadata)


# ============================================================
# V85 - FINIIP FULL ACCOUNTING AI CORE
# ============================================================
# Backend-only upgrade: broad accounting rules, journal suggestion,
# formula solver, risk gate, local knowledge-base answer.

from services.accounting_ai_full import (
    analyze_transaction as v85_analyze_transaction,
    ask_accounting_ai as v85_ask_accounting_ai,
    capability_matrix as v85_capability_matrix,
    journal_totals as v85_journal_totals,
    rule_catalog as v85_rule_catalog,
    solve_formula as v85_solve_formula,
)


class V85AccountingAnalyzeRequest(BaseModel):
    description: str = Field(..., min_length=3)
    amount: Optional[float] = Field(default=None, description="Số tiền giao dịch. Nếu bỏ trống AI sẽ cố parse từ mô tả.")
    vat_rate: Optional[float] = Field(default=None, description="0.1 cho 10%, 0.08 cho 8%. Nếu bỏ trống AI sẽ parse từ mô tả nếu có.")
    amount_includes_vat: bool = Field(default=True)
    has_invoice: Optional[bool] = Field(default=None)
    extra: Dict[str, Any] = Field(default_factory=dict)


class V85AccountingAskRequest(BaseModel):
    question: str = Field(..., min_length=1)
    limit: int = Field(default=5, ge=1, le=12)
    workspace_id: str = Field(default="default")
    conversation_id: str = Field(default="accounting_chat")
    history: str = Field(default="")
    answer_mode: str = Field(default="auto")
    save_memory: bool = Field(default=True)


class V85AccountingFormulaRequest(BaseModel):
    formula: str = Field(..., description="vat | depreciation | prepaid | weighted_average | fifo | profit | cit | payroll")
    amount: Optional[float] = None
    rate: Optional[float] = None
    amount_includes_vat: bool = False
    cost: Optional[float] = None
    residual_value: Optional[float] = None
    months: Optional[int] = None
    used_months: Optional[int] = None
    total_cost: Optional[float] = None
    begin_qty: Optional[float] = None
    begin_value: Optional[float] = None
    import_qty: Optional[float] = None
    import_value: Optional[float] = None
    export_qty: Optional[float] = None
    layers: List[Dict[str, float]] = Field(default_factory=list)
    revenue: Optional[float] = None
    cogs: Optional[float] = None
    selling_expenses: Optional[float] = None
    admin_expenses: Optional[float] = None
    financial_expenses: Optional[float] = None
    other_income: Optional[float] = None
    other_expenses: Optional[float] = None
    taxable_income: Optional[float] = None
    profit_before_tax: Optional[float] = None
    tax_rate: Optional[float] = None
    prepaid_tax: Optional[float] = None
    gross_salary: Optional[float] = None
    employee_insurance_rate: Optional[float] = None
    employer_insurance_rate: Optional[float] = None
    pit: Optional[float] = None


class V85JournalCheckRequest(BaseModel):
    lines: List[Dict[str, Any]] = Field(..., min_length=1)


@app.get("/ai/accounting/full-capabilities")
def ai_accounting_full_capabilities(_: bool = Depends(finiip_api_guard)):
    return v85_capability_matrix()


@app.get("/ai/accounting/rules")
def ai_accounting_rules(_: bool = Depends(finiip_api_guard)):
    return v85_rule_catalog()


@app.post("/ai/accounting/analyze-transaction")
def ai_accounting_analyze_transaction(payload: V85AccountingAnalyzeRequest, _: bool = Depends(finiip_api_guard)):
    return v85_analyze_transaction(
        description=payload.description,
        amount=payload.amount,
        vat_rate=payload.vat_rate,
        amount_includes_vat=payload.amount_includes_vat,
        has_invoice=payload.has_invoice,
        extra=payload.extra,
    )


@app.post("/ai/accounting/suggest-entry")
def ai_accounting_suggest_entry(payload: V85AccountingAnalyzeRequest, _: bool = Depends(finiip_api_guard)):
    result = v85_analyze_transaction(
        description=payload.description,
        amount=payload.amount,
        vat_rate=payload.vat_rate,
        amount_includes_vat=payload.amount_includes_vat,
        has_invoice=payload.has_invoice,
        extra=payload.extra,
    )
    return {
        "version": result.get("version"),
        "description": result.get("description"),
        "category": result.get("category"),
        "confidence": result.get("confidence"),
        "journal_lines": result.get("journal_lines"),
        "journal_check": result.get("journal_check"),
        "risk_review": result.get("risk_review"),
        "decision": result.get("decision"),
        "missing_questions": result.get("missing_questions"),
    }


@app.post("/ai/accounting/solve")
def ai_accounting_solve(payload: V85AccountingFormulaRequest, _: bool = Depends(finiip_api_guard)):
    try:
        return v85_solve_formula(payload.model_dump())
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@app.post("/ai/accounting/ask")
def ai_accounting_ask(payload: V85AccountingAskRequest, _: bool = Depends(finiip_api_guard)):
    """Backward-compatible smart accounting chat endpoint.

    Existing clients can keep sending only question + limit. New clients should
    also keep a stable conversation_id so follow-up questions are resolved.
    """
    return v101_answer_admin_rag_question(
        question=payload.question,
        workspace_id=payload.workspace_id,
        limit=payload.limit,
        history=payload.history,
        answer_mode=payload.answer_mode,
        conversation_id=payload.conversation_id,
        save_memory=payload.save_memory,
    )


@app.post("/ai/accounting/check-journal")
def ai_accounting_check_journal(payload: V85JournalCheckRequest, _: bool = Depends(finiip_api_guard)):
    result = v85_journal_totals(payload.lines)
    issues = []
    if not result.get("is_balanced"):
        issues.append("Bút toán chưa cân Nợ/Có")
    account_pairs = [(x.get("side"), x.get("account_code")) for x in payload.lines]
    debit_accounts = {a for side, a in account_pairs if side == "debit"}
    credit_accounts = {a for side, a in account_pairs if side == "credit"}
    if debit_accounts & credit_accounts and len(payload.lines) == 2:
        issues.append("Tài khoản Nợ/Có đang trùng nhau trong bút toán đơn")
    return {"journal_check": result, "issues": issues, "is_safe_to_post": result.get("is_balanced") and not issues}


@app.post("/api/v1/ai/accounting-preview")
def api_v1_accounting_preview(payload: V85AccountingAnalyzeRequest, _: bool = Depends(finiip_api_guard)):
    """Frontend-friendly accounting AI preview: draft entry + risk + missing fields."""
    result = v85_analyze_transaction(
        description=payload.description,
        amount=payload.amount,
        vat_rate=payload.vat_rate,
        amount_includes_vat=payload.amount_includes_vat,
        has_invoice=payload.has_invoice,
        extra=payload.extra,
    )
    return {
        "stage": API_STAGE,
        "preview_type": "accounting_ai_v85",
        "ai_result": result,
        "frontend_decision": {
            "can_create_draft": result.get("decision") in {"auto_draft_allowed", "review_required"},
            "can_auto_post": result.get("decision") == "auto_draft_allowed",
            "needs_review": result.get("decision") != "auto_draft_allowed",
            "next_api_to_create_transaction": "/ai/create-transaction",
            "next_api_to_feedback": "/ai/feedback",
        },
    }


# ============================================================
# V86-V99 - FINIIP ENTERPRISE ACCOUNTING AI LAYER
# ============================================================
# Adds document/RAG ingestion, invoice parser, journal export,
# review queue, tax risk, agent pipeline, workspace profiles,
# dashboard, database blueprint, frontend contract and production checks.

from services.accounting_ai_enterprise import (
    add_document as v86_add_document,
    add_uploaded_document as v86_add_uploaded_document,
    answer_with_enterprise_rag as v86_answer_with_enterprise_rag,
    closing_checklist as v97_closing_checklist,
    create_journal_entry as v88_create_journal_entry,
    create_or_update_workspace as v93_create_or_update_workspace,
    create_review_item as v89_create_review_item,
    database_schema_blueprint as v95_database_schema_blueprint,
    enterprise_capabilities as v86_v99_enterprise_capabilities,
    export_journal_csv as v88_export_journal_csv,
    export_journal_xlsx as v88_export_journal_xlsx,
    frontend_api_contract as v96_frontend_api_contract,
    list_company_memory as v98_list_company_memory,
    list_journal_entries as v88_list_journal_entries,
    list_review_queue as v89_list_review_queue,
    list_workspaces as v93_list_workspaces,
    monthly_summary_report as v97_monthly_summary_report,
    parse_invoice_text as v87_parse_invoice_text,
    production_readiness_check as v99_production_readiness_check,
    quality_dashboard as v94_quality_dashboard,
    remember_company_fact as v98_remember_company_fact,
    reset_enterprise_store as v86_reset_enterprise_store,
    run_accounting_agent_pipeline as v92_run_accounting_agent_pipeline,
    run_evaluation as v94_run_evaluation,
    search_documents as v86_search_documents,
    smart_followup_questions as v91_smart_followup_questions,
    tax_risk_check as v90_tax_risk_check,
    update_review_item as v89_update_review_item,
)


def _pydantic_payload(payload: BaseModel) -> Dict[str, Any]:
    return payload.model_dump() if hasattr(payload, "model_dump") else payload.dict()


class V86DocumentCreateRequest(BaseModel):
    title: str = Field(..., min_length=2)
    content: str = Field(..., min_length=1)
    workspace_id: str = "default"
    source_type: Optional[str] = None
    metadata: Dict[str, Any] = Field(default_factory=dict)


class V86SearchRequest(BaseModel):
    query: str = Field(..., min_length=1)
    workspace_id: str = "default"
    source_types: List[str] = Field(default_factory=list)
    limit: int = Field(default=6, ge=1, le=20)


class V86RagAskRequest(BaseModel):
    question: str = Field(..., min_length=2)
    workspace_id: str = "default"
    limit: int = Field(default=6, ge=1, le=12)


class V87InvoiceParseRequest(BaseModel):
    text: str = Field(..., min_length=1)


class V88JournalCreateRequest(BaseModel):
    description: str = Field(..., min_length=2)
    amount: Optional[float] = None
    vat_rate: Optional[float] = None
    workspace_id: str = "default"
    source_document_id: Optional[str] = None
    amount_includes_vat: bool = True
    has_invoice: Optional[bool] = None
    extra: Dict[str, Any] = Field(default_factory=dict)


class V88JournalExportRequest(BaseModel):
    workspace_id: str = "default"
    status: Optional[str] = None
    format: str = Field(default="csv", description="csv | xlsx")


class V89ReviewCreateRequest(BaseModel):
    workspace_id: str = "default"
    item_type: str = "manual_review"
    title: str = Field(..., min_length=2)
    payload: Dict[str, Any] = Field(default_factory=dict)
    risk_level: str = "medium"
    priority: Optional[str] = None


class V89ReviewUpdateRequest(BaseModel):
    status: str = Field(..., description="pending | approved | rejected | need_more_info | corrected | posted")
    reviewer_note: Optional[str] = None
    correction: Optional[Dict[str, Any]] = None


class V90RiskCheckRequest(BaseModel):
    workspace_id: str = "default"
    transaction: Dict[str, Any] = Field(default_factory=dict)


class V91FollowupRequest(BaseModel):
    workspace_id: str = "default"
    transaction: Dict[str, Any] = Field(default_factory=dict)


class V92AgentTextRequest(BaseModel):
    text: str = Field(..., min_length=1)
    workspace_id: str = "default"
    filename: Optional[str] = None
    create_review: bool = True


class V93WorkspaceRequest(BaseModel):
    workspace_id: str = Field(..., min_length=1)
    name: Optional[str] = None
    tax_code: Optional[str] = None
    policy: Dict[str, Any] = Field(default_factory=dict)
    chart_of_accounts: Optional[Dict[str, str]] = None
    users: Optional[List[Dict[str, Any]]] = None


class V94EvaluationRequest(BaseModel):
    workspace_id: str = "default"
    cases: Optional[List[Dict[str, Any]]] = None


class V98CompanyMemoryRequest(BaseModel):
    workspace_id: str = "default"
    fact: str = Field(..., min_length=2)
    category: str = "policy"
    source: str = "user"


@app.get("/ai/v86-v99/capabilities")
def ai_v86_v99_capabilities(_: bool = Depends(finiip_api_guard)):
    return v86_v99_enterprise_capabilities()


@app.post("/ai/v86/dev/reset-store")
def ai_v86_dev_reset_store(_: bool = Depends(finiip_api_guard)):
    return v86_reset_enterprise_store()


@app.post("/ai/v86/documents")
def ai_v86_add_document(payload: V86DocumentCreateRequest, _: bool = Depends(finiip_api_guard)):
    return v86_add_document(
        title=payload.title,
        content=payload.content,
        workspace_id=payload.workspace_id,
        source_type=payload.source_type,
        metadata=payload.metadata,
    )


@app.post("/ai/v86/documents/upload")
async def ai_v86_upload_document(
    file: UploadFile = File(...),
    workspace_id: str = Form(default="default"),
    title: Optional[str] = Form(default=None),
    source_type: Optional[str] = Form(default=None),
    _: bool = Depends(finiip_api_guard),
):
    content = await file.read()
    return v86_add_uploaded_document(
        filename=file.filename or "upload.bin",
        content=content,
        workspace_id=workspace_id,
        title=title,
        source_type=source_type,
        metadata={"content_type": file.content_type},
    )


@app.post("/ai/v86/documents/search")
def ai_v86_search_documents(payload: V86SearchRequest, _: bool = Depends(finiip_api_guard)):
    return v86_search_documents(
        query=payload.query,
        workspace_id=payload.workspace_id,
        source_types=payload.source_types,
        limit=payload.limit,
    )


@app.post("/ai/v86/rag/ask")
def ai_v86_rag_ask(payload: V86RagAskRequest, _: bool = Depends(finiip_api_guard)):
    return v86_answer_with_enterprise_rag(payload.question, workspace_id=payload.workspace_id, limit=payload.limit)


@app.post("/ai/v87/invoices/parse")
def ai_v87_parse_invoice(payload: V87InvoiceParseRequest, _: bool = Depends(finiip_api_guard)):
    return v87_parse_invoice_text(payload.text)


@app.post("/ai/v88/journal/create")
def ai_v88_create_journal(payload: V88JournalCreateRequest, _: bool = Depends(finiip_api_guard)):
    return v88_create_journal_entry(
        description=payload.description,
        amount=payload.amount,
        vat_rate=payload.vat_rate,
        workspace_id=payload.workspace_id,
        source_document_id=payload.source_document_id,
        amount_includes_vat=payload.amount_includes_vat,
        has_invoice=payload.has_invoice,
        extra=payload.extra,
    )


@app.get("/ai/v88/journal")
def ai_v88_list_journal(
    workspace_id: str = Query(default="default"),
    status: Optional[str] = Query(default=None),
    limit: int = Query(default=100, ge=1, le=1000),
    _: bool = Depends(finiip_api_guard),
):
    return v88_list_journal_entries(workspace_id=workspace_id, status=status, limit=limit)


@app.post("/ai/v88/journal/export")
def ai_v88_export_journal(payload: V88JournalExportRequest, _: bool = Depends(finiip_api_guard)):
    if payload.format.lower() == "xlsx":
        return v88_export_journal_xlsx(workspace_id=payload.workspace_id, status=payload.status)
    return v88_export_journal_csv(workspace_id=payload.workspace_id, status=payload.status)


@app.post("/ai/v89/review-queue")
def ai_v89_create_review(payload: V89ReviewCreateRequest, _: bool = Depends(finiip_api_guard)):
    return v89_create_review_item(
        workspace_id=payload.workspace_id,
        item_type=payload.item_type,
        title=payload.title,
        payload=payload.payload,
        risk_level=payload.risk_level,
        priority=payload.priority,
    )


@app.get("/ai/v89/review-queue")
def ai_v89_list_review(
    workspace_id: str = Query(default="default"),
    status: Optional[str] = Query(default=None),
    limit: int = Query(default=100, ge=1, le=1000),
    _: bool = Depends(finiip_api_guard),
):
    return v89_list_review_queue(workspace_id=workspace_id, status=status, limit=limit)


@app.post("/ai/v89/review-queue/{review_id}")
def ai_v89_update_review(review_id: str, payload: V89ReviewUpdateRequest, _: bool = Depends(finiip_api_guard)):
    try:
        return v89_update_review_item(review_id, status=payload.status, reviewer_note=payload.reviewer_note, correction=payload.correction)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@app.post("/ai/v90/risk-check")
def ai_v90_risk_check(payload: V90RiskCheckRequest, _: bool = Depends(finiip_api_guard)):
    return v90_tax_risk_check(payload.transaction, workspace_id=payload.workspace_id)


@app.post("/ai/v91/followup-questions")
def ai_v91_followup_questions(payload: V91FollowupRequest, _: bool = Depends(finiip_api_guard)):
    return v91_smart_followup_questions(payload.transaction, workspace_id=payload.workspace_id)


@app.post("/ai/v92/agent/process-text")
def ai_v92_agent_process_text(payload: V92AgentTextRequest, _: bool = Depends(finiip_api_guard)):
    return v92_run_accounting_agent_pipeline(
        text=payload.text,
        workspace_id=payload.workspace_id,
        filename=payload.filename,
        create_review=payload.create_review,
    )


@app.post("/ai/v93/workspaces")
def ai_v93_upsert_workspace(payload: V93WorkspaceRequest, _: bool = Depends(finiip_api_guard)):
    return v93_create_or_update_workspace(
        workspace_id=payload.workspace_id,
        name=payload.name,
        tax_code=payload.tax_code,
        policy=payload.policy,
        chart_of_accounts=payload.chart_of_accounts,
        users=payload.users,
    )


@app.get("/ai/v93/workspaces")
def ai_v93_workspaces(_: bool = Depends(finiip_api_guard)):
    return v93_list_workspaces()


@app.post("/ai/v94/evaluate")
def ai_v94_evaluate(payload: V94EvaluationRequest, _: bool = Depends(finiip_api_guard)):
    return v94_run_evaluation(cases=payload.cases, workspace_id=payload.workspace_id)


@app.get("/ai/v94/dashboard")
def ai_v94_dashboard(workspace_id: str = Query(default="default"), _: bool = Depends(finiip_api_guard)):
    return v94_quality_dashboard(workspace_id=workspace_id)


@app.get("/ai/v95/database-schema")
def ai_v95_database_schema(_: bool = Depends(finiip_api_guard)):
    return v95_database_schema_blueprint()


@app.get("/ai/v96/frontend-contract")
def ai_v96_frontend_contract(_: bool = Depends(finiip_api_guard)):
    return v96_frontend_api_contract()


@app.get("/ai/v97/reports/monthly-summary")
def ai_v97_monthly_summary(workspace_id: str = Query(default="default"), _: bool = Depends(finiip_api_guard)):
    return v97_monthly_summary_report(workspace_id=workspace_id)


@app.get("/ai/v97/reports/closing-checklist")
def ai_v97_closing_checklist(workspace_id: str = Query(default="default"), _: bool = Depends(finiip_api_guard)):
    return v97_closing_checklist(workspace_id=workspace_id)


@app.post("/ai/v98/company-memory")
def ai_v98_remember_company_fact(payload: V98CompanyMemoryRequest, _: bool = Depends(finiip_api_guard)):
    return v98_remember_company_fact(
        workspace_id=payload.workspace_id,
        fact=payload.fact,
        category=payload.category,
        source=payload.source,
    )


@app.get("/ai/v98/company-memory")
def ai_v98_company_memory(workspace_id: str = Query(default="default"), _: bool = Depends(finiip_api_guard)):
    return v98_list_company_memory(workspace_id=workspace_id)


@app.get("/ai/v99/production-readiness")
def ai_v99_production_readiness(_: bool = Depends(finiip_api_guard)):
    return v99_production_readiness_check()


# ==============================================================
# V100 - Backend-only Admin RAG Upload & Document Manager UI
# ============================================================== 
# Purpose: Admin/owner uploads official RAG knowledge here. Normal users
# should upload invoices/reports/files through user-processing endpoints only;
# their files must not be inserted into the official knowledge base.
from services.rag_admin_ui_v100 import (
    V100_VERSION as v100_rag_admin_version,
    SUPABASE_SCHEMA_SQL as v101_supabase_schema_sql,
    admin_key_is_valid as v100_admin_key_is_valid,
    admin_rag_storage_status as v101_admin_rag_storage_status,
    answer_admin_rag_question as v101_answer_admin_rag_question,
    clear_admin_rag_memory as v100_clear_admin_rag_memory,
    delete_admin_rag_document as v100_delete_admin_rag_document,
    get_admin_rag_document as v100_get_admin_rag_document,
    list_admin_rag_documents as v100_list_admin_rag_documents,
    reindex_admin_rag_document as v100_reindex_admin_rag_document,
    render_admin_login_or_unauthorized as v100_render_admin_login,
    render_admin_rag_page as v100_render_admin_rag_page,
    resolve_admin_file_output as v100_resolve_admin_file_output,
    process_admin_file_to_output as v100_process_admin_file_to_output,
    run_admin_rag_eval as v100_run_admin_rag_eval,
    search_documents as v100_search_documents,
    upload_admin_rag_document as v100_upload_admin_rag_document,
)
from services.simple_intents_v101 import (
    V101_INTENT_VERSION as v101_intent_version,
    detect_simple_intent as v101_detect_simple_intent,
    list_simple_intents as v101_list_simple_intents,
)
from services.rag_storage_v101 import (
    V101_VERSION as v106_conversation_version,
    list_supabase_chat_memory as v106_list_chat_memory,
)

from services.file_report_v68_v72 import (
    FILE_REPORT_VERSION as v68_file_report_version,
    FileReportInput as V68FileReportInput,
    capabilities as v68_file_report_capabilities,
    create_and_run_sync as v68_create_and_run_file_report_sync,
    create_file_report_job as v69_create_file_report_job,
    delete_file_report_job as v70_delete_file_report_job,
    get_job_status as v69_get_file_report_job_status,
    list_file_report_history as v70_list_file_report_history,
    resolve_job_output as v69_resolve_file_report_output,
    run_file_report_job as v69_run_file_report_job,
)


@app.get("/admin/rag-ui", response_class=HTMLResponse)
def admin_rag_ui_home(
    key: Optional[str] = Query(default=""),
    workspace_id: str = Query(default="default"),
    message: str = Query(default=""),
    error: str = Query(default=""),
    detail_document_id: Optional[str] = Query(default=None),
):
    if not v100_admin_key_is_valid(key):
        return HTMLResponse(v100_render_admin_login(), status_code=401)
    detail = None
    if detail_document_id:
        try:
            detail = v100_get_admin_rag_document(detail_document_id)
        except Exception as exc:
            error = str(exc)
    return HTMLResponse(v100_render_admin_rag_page(
        admin_key=key or "",
        workspace_id=workspace_id,
        message=message,
        error=error,
        document_detail=detail,
    ))


@app.post("/admin/rag-ui/upload", response_class=HTMLResponse)
async def admin_rag_ui_upload(
    file: UploadFile = File(...),
    admin_key: str = Form(default=""),
    workspace_id: str = Form(default="default"),
    title: Optional[str] = Form(default=None),
    source_type: str = Form(default="knowledge"),
    note: Optional[str] = Form(default=None),
):
    if not v100_admin_key_is_valid(admin_key):
        return HTMLResponse(v100_render_admin_login(), status_code=401)
    try:
        content = await file.read()
        result = v100_upload_admin_rag_document(
            filename=file.filename or "upload.bin",
            content=content,
            workspace_id=workspace_id,
            title=title or file.filename,
            source_type=source_type,
            note=note,
        )
        doc = result.get("document", {})
        message = f"Đã upload & index: {doc.get('title')} ({result.get('chunks_added')} chunks)."
        return HTMLResponse(v100_render_admin_rag_page(admin_key=admin_key, workspace_id=workspace_id, message=message))
    except Exception as exc:
        return HTMLResponse(v100_render_admin_rag_page(admin_key=admin_key, workspace_id=workspace_id, error=str(exc)), status_code=400)


@app.post("/admin/rag-ui/delete", response_class=HTMLResponse)
def admin_rag_ui_delete(
    admin_key: str = Form(default=""),
    workspace_id: str = Form(default="default"),
    document_id: str = Form(...),
):
    if not v100_admin_key_is_valid(admin_key):
        return HTMLResponse(v100_render_admin_login(), status_code=401)
    try:
        result = v100_delete_admin_rag_document(document_id=document_id, hard_delete=False)
        message = f"Đã xóa khỏi RAG: {result.get('document_id')} ({result.get('chunks_removed')} chunks removed)."
        return HTMLResponse(v100_render_admin_rag_page(admin_key=admin_key, workspace_id=workspace_id, message=message))
    except Exception as exc:
        return HTMLResponse(v100_render_admin_rag_page(admin_key=admin_key, workspace_id=workspace_id, error=str(exc)), status_code=400)


@app.post("/admin/rag-ui/reindex", response_class=HTMLResponse)
def admin_rag_ui_reindex(
    admin_key: str = Form(default=""),
    workspace_id: str = Form(default="default"),
    document_id: str = Form(...),
):
    if not v100_admin_key_is_valid(admin_key):
        return HTMLResponse(v100_render_admin_login(), status_code=401)
    try:
        result = v100_reindex_admin_rag_document(document_id=document_id)
        message = f"Đã re-index: {result.get('document_id')} ({result.get('chunks_added')} chunks)."
        return HTMLResponse(v100_render_admin_rag_page(admin_key=admin_key, workspace_id=workspace_id, message=message))
    except Exception as exc:
        return HTMLResponse(v100_render_admin_rag_page(admin_key=admin_key, workspace_id=workspace_id, error=str(exc)), status_code=400)


@app.post("/admin/rag-ui/ask", response_class=HTMLResponse)
def admin_rag_ui_ask(
    admin_key: str = Form(default=""),
    workspace_id: str = Form(default="default"),
    question: str = Form(...),
    history: str = Form(default=""),
    answer_mode: str = Form(default="auto"),
    conversation_id: str = Form(default="admin"),
):
    if not v100_admin_key_is_valid(admin_key):
        return HTMLResponse(v100_render_admin_login(), status_code=401)
    try:
        result = v101_answer_admin_rag_question(question=question, workspace_id=workspace_id, limit=6, history=history, answer_mode=answer_mode, conversation_id=conversation_id)
        return HTMLResponse(v100_render_admin_rag_page(
            admin_key=admin_key,
            workspace_id=workspace_id,
            question=question,
            answer_result=result,
            answer_mode=answer_mode,
        ))
    except Exception as exc:
        return HTMLResponse(v100_render_admin_rag_page(admin_key=admin_key, workspace_id=workspace_id, error=str(exc)), status_code=400)


@app.post("/admin/rag-ui/search", response_class=HTMLResponse)
def admin_rag_ui_search(
    admin_key: str = Form(default=""),
    workspace_id: str = Form(default="default"),
    query: str = Form(...),
):
    if not v100_admin_key_is_valid(admin_key):
        return HTMLResponse(v100_render_admin_login(), status_code=401)
    try:
        result = v100_search_documents(query=query, workspace_id=workspace_id, limit=10)
        return HTMLResponse(v100_render_admin_rag_page(
            admin_key=admin_key,
            workspace_id=workspace_id,
            search_query=query,
            search_result=result,
        ))
    except Exception as exc:
        return HTMLResponse(v100_render_admin_rag_page(admin_key=admin_key, workspace_id=workspace_id, error=str(exc)), status_code=400)


@app.post("/admin/rag-ui/eval", response_class=HTMLResponse)
def admin_rag_ui_eval(
    admin_key: str = Form(default=""),
    workspace_id: str = Form(default="default"),
    cases: str = Form(default=""),
    answer_mode: str = Form(default="short"),
):
    if not v100_admin_key_is_valid(admin_key):
        return HTMLResponse(v100_render_admin_login(), status_code=401)
    try:
        result = v100_run_admin_rag_eval(cases, workspace_id=workspace_id, answer_mode=answer_mode)
        message = f"Đã chạy Test Center: {result.get('passed')}/{result.get('count')} pass, avg={result.get('avg_score')}."
        return HTMLResponse(v100_render_admin_rag_page(
            admin_key=admin_key,
            workspace_id=workspace_id,
            message=message,
            eval_result=result,
            eval_cases=cases,
            answer_mode=answer_mode,
        ))
    except Exception as exc:
        return HTMLResponse(v100_render_admin_rag_page(admin_key=admin_key, workspace_id=workspace_id, error=str(exc), eval_cases=cases, answer_mode=answer_mode), status_code=400)


@app.post("/admin/rag-ui/file/process", response_class=HTMLResponse)
async def admin_rag_ui_file_process(
    file: UploadFile = File(...),
    admin_key: str = Form(default=""),
    workspace_id: str = Form(default="default"),
    task_type: str = Form(default="summary"),
    output_format: str = Form(default="docx"),
    question: str = Form(default=""),
    instruction: str = Form(default=""),
):
    if not v100_admin_key_is_valid(admin_key):
        return HTMLResponse(v100_render_admin_login(), status_code=401)
    try:
        content = await file.read()
        result = v100_process_admin_file_to_output(
            filename=file.filename or "upload.bin",
            content=content,
            workspace_id=workspace_id,
            task_type=task_type,
            output_format=output_format,
            question=question,
            instruction=instruction,
        )
        message = f"Đã đọc file và tạo file kết quả: {result.get('output_filename')}"
        return HTMLResponse(v100_render_admin_rag_page(
            admin_key=admin_key,
            workspace_id=workspace_id,
            message=message,
            file_result=result,
        ))
    except Exception as exc:
        return HTMLResponse(v100_render_admin_rag_page(admin_key=admin_key, workspace_id=workspace_id, error=str(exc)), status_code=400)


@app.get("/admin/rag-ui/file/download")
def admin_rag_ui_file_download(
    job_id: str = Query(...),
    key: Optional[str] = Query(default=""),
):
    if not v100_admin_key_is_valid(key):
        raise HTTPException(status_code=401, detail="Thiếu hoặc sai admin key")
    try:
        resolved = v100_resolve_admin_file_output(job_id)
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    return FileResponse(str(resolved["path"]), filename=resolved.get("filename") or resolved["path"].name)



# ==============================================================
# V68-V72 - Frontend File Reader / Report Generator
# ==============================================================
# Purpose: frontend users upload one or many files, Finiip reads them,
# creates a report, stores a job_id, and returns a downloadable file.
# This is separate from Admin Upload & Index, so user files do not pollute
# the official RAG knowledge base.

async def _v68_collect_uploads(single_file: Optional[UploadFile], files: Optional[List[UploadFile]]) -> List[V68FileReportInput]:
    uploads: List[UploadFile] = []
    if files:
        uploads.extend([f for f in files if f is not None])
    if single_file is not None:
        uploads.append(single_file)
    out: List[V68FileReportInput] = []
    for upload in uploads:
        content = await upload.read()
        out.append(V68FileReportInput(filename=upload.filename or "upload.bin", content=content))
    return out


@app.get("/ai/v68/file-report/capabilities")
def ai_v68_file_report_capabilities(_: bool = Depends(finiip_api_guard)):
    return v68_file_report_capabilities()


@app.post("/ai/v68/file-report/create-sync")
async def ai_v68_file_report_create_sync(
    file: Optional[UploadFile] = File(default=None, description="Single file upload"),
    files: Optional[List[UploadFile]] = File(default=None, description="Multiple file upload; field name files"),
    instruction: str = Form(default="Đọc file và lập báo cáo phân tích chi tiết."),
    question: str = Form(default=""),
    task_type: str = Form(default="auto_report"),
    output_format: str = Form(default="docx"),
    report_style: str = Form(default="detailed"),
    workspace_id: str = Form(default="default"),
    user_id: str = Form(default="anonymous"),
    title: str = Form(default=""),
    return_file: bool = Form(default=False),
    _: bool = Depends(finiip_api_guard),
):
    """V68 synchronous frontend endpoint.

    Good for small files. For large files, use V69 async jobs. If return_file=true,
    the response is the generated file; otherwise it returns JSON with job_id and
    download_url.
    """
    inputs = await _v68_collect_uploads(file, files)
    try:
        job = v68_create_and_run_file_report_sync(
            files=inputs,
            instruction=instruction,
            question=question,
            task_type=task_type,
            output_format=output_format,
            report_style=report_style,
            workspace_id=workspace_id,
            user_id=user_id,
            title=title,
        )
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    if return_file:
        resolved = v69_resolve_file_report_output(job["job_id"])
        return FileResponse(str(resolved["path"]), filename=resolved.get("filename") or resolved["path"].name)
    public = v69_get_file_report_job_status(job["job_id"])
    public["download_url"] = job.get("download_url")
    return public


@app.post("/ai/v69/file-report/jobs")
async def ai_v69_file_report_create_job(
    background_tasks: BackgroundTasks,
    file: Optional[UploadFile] = File(default=None, description="Single file upload"),
    files: Optional[List[UploadFile]] = File(default=None, description="Multiple file upload; field name files"),
    instruction: str = Form(default="Đọc file và lập báo cáo phân tích chi tiết."),
    question: str = Form(default=""),
    task_type: str = Form(default="auto_report"),
    output_format: str = Form(default="docx"),
    report_style: str = Form(default="detailed"),
    workspace_id: str = Form(default="default"),
    user_id: str = Form(default="anonymous"),
    title: str = Form(default=""),
    _: bool = Depends(finiip_api_guard),
):
    """V69 async job endpoint for frontend file-report workflows."""
    inputs = await _v68_collect_uploads(file, files)
    try:
        job = v69_create_file_report_job(
            files=inputs,
            instruction=instruction,
            question=question,
            task_type=task_type,
            output_format=output_format,
            report_style=report_style,
            workspace_id=workspace_id,
            user_id=user_id,
            title=title,
            save_inputs=True,
        )
        background_tasks.add_task(v69_run_file_report_job, job["job_id"])
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    return {
        "version": v68_file_report_version,
        "job_id": job["job_id"],
        "status": job["status"],
        "status_url": f"/ai/v69/file-report/jobs/{job['job_id']}",
        "download_url": f"/ai/v69/file-report/jobs/{job['job_id']}/download",
        "message": "Đã nhận file. Frontend hãy poll status_url cho tới khi status=done rồi tải download_url.",
    }


@app.get("/ai/v69/file-report/jobs/{job_id}")
def ai_v69_file_report_job_status(job_id: str, _: bool = Depends(finiip_api_guard)):
    try:
        return v69_get_file_report_job_status(job_id)
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc))


@app.get("/ai/v69/file-report/jobs/{job_id}/download")
def ai_v69_file_report_job_download(job_id: str, _: bool = Depends(finiip_api_guard)):
    try:
        resolved = v69_resolve_file_report_output(job_id)
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    return FileResponse(str(resolved["path"]), filename=resolved.get("filename") or resolved["path"].name)


@app.get("/ai/v70/file-report/history")
def ai_v70_file_report_history(
    workspace_id: str = Query(default=""),
    user_id: str = Query(default=""),
    limit: int = Query(default=50, ge=1, le=200),
    include_deleted: bool = Query(default=False),
    _: bool = Depends(finiip_api_guard),
):
    return v70_list_file_report_history(workspace_id=workspace_id, user_id=user_id, limit=limit, include_deleted=include_deleted)


@app.delete("/ai/v70/file-report/history/{job_id}")
def ai_v70_file_report_delete_history(
    job_id: str,
    hard_delete: bool = Query(default=False),
    _: bool = Depends(finiip_api_guard),
):
    try:
        return v70_delete_file_report_job(job_id, hard_delete=hard_delete)
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc))


@app.post("/admin/rag-ui/memory/clear")
def admin_rag_ui_memory_clear(
    admin_key: str = Form(default=""),
    workspace_id: str = Form(default="default"),
    conversation_id: str = Form(default="admin"),
):
    if not v100_admin_key_is_valid(admin_key):
        raise HTTPException(status_code=401, detail="Thiếu hoặc sai admin key")
    return v100_clear_admin_rag_memory(workspace_id=workspace_id, conversation_id=conversation_id)


@app.get("/admin/rag-ui/api/documents")
def admin_rag_ui_api_documents(
    key: Optional[str] = Query(default=""),
    workspace_id: Optional[str] = Query(default=None),
):
    if not v100_admin_key_is_valid(key):
        raise HTTPException(status_code=401, detail="Thiếu hoặc sai admin key")
    return v100_list_admin_rag_documents(workspace_id=workspace_id)


@app.get("/admin/rag-ui/api/capabilities")
def admin_rag_ui_api_capabilities(key: Optional[str] = Query(default="")):
    if not v100_admin_key_is_valid(key):
        raise HTTPException(status_code=401, detail="Thiếu hoặc sai admin key")
    return {
        "version": v100_rag_admin_version,
        "ui_path": "/admin/rag-ui",
        "purpose": "Backend-only white Admin UI for official RAG upload/document management.",
        "security": "Set FINIIP_ADMIN_KEY or FINIIP_API_KEY in deployment; open /admin/rag-ui?key=YOUR_KEY.",
        "admin_can": ["upload_official_rag", "list_documents", "view_chunks", "reindex", "delete_from_rag", "test_rag_qa", "search_chunks", "v60_eval_center", "v61_legal_citation", "v62_answer_modes", "v63_accounting_workflow", "v64_document_intelligence", "v65_conflict_checker", "v66_persistent_memory", "v67_file_reader_return_file", "v106_followup_context", "v106_intent_router", "v106_quality_gate", "v106_stable_chat_session"],
        "normal_user_cannot": ["upload_to_official_rag", "modify_knowledge_base", "delete_rag_documents"],
    }


# ==============================================================
# V106 - Conversational RAG: memory, follow-up, intent and quality gate
# ==============================================================

class V106ChatRequest(BaseModel):
    message: str = Field(..., min_length=1)
    workspace_id: str = Field(default="default")
    conversation_id: str = Field(default="default_chat")
    history: str = Field(default="")
    answer_mode: str = Field(default="auto")
    limit: int = Field(default=6, ge=1, le=12)
    save_memory: bool = Field(default=True)


@app.post("/ai/v106/chat")
def ai_v106_chat(payload: V106ChatRequest, _: bool = Depends(finiip_api_guard)):
    return v101_answer_admin_rag_question(
        question=payload.message,
        workspace_id=payload.workspace_id,
        limit=payload.limit,
        history=payload.history,
        answer_mode=payload.answer_mode,
        conversation_id=payload.conversation_id,
        save_memory=payload.save_memory,
    )


@app.get("/ai/v106/chat/history")
def ai_v106_chat_history(
    workspace_id: str = Query(default="default"),
    conversation_id: str = Query(default="default_chat"),
    limit: int = Query(default=20, ge=1, le=50),
    _: bool = Depends(finiip_api_guard),
):
    return v106_list_chat_memory(workspace_id=workspace_id, conversation_id=conversation_id, limit=limit)


@app.delete("/ai/v106/chat/history")
def ai_v106_clear_chat_history(
    workspace_id: str = Query(default="default"),
    conversation_id: str = Query(default="default_chat"),
    _: bool = Depends(finiip_api_guard),
):
    return v100_clear_admin_rag_memory(workspace_id=workspace_id, conversation_id=conversation_id)


@app.get("/ai/v106/capabilities")
def ai_v106_capabilities():
    return {
        "version": v106_conversation_version,
        "chat_endpoint": "/ai/v106/chat",
        "history_endpoint": "/ai/v106/chat/history",
        "features": [
            "persistent_local_or_supabase_memory",
            "followup_question_resolution",
            "accounting_vs_rag_vs_conversation_routing",
            "empathetic_conversation_style",
            "strict_relevance_and_answer_quality_gate",
            "stable_conversation_id",
        ],
        "frontend_note": "Giữ nguyên conversation_id cho cả phiên chat; tạo ID mới khi người dùng bấm New chat.",
    }


# ==============================================================
# V101 - Supabase RAG status + simple intent router
# ==============================================================

class V101IntentDetectRequest(BaseModel):
    message: str = Field(..., description="User/admin message to route")


@app.get("/ai/v101/supabase/status")
def v101_supabase_status():
    return v101_admin_rag_storage_status()


@app.get("/admin/rag-ui/api/storage-status")
def admin_rag_ui_api_storage_status(key: Optional[str] = Query(default="")):
    if not v100_admin_key_is_valid(key):
        raise HTTPException(status_code=401, detail="Thiếu hoặc sai admin key")
    return v101_admin_rag_storage_status()


@app.get("/admin/rag-ui/api/supabase-schema")
def admin_rag_ui_api_supabase_schema(key: Optional[str] = Query(default="")):
    if not v100_admin_key_is_valid(key):
        raise HTTPException(status_code=401, detail="Thiếu hoặc sai admin key")
    return {
        "version": "v101_supabase_rag_storage",
        "storage_status": v101_admin_rag_storage_status(),
        "sql": v101_supabase_schema_sql,
    }


@app.post("/ai/v101/intent/detect")
def v101_intent_detect(payload: V101IntentDetectRequest):
    return v101_detect_simple_intent(payload.message)


@app.get("/ai/v101/intent/catalog")
def v101_intent_catalog():
    return v101_list_simple_intents()


@app.get("/ai/v101/capabilities")
def v101_capabilities():
    return {
        "version": "v101_supabase_rag_and_simple_intents",
        "supabase_rag": {
            "status_endpoint": "/ai/v101/supabase/status",
            "admin_schema_endpoint": "/admin/rag-ui/api/supabase-schema",
            "env": ["RAG_STORAGE_MODE=supabase", "SUPABASE_URL", "SUPABASE_SERVICE_ROLE_KEY", "SUPABASE_RAG_BUCKET=rag-knowledge"],
            "fallback": "If Supabase env is not active, Admin RAG UI keeps using local JSON/file store.",
        },
        "simple_intents": {
            "version": v101_intent_version,
            "detect_endpoint": "/ai/v101/intent/detect",
            "catalog_endpoint": "/ai/v101/intent/catalog",
        },
    }


# =============================================================
# Stable product API consumed by the deployable chat frontend
# =============================================================
from chat_api_v1 import router as chat_api_v1_router
app.include_router(chat_api_v1_router)
