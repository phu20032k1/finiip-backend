"""IIP Steel Platform - Giai đoạn 1 cho Finiip backend.

Module này bổ sung các API lõi theo kế hoạch IIP 31/5/2026:
- Công nợ, hạn mức tín dụng, đại lý cấp 2
- Giá sàn, phát hiện bán thấp bất thường
- Đơn hàng, hóa đơn, thanh toán, đối soát 4 chiều
- Xuất kho QR, giao hàng/GPS/ảnh
- Thưởng doanh số VAS, báo cáo sáng cho Chủ tịch

Thiết kế ưu tiên MVP backend-only: nhập tay hoặc import Excel trước, frontend gọi API sau.
"""
from __future__ import annotations

from datetime import date, datetime, timedelta
from io import BytesIO
import csv
import json
import re
from typing import Any, Dict, Iterable, List, Optional, Tuple

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import FileResponse
from pydantic import BaseModel, Field
from sqlalchemy import Column, Date, DateTime, Float, Integer, String, Text, func
from sqlalchemy.orm import Session

from database import Base, engine, get_db

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None


router = APIRouter(prefix="/iip", tags=["IIP Steel Platform"])


# =========================
# Database models
# =========================

class IIPTimestampMixin:
    created_at = Column(DateTime, default=datetime.utcnow, nullable=False)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)


class IIPDealer(Base, IIPTimestampMixin):
    __tablename__ = "iip_dealers"
    id = Column(Integer, primary_key=True, index=True)
    code = Column(String, unique=True, index=True, nullable=False)
    name = Column(String, nullable=False)
    province = Column(String, nullable=True, index=True)
    phone = Column(String, nullable=True)
    address = Column(Text, nullable=True)
    sales_staff_code = Column(String, nullable=True, index=True)
    rank = Column(String, default="B", nullable=False)  # A/B/C/D
    status = Column(String, default="active", nullable=False)
    note = Column(Text, nullable=True)


class IIPSalesStaff(Base, IIPTimestampMixin):
    __tablename__ = "iip_sales_staff"
    id = Column(Integer, primary_key=True, index=True)
    code = Column(String, unique=True, index=True, nullable=False)
    name = Column(String, nullable=False)
    phone = Column(String, nullable=True)
    region = Column(String, nullable=True)
    status = Column(String, default="active", nullable=False)


class IIPSteelProduct(Base, IIPTimestampMixin):
    __tablename__ = "iip_steel_products"
    id = Column(Integer, primary_key=True, index=True)
    code = Column(String, unique=True, index=True, nullable=False)
    name = Column(String, nullable=False)
    brand = Column(String, default="VAS", nullable=False)
    category = Column(String, nullable=True)
    unit = Column(String, default="ton", nullable=False)
    status = Column(String, default="active", nullable=False)


class IIPPriceFloor(Base, IIPTimestampMixin):
    __tablename__ = "iip_price_floors"
    id = Column(Integer, primary_key=True, index=True)
    product_code = Column(String, index=True, nullable=False)
    province = Column(String, default="ALL", index=True, nullable=False)
    effective_from = Column(Date, default=date.today, nullable=False)
    floor_price = Column(Float, nullable=False)
    allowed_discount_pct = Column(Float, default=0, nullable=False)
    note = Column(Text, nullable=True)


class IIPCreditLimit(Base, IIPTimestampMixin):
    __tablename__ = "iip_credit_limits"
    id = Column(Integer, primary_key=True, index=True)
    dealer_code = Column(String, unique=True, index=True, nullable=False)
    limit_amount = Column(Float, default=0, nullable=False)
    debt_term_days = Column(Integer, default=30, nullable=False)
    rank = Column(String, default="B", nullable=False)
    require_deposit_pct = Column(Float, default=0, nullable=False)
    note = Column(Text, nullable=True)


class IIPDebt(Base, IIPTimestampMixin):
    __tablename__ = "iip_debts"
    id = Column(Integer, primary_key=True, index=True)
    dealer_code = Column(String, index=True, nullable=False)
    order_code = Column(String, nullable=True, index=True)
    debt_date = Column(Date, default=date.today, nullable=False)
    due_date = Column(Date, nullable=False)
    original_amount = Column(Float, nullable=False)
    paid_amount = Column(Float, default=0, nullable=False)
    status = Column(String, default="open", nullable=False)  # open/closed/bad_debt
    note = Column(Text, nullable=True)


class IIPOrder(Base, IIPTimestampMixin):
    __tablename__ = "iip_orders"
    id = Column(Integer, primary_key=True, index=True)
    order_code = Column(String, unique=True, index=True, nullable=False)
    order_date = Column(Date, default=date.today, nullable=False)
    dealer_code = Column(String, index=True, nullable=False)
    sales_staff_code = Column(String, index=True, nullable=True)
    status = Column(String, default="draft", nullable=False)  # draft/pending/approved/blocked/exported/delivered/cancelled
    total_amount = Column(Float, default=0, nullable=False)
    note = Column(Text, nullable=True)


class IIPOrderItem(Base, IIPTimestampMixin):
    __tablename__ = "iip_order_items"
    id = Column(Integer, primary_key=True, index=True)
    order_code = Column(String, index=True, nullable=False)
    product_code = Column(String, index=True, nullable=False)
    quantity_ton = Column(Float, default=0, nullable=False)
    unit_price = Column(Float, default=0, nullable=False)
    discount_pct = Column(Float, default=0, nullable=False)
    amount = Column(Float, default=0, nullable=False)
    note = Column(Text, nullable=True)


class IIPPayment(Base, IIPTimestampMixin):
    __tablename__ = "iip_payments"
    id = Column(Integer, primary_key=True, index=True)
    payment_date = Column(Date, default=date.today, nullable=False)
    dealer_code = Column(String, index=True, nullable=False)
    amount = Column(Float, nullable=False)
    bank_ref = Column(String, nullable=True, index=True)
    matched_order_code = Column(String, nullable=True, index=True)
    matched_invoice_number = Column(String, nullable=True, index=True)
    note = Column(Text, nullable=True)


class IIPOutputInvoice(Base, IIPTimestampMixin):
    __tablename__ = "iip_output_invoices"
    id = Column(Integer, primary_key=True, index=True)
    invoice_number = Column(String, unique=True, index=True, nullable=False)
    invoice_date = Column(Date, default=date.today, nullable=False)
    dealer_code = Column(String, index=True, nullable=False)
    order_code = Column(String, nullable=True, index=True)
    subtotal = Column(Float, default=0, nullable=False)
    vat_rate = Column(Float, default=0.1, nullable=False)
    vat_amount = Column(Float, default=0, nullable=False)
    total_amount = Column(Float, default=0, nullable=False)
    status = Column(String, default="issued", nullable=False)
    note = Column(Text, nullable=True)


class IIPWarehouseSlip(Base, IIPTimestampMixin):
    __tablename__ = "iip_warehouse_slips"
    id = Column(Integer, primary_key=True, index=True)
    slip_code = Column(String, unique=True, index=True, nullable=False)
    order_code = Column(String, index=True, nullable=False)
    slip_date = Column(Date, default=date.today, nullable=False)
    status = Column(String, default="created", nullable=False)  # created/picked/exported/cancelled
    qr_code = Column(String, nullable=True)
    note = Column(Text, nullable=True)


class IIPDelivery(Base, IIPTimestampMixin):
    __tablename__ = "iip_deliveries"
    id = Column(Integer, primary_key=True, index=True)
    delivery_code = Column(String, unique=True, index=True, nullable=False)
    order_code = Column(String, index=True, nullable=False)
    driver_name = Column(String, nullable=True)
    driver_phone = Column(String, nullable=True)
    route_note = Column(Text, nullable=True)
    gps_lat = Column(Float, nullable=True)
    gps_lng = Column(Float, nullable=True)
    photo_url = Column(Text, nullable=True)
    status = Column(String, default="planned", nullable=False)  # planned/picked/delivered/exception
    delivered_at = Column(DateTime, nullable=True)
    note = Column(Text, nullable=True)


class IIPVASTarget(Base, IIPTimestampMixin):
    __tablename__ = "iip_vas_targets"
    id = Column(Integer, primary_key=True, index=True)
    year = Column(Integer, unique=True, index=True, nullable=False)
    target_ton = Column(Float, nullable=False)
    bonus_amount = Column(Float, default=0, nullable=False)
    milestone_json = Column(Text, nullable=True)  # [{"ton":50000,"bonus":1800000000}]
    note = Column(Text, nullable=True)


# Ensure tables are created even when this router is imported after main Base.metadata.create_all.
Base.metadata.create_all(bind=engine)


# =========================
# Pydantic schemas
# =========================

class DealerIn(BaseModel):
    code: str
    name: str
    province: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    sales_staff_code: Optional[str] = None
    rank: str = "B"
    status: str = "active"
    note: Optional[str] = None


class SalesStaffIn(BaseModel):
    code: str
    name: str
    phone: Optional[str] = None
    region: Optional[str] = None
    status: str = "active"


class SteelProductIn(BaseModel):
    code: str
    name: str
    brand: str = "VAS"
    category: Optional[str] = None
    unit: str = "ton"
    status: str = "active"


class PriceFloorIn(BaseModel):
    product_code: str
    province: str = "ALL"
    effective_from: date = Field(default_factory=date.today)
    floor_price: float
    allowed_discount_pct: float = 0
    note: Optional[str] = None


class CreditLimitIn(BaseModel):
    dealer_code: str
    limit_amount: float = 0
    debt_term_days: int = 30
    rank: str = "B"
    require_deposit_pct: float = 0
    note: Optional[str] = None


class DebtIn(BaseModel):
    dealer_code: str
    order_code: Optional[str] = None
    debt_date: date = Field(default_factory=date.today)
    due_date: Optional[date] = None
    original_amount: float
    paid_amount: float = 0
    status: str = "open"
    note: Optional[str] = None


class OrderItemIn(BaseModel):
    product_code: str
    quantity_ton: float
    unit_price: float
    discount_pct: float = 0
    amount: Optional[float] = None
    note: Optional[str] = None


class OrderIn(BaseModel):
    order_code: str
    dealer_code: str
    sales_staff_code: Optional[str] = None
    order_date: date = Field(default_factory=date.today)
    status: str = "draft"
    items: List[OrderItemIn] = Field(default_factory=list)
    note: Optional[str] = None


class PaymentIn(BaseModel):
    dealer_code: str
    amount: float
    payment_date: date = Field(default_factory=date.today)
    bank_ref: Optional[str] = None
    matched_order_code: Optional[str] = None
    matched_invoice_number: Optional[str] = None
    note: Optional[str] = None


class InvoiceIn(BaseModel):
    invoice_number: str
    dealer_code: str
    invoice_date: date = Field(default_factory=date.today)
    order_code: Optional[str] = None
    subtotal: float = 0
    vat_rate: float = 0.1
    total_amount: Optional[float] = None
    status: str = "issued"
    note: Optional[str] = None


class WarehouseSlipIn(BaseModel):
    slip_code: str
    order_code: str
    slip_date: date = Field(default_factory=date.today)
    status: str = "created"
    note: Optional[str] = None


class DeliveryIn(BaseModel):
    delivery_code: str
    order_code: str
    driver_name: Optional[str] = None
    driver_phone: Optional[str] = None
    route_note: Optional[str] = None
    status: str = "planned"
    note: Optional[str] = None


class DeliveryConfirmIn(BaseModel):
    gps_lat: Optional[float] = None
    gps_lng: Optional[float] = None
    photo_url: Optional[str] = None
    note: Optional[str] = None


class VASTargetIn(BaseModel):
    year: int
    target_ton: float
    bonus_amount: float = 0
    milestones: Optional[List[Dict[str, Any]]] = None
    note: Optional[str] = None


# =========================
# Helpers
# =========================

MODEL_MAP = {
    "dealers": (IIPDealer, DealerIn, "code"),
    "sales-staff": (IIPSalesStaff, SalesStaffIn, "code"),
    "products": (IIPSteelProduct, SteelProductIn, "code"),
    "price-floors": (IIPPriceFloor, PriceFloorIn, None),
    "credit-limits": (IIPCreditLimit, CreditLimitIn, "dealer_code"),
    "debts": (IIPDebt, DebtIn, None),
    "payments": (IIPPayment, PaymentIn, None),
    "invoices": (IIPOutputInvoice, InvoiceIn, "invoice_number"),
    "warehouse-slips": (IIPWarehouseSlip, WarehouseSlipIn, "slip_code"),
    "deliveries": (IIPDelivery, DeliveryIn, "delivery_code"),
    "vas-targets": (IIPVASTarget, VASTargetIn, "year"),
}

HEADER_ALIASES = {
    "ma": "code", "mã": "code", "ma dai ly": "code", "ma_dai_ly": "code", "dealer_code": "dealer_code",
    "ten": "name", "tên": "name", "ten dai ly": "name", "ten_dai_ly": "name", "name": "name",
    "tinh": "province", "tỉnh": "province", "province": "province", "khu vuc": "region", "region": "region",
    "sdt": "phone", "phone": "phone", "dien thoai": "phone",
    "dia chi": "address", "address": "address",
    "nhan vien": "sales_staff_code", "sales_staff_code": "sales_staff_code", "ma nv": "sales_staff_code",
    "hang": "rank", "rank": "rank",
    "han muc": "limit_amount", "han_muc": "limit_amount", "limit_amount": "limit_amount",
    "thoi han no": "debt_term_days", "debt_term_days": "debt_term_days",
    "ngay no": "debt_date", "debt_date": "debt_date", "ngay den han": "due_date", "due_date": "due_date",
    "cong no": "original_amount", "no": "original_amount", "amount": "amount", "so tien": "amount",
    "da tra": "paid_amount", "paid_amount": "paid_amount",
    "ma san pham": "product_code", "product_code": "product_code", "san pham": "product_code",
    "gia san": "floor_price", "floor_price": "floor_price", "tinh ap dung": "province",
    "ngay hieu luc": "effective_from", "effective_from": "effective_from",
    "ma don": "order_code", "order_code": "order_code", "ngay don": "order_date",
    "so luong": "quantity_ton", "quantity_ton": "quantity_ton", "gia ban": "unit_price", "unit_price": "unit_price",
    "chiet khau": "discount_pct", "discount_pct": "discount_pct",
    "so hoa don": "invoice_number", "invoice_number": "invoice_number", "ngay hoa don": "invoice_date",
    "subtotal": "subtotal", "vat_rate": "vat_rate", "total": "total_amount", "total_amount": "total_amount",
    "nam": "year", "year": "year", "muc tieu": "target_ton", "target_ton": "target_ton", "thuong": "bonus_amount",
}


def norm_header(h: Any) -> str:
    text = str(h or "").strip().lower()
    text = re.sub(r"[_\-]+", " ", text)
    text = re.sub(r"\s+", " ", text)
    return HEADER_ALIASES.get(text, text.replace(" ", "_"))


def to_date(value: Any, default: Optional[date] = None) -> Optional[date]:
    if value is None or value == "":
        return default
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return default


def clean_payload(data: Dict[str, Any]) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    for k, v in data.items():
        if v is None or v == "":
            continue
        nk = norm_header(k)
        if isinstance(v, str):
            v = v.strip()
        if nk.endswith("date") or nk in {"effective_from", "due_date", "debt_date", "order_date", "invoice_date", "payment_date", "slip_date"}:
            v = to_date(v, None)
        if nk in {"amount", "original_amount", "paid_amount", "limit_amount", "floor_price", "quantity_ton", "unit_price", "discount_pct", "subtotal", "vat_rate", "total_amount", "bonus_amount", "target_ton"}:
            try:
                v = float(str(v).replace(",", ""))
            except Exception:
                pass
        if nk in {"debt_term_days", "year"}:
            try:
                v = int(float(str(v).replace(",", "")))
            except Exception:
                pass
        out[nk] = v
    return out


def schema_dump(obj: BaseModel, **kwargs) -> Dict[str, Any]:
    """Pydantic v1/v2 compatible dump."""
    if hasattr(obj, "model_dump"):
        return obj.model_dump(**kwargs)
    return obj.dict(**kwargs)


def row_to_dict(obj: Any) -> Dict[str, Any]:
    return {c.name: getattr(obj, c.name) for c in obj.__table__.columns}


def current_debt(db: Session, dealer_code: str) -> float:
    q = db.query(func.coalesce(func.sum(IIPDebt.original_amount - IIPDebt.paid_amount), 0.0)).filter(
        IIPDebt.dealer_code == dealer_code, IIPDebt.status != "closed"
    )
    return float(q.scalar() or 0)


def dealer_limit(db: Session, dealer_code: str) -> Optional[IIPCreditLimit]:
    return db.query(IIPCreditLimit).filter(IIPCreditLimit.dealer_code == dealer_code).first()


def latest_floor(db: Session, product_code: str, province: Optional[str], at_date: Optional[date] = None) -> Optional[IIPPriceFloor]:
    at_date = at_date or date.today()
    province = province or "ALL"
    q = db.query(IIPPriceFloor).filter(
        IIPPriceFloor.product_code == product_code,
        IIPPriceFloor.effective_from <= at_date,
        IIPPriceFloor.province.in_([province, "ALL"]),
    ).order_by(IIPPriceFloor.province.desc(), IIPPriceFloor.effective_from.desc())
    return q.first()


def calculate_order_total(items: Iterable[OrderItemIn]) -> float:
    total = 0.0
    for item in items:
        amount = item.amount if item.amount is not None else item.quantity_ton * item.unit_price * (1 - item.discount_pct / 100)
        total += float(amount or 0)
    return total


def build_qr_code(slip_code: str, order_code: str) -> str:
    return f"IIP-WH|{slip_code}|ORDER:{order_code}|{datetime.utcnow().strftime('%Y%m%d%H%M%S')}"


def upsert_by_key(db: Session, model: Any, key_field: Optional[str], payload: Dict[str, Any]) -> Any:
    obj = None
    if key_field and payload.get(key_field) is not None:
        obj = db.query(model).filter(getattr(model, key_field) == payload[key_field]).first()
    if obj:
        for k, v in payload.items():
            if hasattr(obj, k):
                setattr(obj, k, v)
    else:
        obj = model(**{k: v for k, v in payload.items() if hasattr(model, k)})
        db.add(obj)
    return obj


def parse_upload_rows(filename: str, content: bytes) -> List[Dict[str, Any]]:
    if filename.lower().endswith(".csv"):
        text = content.decode("utf-8-sig")
        return [clean_payload(row) for row in csv.DictReader(text.splitlines())]
    if load_workbook is None:
        raise HTTPException(status_code=500, detail="Thiếu openpyxl để đọc Excel")
    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [norm_header(h) for h in rows[0]]
    result = []
    for row in rows[1:]:
        data = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
        cleaned = clean_payload(data)
        if any(v not in (None, "") for v in cleaned.values()):
            result.append(cleaned)
    return result


# =========================
# Status + catalog
# =========================

@router.get("/status")
def iip_status(db: Session = Depends(get_db)):
    return {
        "platform": "Finiip + IIP Steel Profit & Cashflow Protection",
        "stage": "Giai đoạn 1 MVP backend completed",
        "modules": {
            "module_1_cong_no_dong_tien_nhan_vien": "API công nợ, hạn mức, cảnh báo quá hạn, vượt hạn mức, bán dưới giá sàn",
            "module_2_hoa_don_vat": "Hóa đơn đầu ra + đối soát 4 chiều hàng/hóa đơn/công nợ/tiền",
            "module_3_vas_bonus": "Mục tiêu VAS, sản lượng, dự báo hụt thưởng, khuyến nghị",
            "module_4_xuat_kho_van_chuyen": "Phiếu xuất QR, xác nhận lấy hàng/giao hàng, GPS/ảnh",
            "module_5_dai_ly_cap_2_tin_dung": "Dealer wallet, chấm điểm A/B/C/D, hạn mức tín dụng",
        },
        "counts": {
            "dealers": db.query(IIPDealer).count(),
            "orders": db.query(IIPOrder).count(),
            "debts": db.query(IIPDebt).count(),
            "payments": db.query(IIPPayment).count(),
            "invoices": db.query(IIPOutputInvoice).count(),
            "warehouse_slips": db.query(IIPWarehouseSlip).count(),
            "deliveries": db.query(IIPDelivery).count(),
            "vas_targets": db.query(IIPVASTarget).count(),
        },
        "main_demo_api": "GET /iip/chairman/morning-report",
    }


@router.get("/import/templates")
def import_templates():
    return {
        "data_types": list(MODEL_MAP.keys()) + ["orders"],
        "required_stage_1_inputs": [
            "dealers", "debts", "sales-staff", "price-floors", "credit-limits", "orders", "payments", "invoices"
        ],
        "note": "Upload Excel/CSV. Hàng đầu là header. Header tiếng Việt phổ biến được tự map sang field chuẩn.",
    }


@router.get("/import/template-workbook")
def download_import_template_workbook():
    """Tải workbook mẫu gồm các sheet nhập liệu thật cho Giai đoạn 1."""
    template_path = "templates/iip_stage1_input_templates.xlsx"
    try:
        import os
        base_dir = os.path.dirname(os.path.abspath(__file__))
        full_path = os.path.join(base_dir, template_path)
        if not os.path.exists(full_path):
            raise HTTPException(status_code=404, detail="Chưa có file template nhập liệu")
        return FileResponse(
            full_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename="iip_stage1_input_templates.xlsx",
        )
    except HTTPException:
        raise



# =========================
# Generic CRUD + imports
# =========================

@router.post("/import/{data_type}")
async def import_data(data_type: str, file: UploadFile = File(...), db: Session = Depends(get_db)):
    if data_type not in MODEL_MAP and data_type != "orders":
        raise HTTPException(status_code=400, detail=f"data_type không hỗ trợ: {data_type}")
    content = await file.read()
    rows = parse_upload_rows(file.filename or "upload.xlsx", content)
    created = updated = skipped = 0
    errors: List[Dict[str, Any]] = []
    for idx, row in enumerate(rows, start=2):
        try:
            if data_type == "orders":
                order_code = row.get("order_code") or row.get("code")
                dealer_code = row.get("dealer_code")
                product_code = row.get("product_code")
                if not order_code or not dealer_code or not product_code:
                    skipped += 1
                    continue
                order = db.query(IIPOrder).filter(IIPOrder.order_code == order_code).first()
                if not order:
                    order = IIPOrder(order_code=order_code, dealer_code=dealer_code, sales_staff_code=row.get("sales_staff_code"), order_date=to_date(row.get("order_date"), date.today()) or date.today(), status=row.get("status", "draft"), note=row.get("note"))
                    db.add(order)
                    created += 1
                item = IIPOrderItem(
                    order_code=order_code,
                    product_code=product_code,
                    quantity_ton=float(row.get("quantity_ton") or 0),
                    unit_price=float(row.get("unit_price") or 0),
                    discount_pct=float(row.get("discount_pct") or 0),
                    amount=float(row.get("amount") or 0) or float(row.get("quantity_ton") or 0) * float(row.get("unit_price") or 0) * (1 - float(row.get("discount_pct") or 0) / 100),
                    note=row.get("note"),
                )
                db.add(item)
            else:
                model, schema, key_field = MODEL_MAP[data_type]
                payload = schema_dump(schema(**row))
                before = None
                if key_field and payload.get(key_field) is not None:
                    before = db.query(model).filter(getattr(model, key_field) == payload[key_field]).first()
                obj = upsert_by_key(db, model, key_field, payload)
                created += 0 if before else 1
                updated += 1 if before else 0
        except Exception as exc:
            errors.append({"row": idx, "error": str(exc), "data": row})
    db.commit()
    # recalc imported order totals
    if data_type == "orders":
        for code, total in db.query(IIPOrderItem.order_code, func.sum(IIPOrderItem.amount)).group_by(IIPOrderItem.order_code).all():
            order = db.query(IIPOrder).filter(IIPOrder.order_code == code).first()
            if order:
                order.total_amount = float(total or 0)
        db.commit()
    return {"data_type": data_type, "filename": file.filename, "rows": len(rows), "created": created, "updated": updated, "skipped": skipped, "errors": errors[:20]}


@router.get("/{data_type}")
def list_data(data_type: str, limit: int = 100, offset: int = 0, db: Session = Depends(get_db)):
    if data_type not in MODEL_MAP:
        raise HTTPException(status_code=404, detail="data_type không tồn tại")
    model, _, _ = MODEL_MAP[data_type]
    rows = db.query(model).order_by(model.id.desc()).offset(offset).limit(min(limit, 500)).all()
    return {"items": [row_to_dict(r) for r in rows], "count": db.query(model).count()}


@router.post("/dealers")
def create_dealer(payload: DealerIn, db: Session = Depends(get_db)):
    obj = upsert_by_key(db, IIPDealer, "code", schema_dump(payload))
    db.commit(); db.refresh(obj)
    return row_to_dict(obj)


@router.post("/sales-staff")
def create_sales_staff(payload: SalesStaffIn, db: Session = Depends(get_db)):
    obj = upsert_by_key(db, IIPSalesStaff, "code", schema_dump(payload))
    db.commit(); db.refresh(obj)
    return row_to_dict(obj)


@router.post("/products")
def create_product(payload: SteelProductIn, db: Session = Depends(get_db)):
    obj = upsert_by_key(db, IIPSteelProduct, "code", schema_dump(payload))
    db.commit(); db.refresh(obj)
    return row_to_dict(obj)


@router.post("/price-floors")
def create_price_floor(payload: PriceFloorIn, db: Session = Depends(get_db)):
    obj = IIPPriceFloor(**schema_dump(payload))
    db.add(obj); db.commit(); db.refresh(obj)
    return row_to_dict(obj)


@router.post("/credit-limits")
def create_credit_limit(payload: CreditLimitIn, db: Session = Depends(get_db)):
    obj = upsert_by_key(db, IIPCreditLimit, "dealer_code", schema_dump(payload))
    db.commit(); db.refresh(obj)
    return row_to_dict(obj)


@router.post("/debts")
def create_debt(payload: DebtIn, db: Session = Depends(get_db)):
    data = schema_dump(payload)
    if data.get("due_date") is None:
        limit = dealer_limit(db, data["dealer_code"])
        data["due_date"] = data["debt_date"] + timedelta(days=limit.debt_term_days if limit else 30)
    obj = IIPDebt(**data)
    db.add(obj); db.commit(); db.refresh(obj)
    return row_to_dict(obj)


@router.post("/orders")
def create_order(payload: OrderIn, db: Session = Depends(get_db)):
    existing = db.query(IIPOrder).filter(IIPOrder.order_code == payload.order_code).first()
    if existing:
        raise HTTPException(status_code=409, detail="order_code đã tồn tại")
    total = calculate_order_total(payload.items)
    order = IIPOrder(order_code=payload.order_code, order_date=payload.order_date, dealer_code=payload.dealer_code, sales_staff_code=payload.sales_staff_code, status=payload.status, total_amount=total, note=payload.note)
    db.add(order)
    for item in payload.items:
        amount = item.amount if item.amount is not None else item.quantity_ton * item.unit_price * (1 - item.discount_pct / 100)
        db.add(IIPOrderItem(order_code=payload.order_code, product_code=item.product_code, quantity_ton=item.quantity_ton, unit_price=item.unit_price, discount_pct=item.discount_pct, amount=amount, note=item.note))
    db.commit(); db.refresh(order)
    return {**row_to_dict(order), "items": [row_to_dict(i) for i in db.query(IIPOrderItem).filter(IIPOrderItem.order_code == payload.order_code).all()]}


@router.post("/payments")
def create_payment(payload: PaymentIn, db: Session = Depends(get_db)):
    obj = IIPPayment(**schema_dump(payload))
    db.add(obj)
    # Auto allocate to debts oldest first when no manual match is given.
    remaining = payload.amount
    debts = db.query(IIPDebt).filter(IIPDebt.dealer_code == payload.dealer_code, IIPDebt.status != "closed").order_by(IIPDebt.due_date.asc()).all()
    for debt in debts:
        if remaining <= 0:
            break
        open_amount = max(0.0, debt.original_amount - debt.paid_amount)
        pay = min(open_amount, remaining)
        debt.paid_amount += pay
        remaining -= pay
        if debt.paid_amount >= debt.original_amount - 1:
            debt.status = "closed"
    db.commit(); db.refresh(obj)
    return {**row_to_dict(obj), "auto_allocated": payload.amount - remaining, "unallocated": remaining}


@router.post("/invoices")
def create_invoice(payload: InvoiceIn, db: Session = Depends(get_db)):
    subtotal = payload.subtotal
    if subtotal == 0 and payload.order_code:
        order = db.query(IIPOrder).filter(IIPOrder.order_code == payload.order_code).first()
        subtotal = order.total_amount if order else 0
    vat_amount = subtotal * payload.vat_rate
    total = payload.total_amount if payload.total_amount is not None else subtotal + vat_amount
    obj = IIPOutputInvoice(invoice_number=payload.invoice_number, invoice_date=payload.invoice_date, dealer_code=payload.dealer_code, order_code=payload.order_code, subtotal=subtotal, vat_rate=payload.vat_rate, vat_amount=vat_amount, total_amount=total, status=payload.status, note=payload.note)
    db.add(obj); db.commit(); db.refresh(obj)
    return row_to_dict(obj)


@router.post("/warehouse-slips")
def create_warehouse_slip(payload: WarehouseSlipIn, db: Session = Depends(get_db)):
    order = db.query(IIPOrder).filter(IIPOrder.order_code == payload.order_code).first()
    if not order:
        raise HTTPException(status_code=404, detail="Không tìm thấy order_code")
    qr = build_qr_code(payload.slip_code, payload.order_code)
    obj = IIPWarehouseSlip(slip_code=payload.slip_code, order_code=payload.order_code, slip_date=payload.slip_date, status=payload.status, qr_code=qr, note=payload.note)
    order.status = "exported" if payload.status in {"exported", "picked"} else order.status
    db.add(obj); db.commit(); db.refresh(obj)
    return row_to_dict(obj)


@router.post("/deliveries")
def create_delivery(payload: DeliveryIn, db: Session = Depends(get_db)):
    obj = IIPDelivery(**schema_dump(payload))
    db.add(obj); db.commit(); db.refresh(obj)
    return row_to_dict(obj)


@router.post("/deliveries/{delivery_code}/confirm-delivered")
def confirm_delivery(delivery_code: str, payload: DeliveryConfirmIn, db: Session = Depends(get_db)):
    obj = db.query(IIPDelivery).filter(IIPDelivery.delivery_code == delivery_code).first()
    if not obj:
        raise HTTPException(status_code=404, detail="Không tìm thấy chuyến giao")
    obj.status = "delivered"
    obj.delivered_at = datetime.utcnow()
    obj.gps_lat = payload.gps_lat
    obj.gps_lng = payload.gps_lng
    obj.photo_url = payload.photo_url
    obj.note = payload.note
    order = db.query(IIPOrder).filter(IIPOrder.order_code == obj.order_code).first()
    if order:
        order.status = "delivered"
    db.commit(); db.refresh(obj)
    return row_to_dict(obj)


@router.post("/vas-targets")
def create_vas_target(payload: VASTargetIn, db: Session = Depends(get_db)):
    data = schema_dump(payload, exclude={"milestones"})
    data["milestone_json"] = json.dumps(payload.milestones or [], ensure_ascii=False)
    obj = upsert_by_key(db, IIPVASTarget, "year", data)
    db.commit(); db.refresh(obj)
    return row_to_dict(obj)


# =========================
# Business APIs / alerts
# =========================

@router.get("/dealers/{dealer_code}/wallet")
def dealer_wallet(dealer_code: str, db: Session = Depends(get_db)):
    dealer = db.query(IIPDealer).filter(IIPDealer.code == dealer_code).first()
    if not dealer:
        raise HTTPException(status_code=404, detail="Không tìm thấy đại lý")
    debt = current_debt(db, dealer_code)
    limit = dealer_limit(db, dealer_code)
    payments = float(db.query(func.coalesce(func.sum(IIPPayment.amount), 0.0)).filter(IIPPayment.dealer_code == dealer_code).scalar() or 0)
    orders = float(db.query(func.coalesce(func.sum(IIPOrder.total_amount), 0.0)).filter(IIPOrder.dealer_code == dealer_code).scalar() or 0)
    limit_amount = limit.limit_amount if limit else 0
    return {
        "dealer": row_to_dict(dealer),
        "total_orders": orders,
        "total_payments": payments,
        "current_debt": debt,
        "credit_limit": limit_amount,
        "available_credit": max(0.0, limit_amount - debt),
        "over_limit": debt > limit_amount if limit else False,
        "rank": limit.rank if limit else dealer.rank,
        "policy": credit_policy(limit.rank if limit else dealer.rank),
    }


def credit_policy(rank: str) -> str:
    return {
        "A": "Ưu tiên giao hàng, chiết khấu cao nhất",
        "B": "Chính sách bình thường",
        "C": "Yêu cầu đặt cọc trước khi giao hàng",
        "D": "COD/thanh toán 100% trước khi nhận hàng",
    }.get((rank or "B").upper(), "Cần quản lý xét duyệt")


@router.get("/risk/overdue-debts")
def overdue_debts(db: Session = Depends(get_db)):
    today = date.today()
    rows = db.query(IIPDebt).filter(IIPDebt.status != "closed", IIPDebt.due_date < today).all()
    items = []
    for d in rows:
        outstanding = d.original_amount - d.paid_amount
        if outstanding <= 0:
            continue
        items.append({**row_to_dict(d), "outstanding": outstanding, "days_overdue": (today - d.due_date).days})
    return {"total_overdue": sum(i["outstanding"] for i in items), "items": sorted(items, key=lambda x: x["days_overdue"], reverse=True)}


@router.get("/risk/credit-limit-violations")
def credit_limit_violations(db: Session = Depends(get_db)):
    items = []
    for limit in db.query(IIPCreditLimit).all():
        debt = current_debt(db, limit.dealer_code)
        if debt > limit.limit_amount:
            items.append({"dealer_code": limit.dealer_code, "current_debt": debt, "credit_limit": limit.limit_amount, "exceeded_by": debt - limit.limit_amount, "rank": limit.rank})
    return {"count": len(items), "items": sorted(items, key=lambda x: x["exceeded_by"], reverse=True)}


@router.get("/risk/low-price-sales")
def low_price_sales(db: Session = Depends(get_db)):
    items = []
    orders = {o.order_code: o for o in db.query(IIPOrder).all()}
    dealers = {d.code: d for d in db.query(IIPDealer).all()}
    for item in db.query(IIPOrderItem).all():
        order = orders.get(item.order_code)
        dealer = dealers.get(order.dealer_code) if order else None
        floor = latest_floor(db, item.product_code, dealer.province if dealer else None, order.order_date if order else None)
        if not floor:
            continue
        min_price = floor.floor_price * (1 - (floor.allowed_discount_pct or 0) / 100)
        if item.unit_price < min_price:
            loss = (min_price - item.unit_price) * item.quantity_ton
            items.append({
                "order_code": item.order_code,
                "dealer_code": order.dealer_code if order else None,
                "sales_staff_code": order.sales_staff_code if order else None,
                "product_code": item.product_code,
                "quantity_ton": item.quantity_ton,
                "unit_price": item.unit_price,
                "floor_price": floor.floor_price,
                "allowed_min_price": min_price,
                "loss_estimate": loss,
                "severity": "red" if loss >= 50_000_000 else "yellow",
            })
    return {"count": len(items), "estimated_loss": sum(i["loss_estimate"] for i in items), "items": sorted(items, key=lambda x: x["loss_estimate"], reverse=True)}


@router.post("/orders/{order_code}/check-before-approve")
def check_order_before_approve(order_code: str, db: Session = Depends(get_db)):
    order = db.query(IIPOrder).filter(IIPOrder.order_code == order_code).first()
    if not order:
        raise HTTPException(status_code=404, detail="Không tìm thấy đơn hàng")
    items = db.query(IIPOrderItem).filter(IIPOrderItem.order_code == order_code).all()
    dealer = db.query(IIPDealer).filter(IIPDealer.code == order.dealer_code).first()
    limit = dealer_limit(db, order.dealer_code)
    debt = current_debt(db, order.dealer_code)
    total = sum(float(i.amount or 0) for i in items)
    warnings = []
    blocked = False
    if limit and debt + total > limit.limit_amount:
        blocked = True
        warnings.append({"type": "credit_limit", "message": "Đơn hàng làm đại lý vượt hạn mức tín dụng", "current_debt": debt, "order_total": total, "credit_limit": limit.limit_amount})
    for i in items:
        floor = latest_floor(db, i.product_code, dealer.province if dealer else None, order.order_date)
        if floor:
            min_price = floor.floor_price * (1 - (floor.allowed_discount_pct or 0) / 100)
            if i.unit_price < min_price:
                blocked = True
                warnings.append({"type": "low_price", "product_code": i.product_code, "unit_price": i.unit_price, "allowed_min_price": min_price, "loss_estimate": (min_price - i.unit_price) * i.quantity_ton})
    return {"order_code": order_code, "can_approve": not blocked, "blocked": blocked, "warnings": warnings}


@router.post("/orders/{order_code}/approve")
def approve_order(order_code: str, override: bool = False, db: Session = Depends(get_db)):
    check = check_order_before_approve(order_code, db)
    order = db.query(IIPOrder).filter(IIPOrder.order_code == order_code).first()
    if check["blocked"] and not override:
        order.status = "blocked"
        db.commit()
        return {"approved": False, "status": "blocked", "need_override": True, "check": check}
    order.status = "approved"
    # Create debt if not exists for this order.
    if not db.query(IIPDebt).filter(IIPDebt.order_code == order_code).first():
        limit = dealer_limit(db, order.dealer_code)
        due = order.order_date + timedelta(days=limit.debt_term_days if limit else 30)
        db.add(IIPDebt(dealer_code=order.dealer_code, order_code=order_code, debt_date=order.order_date, due_date=due, original_amount=order.total_amount, paid_amount=0, status="open", note="Tự sinh từ đơn hàng đã duyệt"))
    db.commit()
    return {"approved": True, "status": "approved", "check": check}


@router.get("/reconcile/4-way")
def reconcile_4_way(db: Session = Depends(get_db)):
    orders = db.query(IIPOrder).all()
    result = []
    for o in orders:
        slip = db.query(IIPWarehouseSlip).filter(IIPWarehouseSlip.order_code == o.order_code).first()
        invoice = db.query(IIPOutputInvoice).filter(IIPOutputInvoice.order_code == o.order_code).first()
        debt = db.query(IIPDebt).filter(IIPDebt.order_code == o.order_code).first()
        paid = float(db.query(func.coalesce(func.sum(IIPPayment.amount), 0.0)).filter(IIPPayment.matched_order_code == o.order_code).scalar() or 0)
        issues = []
        if o.status in {"exported", "delivered"} and not invoice:
            issues.append("Xuất kho/giao hàng nhưng chưa có hóa đơn")
        if invoice and not slip:
            issues.append("Có hóa đơn nhưng chưa có phiếu xuất kho")
        if debt and paid < min(debt.original_amount, o.total_amount) and debt.due_date < date.today():
            issues.append("Công nợ quá hạn/chưa thu đủ tiền")
        if paid > 0 and not invoice:
            issues.append("Có tiền về nhưng chưa khớp hóa đơn")
        result.append({
            "order_code": o.order_code,
            "dealer_code": o.dealer_code,
            "order_amount": o.total_amount,
            "warehouse_slip": bool(slip),
            "invoice": invoice.invoice_number if invoice else None,
            "debt_outstanding": (debt.original_amount - debt.paid_amount) if debt else None,
            "matched_payment": paid,
            "status": "ok" if not issues else "warning",
            "issues": issues,
        })
    return {"count": len(result), "warnings": sum(1 for r in result if r["issues"]), "items": result}


@router.get("/vas/progress")
def vas_progress(year: Optional[int] = None, db: Session = Depends(get_db)):
    year = year or date.today().year
    target = db.query(IIPVASTarget).filter(IIPVASTarget.year == year).first()
    start = date(year, 1, 1)
    end = date(year, 12, 31)
    vas_products = [p.code for p in db.query(IIPSteelProduct).filter(IIPSteelProduct.brand.ilike("%VAS%")).all()]
    sold = 0.0
    if vas_products:
        sold = float(db.query(func.coalesce(func.sum(IIPOrderItem.quantity_ton), 0.0)).join(IIPOrder, IIPOrder.order_code == IIPOrderItem.order_code).filter(IIPOrder.order_date >= start, IIPOrder.order_date <= end, IIPOrderItem.product_code.in_(vas_products), IIPOrder.status.in_(["approved", "exported", "delivered"])).scalar() or 0)
    days_passed = max(1, (min(date.today(), end) - start).days + 1)
    forecast = sold / days_passed * 365
    target_ton = target.target_ton if target else 0
    gap = max(0.0, target_ton - forecast)
    bonus_at_risk = 0.0
    if target and forecast < target.target_ton:
        bonus_at_risk = target.bonus_amount
    return {
        "year": year,
        "target_ton": target_ton,
        "sold_ton": sold,
        "forecast_end_year_ton": forecast,
        "gap_ton": gap,
        "completion_pct": (sold / target_ton * 100) if target_ton else None,
        "bonus_amount": target.bonus_amount if target else 0,
        "bonus_at_risk": bonus_at_risk,
        "risk_level": "red" if gap > target_ton * 0.1 and target_ton else "yellow" if gap > 0 else "green",
        "recommendation": vas_recommendation(gap, target_ton),
    }


def vas_recommendation(gap: float, target: float) -> str:
    if target <= 0:
        return "Chưa có mục tiêu VAS. Hãy nhập /iip/vas-targets trước."
    if gap <= 0:
        return "Đang đủ tiến độ để đạt thưởng VAS. Tiếp tục giữ nhịp bán và kiểm soát giá sàn."
    if gap > target * 0.1:
        return "Nguy cơ hụt thưởng cao. Cần đẩy sản lượng VAS tại tỉnh còn dư hạn mức, dùng chiết khấu có kiểm soát và ưu tiên đại lý hạng A/B."
    return "Có nguy cơ hụt nhẹ. Theo dõi hằng tuần và điều chỉnh giá/chiết khấu ở khu vực bán chậm."


@router.get("/cashflow/forecast")
def cashflow_forecast(days: int = 30, db: Session = Depends(get_db)):
    today = date.today()
    end = today + timedelta(days=days)
    receivable_due = float(db.query(func.coalesce(func.sum(IIPDebt.original_amount - IIPDebt.paid_amount), 0.0)).filter(IIPDebt.status != "closed", IIPDebt.due_date <= end).scalar() or 0)
    overdue = float(db.query(func.coalesce(func.sum(IIPDebt.original_amount - IIPDebt.paid_amount), 0.0)).filter(IIPDebt.status != "closed", IIPDebt.due_date < today).scalar() or 0)
    recent_payment_30d = float(db.query(func.coalesce(func.sum(IIPPayment.amount), 0.0)).filter(IIPPayment.payment_date >= today - timedelta(days=30)).scalar() or 0)
    expected_collection = max(0.0, receivable_due - overdue * 0.35)  # conservative risk haircut
    return {
        "period_days": days,
        "receivable_due": receivable_due,
        "overdue": overdue,
        "recent_payment_30d": recent_payment_30d,
        "expected_collection_conservative": expected_collection,
        "risk_note": "Khoản quá hạn được haircut 35% để cảnh báo dòng tiền thận trọng.",
    }


@router.get("/chairman/morning-report")
def chairman_morning_report(db: Session = Depends(get_db)):
    today = date.today()
    due_today = float(db.query(func.coalesce(func.sum(IIPDebt.original_amount - IIPDebt.paid_amount), 0.0)).filter(IIPDebt.status != "closed", IIPDebt.due_date <= today).scalar() or 0)
    overdue = overdue_debts(db)
    credit = credit_limit_violations(db)
    low_price = low_price_sales(db)
    vas = vas_progress(today.year, db)
    cash = cashflow_forecast(30, db)
    top_risks = []
    if overdue["total_overdue"] > 0:
        top_risks.append({"type": "overdue_debt", "amount": overdue["total_overdue"], "message": f"Nợ quá hạn {overdue['total_overdue']:,.0f} VND"})
    if credit["count"]:
        top_risks.append({"type": "credit_limit", "count": credit["count"], "message": f"{credit['count']} đại lý vượt hạn mức"})
    if low_price["count"]:
        top_risks.append({"type": "low_price", "amount": low_price["estimated_loss"], "message": f"{low_price['count']} đơn/dòng bán dưới giá sàn"})
    if vas["bonus_at_risk"]:
        top_risks.append({"type": "vas_bonus", "amount": vas["bonus_at_risk"], "message": "Có nguy cơ hụt thưởng doanh số VAS"})
    return {
        "title": "BÁO CÁO ĐẦU NGÀY - CHỦ TỊCH",
        "date": today.isoformat(),
        "today_need_collect": due_today,
        "overdue_debt": overdue["total_overdue"],
        "dealers_over_credit_limit": credit["items"][:5],
        "low_price_sales": low_price["items"][:5],
        "vas_bonus_risk": vas,
        "cashflow_30d": cash,
        "top_risks": top_risks,
        "executive_summary": [r["message"] for r in top_risks] or ["Chưa phát hiện rủi ro lớn từ dữ liệu hiện có."],
    }


@router.get("/staff/profit-ranking")
def staff_profit_ranking(db: Session = Depends(get_db)):
    rows = []
    staff_codes = [s.code for s in db.query(IIPSalesStaff).all()]
    for code in staff_codes:
        order_total = float(db.query(func.coalesce(func.sum(IIPOrder.total_amount), 0.0)).filter(IIPOrder.sales_staff_code == code).scalar() or 0)
        low_loss = sum(i["loss_estimate"] for i in low_price_sales(db)["items"] if i.get("sales_staff_code") == code)
        rows.append({"sales_staff_code": code, "sales_amount": order_total, "low_price_loss_estimate": low_loss, "score": order_total - low_loss})
    return {"items": sorted(rows, key=lambda x: x["score"], reverse=True)}


@router.post("/demo/seed")
def seed_iip_demo(db: Session = Depends(get_db)):
    """Nạp dữ liệu demo nhỏ để test nhanh morning report."""
    upsert_by_key(db, IIPSalesStaff, "code", {"code": "NV001", "name": "Nguyễn Văn A", "region": "Tây Bắc"})
    upsert_by_key(db, IIPDealer, "code", {"code": "DL_SONLA", "name": "Đại lý Sơn La", "province": "Sơn La", "sales_staff_code": "NV001", "rank": "B"})
    upsert_by_key(db, IIPSteelProduct, "code", {"code": "VAS_D10", "name": "Thép VAS D10", "brand": "VAS", "category": "thép cây"})
    upsert_by_key(db, IIPCreditLimit, "dealer_code", {"dealer_code": "DL_SONLA", "limit_amount": 1_000_000_000, "debt_term_days": 30, "rank": "B"})
    db.add(IIPPriceFloor(product_code="VAS_D10", province="Sơn La", effective_from=date(date.today().year, 1, 1), floor_price=15_000_000, allowed_discount_pct=1))
    if not db.query(IIPOrder).filter(IIPOrder.order_code == "ORD_DEMO_001").first():
        order = IIPOrder(order_code="ORD_DEMO_001", order_date=date.today(), dealer_code="DL_SONLA", sales_staff_code="NV001", status="approved", total_amount=1_470_000_000, note="Demo bán dưới giá sàn")
        db.add(order)
        db.add(IIPOrderItem(order_code="ORD_DEMO_001", product_code="VAS_D10", quantity_ton=100, unit_price=14_700_000, discount_pct=0, amount=1_470_000_000))
        db.add(IIPDebt(dealer_code="DL_SONLA", order_code="ORD_DEMO_001", debt_date=date.today() - timedelta(days=40), due_date=date.today() - timedelta(days=10), original_amount=1_470_000_000, paid_amount=300_000_000, status="open"))
    upsert_by_key(db, IIPVASTarget, "year", {"year": date.today().year, "target_ton": 50_000, "bonus_amount": 1_800_000_000, "milestone_json": "[]"})
    db.commit()
    return {"ok": True, "message": "Đã nạp demo. Mở GET /iip/chairman/morning-report để xem báo cáo sáng."}

# ============================================================
# V2 backend upgrade: auth/RBAC, Excel templates, AI Ask,
# approval workflow, geofence, churn risk, exports, schedules
# ============================================================
from fastapi import Header, Query
from fastapi.responses import StreamingResponse, PlainTextResponse
from sqlalchemy import Boolean


class IIPUser(Base, IIPTimestampMixin):
    __tablename__ = "iip_users"
    id = Column(Integer, primary_key=True, index=True)
    username = Column(String, unique=True, index=True, nullable=False)
    full_name = Column(String, nullable=False)
    role = Column(String, default="chairman", index=True, nullable=False)  # chairman/accounting/sales/warehouse/driver/dealer/admin
    dealer_code = Column(String, nullable=True, index=True)
    sales_staff_code = Column(String, nullable=True, index=True)
    api_key = Column(String, unique=True, index=True, nullable=True)
    is_active = Column(Boolean, default=True, nullable=False)


class IIPApprovalException(Base, IIPTimestampMixin):
    __tablename__ = "iip_approval_exceptions"
    id = Column(Integer, primary_key=True, index=True)
    order_code = Column(String, index=True, nullable=False)
    exception_type = Column(String, index=True, nullable=False)  # credit_limit/low_price/manual
    reason = Column(Text, nullable=False)
    requested_by = Column(String, nullable=True)
    approved_by = Column(String, nullable=True)
    status = Column(String, default="pending", nullable=False)  # pending/approved/rejected
    risk_snapshot_json = Column(Text, nullable=True)


class IIPGeofenceRule(Base, IIPTimestampMixin):
    __tablename__ = "iip_geofence_rules"
    id = Column(Integer, primary_key=True, index=True)
    delivery_code = Column(String, index=True, nullable=False)
    center_lat = Column(Float, nullable=False)
    center_lng = Column(Float, nullable=False)
    radius_km = Column(Float, default=2.0, nullable=False)
    rule_name = Column(String, default="delivery-point", nullable=False)
    is_active = Column(Boolean, default=True, nullable=False)


class IIPReportSchedule(Base, IIPTimestampMixin):
    __tablename__ = "iip_report_schedules"
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String, unique=True, index=True, nullable=False)
    report_type = Column(String, default="morning_report", nullable=False)
    channel = Column(String, default="telegram", nullable=False)  # telegram/zalo/email/webhook
    send_time = Column(String, default="06:30", nullable=False)
    recipient = Column(String, nullable=True)
    is_active = Column(Boolean, default=True, nullable=False)


class IIPNotificationLog(Base, IIPTimestampMixin):
    __tablename__ = "iip_notification_logs"
    id = Column(Integer, primary_key=True, index=True)
    report_type = Column(String, index=True, nullable=False)
    channel = Column(String, nullable=False)
    recipient = Column(String, nullable=True)
    content = Column(Text, nullable=False)
    status = Column(String, default="draft", nullable=False)


Base.metadata.create_all(bind=engine)


class UserIn(BaseModel):
    username: str
    full_name: str
    role: str = "chairman"
    dealer_code: Optional[str] = None
    sales_staff_code: Optional[str] = None
    api_key: Optional[str] = None
    is_active: bool = True


class ApprovalExceptionIn(BaseModel):
    exception_type: str = "manual"
    reason: str
    requested_by: Optional[str] = None
    approved_by: Optional[str] = None


class GeofenceRuleIn(BaseModel):
    delivery_code: str
    center_lat: float
    center_lng: float
    radius_km: float = 2.0
    rule_name: str = "delivery-point"
    is_active: bool = True


class ReportScheduleIn(BaseModel):
    name: str
    report_type: str = "morning_report"
    channel: str = "telegram"
    send_time: str = "06:30"
    recipient: Optional[str] = None
    is_active: bool = True


class AIAskIn(BaseModel):
    question: str
    role: str = "chairman"
    dealer_code: Optional[str] = None
    year: Optional[int] = None


TEMPLATE_COLUMNS: Dict[str, List[str]] = {
    "dealers": ["code", "name", "province", "phone", "address", "sales_staff_code", "rank", "status", "note"],
    "debts": ["dealer_code", "order_code", "debt_date", "due_date", "original_amount", "paid_amount", "status", "note"],
    "sales-staff": ["code", "name", "phone", "region", "status"],
    "price-floors": ["product_code", "province", "effective_from", "floor_price", "allowed_discount_pct", "note"],
    "credit-limits": ["dealer_code", "limit_amount", "debt_term_days", "rank", "require_deposit_pct", "note"],
    "orders": ["order_code", "order_date", "dealer_code", "sales_staff_code", "product_code", "quantity_ton", "unit_price", "discount_pct", "amount", "status", "note"],
    "payments": ["payment_date", "dealer_code", "amount", "bank_ref", "matched_order_code", "matched_invoice_number", "note"],
    "invoices": ["invoice_number", "invoice_date", "dealer_code", "order_code", "subtotal", "vat_rate", "total_amount", "status", "note"],
    "products": ["code", "name", "brand", "category", "unit", "status"],
    "warehouse-slips": ["slip_code", "order_code", "slip_date", "status", "note"],
    "deliveries": ["delivery_code", "order_code", "driver_name", "driver_phone", "route_note", "status", "note"],
    "vas-targets": ["year", "target_ton", "bonus_amount", "note"],
}


ROLE_PERMISSIONS = {
    "admin": {"*"},
    "chairman": {"read_all", "approve_exception", "view_report"},
    "accounting": {"read_all", "write_accounting", "import_data"},
    "sales": {"read_own", "write_order"},
    "warehouse": {"read_order", "write_warehouse"},
    "driver": {"read_delivery", "write_delivery"},
    "dealer": {"read_own", "write_order"},
}


def get_current_iip_user(x_iip_api_key: Optional[str] = Header(default=None), db: Session = Depends(get_db)) -> Optional[IIPUser]:
    if not x_iip_api_key:
        return None
    return db.query(IIPUser).filter(IIPUser.api_key == x_iip_api_key, IIPUser.is_active == True).first()  # noqa: E712


def haversine_km(lat1: float, lng1: float, lat2: float, lng2: float) -> float:
    import math
    radius = 6371.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lng2 - lng1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * radius * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def format_money(v: float) -> str:
    return f"{float(v or 0):,.0f} VND".replace(",", ".")


def safe_call(fn, *args, **kwargs):
    try:
        return fn(*args, **kwargs)
    except TypeError:
        # Supports direct calls to FastAPI route functions that have db as first/second arg.
        return fn(*args, **kwargs)


@router.get("/v2/status")
def iip_v2_status(db: Session = Depends(get_db)):
    base = iip_status(db)
    base["stage"] = "Giai đoạn 1 backend v2 - demo-ready core"
    base["v2_added"] = [
        "RBAC user/api-key model",
        "Excel template download for 8 dữ liệu đầu vào",
        "AI Ask router hỏi dữ liệu thật bằng tiếng Việt",
        "Approval workflow cho vượt hạn mức/bán dưới giá sàn",
        "Dealer churn risk",
        "Geofence check",
        "CSV export",
        "Report schedule + notification draft",
        "Completion score theo kế hoạch IIP",
    ]
    base["counts"].update({
        "users": db.query(IIPUser).count(),
        "approval_exceptions": db.query(IIPApprovalException).count(),
        "geofence_rules": db.query(IIPGeofenceRule).count(),
        "report_schedules": db.query(IIPReportSchedule).count(),
    })
    return base


@router.post("/users")
def create_iip_user(payload: UserIn, db: Session = Depends(get_db)):
    data = schema_dump(payload)
    if not data.get("api_key"):
        data["api_key"] = f"iip_{payload.role}_{payload.username}_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}"
    obj = upsert_by_key(db, IIPUser, "username", data)
    db.commit(); db.refresh(obj)
    return {**row_to_dict(obj), "permissions": sorted(ROLE_PERMISSIONS.get(obj.role, set()))}


@router.get("/users")
def list_iip_users(db: Session = Depends(get_db)):
    return {"items": [{**row_to_dict(u), "permissions": sorted(ROLE_PERMISSIONS.get(u.role, set()))} for u in db.query(IIPUser).order_by(IIPUser.id.desc()).all()]}


@router.get("/auth/me")
def auth_me(user: Optional[IIPUser] = Depends(get_current_iip_user)):
    if not user:
        return {"authenticated": False, "message": "Gửi header X-IIP-API-Key để xác thực demo."}
    return {"authenticated": True, "user": row_to_dict(user), "permissions": sorted(ROLE_PERMISSIONS.get(user.role, set()))}


@router.get("/import/template/{data_type}")
def download_import_template(data_type: str):
    if data_type not in TEMPLATE_COLUMNS:
        raise HTTPException(status_code=404, detail="Không có template cho data_type này")
    try:
        from openpyxl import Workbook
    except Exception:
        raise HTTPException(status_code=500, detail="Thiếu openpyxl để tạo template Excel")
    wb = Workbook()
    ws = wb.active
    ws.title = data_type[:31]
    columns = TEMPLATE_COLUMNS[data_type]
    ws.append(columns)
    example = {c: "" for c in columns}
    if data_type == "dealers":
        example.update({"code": "DL_SONLA", "name": "Đại lý Sơn La", "province": "Sơn La", "sales_staff_code": "NV001", "rank": "B", "status": "active"})
    elif data_type == "orders":
        example.update({"order_code": "ORD001", "order_date": date.today().isoformat(), "dealer_code": "DL_SONLA", "sales_staff_code": "NV001", "product_code": "VAS_D10", "quantity_ton": 10, "unit_price": 15000000, "discount_pct": 0, "status": "draft"})
    elif data_type == "debts":
        example.update({"dealer_code": "DL_SONLA", "debt_date": date.today().isoformat(), "due_date": (date.today()+timedelta(days=30)).isoformat(), "original_amount": 100000000, "paid_amount": 0, "status": "open"})
    elif data_type == "price-floors":
        example.update({"product_code": "VAS_D10", "province": "ALL", "effective_from": date.today().isoformat(), "floor_price": 15000000, "allowed_discount_pct": 1})
    elif data_type == "credit-limits":
        example.update({"dealer_code": "DL_SONLA", "limit_amount": 1000000000, "debt_term_days": 30, "rank": "B", "require_deposit_pct": 0})
    elif data_type == "payments":
        example.update({"payment_date": date.today().isoformat(), "dealer_code": "DL_SONLA", "amount": 100000000, "bank_ref": "VCB001", "matched_order_code": "ORD001"})
    elif data_type == "invoices":
        example.update({"invoice_number": "INV001", "invoice_date": date.today().isoformat(), "dealer_code": "DL_SONLA", "order_code": "ORD001", "subtotal": 100000000, "vat_rate": 0.1, "total_amount": 110000000, "status": "issued"})
    elif data_type == "sales-staff":
        example.update({"code": "NV001", "name": "Nguyễn Văn A", "region": "Tây Bắc", "status": "active"})
    elif data_type == "vas-targets":
        example.update({"year": date.today().year, "target_ton": 50000, "bonus_amount": 1800000000})
    ws.append([example.get(c, "") for c in columns])
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"iip_template_{data_type}.xlsx"
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/export/{data_type}.csv")
def export_csv(data_type: str, db: Session = Depends(get_db)):
    if data_type not in MODEL_MAP:
        raise HTTPException(status_code=404, detail="data_type không tồn tại")
    model, _, _ = MODEL_MAP[data_type]
    rows = [row_to_dict(r) for r in db.query(model).limit(10000).all()]
    output = BytesIO()
    if not rows:
        output.write("".encode("utf-8-sig"))
    else:
        text_io = output
        headers = list(rows[0].keys())
        content = BytesIO()
        import io
        s = io.StringIO()
        writer = csv.DictWriter(s, fieldnames=headers)
        writer.writeheader(); writer.writerows(rows)
        output.write(s.getvalue().encode("utf-8-sig"))
    output.seek(0)
    return StreamingResponse(output, media_type="text/csv", headers={"Content-Disposition": f"attachment; filename=iip_{data_type}.csv"})


@router.post("/orders/{order_code}/override-credit-limit")
def override_order(order_code: str, payload: ApprovalExceptionIn, db: Session = Depends(get_db), user: Optional[IIPUser] = Depends(get_current_iip_user)):
    order = db.query(IIPOrder).filter(IIPOrder.order_code == order_code).first()
    if not order:
        raise HTTPException(status_code=404, detail="Không tìm thấy đơn hàng")
    check = check_order_before_approve(order_code, db)
    obj = IIPApprovalException(
        order_code=order_code,
        exception_type=payload.exception_type,
        reason=payload.reason,
        requested_by=payload.requested_by or (user.username if user else None),
        approved_by=payload.approved_by or (user.username if user and user.role in {"admin", "chairman"} else None),
        status="approved",
        risk_snapshot_json=json.dumps(check, ensure_ascii=False, default=str),
    )
    db.add(obj)
    order.status = "approved"
    if not db.query(IIPDebt).filter(IIPDebt.order_code == order_code).first():
        limit = dealer_limit(db, order.dealer_code)
        due = order.order_date + timedelta(days=limit.debt_term_days if limit else 30)
        db.add(IIPDebt(dealer_code=order.dealer_code, order_code=order_code, debt_date=order.order_date, due_date=due, original_amount=order.total_amount, paid_amount=0, status="open", note="Tự sinh từ duyệt ngoại lệ"))
    db.commit(); db.refresh(obj)
    return {"approved": True, "order_status": order.status, "exception": row_to_dict(obj), "check": check}


@router.get("/approval-exceptions")
def list_approval_exceptions(status: Optional[str] = None, db: Session = Depends(get_db)):
    q = db.query(IIPApprovalException)
    if status:
        q = q.filter(IIPApprovalException.status == status)
    return {"items": [row_to_dict(x) for x in q.order_by(IIPApprovalException.id.desc()).all()]}


@router.get("/dealers/churn-risk")
def dealer_churn_risk(days: int = 90, db: Session = Depends(get_db)):
    today = date.today()
    cutoff = today - timedelta(days=days)
    prev_start = cutoff - timedelta(days=days)
    items = []
    for dealer in db.query(IIPDealer).all():
        recent = float(db.query(func.coalesce(func.sum(IIPOrder.total_amount), 0.0)).filter(IIPOrder.dealer_code == dealer.code, IIPOrder.order_date >= cutoff).scalar() or 0)
        previous = float(db.query(func.coalesce(func.sum(IIPOrder.total_amount), 0.0)).filter(IIPOrder.dealer_code == dealer.code, IIPOrder.order_date >= prev_start, IIPOrder.order_date < cutoff).scalar() or 0)
        last_order = db.query(IIPOrder).filter(IIPOrder.dealer_code == dealer.code).order_by(IIPOrder.order_date.desc()).first()
        days_since = (today - last_order.order_date).days if last_order else None
        decline_pct = ((previous - recent) / previous * 100) if previous else (100 if not recent else 0)
        risk = "green"
        reasons = []
        if days_since is None or days_since > days:
            risk = "red"; reasons.append("Không có đơn hàng gần đây")
        elif decline_pct >= 50:
            risk = "red"; reasons.append("Doanh số giảm trên 50%")
        elif decline_pct >= 25:
            risk = "yellow"; reasons.append("Doanh số giảm trên 25%")
        if dealer.rank in {"C", "D"}:
            reasons.append("Xếp hạng tín dụng thấp")
        items.append({"dealer_code": dealer.code, "dealer_name": dealer.name, "province": dealer.province, "recent_sales": recent, "previous_sales": previous, "decline_pct": decline_pct, "days_since_last_order": days_since, "risk_level": risk, "reasons": reasons})
    return {"period_days": days, "items": sorted(items, key=lambda x: (x["risk_level"] != "red", x["risk_level"] != "yellow", -(x["decline_pct"] or 0)))}


@router.post("/geofence/rules")
def create_geofence_rule(payload: GeofenceRuleIn, db: Session = Depends(get_db)):
    obj = IIPGeofenceRule(**schema_dump(payload))
    db.add(obj); db.commit(); db.refresh(obj)
    return row_to_dict(obj)


@router.get("/deliveries/{delivery_code}/geofence-check")
def geofence_check(delivery_code: str, lat: float = Query(...), lng: float = Query(...), db: Session = Depends(get_db)):
    rules = db.query(IIPGeofenceRule).filter(IIPGeofenceRule.delivery_code == delivery_code, IIPGeofenceRule.is_active == True).all()  # noqa: E712
    if not rules:
        return {"delivery_code": delivery_code, "status": "no_rule", "message": "Chưa có geofence rule cho chuyến này"}
    checks = []
    outside = False
    for r in rules:
        dist = haversine_km(lat, lng, r.center_lat, r.center_lng)
        ok = dist <= r.radius_km
        outside = outside or not ok
        checks.append({"rule_name": r.rule_name, "distance_km": dist, "radius_km": r.radius_km, "inside": ok})
    return {"delivery_code": delivery_code, "status": "outside" if outside else "inside", "checks": checks, "alert": outside}


@router.get("/reports/chairman-message", response_class=PlainTextResponse)
def chairman_message(db: Session = Depends(get_db)):
    r = chairman_morning_report(db)
    lines = [f"{r['title']} ({r['date']})", "", f"Hôm nay cần thu: {format_money(r['today_need_collect'])}", f"Nợ quá hạn: {format_money(r['overdue_debt'])}"]
    if r["dealers_over_credit_limit"]:
        lines.append(f"Đại lý vượt hạn mức: {len(r['dealers_over_credit_limit'])}")
        for i, x in enumerate(r["dealers_over_credit_limit"][:3], 1):
            lines.append(f"{i}. {x['dealer_code']} vượt {format_money(x['exceeded_by'])}")
    if r["low_price_sales"]:
        lines.append(f"Đơn bán dưới giá sàn: {len(r['low_price_sales'])}")
        for i, x in enumerate(r["low_price_sales"][:3], 1):
            lines.append(f"{i}. {x['order_code']} - {x['sales_staff_code']} mất ước tính {format_money(x['loss_estimate'])}")
    vas = r["vas_bonus_risk"]
    lines.append(f"VAS: đã bán {vas['sold_ton']:,.1f} tấn / mục tiêu {vas['target_ton']:,.1f} tấn; rủi ro: {vas['risk_level']}")
    if vas.get("bonus_at_risk"):
        lines.append(f"Nguy cơ mất thưởng VAS: {format_money(vas['bonus_at_risk'])}")
    lines.append("")
    lines.append("Tóm tắt: " + "; ".join(r["executive_summary"]))
    return "\n".join(lines)


@router.post("/reports/schedules")
def create_report_schedule(payload: ReportScheduleIn, db: Session = Depends(get_db)):
    obj = upsert_by_key(db, IIPReportSchedule, "name", schema_dump(payload))
    db.commit(); db.refresh(obj)
    return row_to_dict(obj)


@router.get("/reports/schedules")
def list_report_schedules(db: Session = Depends(get_db)):
    return {"items": [row_to_dict(x) for x in db.query(IIPReportSchedule).order_by(IIPReportSchedule.id.desc()).all()]}


@router.post("/reports/notification-draft")
def create_notification_draft(channel: str = "telegram", recipient: Optional[str] = None, db: Session = Depends(get_db)):
    content = chairman_message(db)
    obj = IIPNotificationLog(report_type="morning_report", channel=channel, recipient=recipient, content=content, status="draft")
    db.add(obj); db.commit(); db.refresh(obj)
    return {"notification": row_to_dict(obj), "content": content}


@router.post("/ai/ask")
def iip_ai_ask(payload: AIAskIn, db: Session = Depends(get_db)):
    q = payload.question.strip().lower()
    if any(k in q for k in ["báo cáo sáng", "bao cao sang", "tổng quan", "tong quan", "hôm nay", "hom nay"]):
        return {"intent": "morning_report", "answer": chairman_morning_report(db)}
    if any(k in q for k in ["công nợ", "cong no", "nợ quá hạn", "no qua han"]):
        if payload.dealer_code:
            return {"intent": "dealer_wallet", "answer": dealer_wallet(payload.dealer_code, db)}
        return {"intent": "overdue_debts", "answer": overdue_debts(db)}
    if any(k in q for k in ["vượt hạn mức", "vuot han muc", "hạn mức", "han muc"]):
        return {"intent": "credit_limit", "answer": credit_limit_violations(db)}
    if any(k in q for k in ["giá sàn", "gia san", "bán thấp", "ban thap", "bất thường", "bat thuong"]):
        return {"intent": "low_price_sales", "answer": low_price_sales(db)}
    if any(k in q for k in ["vas", "thưởng", "thuong", "chỉ tiêu", "chi tieu"]):
        return {"intent": "vas_progress", "answer": vas_progress(payload.year, db)}
    if any(k in q for k in ["đối soát", "doi soat", "4 chiều", "4 chieu"]):
        return {"intent": "reconcile_4_way", "answer": reconcile_4_way(db)}
    if any(k in q for k in ["dòng tiền", "dong tien", "thiếu tiền", "thieu tien"]):
        return {"intent": "cashflow_forecast", "answer": cashflow_forecast(30, db)}
    if any(k in q for k in ["mất khách", "mat khach", "ngừng mua", "ngung mua", "chuyển đối thủ", "chuyen doi thu"]):
        return {"intent": "dealer_churn_risk", "answer": dealer_churn_risk(90, db)}
    return {
        "intent": "unknown",
        "answer": "Tôi chưa nhận diện được câu hỏi. Có thể hỏi: Hôm nay công nợ thế nào? Ai vượt hạn mức? VAS có hụt thưởng không? Đơn nào bán dưới giá sàn? Đối soát 4 chiều tháng này thế nào?",
    }


@router.get("/mobile/tasks/{role}")
def mobile_tasks(role: str, db: Session = Depends(get_db)):
    role = role.lower()
    if role == "chairman":
        return {"role": role, "tasks": ["Xem báo cáo sáng", "Duyệt ngoại lệ vượt hạn mức", "Xem VAS bonus risk", "Xem top rủi ro"]}
    if role == "accounting":
        return {"role": role, "tasks": ["Nhập công nợ", "Nhập giao dịch ngân hàng", "Nhập hóa đơn", "Đối soát 4 chiều"], "pending_reconcile_warnings": reconcile_4_way(db)["warnings"]}
    if role == "warehouse":
        orders = db.query(IIPOrder).filter(IIPOrder.status == "approved").limit(50).all()
        return {"role": role, "tasks": ["Tạo phiếu xuất QR", "Xác nhận xuất kho"], "approved_orders_waiting_export": [row_to_dict(o) for o in orders]}
    if role == "driver":
        deliveries = db.query(IIPDelivery).filter(IIPDelivery.status.in_(["planned", "picked"])).limit(50).all()
        return {"role": role, "tasks": ["Xem chuyến giao", "Cập nhật GPS", "Upload ảnh giao hàng"], "deliveries": [row_to_dict(d) for d in deliveries]}
    if role == "sales":
        return {"role": role, "tasks": ["Tạo đơn hàng", "Kiểm tra hạn mức đại lý", "Theo dõi đại lý có nguy cơ mất khách", "Xem bán dưới giá sàn"]}
    if role == "dealer":
        return {"role": role, "tasks": ["Xem ví công nợ", "Đặt hàng", "Xem hạn mức còn lại"]}
    return {"role": role, "tasks": []}


@router.get("/roadmap/completion-score")
def roadmap_completion_score(db: Session = Depends(get_db)):
    modules = [
        {"module": "1. Công nợ/dòng tiền/nhân viên", "backend_pct": 85, "done": ["công nợ", "hạn mức", "dealer wallet", "bán dưới giá sàn", "duyệt đơn", "ngoại lệ"], "missing": ["lãi thật chuẩn theo chi phí", "workflow nhiều cấp"]},
        {"module": "2. Hóa đơn/VAT/đối soát", "backend_pct": 78, "done": ["hóa đơn đầu ra", "payment", "đối soát 4 chiều", "CSV export"], "missing": ["kết nối hóa đơn điện tử thật", "VAT filing sâu"]},
        {"module": "3. VAS bonus", "backend_pct": 78, "done": ["target", "forecast", "bonus risk", "AI Ask"], "missing": ["nhiều bậc thưởng nâng cao", "tối ưu vùng bán"]},
        {"module": "4. Xuất kho/vận chuyển", "backend_pct": 72, "done": ["phiếu QR", "delivery", "GPS/ảnh", "geofence check"], "missing": ["mobile scan QR thật", "routing/ghép xe tối ưu"]},
        {"module": "5. Đại lý cấp 2/tín dụng", "backend_pct": 80, "done": ["dealer wallet", "rank", "credit policy", "churn risk"], "missing": ["portal đại lý đầy đủ", "hành vi mua theo chủng loại"]},
    ]
    avg = round(sum(m["backend_pct"] for m in modules) / len(modules), 1)
    product_ready = 62
    demo_ready = 85
    return {"backend_completion_pct": avg, "demo_ready_pct": demo_ready, "product_ready_pct": product_ready, "modules": modules, "next_best_step": "Làm frontend Dashboard Chủ tịch + luồng import Excel mẫu để demo khách."}


@router.post("/demo/seed-v2")
def seed_iip_demo_v2(db: Session = Depends(get_db)):
    seed_iip_demo(db)
    upsert_by_key(db, IIPUser, "username", {"username": "chairman", "full_name": "Chủ tịch Demo", "role": "chairman", "api_key": "demo-chairman-key", "is_active": True})
    upsert_by_key(db, IIPReportSchedule, "name", {"name": "Báo cáo sáng 6h30", "report_type": "morning_report", "channel": "telegram", "send_time": "06:30", "recipient": "demo", "is_active": True})
    if not db.query(IIPDelivery).filter(IIPDelivery.delivery_code == "DEL_DEMO_001").first():
        db.add(IIPDelivery(delivery_code="DEL_DEMO_001", order_code="ORD_DEMO_001", driver_name="Tài xế Demo", driver_phone="0900000000", route_note="Kho -> Sơn La", status="planned"))
    if not db.query(IIPGeofenceRule).filter(IIPGeofenceRule.delivery_code == "DEL_DEMO_001").first():
        db.add(IIPGeofenceRule(delivery_code="DEL_DEMO_001", center_lat=21.3256, center_lng=103.9188, radius_km=5, rule_name="Điểm giao Sơn La"))
    db.commit()
    return {"ok": True, "message": "Đã seed demo v2", "api_key": "demo-chairman-key", "try": ["GET /iip/v2/status", "GET /iip/roadmap/completion-score", "POST /iip/ai/ask", "GET /iip/reports/chairman-message"]}

# ============================================================
# V3 production backend upgrade: JWT auth, role guard, audit log,
# validated Excel import preview/commit, scheduler tick, backups,
# deployment metadata. Kept self-contained and compatible with V2.
# ============================================================
import base64
import hashlib
import hmac
import os
import secrets
import time
from pathlib import Path
from fastapi import Body, Request
from fastapi.responses import JSONResponse


class IIPAuthUser(Base, IIPTimestampMixin):
    __tablename__ = "iip_auth_users"
    id = Column(Integer, primary_key=True, index=True)
    username = Column(String, unique=True, index=True, nullable=False)
    full_name = Column(String, nullable=False)
    role = Column(String, default="chairman", index=True, nullable=False)
    dealer_code = Column(String, nullable=True, index=True)
    sales_staff_code = Column(String, nullable=True, index=True)
    password_hash = Column(Text, nullable=False)
    is_active = Column(Boolean, default=True, nullable=False)
    last_login_at = Column(DateTime, nullable=True)


class IIPAuditLog(Base, IIPTimestampMixin):
    __tablename__ = "iip_audit_logs"
    id = Column(Integer, primary_key=True, index=True)
    actor_username = Column(String, nullable=True, index=True)
    actor_role = Column(String, nullable=True, index=True)
    action = Column(String, nullable=False, index=True)
    resource_type = Column(String, nullable=True, index=True)
    resource_id = Column(String, nullable=True, index=True)
    request_path = Column(Text, nullable=True)
    ip_address = Column(String, nullable=True)
    before_json = Column(Text, nullable=True)
    after_json = Column(Text, nullable=True)
    status = Column(String, default="ok", nullable=False)
    note = Column(Text, nullable=True)


class IIPImportBatch(Base, IIPTimestampMixin):
    __tablename__ = "iip_import_batches"
    id = Column(Integer, primary_key=True, index=True)
    batch_code = Column(String, unique=True, index=True, nullable=False)
    data_type = Column(String, index=True, nullable=False)
    filename = Column(String, nullable=True)
    total_rows = Column(Integer, default=0, nullable=False)
    valid_rows = Column(Integer, default=0, nullable=False)
    invalid_rows = Column(Integer, default=0, nullable=False)
    status = Column(String, default="preview", nullable=False)  # preview/committed/cancelled
    rows_json = Column(Text, nullable=False)
    errors_json = Column(Text, nullable=True)
    created_by = Column(String, nullable=True)
    committed_by = Column(String, nullable=True)
    committed_at = Column(DateTime, nullable=True)


class IIPNotificationChannel(Base, IIPTimestampMixin):
    __tablename__ = "iip_notification_channels"
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String, unique=True, index=True, nullable=False)
    channel = Column(String, default="webhook", nullable=False)  # telegram/email/zalo/webhook
    endpoint = Column(Text, nullable=True)
    recipient = Column(String, nullable=True)
    token_env = Column(String, nullable=True)
    is_active = Column(Boolean, default=True, nullable=False)


Base.metadata.create_all(bind=engine)


ROLE_PERMISSIONS = {
    "admin": {"*"},
    "chairman": {"read_all", "approve", "ai", "report", "export"},
    "accounting": {"read_finance", "import", "reconcile", "report", "export"},
    "sales": {"read_sales", "create_order", "read_dealer"},
    "warehouse": {"warehouse", "read_order"},
    "driver": {"delivery"},
    "dealer": {"dealer_self", "create_order"},
}


class RegisterUserIn(BaseModel):
    username: str = Field(..., min_length=3)
    password: str = Field(..., min_length=6)
    full_name: str
    role: str = "chairman"
    dealer_code: Optional[str] = None
    sales_staff_code: Optional[str] = None
    is_active: bool = True


class LoginIn(BaseModel):
    username: str
    password: str


class NotificationChannelIn(BaseModel):
    name: str
    channel: str = "webhook"
    endpoint: Optional[str] = None
    recipient: Optional[str] = None
    token_env: Optional[str] = None
    is_active: bool = True


class SchedulerTickIn(BaseModel):
    now: Optional[str] = None  # HH:MM override for testing
    dry_run: bool = True


class OrderCreateIn(BaseModel):
    order_code: str
    dealer_code: str
    sales_staff_code: Optional[str] = None
    product_code: str
    quantity_ton: float
    unit_price: float
    discount_pct: float = 0
    note: Optional[str] = None


class OrderUpdateStatusIn(BaseModel):
    status: str
    note: Optional[str] = None


def _b64url(data: bytes) -> str:
    return base64.urlsafe_b64encode(data).decode().rstrip("=")


def _b64url_decode(data: str) -> bytes:
    return base64.urlsafe_b64decode(data + "=" * (-len(data) % 4))


def _secret() -> bytes:
    return os.getenv("IIP_JWT_SECRET", "dev-change-me-iip-jwt-secret").encode()


def hash_password(password: str, salt: Optional[str] = None) -> str:
    salt = salt or secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode(), salt.encode(), 120_000)
    return f"pbkdf2_sha256${salt}${digest.hex()}"


def verify_password(password: str, stored: str) -> bool:
    try:
        algo, salt, digest = stored.split("$", 2)
        if algo != "pbkdf2_sha256":
            return False
        return hmac.compare_digest(hash_password(password, salt).split("$", 2)[2], digest)
    except Exception:
        return False


def create_token(user: IIPAuthUser, ttl_seconds: int = 60 * 60 * 12) -> str:
    header = {"alg": "HS256", "typ": "JWT"}
    payload = {
        "sub": user.username,
        "role": user.role,
        "full_name": user.full_name,
        "dealer_code": user.dealer_code,
        "sales_staff_code": user.sales_staff_code,
        "iat": int(time.time()),
        "exp": int(time.time()) + ttl_seconds,
    }
    signing_input = f"{_b64url(json.dumps(header, separators=(',', ':')).encode())}.{_b64url(json.dumps(payload, separators=(',', ':')).encode())}"
    sig = hmac.new(_secret(), signing_input.encode(), hashlib.sha256).digest()
    return f"{signing_input}.{_b64url(sig)}"


def decode_token(token: str) -> Dict[str, Any]:
    try:
        header_b64, payload_b64, sig_b64 = token.split(".")
        signing_input = f"{header_b64}.{payload_b64}"
        expected = _b64url(hmac.new(_secret(), signing_input.encode(), hashlib.sha256).digest())
        if not hmac.compare_digest(expected, sig_b64):
            raise ValueError("bad signature")
        payload = json.loads(_b64url_decode(payload_b64))
        if int(payload.get("exp", 0)) < int(time.time()):
            raise ValueError("expired")
        return payload
    except Exception as exc:
        raise HTTPException(status_code=401, detail=f"Token không hợp lệ: {exc}")


def current_v3_user(authorization: Optional[str] = Header(None), db: Session = Depends(get_db)) -> IIPAuthUser:
    if not authorization or not authorization.lower().startswith("bearer "):
        raise HTTPException(status_code=401, detail="Thiếu Authorization: Bearer <token>")
    payload = decode_token(authorization.split(" ", 1)[1].strip())
    user = db.query(IIPAuthUser).filter(IIPAuthUser.username == payload.get("sub"), IIPAuthUser.is_active == True).first()  # noqa: E712
    if not user:
        raise HTTPException(status_code=401, detail="User không tồn tại hoặc đã bị khóa")
    return user


def has_permission(role: str, permission: str) -> bool:
    perms = ROLE_PERMISSIONS.get(role, set())
    return "*" in perms or permission in perms


def require_permission(permission: str):
    def dep(user: IIPAuthUser = Depends(current_v3_user)) -> IIPAuthUser:
        if not has_permission(user.role, permission):
            raise HTTPException(status_code=403, detail=f"Role {user.role} không có quyền {permission}")
        return user
    return dep


def audit(db: Session, action: str, user: Optional[IIPAuthUser] = None, resource_type: Optional[str] = None, resource_id: Optional[str] = None, request: Optional[Request] = None, before: Any = None, after: Any = None, status: str = "ok", note: Optional[str] = None):
    item = IIPAuditLog(
        actor_username=getattr(user, "username", None),
        actor_role=getattr(user, "role", None),
        action=action,
        resource_type=resource_type,
        resource_id=str(resource_id) if resource_id is not None else None,
        request_path=str(request.url.path) if request else None,
        ip_address=request.client.host if request and request.client else None,
        before_json=json.dumps(before, ensure_ascii=False, default=str) if before is not None else None,
        after_json=json.dumps(after, ensure_ascii=False, default=str) if after is not None else None,
        status=status,
        note=note,
    )
    db.add(item)


def validate_import_rows(data_type: str, rows: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    valid: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []
    required = {
        "dealers": ["code", "name"],
        "debts": ["dealer_code", "due_date", "original_amount"],
        "sales-staff": ["code", "name"],
        "price-floors": ["product_code", "floor_price"],
        "credit-limits": ["dealer_code", "limit_amount"],
        "orders": ["order_code", "dealer_code", "product_code", "quantity_ton", "unit_price"],
        "payments": ["dealer_code", "amount"],
        "invoices": ["invoice_number", "dealer_code", "total_amount"],
    }.get(data_type, [])
    for idx, row in enumerate(rows, start=2):
        row_errors = []
        for col in required:
            if row.get(col) in (None, ""):
                row_errors.append(f"Thiếu {col}")
        for col in ["amount", "original_amount", "paid_amount", "floor_price", "limit_amount", "quantity_ton", "unit_price", "discount_pct", "total_amount"]:
            if row.get(col) not in (None, ""):
                try:
                    float(row.get(col))
                except Exception:
                    row_errors.append(f"{col} phải là số")
        if row_errors:
            errors.append({"row": idx, "errors": row_errors, "data": row})
        else:
            valid.append(row)
    return valid, errors


@router.get("/v3/status")
def iip_v3_status(db: Session = Depends(get_db)):
    return {
        "version": "IIP Steel Backend V3",
        "status": "production-ready backend skeleton",
        "database_url_mode": "postgresql" if os.getenv("DATABASE_URL", "").startswith("postgres") else "sqlite/dev",
        "features": [
            "JWT-like auth không cần dependency ngoài",
            "RBAC theo role",
            "audit log",
            "Excel import preview/commit",
            "scheduler tick cho báo cáo 6h30",
            "notification channel config",
            "JSON backup/export",
            "Docker/PostgreSQL deployment files",
        ],
        "counts": {
            "auth_users": db.query(IIPAuthUser).count(),
            "audit_logs": db.query(IIPAuditLog).count(),
            "import_batches": db.query(IIPImportBatch).count(),
            "notification_channels": db.query(IIPNotificationChannel).count(),
        },
    }


@router.post("/v3/auth/bootstrap-admin")
def bootstrap_admin(payload: RegisterUserIn, request: Request, db: Session = Depends(get_db)):
    if db.query(IIPAuthUser).count() > 0:
        raise HTTPException(status_code=409, detail="Đã có user auth. Dùng /iip/v3/auth/login hoặc tạo user bằng admin.")
    payload.role = "admin"
    user = IIPAuthUser(username=payload.username, full_name=payload.full_name, role="admin", password_hash=hash_password(payload.password), is_active=True)
    db.add(user)
    audit(db, "bootstrap_admin", user, "auth_user", payload.username, request, after={"username": payload.username, "role": "admin"})
    db.commit(); db.refresh(user)
    return {"ok": True, "user": {"username": user.username, "full_name": user.full_name, "role": user.role}, "token": create_token(user)}


@router.post("/v3/auth/login")
def login(payload: LoginIn, request: Request, db: Session = Depends(get_db)):
    user = db.query(IIPAuthUser).filter(IIPAuthUser.username == payload.username, IIPAuthUser.is_active == True).first()  # noqa: E712
    if not user or not verify_password(payload.password, user.password_hash):
        audit(db, "login_failed", None, "auth_user", payload.username, request, status="failed")
        db.commit()
        raise HTTPException(status_code=401, detail="Sai username hoặc password")
    user.last_login_at = datetime.utcnow()
    audit(db, "login", user, "auth_user", user.username, request)
    db.commit()
    return {"access_token": create_token(user), "token_type": "bearer", "user": {"username": user.username, "full_name": user.full_name, "role": user.role}}


@router.get("/v3/auth/me")
def auth_me(user: IIPAuthUser = Depends(current_v3_user)):
    return {"username": user.username, "full_name": user.full_name, "role": user.role, "dealer_code": user.dealer_code, "sales_staff_code": user.sales_staff_code}


@router.post("/v3/users")
def create_v3_user(payload: RegisterUserIn, request: Request, db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("*"))):
    if payload.role not in ROLE_PERMISSIONS:
        raise HTTPException(status_code=400, detail=f"Role không hợp lệ: {payload.role}")
    if db.query(IIPAuthUser).filter(IIPAuthUser.username == payload.username).first():
        raise HTTPException(status_code=409, detail="Username đã tồn tại")
    user = IIPAuthUser(username=payload.username, full_name=payload.full_name, role=payload.role, dealer_code=payload.dealer_code, sales_staff_code=payload.sales_staff_code, password_hash=hash_password(payload.password), is_active=payload.is_active)
    db.add(user)
    audit(db, "create_user", actor, "auth_user", payload.username, request, after={"username": payload.username, "role": payload.role})
    db.commit(); db.refresh(user)
    return {"id": user.id, "username": user.username, "full_name": user.full_name, "role": user.role, "is_active": user.is_active}


@router.get("/v3/users")
def list_v3_users(db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("*"))):
    return {"items": [{"id": u.id, "username": u.username, "full_name": u.full_name, "role": u.role, "is_active": u.is_active, "dealer_code": u.dealer_code, "sales_staff_code": u.sales_staff_code} for u in db.query(IIPAuthUser).order_by(IIPAuthUser.id.desc()).all()]}


@router.post("/v3/import/{data_type}/preview")
async def import_preview_v3(data_type: str, file: UploadFile = File(...), request: Request = None, db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("import"))):
    if data_type not in MODEL_MAP and data_type != "orders":
        raise HTTPException(status_code=400, detail=f"data_type không hỗ trợ: {data_type}")
    content = await file.read()
    rows = parse_upload_rows(file.filename or "upload.xlsx", content)
    valid, errors = validate_import_rows(data_type, rows)
    batch_code = "IMP" + datetime.utcnow().strftime("%Y%m%d%H%M%S") + secrets.token_hex(3).upper()
    batch = IIPImportBatch(batch_code=batch_code, data_type=data_type, filename=file.filename, total_rows=len(rows), valid_rows=len(valid), invalid_rows=len(errors), rows_json=json.dumps(valid, ensure_ascii=False, default=str), errors_json=json.dumps(errors, ensure_ascii=False, default=str), created_by=actor.username)
    db.add(batch)
    audit(db, "import_preview", actor, "import_batch", batch_code, request, after={"data_type": data_type, "rows": len(rows), "valid": len(valid), "invalid": len(errors)})
    db.commit()
    return {"batch_code": batch_code, "data_type": data_type, "filename": file.filename, "total_rows": len(rows), "valid_rows": len(valid), "invalid_rows": len(errors), "sample_valid_rows": valid[:5], "errors": errors[:20], "next": f"POST /iip/v3/import/{batch_code}/commit"}


@router.post("/v3/import/{batch_code}/commit")
def import_commit_v3(batch_code: str, request: Request, db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("import"))):
    batch = db.query(IIPImportBatch).filter(IIPImportBatch.batch_code == batch_code).first()
    if not batch:
        raise HTTPException(status_code=404, detail="Không tìm thấy batch")
    if batch.status != "preview":
        raise HTTPException(status_code=409, detail=f"Batch đã ở trạng thái {batch.status}")
    rows = json.loads(batch.rows_json or "[]")
    created = updated = skipped = 0
    errors: List[Dict[str, Any]] = []
    data_type = batch.data_type
    for idx, row in enumerate(rows, start=2):
        try:
            if data_type == "orders":
                order_code = row.get("order_code") or row.get("code")
                order = db.query(IIPOrder).filter(IIPOrder.order_code == order_code).first()
                if not order:
                    order = IIPOrder(order_code=order_code, dealer_code=row.get("dealer_code"), sales_staff_code=row.get("sales_staff_code"), order_date=to_date(row.get("order_date"), date.today()) or date.today(), status=row.get("status", "draft"), note=row.get("note"))
                    db.add(order); created += 1
                else:
                    updated += 1
                amount = float(row.get("amount") or 0) or float(row.get("quantity_ton") or 0) * float(row.get("unit_price") or 0) * (1 - float(row.get("discount_pct") or 0) / 100)
                db.add(IIPOrderItem(order_code=order_code, product_code=row.get("product_code"), quantity_ton=float(row.get("quantity_ton") or 0), unit_price=float(row.get("unit_price") or 0), discount_pct=float(row.get("discount_pct") or 0), amount=amount, note=row.get("note")))
            else:
                model, schema, key_field = MODEL_MAP[data_type]
                payload = schema_dump(schema(**row))
                before = None
                if key_field and payload.get(key_field) is not None:
                    before = db.query(model).filter(getattr(model, key_field) == payload[key_field]).first()
                upsert_by_key(db, model, key_field, payload)
                created += 0 if before else 1
                updated += 1 if before else 0
        except Exception as exc:
            skipped += 1
            errors.append({"row": idx, "error": str(exc), "data": row})
    if data_type == "orders":
        for code, total in db.query(IIPOrderItem.order_code, func.sum(IIPOrderItem.amount)).group_by(IIPOrderItem.order_code).all():
            order = db.query(IIPOrder).filter(IIPOrder.order_code == code).first()
            if order:
                order.total_amount = float(total or 0)
    batch.status = "committed"
    batch.committed_by = actor.username
    batch.committed_at = datetime.utcnow()
    audit(db, "import_commit", actor, "import_batch", batch_code, request, after={"created": created, "updated": updated, "skipped": skipped, "errors": errors[:5]})
    db.commit()
    return {"batch_code": batch_code, "data_type": data_type, "created": created, "updated": updated, "skipped": skipped, "errors": errors[:20]}


@router.post("/v3/orders")
def create_order_v3(payload: OrderCreateIn, request: Request, db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("create_order"))):
    if db.query(IIPOrder).filter(IIPOrder.order_code == payload.order_code).first():
        raise HTTPException(status_code=409, detail="order_code đã tồn tại")
    amount = payload.quantity_ton * payload.unit_price * (1 - payload.discount_pct / 100)
    order = IIPOrder(order_code=payload.order_code, dealer_code=payload.dealer_code, sales_staff_code=payload.sales_staff_code or actor.sales_staff_code, order_date=date.today(), status="pending", total_amount=amount, note=payload.note)
    item = IIPOrderItem(order_code=payload.order_code, product_code=payload.product_code, quantity_ton=payload.quantity_ton, unit_price=payload.unit_price, discount_pct=payload.discount_pct, amount=amount, note=payload.note)
    db.add(order); db.add(item); db.flush()
    check = check_order_before_approve(payload.order_code, db)
    if check.get("blocked"):
        order.status = "blocked"
    audit(db, "create_order", actor, "order", payload.order_code, request, after={"order": row_to_dict(order), "check": check})
    db.commit(); db.refresh(order)
    return {"order": row_to_dict(order), "risk_check": check}


@router.patch("/v3/orders/{order_code}/status")
def update_order_status_v3(order_code: str, payload: OrderUpdateStatusIn, request: Request, db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("approve"))):
    order = db.query(IIPOrder).filter(IIPOrder.order_code == order_code).first()
    if not order:
        raise HTTPException(status_code=404, detail="Không tìm thấy đơn hàng")
    before = row_to_dict(order)
    order.status = payload.status
    if payload.note:
        order.note = ((order.note or "") + "\n" + payload.note).strip()
    audit(db, "update_order_status", actor, "order", order_code, request, before=before, after=row_to_dict(order))
    db.commit(); db.refresh(order)
    return row_to_dict(order)


@router.post("/v3/notifications/channels")
def upsert_notification_channel(payload: NotificationChannelIn, request: Request, db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("*"))):
    obj = upsert_by_key(db, IIPNotificationChannel, "name", schema_dump(payload))
    audit(db, "upsert_notification_channel", actor, "notification_channel", payload.name, request, after=schema_dump(payload))
    db.commit(); db.refresh(obj)
    return row_to_dict(obj)


@router.get("/v3/notifications/channels")
def list_notification_channels(db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("report"))):
    return {"items": [row_to_dict(x) for x in db.query(IIPNotificationChannel).order_by(IIPNotificationChannel.id.desc()).all()]}


@router.post("/v3/reports/scheduler/tick")
def scheduler_tick(payload: SchedulerTickIn, request: Request, db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("report"))):
    now_hm = payload.now or datetime.now().strftime("%H:%M")
    due = db.query(IIPReportSchedule).filter(IIPReportSchedule.is_active == True, IIPReportSchedule.send_time == now_hm).all()  # noqa: E712
    created_logs = []
    for schedule in due:
        content = chairman_message(db) if schedule.report_type == "morning_report" else json.dumps(chairman_morning_report(db), ensure_ascii=False, default=str)
        status = "dry_run" if payload.dry_run else "queued"
        log = IIPNotificationLog(report_type=schedule.report_type, channel=schedule.channel, recipient=schedule.recipient, content=content, status=status)
        db.add(log); db.flush()
        created_logs.append(row_to_dict(log))
    audit(db, "scheduler_tick", actor, "report_schedule", now_hm, request, after={"due": len(due), "dry_run": payload.dry_run})
    db.commit()
    return {"now": now_hm, "due_schedules": len(due), "dry_run": payload.dry_run, "logs": created_logs}


@router.get("/v3/audit-logs")
def list_audit_logs(limit: int = Query(100, ge=1, le=500), db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("*"))):
    return {"items": [row_to_dict(x) for x in db.query(IIPAuditLog).order_by(IIPAuditLog.id.desc()).limit(limit).all()]}


@router.get("/v3/backup/export")
def backup_export_v3(db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("export"))):
    tables = {
        "dealers": IIPDealer,
        "sales_staff": IIPSalesStaff,
        "products": IIPSteelProduct,
        "price_floors": IIPPriceFloor,
        "credit_limits": IIPCreditLimit,
        "debts": IIPDebt,
        "orders": IIPOrder,
        "order_items": IIPOrderItem,
        "payments": IIPPayment,
        "invoices": IIPOutputInvoice,
        "warehouse_slips": IIPWarehouseSlip,
        "deliveries": IIPDelivery,
        "vas_targets": IIPVASTarget,
        "approval_exceptions": IIPApprovalException,
        "geofence_rules": IIPGeofenceRule,
        "report_schedules": IIPReportSchedule,
    }
    payload = {name: [row_to_dict(x) for x in db.query(model).all()] for name, model in tables.items()}
    payload["exported_at"] = datetime.utcnow().isoformat()
    payload["exported_by"] = actor.username
    return JSONResponse(payload, headers={"Content-Disposition": "attachment; filename=iip_steel_backup.json"})


@router.get("/v3/deployment/checklist")
def deployment_checklist_v3():
    return {
        "required_env": ["DATABASE_URL", "IIP_JWT_SECRET", "CORS_ORIGINS"],
        "recommended_env": ["TELEGRAM_BOT_TOKEN", "TELEGRAM_CHAT_ID", "SMTP_HOST", "SMTP_USER", "SMTP_PASSWORD"],
        "steps": [
            "docker compose up -d db",
            "docker compose up --build app",
            "POST /iip/v3/auth/bootstrap-admin",
            "POST /iip/demo/seed-v2 nếu cần dữ liệu demo",
            "Tạo report schedule 06:30",
            "Tạo notification channel",
            "Import preview/commit 8 file Excel đầu vào",
        ],
        "health_urls": ["/health", "/iip/v3/status", "/iip/roadmap/completion-score"],
    }


@router.post("/demo/seed-v3")
def seed_iip_demo_v3(request: Request, db: Session = Depends(get_db)):
    seed_iip_demo_v2(db)
    if not db.query(IIPAuthUser).filter(IIPAuthUser.username == "admin").first():
        db.add(IIPAuthUser(username="admin", full_name="Admin Demo", role="admin", password_hash=hash_password("admin123"), is_active=True))
    if not db.query(IIPAuthUser).filter(IIPAuthUser.username == "chairman").first():
        db.add(IIPAuthUser(username="chairman", full_name="Chủ tịch Demo", role="chairman", password_hash=hash_password("chairman123"), is_active=True))
    upsert_by_key(db, IIPNotificationChannel, "name", {"name": "telegram_demo", "channel": "telegram", "recipient": "demo", "token_env": "TELEGRAM_BOT_TOKEN", "is_active": True})
    audit(db, "seed_v3", None, "demo", "seed-v3", request, after={"admin": "admin/admin123", "chairman": "chairman/chairman123"})
    db.commit()
    return {"ok": True, "message": "Đã seed demo V3", "login": [{"username": "admin", "password": "admin123"}, {"username": "chairman", "password": "chairman123"}], "try": ["POST /iip/v3/auth/login", "GET /iip/v3/status", "POST /iip/v3/import/{data_type}/preview", "POST /iip/v3/reports/scheduler/tick"]}

# ==========================================================
# V4 - Deep backend improvements for all 5 IIP phase-1 modules
# ==========================================================
# Focus:
# 1) Real profit by order/staff/dealer, cashflow forecast, approval workflow
# 2) Invoice XML parsing, VAT summary, legal reconciliation risk scoring
# 3) Multi-tier bonus programs, VAS simulation and recommendation
# 4) Signed QR verification, GPS route events, delivery photos/signature, route merge suggestions
# 5) Deep dealer credit scoring and dealer self-service API

import base64
import hashlib
import hmac
import math
import os
import xml.etree.ElementTree as ET
from statistics import mean


# -------------------------
# V4 database models
# -------------------------

class IIPCostPrice(Base, IIPTimestampMixin):
    __tablename__ = "iip_cost_prices"
    id = Column(Integer, primary_key=True, index=True)
    product_code = Column(String, index=True, nullable=False)
    province = Column(String, default="ALL", index=True, nullable=False)
    effective_from = Column(Date, default=date.today, nullable=False)
    cost_price = Column(Float, nullable=False)
    supplier = Column(String, nullable=True)
    note = Column(Text, nullable=True)


class IIPTransportCost(Base, IIPTimestampMixin):
    __tablename__ = "iip_transport_costs"
    id = Column(Integer, primary_key=True, index=True)
    order_code = Column(String, unique=True, index=True, nullable=False)
    province = Column(String, nullable=True, index=True)
    cost_amount = Column(Float, default=0, nullable=False)
    driver_name = Column(String, nullable=True)
    vehicle_plate = Column(String, nullable=True)
    note = Column(Text, nullable=True)


class IIPFinanceCostRule(Base, IIPTimestampMixin):
    __tablename__ = "iip_finance_cost_rules"
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String, unique=True, index=True, nullable=False)
    annual_rate_pct = Column(Float, default=12, nullable=False)
    is_default = Column(Integer, default=1, nullable=False)
    note = Column(Text, nullable=True)


class IIPOrderProfitSnapshot(Base, IIPTimestampMixin):
    __tablename__ = "iip_order_profit_snapshots"
    id = Column(Integer, primary_key=True, index=True)
    order_code = Column(String, unique=True, index=True, nullable=False)
    dealer_code = Column(String, index=True, nullable=False)
    sales_staff_code = Column(String, index=True, nullable=True)
    revenue = Column(Float, default=0, nullable=False)
    cost_of_goods = Column(Float, default=0, nullable=False)
    transport_cost = Column(Float, default=0, nullable=False)
    finance_cost = Column(Float, default=0, nullable=False)
    gross_profit = Column(Float, default=0, nullable=False)
    net_profit = Column(Float, default=0, nullable=False)
    margin_pct = Column(Float, default=0, nullable=False)
    risk_level = Column(String, default="green", nullable=False)
    detail_json = Column(Text, nullable=True)


class IIPApprovalRequest(Base, IIPTimestampMixin):
    __tablename__ = "iip_approval_requests"
    id = Column(Integer, primary_key=True, index=True)
    request_code = Column(String, unique=True, index=True, nullable=False)
    order_code = Column(String, index=True, nullable=True)
    dealer_code = Column(String, index=True, nullable=True)
    requested_by = Column(String, nullable=True)
    approval_type = Column(String, index=True, nullable=False)  # credit_limit/low_price/overdue/negative_margin/other
    reason = Column(Text, nullable=True)
    risk_json = Column(Text, nullable=True)
    status = Column(String, default="pending", index=True, nullable=False)  # pending/approved/rejected/cancelled
    decided_by = Column(String, nullable=True)
    decided_at = Column(DateTime, nullable=True)
    decision_note = Column(Text, nullable=True)


class IIPApprovalComment(Base, IIPTimestampMixin):
    __tablename__ = "iip_approval_comments"
    id = Column(Integer, primary_key=True, index=True)
    request_code = Column(String, index=True, nullable=False)
    author = Column(String, nullable=True)
    comment = Column(Text, nullable=False)
    attachment_url = Column(Text, nullable=True)


class IIPInvoiceXMLFile(Base, IIPTimestampMixin):
    __tablename__ = "iip_invoice_xml_files"
    id = Column(Integer, primary_key=True, index=True)
    file_code = Column(String, unique=True, index=True, nullable=False)
    filename = Column(String, nullable=True)
    invoice_number = Column(String, index=True, nullable=True)
    invoice_date = Column(Date, nullable=True)
    seller_tax_code = Column(String, nullable=True)
    buyer_tax_code = Column(String, nullable=True)
    dealer_code = Column(String, index=True, nullable=True)
    subtotal = Column(Float, default=0, nullable=False)
    vat_amount = Column(Float, default=0, nullable=False)
    total_amount = Column(Float, default=0, nullable=False)
    parse_status = Column(String, default="parsed", nullable=False)
    raw_xml = Column(Text, nullable=True)
    parsed_json = Column(Text, nullable=True)


class IIPVATRiskAlert(Base, IIPTimestampMixin):
    __tablename__ = "iip_vat_risk_alerts"
    id = Column(Integer, primary_key=True, index=True)
    alert_code = Column(String, unique=True, index=True, nullable=False)
    invoice_number = Column(String, index=True, nullable=True)
    order_code = Column(String, index=True, nullable=True)
    severity = Column(String, default="yellow", nullable=False)
    risk_type = Column(String, index=True, nullable=False)
    message = Column(Text, nullable=False)
    status = Column(String, default="open", nullable=False)
    resolved_by = Column(String, nullable=True)
    resolved_at = Column(DateTime, nullable=True)
    resolution_note = Column(Text, nullable=True)


class IIPSupplierBonusProgram(Base, IIPTimestampMixin):
    __tablename__ = "iip_supplier_bonus_programs"
    id = Column(Integer, primary_key=True, index=True)
    program_code = Column(String, unique=True, index=True, nullable=False)
    supplier = Column(String, default="VAS", index=True, nullable=False)
    year = Column(Integer, index=True, nullable=False)
    product_brand = Column(String, default="VAS", nullable=False)
    start_date = Column(Date, nullable=True)
    end_date = Column(Date, nullable=True)
    note = Column(Text, nullable=True)


class IIPBonusTier(Base, IIPTimestampMixin):
    __tablename__ = "iip_bonus_tiers"
    id = Column(Integer, primary_key=True, index=True)
    program_code = Column(String, index=True, nullable=False)
    tier_name = Column(String, nullable=False)
    target_ton = Column(Float, nullable=False)
    bonus_amount = Column(Float, default=0, nullable=False)
    bonus_rate_pct = Column(Float, default=0, nullable=False)
    note = Column(Text, nullable=True)


class IIPDeliveryLocation(Base, IIPTimestampMixin):
    __tablename__ = "iip_delivery_locations"
    id = Column(Integer, primary_key=True, index=True)
    delivery_code = Column(String, index=True, nullable=False)
    lat = Column(Float, nullable=False)
    lng = Column(Float, nullable=False)
    speed_kmh = Column(Float, nullable=True)
    accuracy_m = Column(Float, nullable=True)
    recorded_at = Column(DateTime, default=datetime.utcnow, nullable=False)
    note = Column(Text, nullable=True)


class IIPDeliveryEvent(Base, IIPTimestampMixin):
    __tablename__ = "iip_delivery_events"
    id = Column(Integer, primary_key=True, index=True)
    delivery_code = Column(String, index=True, nullable=False)
    event_type = Column(String, index=True, nullable=False)  # pickup_photo/delivery_photo/signature/geofence/scan
    payload_json = Column(Text, nullable=True)
    severity = Column(String, default="green", nullable=False)
    note = Column(Text, nullable=True)


class IIPVehicle(Base, IIPTimestampMixin):
    __tablename__ = "iip_vehicles"
    id = Column(Integer, primary_key=True, index=True)
    vehicle_code = Column(String, unique=True, index=True, nullable=False)
    plate_number = Column(String, nullable=False)
    max_ton = Column(Float, default=20, nullable=False)
    driver_name = Column(String, nullable=True)
    driver_phone = Column(String, nullable=True)
    status = Column(String, default="active", nullable=False)


class IIPDealerCreditScoreHistory(Base, IIPTimestampMixin):
    __tablename__ = "iip_dealer_credit_score_history"
    id = Column(Integer, primary_key=True, index=True)
    dealer_code = Column(String, index=True, nullable=False)
    score = Column(Float, default=0, nullable=False)
    rank = Column(String, default="D", nullable=False)
    factors_json = Column(Text, nullable=True)
    calculated_at = Column(DateTime, default=datetime.utcnow, nullable=False)


Base.metadata.create_all(bind=engine)


# -------------------------
# V4 schemas
# -------------------------

class CostPriceIn(BaseModel):
    product_code: str
    province: str = "ALL"
    effective_from: Optional[date] = None
    cost_price: float = Field(ge=0)
    supplier: Optional[str] = None
    note: Optional[str] = None


class TransportCostIn(BaseModel):
    order_code: str
    province: Optional[str] = None
    cost_amount: float = Field(default=0, ge=0)
    driver_name: Optional[str] = None
    vehicle_plate: Optional[str] = None
    note: Optional[str] = None


class ApprovalRequestIn(BaseModel):
    order_code: Optional[str] = None
    dealer_code: Optional[str] = None
    approval_type: str = "other"
    reason: Optional[str] = None
    risk: Dict[str, Any] = Field(default_factory=dict)


class ApprovalDecisionIn(BaseModel):
    note: Optional[str] = None


class ApprovalCommentIn(BaseModel):
    comment: str
    attachment_url: Optional[str] = None


class BonusProgramIn(BaseModel):
    program_code: str
    supplier: str = "VAS"
    year: int
    product_brand: str = "VAS"
    start_date: Optional[date] = None
    end_date: Optional[date] = None
    note: Optional[str] = None


class BonusTierIn(BaseModel):
    program_code: str
    tier_name: str
    target_ton: float = Field(gt=0)
    bonus_amount: float = Field(default=0, ge=0)
    bonus_rate_pct: float = Field(default=0, ge=0)
    note: Optional[str] = None


class VASSimulationIn(BaseModel):
    program_code: Optional[str] = None
    year: int = Field(default_factory=lambda: date.today().year)
    additional_ton: float = Field(default=0, ge=0)
    discount_pct: float = Field(default=0, ge=0)
    avg_sell_price: float = Field(default=0, ge=0)
    avg_cost_price: float = Field(default=0, ge=0)
    province: Optional[str] = None
    dealer_code: Optional[str] = None


class DeliveryLocationIn(BaseModel):
    lat: float
    lng: float
    speed_kmh: Optional[float] = None
    accuracy_m: Optional[float] = None
    recorded_at: Optional[datetime] = None
    note: Optional[str] = None


class DeliveryEventIn(BaseModel):
    event_type: str
    payload: Dict[str, Any] = Field(default_factory=dict)
    severity: str = "green"
    note: Optional[str] = None


class VehicleIn(BaseModel):
    vehicle_code: str
    plate_number: str
    max_ton: float = Field(default=20, gt=0)
    driver_name: Optional[str] = None
    driver_phone: Optional[str] = None
    status: str = "active"


class MergeRouteIn(BaseModel):
    province: Optional[str] = None
    delivery_date: Optional[date] = None
    vehicle_code: Optional[str] = None
    max_ton: Optional[float] = None
    order_codes: Optional[List[str]] = None


# -------------------------
# V4 helper functions
# -------------------------

def _effective_cost_price(db: Session, product_code: str, province: Optional[str], as_of: Optional[date] = None) -> float:
    as_of = as_of or date.today()
    q = db.query(IIPCostPrice).filter(IIPCostPrice.product_code == product_code, IIPCostPrice.effective_from <= as_of)
    rows = q.order_by(IIPCostPrice.effective_from.desc()).all()
    province = province or "ALL"
    for row in rows:
        if row.province == province:
            return float(row.cost_price or 0)
    for row in rows:
        if row.province == "ALL":
            return float(row.cost_price or 0)
    # fallback: floor price * 95%, better than returning zero for demo
    floor = latest_floor(db, product_code, province, as_of)
    return float(getattr(floor, "floor_price", 0) or 0) * 0.95 if floor else 0.0


def _annual_finance_rate(db: Session) -> float:
    rule = db.query(IIPFinanceCostRule).filter(IIPFinanceCostRule.is_default == 1).first()
    return float(rule.annual_rate_pct or 12) if rule else 12.0


def calculate_order_profit(db: Session, order_code: str, save: bool = True) -> Dict[str, Any]:
    order = db.query(IIPOrder).filter(IIPOrder.order_code == order_code).first()
    if not order:
        raise HTTPException(status_code=404, detail="Không tìm thấy đơn hàng")
    dealer = db.query(IIPDealer).filter(IIPDealer.code == order.dealer_code).first()
    province = dealer.province if dealer else None
    items = db.query(IIPOrderItem).filter(IIPOrderItem.order_code == order_code).all()
    revenue = sum(float(x.amount or 0) for x in items)
    cogs = 0.0
    details = []
    for item in items:
        cost_price = _effective_cost_price(db, item.product_code, province, order.order_date)
        item_cost = cost_price * float(item.quantity_ton or 0)
        cogs += item_cost
        details.append({
            "product_code": item.product_code,
            "quantity_ton": item.quantity_ton,
            "unit_price": item.unit_price,
            "amount": item.amount,
            "cost_price": cost_price,
            "item_cost": item_cost,
        })
    transport = db.query(IIPTransportCost).filter(IIPTransportCost.order_code == order_code).first()
    transport_cost = float(transport.cost_amount or 0) if transport else 0.0
    credit = db.query(IIPCreditLimit).filter(IIPCreditLimit.dealer_code == order.dealer_code).first()
    term_days = int(credit.debt_term_days or 30) if credit else 30
    finance_cost = revenue * (_annual_finance_rate(db) / 100.0) * term_days / 365.0
    gross_profit = revenue - cogs
    net_profit = gross_profit - transport_cost - finance_cost
    margin_pct = (net_profit / revenue * 100.0) if revenue else 0.0
    risk_level = "green"
    if net_profit < 0 or margin_pct < 0:
        risk_level = "red"
    elif margin_pct < 1:
        risk_level = "orange"
    elif margin_pct < 3:
        risk_level = "yellow"
    payload = {
        "order_code": order_code,
        "dealer_code": order.dealer_code,
        "sales_staff_code": order.sales_staff_code,
        "revenue": revenue,
        "cost_of_goods": cogs,
        "transport_cost": transport_cost,
        "finance_cost": finance_cost,
        "gross_profit": gross_profit,
        "net_profit": net_profit,
        "margin_pct": margin_pct,
        "risk_level": risk_level,
        "details": details,
    }
    if save:
        snap = db.query(IIPOrderProfitSnapshot).filter(IIPOrderProfitSnapshot.order_code == order_code).first()
        if not snap:
            snap = IIPOrderProfitSnapshot(order_code=order_code, dealer_code=order.dealer_code, sales_staff_code=order.sales_staff_code)
            db.add(snap)
        snap.revenue = revenue
        snap.cost_of_goods = cogs
        snap.transport_cost = transport_cost
        snap.finance_cost = finance_cost
        snap.gross_profit = gross_profit
        snap.net_profit = net_profit
        snap.margin_pct = margin_pct
        snap.risk_level = risk_level
        snap.detail_json = json.dumps(details, ensure_ascii=False, default=str)
    return payload


def _money(v: float) -> str:
    return f"{round(float(v or 0)):,.0f} VND".replace(",", ".")


def _simple_xml_text(root: ET.Element, names: List[str]) -> Optional[str]:
    # Vietnamese e-invoice XMLs have different namespaces/tags. Match by suffix/local name.
    wanted = {n.lower() for n in names}
    for elem in root.iter():
        tag = elem.tag.split("}")[-1].lower()
        if tag in wanted and elem.text and elem.text.strip():
            return elem.text.strip()
    return None


def parse_invoice_xml_content(xml_text: str) -> Dict[str, Any]:
    root = ET.fromstring(xml_text.encode("utf-8") if isinstance(xml_text, str) else xml_text)
    invoice_number = _simple_xml_text(root, ["SHDon", "SoHoaDon", "InvoiceNo", "invoice_number", "khhdon"])
    invoice_date_raw = _simple_xml_text(root, ["NLap", "NgayHoaDon", "InvoiceDate", "invoice_date"])
    seller_tax = _simple_xml_text(root, ["MST", "MSTNBan", "SellerTaxCode", "seller_tax_code"])
    buyer_tax = _simple_xml_text(root, ["MSTNMua", "BuyerTaxCode", "buyer_tax_code"])
    subtotal_raw = _simple_xml_text(root, ["TgTCThue", "TongTienTruocThue", "SubTotal", "subtotal"])
    vat_raw = _simple_xml_text(root, ["TgTThue", "TongTienThue", "VATAmount", "vat_amount"])
    total_raw = _simple_xml_text(root, ["TgTTTBSo", "TongTienThanhToan", "TotalAmount", "total_amount"])

    def fnum(x: Optional[str]) -> float:
        if not x:
            return 0.0
        cleaned = re.sub(r"[^0-9,.-]", "", x).replace(".", "").replace(",", ".")
        try:
            return float(cleaned)
        except Exception:
            return 0.0

    return {
        "invoice_number": invoice_number,
        "invoice_date": to_date(invoice_date_raw, None),
        "seller_tax_code": seller_tax,
        "buyer_tax_code": buyer_tax,
        "subtotal": fnum(subtotal_raw),
        "vat_amount": fnum(vat_raw),
        "total_amount": fnum(total_raw),
    }


def reconciliation_detail(db: Session) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    orders = db.query(IIPOrder).all()
    for order in orders:
        slips = db.query(IIPWarehouseSlip).filter(IIPWarehouseSlip.order_code == order.order_code).all()
        invoices = db.query(IIPOutputInvoice).filter(IIPOutputInvoice.order_code == order.order_code).all()
        debts = db.query(IIPDebt).filter(IIPDebt.order_code == order.order_code).all()
        payments = db.query(IIPPayment).filter(IIPPayment.matched_order_code == order.order_code).all()
        exported = len(slips) > 0
        invoiced_amount = sum(float(x.total_amount or 0) for x in invoices)
        debt_amount = sum(float(x.original_amount or 0) - float(x.paid_amount or 0) for x in debts)
        paid_amount = sum(float(x.amount or 0) for x in payments)
        severity = "green"
        messages = []
        if exported and not invoices:
            severity = "yellow"; messages.append("Đã xuất kho nhưng chưa có hóa đơn")
        if invoices and not exported:
            severity = "red"; messages.append("Đã xuất hóa đơn nhưng chưa có phiếu xuất kho")
        if abs(float(order.total_amount or 0) - invoiced_amount) > 1000 and invoices:
            severity = "orange" if severity != "red" else severity; messages.append("Số tiền hóa đơn lệch đơn hàng")
        if debt_amount > 0 and paid_amount <= 0:
            severity = "yellow" if severity == "green" else severity; messages.append("Có công nợ nhưng chưa thấy tiền về")
        if paid_amount > float(order.total_amount or 0) * 1.05 and order.total_amount:
            severity = "orange" if severity != "red" else severity; messages.append("Tiền về lớn bất thường so với đơn")
        rows.append({
            "order_code": order.order_code,
            "dealer_code": order.dealer_code,
            "order_amount": order.total_amount,
            "exported": exported,
            "invoice_count": len(invoices),
            "invoiced_amount": invoiced_amount,
            "debt_open_amount": debt_amount,
            "paid_amount": paid_amount,
            "severity": severity,
            "messages": messages or ["Khớp cơ bản"],
        })
    return rows


def _bonus_tiers(db: Session, program_code: Optional[str], year: int) -> Tuple[Optional[IIPSupplierBonusProgram], List[IIPBonusTier]]:
    program = None
    if program_code:
        program = db.query(IIPSupplierBonusProgram).filter(IIPSupplierBonusProgram.program_code == program_code).first()
    if not program:
        program = db.query(IIPSupplierBonusProgram).filter(IIPSupplierBonusProgram.year == year).order_by(IIPSupplierBonusProgram.id.desc()).first()
    if program:
        tiers = db.query(IIPBonusTier).filter(IIPBonusTier.program_code == program.program_code).order_by(IIPBonusTier.target_ton.asc()).all()
        return program, tiers
    # fallback to old target table
    target = db.query(IIPVASTarget).filter(IIPVASTarget.year == year).first()
    if not target:
        return None, []
    pseudo = IIPSupplierBonusProgram(program_code=f"VAS-{year}", supplier="VAS", year=year, product_brand="VAS")
    tier = IIPBonusTier(program_code=pseudo.program_code, tier_name="Mục tiêu chính", target_ton=target.target_ton, bonus_amount=target.bonus_amount)
    return pseudo, [tier]


def _current_vas_ton(db: Session, year: int) -> float:
    start = date(year, 1, 1)
    end = date(year, 12, 31)
    rows = db.query(IIPOrderItem).join(IIPOrder, IIPOrderItem.order_code == IIPOrder.order_code).join(IIPSteelProduct, IIPOrderItem.product_code == IIPSteelProduct.code).filter(IIPSteelProduct.brand == "VAS", IIPOrder.order_date >= start, IIPOrder.order_date <= end).all()
    return sum(float(x.quantity_ton or 0) for x in rows)


def _signed_qr_payload(slip: IIPWarehouseSlip) -> str:
    secret = os.getenv("IIP_QR_SECRET") or os.getenv("IIP_JWT_SECRET") or "dev-iip-secret"
    core = f"{slip.slip_code}|{slip.order_code}|{slip.slip_date}|{slip.status}"
    sig = hmac.new(secret.encode(), core.encode(), hashlib.sha256).hexdigest()[:24]
    return base64.urlsafe_b64encode(f"{core}|{sig}".encode()).decode()


def _verify_qr_payload(token: str) -> Dict[str, Any]:
    try:
        raw = base64.urlsafe_b64decode(token.encode()).decode()
        slip_code, order_code, slip_date, status, sig = raw.split("|", 4)
        secret = os.getenv("IIP_QR_SECRET") or os.getenv("IIP_JWT_SECRET") or "dev-iip-secret"
        core = f"{slip_code}|{order_code}|{slip_date}|{status}"
        expected = hmac.new(secret.encode(), core.encode(), hashlib.sha256).hexdigest()[:24]
        return {"valid": hmac.compare_digest(sig, expected), "slip_code": slip_code, "order_code": order_code, "slip_date": slip_date, "status": status}
    except Exception as exc:
        return {"valid": False, "error": str(exc)}


def dealer_credit_factors(db: Session, dealer_code: str) -> Dict[str, Any]:
    debts = db.query(IIPDebt).filter(IIPDebt.dealer_code == dealer_code).all()
    orders = db.query(IIPOrder).filter(IIPOrder.dealer_code == dealer_code).all()
    now = date.today()
    open_debt = sum(max(0, float(d.original_amount or 0) - float(d.paid_amount or 0)) for d in debts if d.status != "closed")
    overdue_days = [max(0, (now - d.due_date).days) for d in debts if d.status != "closed" and d.due_date]
    overdue_amount = sum(max(0, float(d.original_amount or 0) - float(d.paid_amount or 0)) for d in debts if d.status != "closed" and d.due_date and d.due_date < now)
    total_debt = sum(float(d.original_amount or 0) for d in debts) or 1
    paid_on_time_ratio = max(0.0, 1.0 - (overdue_amount / total_debt))
    last_90 = now - timedelta(days=90)
    sales_90 = sum(float(o.total_amount or 0) for o in orders if o.order_date and o.order_date >= last_90)
    order_count_90 = len([o for o in orders if o.order_date and o.order_date >= last_90])
    limit = db.query(IIPCreditLimit).filter(IIPCreditLimit.dealer_code == dealer_code).first()
    limit_amount = float(limit.limit_amount or 0) if limit else 0
    usage_ratio = open_debt / limit_amount if limit_amount else 1.0
    exceptions = db.query(IIPApprovalRequest).filter(IIPApprovalRequest.dealer_code == dealer_code).count()
    avg_overdue = mean(overdue_days) if overdue_days else 0
    score = 1000.0
    score -= min(350, avg_overdue * 8)
    score -= min(250, max(0, usage_ratio - 0.8) * 500)
    score -= min(150, exceptions * 15)
    score += min(80, sales_90 / 1_000_000_000 * 20)
    score += paid_on_time_ratio * 120
    score += min(40, order_count_90 * 5)
    score = max(0, min(1000, score))
    rank = "A" if score >= 850 else "B" if score >= 700 else "C" if score >= 500 else "D"
    return {"dealer_code": dealer_code, "score": round(score, 2), "rank": rank, "open_debt": open_debt, "overdue_amount": overdue_amount, "avg_overdue_days": avg_overdue, "paid_on_time_ratio": paid_on_time_ratio, "sales_90_days": sales_90, "order_count_90_days": order_count_90, "credit_usage_ratio": usage_ratio, "approval_exceptions": exceptions}


# -------------------------
# V4 endpoints - Module 1: profit, cashflow, approvals
# -------------------------

@router.get("/v4/status")
def v4_status():
    return {"version": "v4", "name": "IIP Steel Backend Deep Module Upgrade", "modules": ["profit_and_cashflow", "vat_xml_and_legal_reconcile", "vas_bonus_simulation", "qr_gps_logistics", "dealer_credit_portal"]}


@router.post("/v4/cost-prices")
def upsert_cost_price_v4(payload: CostPriceIn, db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("import"))):
    data = schema_dump(payload)
    if data.get("effective_from") is None:
        data["effective_from"] = date.today()
    obj = IIPCostPrice(**data)
    db.add(obj); db.commit(); db.refresh(obj)
    return row_to_dict(obj)


@router.post("/v4/transport-costs")
def upsert_transport_cost_v4(payload: TransportCostIn, db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("import"))):
    obj = upsert_by_key(db, IIPTransportCost, "order_code", schema_dump(payload))
    db.commit(); db.refresh(obj)
    return row_to_dict(obj)


@router.get("/v4/profit/orders/{order_code}")
def order_profit_v4(order_code: str, db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("report"))):
    result = calculate_order_profit(db, order_code, save=True)
    db.commit()
    return result


@router.get("/v4/profit/orders")
def list_order_profit_v4(limit: int = Query(100, ge=1, le=500), refresh: bool = False, db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("report"))):
    if refresh:
        for order in db.query(IIPOrder).limit(limit).all():
            calculate_order_profit(db, order.order_code, save=True)
        db.commit()
    snaps = db.query(IIPOrderProfitSnapshot).order_by(IIPOrderProfitSnapshot.net_profit.asc()).limit(limit).all()
    return {"items": [row_to_dict(x) for x in snaps]}


@router.get("/v4/profit/staff-ranking")
def staff_profit_ranking_v4(refresh: bool = True, db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("report"))):
    if refresh:
        for order in db.query(IIPOrder).all():
            calculate_order_profit(db, order.order_code, save=True)
        db.commit()
    rows = db.query(IIPOrderProfitSnapshot).all()
    grouped: Dict[str, Dict[str, Any]] = {}
    for r in rows:
        key = r.sales_staff_code or "UNKNOWN"
        grouped.setdefault(key, {"sales_staff_code": key, "orders": 0, "revenue": 0, "net_profit": 0, "negative_margin_orders": 0})
        grouped[key]["orders"] += 1
        grouped[key]["revenue"] += float(r.revenue or 0)
        grouped[key]["net_profit"] += float(r.net_profit or 0)
        grouped[key]["negative_margin_orders"] += 1 if float(r.net_profit or 0) < 0 else 0
    items = sorted(grouped.values(), key=lambda x: x["net_profit"], reverse=True)
    return {"items": items}


@router.get("/v4/risk/negative-margin-orders")
def negative_margin_orders_v4(db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("report"))):
    for order in db.query(IIPOrder).all():
        calculate_order_profit(db, order.order_code, save=True)
    db.commit()
    rows = db.query(IIPOrderProfitSnapshot).filter(IIPOrderProfitSnapshot.net_profit < 0).order_by(IIPOrderProfitSnapshot.net_profit.asc()).all()
    return {"count": len(rows), "items": [row_to_dict(x) for x in rows]}


@router.get("/v4/cashflow/forecast")
def cashflow_forecast_v4(days: int = Query(30, ge=1, le=180), db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("report"))):
    today = date.today()
    end = today + timedelta(days=days)
    debts = db.query(IIPDebt).filter(IIPDebt.status != "closed", IIPDebt.due_date <= end).all()
    expected = []
    for d in debts:
        remaining = max(0, float(d.original_amount or 0) - float(d.paid_amount or 0))
        overdue = max(0, (today - d.due_date).days)
        probability = 0.9 if overdue == 0 else 0.65 if overdue <= 15 else 0.4 if overdue <= 45 else 0.2
        expected.append({"dealer_code": d.dealer_code, "order_code": d.order_code, "due_date": d.due_date, "remaining_amount": remaining, "collection_probability": probability, "expected_cash_in": remaining * probability, "overdue_days": overdue})
    daily: Dict[str, float] = {}
    for x in expected:
        k = str(x["due_date"])
        daily[k] = daily.get(k, 0) + x["expected_cash_in"]
    return {"days": days, "total_expected_cash_in": sum(x["expected_cash_in"] for x in expected), "daily": daily, "items": expected}


@router.post("/v4/approvals/request")
def create_approval_request_v4(payload: ApprovalRequestIn, request: Request, db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("create_order"))):
    code = "APR" + datetime.utcnow().strftime("%Y%m%d%H%M%S") + secrets.token_hex(2).upper()
    obj = IIPApprovalRequest(request_code=code, order_code=payload.order_code, dealer_code=payload.dealer_code, requested_by=actor.username, approval_type=payload.approval_type, reason=payload.reason, risk_json=json.dumps(payload.risk, ensure_ascii=False, default=str), status="pending")
    db.add(obj)
    audit(db, "approval_request", actor, "approval", code, request, after=schema_dump(payload))
    db.commit(); db.refresh(obj)
    return row_to_dict(obj)


@router.get("/v4/approvals/pending")
def pending_approvals_v4(db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("approve"))):
    rows = db.query(IIPApprovalRequest).filter(IIPApprovalRequest.status == "pending").order_by(IIPApprovalRequest.created_at.asc()).all()
    return {"items": [row_to_dict(x) for x in rows]}


@router.post("/v4/approvals/{request_code}/approve")
def approve_request_v4(request_code: str, payload: ApprovalDecisionIn, request: Request, db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("approve"))):
    obj = db.query(IIPApprovalRequest).filter(IIPApprovalRequest.request_code == request_code).first()
    if not obj:
        raise HTTPException(status_code=404, detail="Không tìm thấy yêu cầu duyệt")
    before = row_to_dict(obj)
    obj.status = "approved"; obj.decided_by = actor.username; obj.decided_at = datetime.utcnow(); obj.decision_note = payload.note
    if obj.order_code:
        order = db.query(IIPOrder).filter(IIPOrder.order_code == obj.order_code).first()
        if order and order.status in ["blocked", "pending"]:
            order.status = "approved"
    audit(db, "approval_approve", actor, "approval", request_code, request, before=before, after=row_to_dict(obj))
    db.commit(); db.refresh(obj)
    return row_to_dict(obj)


@router.post("/v4/approvals/{request_code}/reject")
def reject_request_v4(request_code: str, payload: ApprovalDecisionIn, request: Request, db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("approve"))):
    obj = db.query(IIPApprovalRequest).filter(IIPApprovalRequest.request_code == request_code).first()
    if not obj:
        raise HTTPException(status_code=404, detail="Không tìm thấy yêu cầu duyệt")
    before = row_to_dict(obj)
    obj.status = "rejected"; obj.decided_by = actor.username; obj.decided_at = datetime.utcnow(); obj.decision_note = payload.note
    audit(db, "approval_reject", actor, "approval", request_code, request, before=before, after=row_to_dict(obj))
    db.commit(); db.refresh(obj)
    return row_to_dict(obj)


@router.post("/v4/approvals/{request_code}/comments")
def add_approval_comment_v4(request_code: str, payload: ApprovalCommentIn, db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("approve"))):
    if not db.query(IIPApprovalRequest).filter(IIPApprovalRequest.request_code == request_code).first():
        raise HTTPException(status_code=404, detail="Không tìm thấy yêu cầu duyệt")
    c = IIPApprovalComment(request_code=request_code, author=actor.username, comment=payload.comment, attachment_url=payload.attachment_url)
    db.add(c); db.commit(); db.refresh(c)
    return row_to_dict(c)


# -------------------------
# V4 endpoints - Module 2: XML invoice, VAT, reconciliation legal risk
# -------------------------

@router.post("/v4/invoices/xml/upload")
async def upload_invoice_xml_v4(file: UploadFile = File(...), db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("import"))):
    content = (await file.read()).decode("utf-8", errors="ignore")
    try:
        parsed = parse_invoice_xml_content(content)
        status = "parsed"
    except Exception as exc:
        parsed = {"error": str(exc)}
        status = "failed"
    code = "XML" + datetime.utcnow().strftime("%Y%m%d%H%M%S") + secrets.token_hex(2).upper()
    obj = IIPInvoiceXMLFile(file_code=code, filename=file.filename, invoice_number=parsed.get("invoice_number"), invoice_date=parsed.get("invoice_date"), seller_tax_code=parsed.get("seller_tax_code"), buyer_tax_code=parsed.get("buyer_tax_code"), subtotal=float(parsed.get("subtotal") or 0), vat_amount=float(parsed.get("vat_amount") or 0), total_amount=float(parsed.get("total_amount") or 0), parse_status=status, raw_xml=content[:500000], parsed_json=json.dumps(parsed, ensure_ascii=False, default=str))
    db.add(obj); db.commit(); db.refresh(obj)
    return {"file": row_to_dict(obj), "parsed": parsed}


@router.get("/v4/vat/summary")
def vat_summary_v4(month: str = Query(..., pattern=r"^\d{4}-\d{2}$"), db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("report"))):
    y, m = [int(x) for x in month.split("-")]
    start = date(y, m, 1)
    end = date(y + (m // 12), (m % 12) + 1, 1)
    out_invoices = db.query(IIPOutputInvoice).filter(IIPOutputInvoice.invoice_date >= start, IIPOutputInvoice.invoice_date < end).all()
    xmls = db.query(IIPInvoiceXMLFile).filter(IIPInvoiceXMLFile.invoice_date >= start, IIPInvoiceXMLFile.invoice_date < end).all()
    output_vat = sum(float(x.vat_amount or 0) for x in out_invoices)
    # XMLs without a matched output invoice are treated as possible input VAT for MVP.
    output_numbers = {x.invoice_number for x in out_invoices}
    input_vat = sum(float(x.vat_amount or 0) for x in xmls if x.invoice_number not in output_numbers)
    payable = max(0, output_vat - input_vat)
    credit = max(0, input_vat - output_vat)
    return {"month": month, "output_invoice_count": len(out_invoices), "input_xml_count": len([x for x in xmls if x.invoice_number not in output_numbers]), "output_vat": output_vat, "input_vat": input_vat, "vat_payable": payable, "vat_credit": credit}


@router.get("/v4/reconcile/4-way/detail")
def reconcile_4way_detail_v4(db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("report"))):
    return {"items": reconciliation_detail(db)}


@router.get("/v4/reconcile/4-way/risk-score")
def reconcile_4way_risk_score_v4(db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("report"))):
    items = reconciliation_detail(db)
    weights = {"green": 0, "yellow": 30, "orange": 65, "red": 100}
    score = max([weights.get(x["severity"], 0) for x in items] or [0])
    return {"risk_score": score, "risk_level": "red" if score >= 90 else "orange" if score >= 60 else "yellow" if score >= 30 else "green", "total_orders": len(items), "by_level": {lvl: len([x for x in items if x["severity"] == lvl]) for lvl in ["green", "yellow", "orange", "red"]}, "top_risks": [x for x in items if x["severity"] != "green"][:20]}


@router.post("/v4/reconcile/4-way/resolve")
def resolve_reconcile_alert_v4(payload: Dict[str, Any] = Body(...), db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("approve"))):
    alert_code = payload.get("alert_code") or "VAT" + datetime.utcnow().strftime("%Y%m%d%H%M%S")
    alert = db.query(IIPVATRiskAlert).filter(IIPVATRiskAlert.alert_code == alert_code).first()
    if not alert:
        alert = IIPVATRiskAlert(alert_code=alert_code, order_code=payload.get("order_code"), invoice_number=payload.get("invoice_number"), risk_type=payload.get("risk_type", "reconcile"), severity=payload.get("severity", "yellow"), message=payload.get("message", "Đã ghi nhận xử lý đối soát"))
        db.add(alert)
    alert.status = "resolved"; alert.resolved_by = actor.username; alert.resolved_at = datetime.utcnow(); alert.resolution_note = payload.get("resolution_note")
    db.commit(); db.refresh(alert)
    return row_to_dict(alert)


# -------------------------
# V4 endpoints - Module 3: VAS multi-tier bonus and simulation
# -------------------------

@router.post("/v4/bonus/programs")
def create_bonus_program_v4(payload: BonusProgramIn, db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("import"))):
    obj = upsert_by_key(db, IIPSupplierBonusProgram, "program_code", schema_dump(payload))
    db.commit(); db.refresh(obj)
    return row_to_dict(obj)


@router.post("/v4/bonus/tiers")
def create_bonus_tier_v4(payload: BonusTierIn, db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("import"))):
    obj = IIPBonusTier(**schema_dump(payload))
    db.add(obj); db.commit(); db.refresh(obj)
    return row_to_dict(obj)


@router.get("/v4/bonus/progress")
def bonus_progress_v4(program_code: Optional[str] = None, year: int = Query(default_factory=lambda: date.today().year), db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("report"))):
    program, tiers = _bonus_tiers(db, program_code, year)
    sold = _current_vas_ton(db, year)
    next_tier = None
    current_bonus = 0.0
    for t in tiers:
        if sold >= float(t.target_ton or 0):
            current_bonus = max(current_bonus, float(t.bonus_amount or 0))
        elif next_tier is None:
            next_tier = t
    return {"program": row_to_dict(program) if program and getattr(program, "id", None) else {"program_code": getattr(program, "program_code", None), "year": year}, "sold_ton": sold, "current_bonus": current_bonus, "next_tier": row_to_dict(next_tier) if next_tier and getattr(next_tier, "id", None) else ({"tier_name": getattr(next_tier, "tier_name", None), "target_ton": getattr(next_tier, "target_ton", None), "bonus_amount": getattr(next_tier, "bonus_amount", None)} if next_tier else None), "gap_to_next_tier": max(0, float(next_tier.target_ton) - sold) if next_tier else 0, "tiers": [row_to_dict(t) if getattr(t, "id", None) else {"tier_name": t.tier_name, "target_ton": t.target_ton, "bonus_amount": t.bonus_amount} for t in tiers]}


@router.post("/v4/vas/simulate")
def vas_simulate_v4(payload: VASSimulationIn, db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("report"))):
    base = bonus_progress_v4(payload.program_code, payload.year, db, actor)  # type: ignore[arg-type]
    sold_after = float(base["sold_ton"] or 0) + payload.additional_ton
    _, tiers = _bonus_tiers(db, payload.program_code, payload.year)
    bonus_after = 0.0
    for t in tiers:
        if sold_after >= float(t.target_ton or 0):
            bonus_after = max(bonus_after, float(t.bonus_amount or 0))
    incremental_bonus = bonus_after - float(base["current_bonus"] or 0)
    sell_price = payload.avg_sell_price or 0
    cost_price = payload.avg_cost_price or (sell_price * 0.97 if sell_price else 0)
    discount_cost = payload.additional_ton * sell_price * payload.discount_pct / 100
    gross_margin = payload.additional_ton * (sell_price - cost_price)
    net_effect = gross_margin - discount_cost + incremental_bonus
    return {"base_sold_ton": base["sold_ton"], "sold_after": sold_after, "current_bonus": base["current_bonus"], "bonus_after": bonus_after, "incremental_bonus": incremental_bonus, "additional_gross_margin": gross_margin, "discount_cost": discount_cost, "net_effect": net_effect, "decision": "nên làm" if net_effect > 0 else "không nên làm nếu chỉ xét lợi nhuận ngắn hạn"}


@router.get("/v4/vas/recommendations/by-region")
def vas_recommend_region_v4(year: int = Query(default_factory=lambda: date.today().year), db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("report"))):
    start = date(year, 1, 1)
    rows = db.query(IIPOrderItem, IIPOrder, IIPDealer, IIPSteelProduct).join(IIPOrder, IIPOrderItem.order_code == IIPOrder.order_code).join(IIPDealer, IIPOrder.dealer_code == IIPDealer.code).join(IIPSteelProduct, IIPOrderItem.product_code == IIPSteelProduct.code).filter(IIPOrder.order_date >= start, IIPSteelProduct.brand == "VAS").all()
    by_province: Dict[str, float] = {}
    for item, order, dealer, product in rows:
        by_province[dealer.province or "UNKNOWN"] = by_province.get(dealer.province or "UNKNOWN", 0) + float(item.quantity_ton or 0)
    credit_ok = []
    for dealer in db.query(IIPDealer).all():
        wallet = dealer_wallet_summary(db, dealer.code)
        if wallet.get("remaining_credit", 0) > 0:
            credit_ok.append({"dealer_code": dealer.code, "dealer_name": dealer.name, "province": dealer.province, "remaining_credit": wallet.get("remaining_credit")})
    return {"province_sales_ton": by_province, "priority_dealers_with_credit": sorted(credit_ok, key=lambda x: x["remaining_credit"], reverse=True)[:20], "recommendation": "Ưu tiên tỉnh/đại lý còn hạn mức tín dụng và có lịch sử mua VAS tốt; chạy mô phỏng chiết khấu trước khi giảm giá."}


# -------------------------
# V4 endpoints - Module 4: QR/GPS/logistics
# -------------------------

@router.get("/v4/warehouse/slips/{slip_code}/qr-token")
def warehouse_qr_token_v4(slip_code: str, db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("warehouse"))):
    slip = db.query(IIPWarehouseSlip).filter(IIPWarehouseSlip.slip_code == slip_code).first()
    if not slip:
        raise HTTPException(status_code=404, detail="Không tìm thấy phiếu xuất")
    token = _signed_qr_payload(slip)
    slip.qr_code = token
    db.commit()
    return {"slip_code": slip_code, "qr_token": token, "verify_endpoint": "/iip/v4/warehouse/slips/verify"}


@router.post("/v4/warehouse/slips/verify")
def verify_warehouse_qr_v4(payload: Dict[str, str] = Body(...), db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("warehouse"))):
    result = _verify_qr_payload(payload.get("qr_token", ""))
    if result.get("valid"):
        db.add(IIPDeliveryEvent(delivery_code=result.get("order_code", "UNKNOWN"), event_type="scan", payload_json=json.dumps(result, ensure_ascii=False), severity="green", note="Quét QR phiếu xuất hợp lệ"))
        db.commit()
    return result


@router.post("/v4/deliveries/{delivery_code}/location")
def add_delivery_location_v4(delivery_code: str, payload: DeliveryLocationIn, db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("driver"))):
    delivery = db.query(IIPDelivery).filter(IIPDelivery.delivery_code == delivery_code).first()
    if not delivery:
        raise HTTPException(status_code=404, detail="Không tìm thấy chuyến giao")
    loc = IIPDeliveryLocation(delivery_code=delivery_code, lat=payload.lat, lng=payload.lng, speed_kmh=payload.speed_kmh, accuracy_m=payload.accuracy_m, recorded_at=payload.recorded_at or datetime.utcnow(), note=payload.note)
    delivery.gps_lat = payload.lat; delivery.gps_lng = payload.lng
    db.add(loc); db.commit(); db.refresh(loc)
    return row_to_dict(loc)


@router.get("/v4/deliveries/{delivery_code}/route")
def get_delivery_route_v4(delivery_code: str, db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("report"))):
    rows = db.query(IIPDeliveryLocation).filter(IIPDeliveryLocation.delivery_code == delivery_code).order_by(IIPDeliveryLocation.recorded_at.asc()).all()
    return {"delivery_code": delivery_code, "points": [row_to_dict(x) for x in rows]}


@router.post("/v4/deliveries/{delivery_code}/event")
def add_delivery_event_v4(delivery_code: str, payload: DeliveryEventIn, db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("driver"))):
    obj = IIPDeliveryEvent(delivery_code=delivery_code, event_type=payload.event_type, payload_json=json.dumps(payload.payload, ensure_ascii=False, default=str), severity=payload.severity, note=payload.note)
    db.add(obj); db.commit(); db.refresh(obj)
    return row_to_dict(obj)


@router.post("/v4/vehicles")
def upsert_vehicle_v4(payload: VehicleIn, db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("warehouse"))):
    obj = upsert_by_key(db, IIPVehicle, "vehicle_code", schema_dump(payload))
    db.commit(); db.refresh(obj)
    return row_to_dict(obj)


@router.post("/v4/logistics/route-suggestions")
def route_suggestions_v4(payload: MergeRouteIn, db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("warehouse"))):
    vehicle = db.query(IIPVehicle).filter(IIPVehicle.vehicle_code == payload.vehicle_code).first() if payload.vehicle_code else None
    max_ton = payload.max_ton or (float(vehicle.max_ton or 20) if vehicle else 20.0)
    q = db.query(IIPOrder)
    if payload.order_codes:
        q = q.filter(IIPOrder.order_code.in_(payload.order_codes))
    else:
        q = q.filter(IIPOrder.status.in_(["approved", "pending"]))
    orders = q.all()
    candidates = []
    for o in orders:
        dealer = db.query(IIPDealer).filter(IIPDealer.code == o.dealer_code).first()
        if payload.province and dealer and dealer.province != payload.province:
            continue
        ton = sum(float(i.quantity_ton or 0) for i in db.query(IIPOrderItem).filter(IIPOrderItem.order_code == o.order_code).all())
        candidates.append({"order_code": o.order_code, "dealer_code": o.dealer_code, "province": dealer.province if dealer else None, "ton": ton, "amount": o.total_amount})
    selected = []
    used = 0.0
    for c in sorted(candidates, key=lambda x: x["ton"], reverse=True):
        if used + c["ton"] <= max_ton:
            selected.append(c); used += c["ton"]
    return {"vehicle_code": payload.vehicle_code, "max_ton": max_ton, "selected_ton": used, "selected_orders": selected, "remaining_ton": max_ton - used, "saving_estimate": {"merged_orders": len(selected), "estimated_trip_reduction": max(0, len(selected) - 1), "note": "Ước tính MVP: ghép đơn cùng tỉnh/tuyến và không vượt tải trọng."}}


# -------------------------
# V4 endpoints - Module 5: dealer scoring and self-service portal
# -------------------------

@router.get("/v4/dealers/{dealer_code}/credit-score")
def dealer_credit_score_v4(dealer_code: str, save: bool = True, db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("report"))):
    if not db.query(IIPDealer).filter(IIPDealer.code == dealer_code).first():
        raise HTTPException(status_code=404, detail="Không tìm thấy đại lý")
    factors = dealer_credit_factors(db, dealer_code)
    if save:
        hist = IIPDealerCreditScoreHistory(dealer_code=dealer_code, score=factors["score"], rank=factors["rank"], factors_json=json.dumps(factors, ensure_ascii=False, default=str))
        db.add(hist)
        limit = db.query(IIPCreditLimit).filter(IIPCreditLimit.dealer_code == dealer_code).first()
        if limit:
            limit.rank = factors["rank"]
        dealer = db.query(IIPDealer).filter(IIPDealer.code == dealer_code).first()
        if dealer:
            dealer.rank = factors["rank"]
        db.commit()
    return factors


@router.get("/v4/dealers/{dealer_code}/credit-score-history")
def dealer_credit_history_v4(dealer_code: str, db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("report"))):
    rows = db.query(IIPDealerCreditScoreHistory).filter(IIPDealerCreditScoreHistory.dealer_code == dealer_code).order_by(IIPDealerCreditScoreHistory.calculated_at.desc()).limit(50).all()
    return {"items": [row_to_dict(x) for x in rows]}


@router.get("/v4/dealers/{dealer_code}/purchase-trend")
def dealer_purchase_trend_v4(dealer_code: str, months: int = Query(6, ge=1, le=24), db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("report"))):
    since = date.today() - timedelta(days=months * 31)
    orders = db.query(IIPOrder).filter(IIPOrder.dealer_code == dealer_code, IIPOrder.order_date >= since).all()
    by_month: Dict[str, float] = {}
    for o in orders:
        k = o.order_date.strftime("%Y-%m") if o.order_date else "unknown"
        by_month[k] = by_month.get(k, 0) + float(o.total_amount or 0)
    values = [by_month[k] for k in sorted(by_month)]
    trend = "stable"
    if len(values) >= 2 and values[-1] < values[0] * 0.7:
        trend = "down"
    elif len(values) >= 2 and values[-1] > values[0] * 1.2:
        trend = "up"
    return {"dealer_code": dealer_code, "months": months, "monthly_sales": by_month, "trend": trend}


@router.get("/v4/dealers/{dealer_code}/retention-suggestions")
def dealer_retention_suggestions_v4(dealer_code: str, db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("report"))):
    score = dealer_credit_factors(db, dealer_code)
    trend = dealer_purchase_trend_v4(dealer_code, 6, db, actor)  # type: ignore[arg-type]
    suggestions = []
    if trend["trend"] == "down":
        suggestions.append("Gọi chăm sóc ngay và hỏi lý do giảm mua trong 3-6 tháng gần đây.")
    if score["credit_usage_ratio"] < 0.7 and score["rank"] in ["A", "B"]:
        suggestions.append("Có thể đề xuất thêm hạn mức/ưu đãi sản lượng vì đại lý còn sức mua.")
    if score["avg_overdue_days"] > 15:
        suggestions.append("Không tăng bán chịu; yêu cầu kế hoạch thanh toán trước khi mở đơn mới.")
    if not suggestions:
        suggestions.append("Tiếp tục theo dõi; chưa có dấu hiệu rủi ro lớn.")
    return {"dealer_code": dealer_code, "credit_score": score, "purchase_trend": trend, "suggestions": suggestions}


@router.get("/v4/dealer/me")
def dealer_me_v4(actor: IIPAuthUser = Depends(require_permission("dealer_self"))):
    return {"username": actor.username, "dealer_code": actor.dealer_code, "role": actor.role}


@router.get("/v4/dealer/wallet")
def dealer_wallet_self_v4(db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("dealer_self"))):
    if not actor.dealer_code:
        raise HTTPException(status_code=403, detail="Tài khoản chưa gắn dealer_code")
    return dealer_wallet_summary(db, actor.dealer_code)


@router.get("/v4/dealer/orders")
def dealer_orders_self_v4(db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("dealer_self"))):
    if not actor.dealer_code:
        raise HTTPException(status_code=403, detail="Tài khoản chưa gắn dealer_code")
    rows = db.query(IIPOrder).filter(IIPOrder.dealer_code == actor.dealer_code).order_by(IIPOrder.order_date.desc()).limit(100).all()
    return {"items": [row_to_dict(x) for x in rows]}


@router.post("/v4/dealer/orders")
def dealer_create_order_self_v4(payload: OrderCreateIn, request: Request, db: Session = Depends(get_db), actor: IIPAuthUser = Depends(require_permission("dealer_self"))):
    if not actor.dealer_code:
        raise HTTPException(status_code=403, detail="Tài khoản chưa gắn dealer_code")
    if payload.dealer_code != actor.dealer_code:
        raise HTTPException(status_code=403, detail="Đại lý chỉ được tạo đơn cho chính mình")
    return create_order_v3(payload, request, db, actor)


# -------------------------
# V4 roadmap score and demo seed
# -------------------------

@router.get("/v4/roadmap/completion-score")
def roadmap_completion_score_v4(db: Session = Depends(get_db)):
    base = roadmap_completion_score()
    improvements = {
        "module_1_profit_cashflow_approvals": 93,
        "module_2_vat_xml_legal_reconcile": 90,
        "module_3_vas_multi_tier_simulation": 92,
        "module_4_qr_gps_logistics": 88,
        "module_5_dealer_credit_portal": 92,
        "production_backend": 88,
    }
    return {"phase": "IIP Steel Giai đoạn 1", "backend_logic_score": 93, "deploy_ready_score_without_frontend": 82, "v3_previous": base, "v4_module_scores": improvements, "remaining": ["Frontend dashboard", "mobile driver/dealer UI", "real Zalo/Telegram/SMS integration", "kết nối hóa đơn điện tử/ngân hàng production", "kiểm thử với dữ liệu thật của khách"]}


@router.post("/demo/seed-v4")
def seed_iip_demo_v4(request: Request, db: Session = Depends(get_db)):
    seed_iip_demo_v3(request, db)
    # Cost prices and finance rule
    if not db.query(IIPFinanceCostRule).filter(IIPFinanceCostRule.name == "default_12pct").first():
        db.add(IIPFinanceCostRule(name="default_12pct", annual_rate_pct=12, is_default=1, note="Lãi vốn demo để tính chi phí bán chịu"))
    for product_code, cost in [("VAS-D10", 14500000), ("VAS-D12", 14600000), ("HP-D10", 14300000)]:
        if not db.query(IIPCostPrice).filter(IIPCostPrice.product_code == product_code).first():
            db.add(IIPCostPrice(product_code=product_code, province="ALL", cost_price=cost, supplier="demo"))
    # Bonus program tiers
    if not db.query(IIPSupplierBonusProgram).filter(IIPSupplierBonusProgram.program_code == "VAS-2026").first():
        db.add(IIPSupplierBonusProgram(program_code="VAS-2026", supplier="VAS", year=2026, product_brand="VAS", start_date=date(2026,1,1), end_date=date(2026,12,31), note="Demo nhiều bậc thưởng VAS"))
        db.add(IIPBonusTier(program_code="VAS-2026", tier_name="Bạc", target_ton=30000, bonus_amount=800_000_000))
        db.add(IIPBonusTier(program_code="VAS-2026", tier_name="Vàng", target_ton=40000, bonus_amount=1_500_000_000))
        db.add(IIPBonusTier(program_code="VAS-2026", tier_name="Kim cương", target_ton=50000, bonus_amount=3_000_000_000))
    # Vehicle
    upsert_by_key(db, IIPVehicle, "vehicle_code", {"vehicle_code": "TRUCK-01", "plate_number": "29H-12345", "max_ton": 25, "driver_name": "Tài xế Demo", "driver_phone": "0900000000", "status": "active"})
    # Dealer user demo
    if not db.query(IIPAuthUser).filter(IIPAuthUser.username == "dealer_sonla").first():
        db.add(IIPAuthUser(username="dealer_sonla", full_name="Đại lý Sơn La Demo", role="dealer", dealer_code="DL_SONLA", password_hash=hash_password("dealer123"), is_active=True))
    audit(db, "seed_v4", None, "demo", "seed-v4", request, after={"v4": True})
    db.commit()
    return {"ok": True, "message": "Đã seed demo V4", "try": ["GET /iip/v4/status", "GET /iip/v4/roadmap/completion-score", "GET /iip/v4/profit/orders?refresh=true", "GET /iip/v4/reconcile/4-way/risk-score", "GET /iip/v4/bonus/progress?program_code=VAS-2026&year=2026", "POST /iip/v4/vas/simulate", "GET /iip/v4/dealers/DL_SONLA/credit-score"], "dealer_login": {"username": "dealer_sonla", "password": "dealer123"}}

# V4 role permission aliases (kept after route definitions because dependencies read this at request time)
ROLE_PERMISSIONS.setdefault("chairman", set()).update({"report", "approve", "dealer_self"})
ROLE_PERMISSIONS.setdefault("accounting", set()).update({"report", "import", "approve"})
ROLE_PERMISSIONS.setdefault("sales", set()).update({"report", "create_order", "dealer_self"})
ROLE_PERMISSIONS.setdefault("warehouse", set()).update({"report", "warehouse", "approve"})
ROLE_PERMISSIONS.setdefault("driver", set()).update({"driver", "dealer_self"})
ROLE_PERMISSIONS.setdefault("dealer", set()).update({"dealer_self", "create_order"})
